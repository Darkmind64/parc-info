"""Exports CSV et Excel — vérifie que les deux formats servis en //appareils
et /peripheriques contiennent bien les mêmes données, et qu'un client ne
peut pas se retrouver avec les lignes d'un autre dans son export."""
import io

from openpyxl import load_workbook

from conftest import login_session


def test_export_appareils_xlsx_contient_les_bonnes_lignes(client, make_user, make_client, make_appareil):
    uid, _l, _p = make_user(role='admin')
    cid = make_client(auth_user_id=uid)
    make_appareil(cid, nom_machine='POSTE-EXPORT-XLSX', marque='Dell')

    login_session(client, uid, cid)
    resp = client.get('/appareils/export.xlsx')
    assert resp.status_code == 200
    assert resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == 'nom_machine'
    noms = [row[header.index('nom_machine')].value for row in ws.iter_rows(min_row=2)]
    marques = [row[header.index('marque')].value for row in ws.iter_rows(min_row=2)]
    assert 'POSTE-EXPORT-XLSX' in noms
    assert marques[noms.index('POSTE-EXPORT-XLSX')] == 'Dell'


def test_export_xlsx_isole_par_client(client, make_user, make_client, make_appareil):
    uid, _l, _p = make_user(role='admin')
    cid_a = make_client(auth_user_id=uid)
    cid_b = make_client(auth_user_id=uid)
    make_appareil(cid_a, nom_machine='VISIBLE-CLIENT-A')
    make_appareil(cid_b, nom_machine='INVISIBLE-DEPUIS-A')

    login_session(client, uid, cid_a)
    resp = client.get('/appareils/export.xlsx')
    wb = load_workbook(io.BytesIO(resp.data))
    noms = [row[0].value for row in wb.active.iter_rows(min_row=2)]
    assert 'VISIBLE-CLIENT-A' in noms
    assert 'INVISIBLE-DEPUIS-A' not in noms


def test_export_peripheriques_xlsx(client, make_user, make_client, conn):
    uid, _l, _p = make_user(role='admin')
    cid = make_client(auth_user_id=uid)
    conn.execute(
        "INSERT INTO peripheriques (client_id, categorie, marque, modele) VALUES (?,?,?,?)",
        (cid, 'Ecran', 'Dell', 'U2415'))
    conn.commit()

    login_session(client, uid, cid)
    resp = client.get('/peripheriques/export.xlsx')
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    header = [c.value for c in wb.active[1]]
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert ('Ecran', 'Dell', 'U2415') == rows[0][:3]
    assert header[:3] == ['categorie', 'marque', 'modele']


def test_export_csv_et_xlsx_ont_les_memes_donnees(client, make_user, make_client, make_appareil):
    uid, _l, _p = make_user(role='admin')
    cid = make_client(auth_user_id=uid)
    make_appareil(cid, nom_machine='POSTE-COHERENCE', numero_serie='SN-123')

    login_session(client, uid, cid)
    csv_text = client.get('/appareils/export.csv').get_data(as_text=True)
    wb = load_workbook(io.BytesIO(client.get('/appareils/export.xlsx').data))
    xlsx_noms = [row[0].value for row in wb.active.iter_rows(min_row=2)]

    assert 'POSTE-COHERENCE' in csv_text
    assert 'SN-123' in csv_text
    assert 'POSTE-COHERENCE' in xlsx_noms
