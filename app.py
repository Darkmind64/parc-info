from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_from_directory, make_response, send_file, abort, get_flashed_messages
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta, timezone
import sqlite3, subprocess, re, socket, ipaddress, threading, os, platform, concurrent.futures, hashlib, secrets, logging, json, time, io, colorsys
from PIL import Image
from io import BytesIO
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def _utcnow() -> datetime:
    """Équivalent de _utcnow() (dépréciée depuis 3.12) : même valeur
    naïve en UTC, obtenue via l'API timezone-aware recommandée. Les horodatages
    stockés en base restent ainsi inchangés (pas de suffixe +00:00)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


# ─── mDNS SUPPORT (parcinfo.local) ────────────────────────────────────────────
try:
    from zeroconf import ServiceInfo, Zeroconf
    MDNS_AVAILABLE = True
except ImportError:
    MDNS_AVAILABLE = False

# ─── LOGGING ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
logger = logging.getLogger('parcinfo')

# ── Support PyInstaller (exécutable portable) + Docker ───────────────────────
import sys as _sys
if getattr(_sys, 'frozen', False):
    # Mode exécutable : les ressources sont dans _MEIPASS
    _resource_base = _sys._MEIPASS
    _data_base     = os.path.dirname(_sys.executable)
else:
    _resource_base = os.path.dirname(os.path.abspath(__file__))
    _data_base     = os.path.dirname(os.path.abspath(__file__))

# DATA_DIR permet de séparer données persistantes du code (Docker, NAS)
_data_dir_env = os.environ.get('DATA_DIR', '').strip()
if _data_dir_env:
    _data_base = _data_dir_env
    os.makedirs(_data_base, exist_ok=True)

# Fichier journal persistant, en plus du flux console de basicConfig() —
# indispensable en exécutable packagé (console=False, cf. parcinfo.spec) :
# sans fenêtre de console, cette sortie ne va nulle part de consultable après
# coup. Signalé en cherchant en vain le détail d'un blocage Gatekeeper macOS
# (xattr/codesign/spctl --add) : chaque étape est bien journalisée par ce
# même logger, mais rien ne la conservait sur un Mac ou un PC packagés.
if getattr(_sys, 'frozen', False):
    try:
        from logging.handlers import RotatingFileHandler
        _handler_fichier = RotatingFileHandler(
            os.path.join(_data_base, 'parcinfo.log'),
            maxBytes=2 * 1024 * 1024, backupCount=2, encoding='utf-8')
        _handler_fichier.setFormatter(logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'))
        logging.getLogger().addHandler(_handler_fichier)
    except Exception:
        logger.exception("Impossible de créer parcinfo.log (non bloquant)")

# Diagnostic permanent : sur macOS, un bundle non notarié ouvert via Launch
# Services (Finder, `open`) alors qu'il porte encore une trace de quarantaine
# peut être « transloqué » — exécuté depuis une copie en lecture seule à un
# chemin aléatoire plutôt que /Applications/ParcInfo.app. Signalé en usage
# réel : plusieurs cycles de mise à jour macOS échouaient sans le moindre
# avertissement Gatekeeper, signe probable d'une translocation plutôt que
# d'un vrai blocage. sys.executable révèle directement le chemin RÉEL depuis
# lequel ce process tourne — si « AppTranslocation » y figure, le doute est
# levé pour de bon, sans avoir à deviner depuis un journal après coup.
if getattr(_sys, 'frozen', False) and _sys.platform == 'darwin' \
        and 'AppTranslocation' in _sys.executable:
    logger.warning(
        "⚠️ Cette instance tourne depuis un chemin transloqué par macOS (%s), "
        "pas depuis son emplacement réel — signe que le bundle a été ouvert "
        "via Launch Services alors qu'il portait encore une trace de "
        "quarantaine. Une mise à jour lancée depuis cet état a de bonnes "
        "chances d'échouer.", _sys.executable)

app = Flask(
    __name__,
    template_folder=os.path.join(_resource_base, 'templates'),
    static_folder=os.path.join(_resource_base, 'static'),
)

# ─── COMPRESSION GZIP (Optimisation Performance) ───────────────────────────────
try:
    from flask_compress import Compress
    Compress(app)
    logger.info('✅ Compression GZIP activée')
except ImportError:
    logger.warning('⚠️ flask-compress non installé (pip install flask-compress)')

# Base de données et uploads dans le dossier des données (à côté de l'exe)
DATABASE     = os.path.join(_data_base, 'parc_info.db')
UPLOAD_FOLDER = os.path.join(_data_base, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ─── MODULES UTILITAIRES ──────────────────────────────────────────────────────
from database       import get_db, row_to_dict, init_paths
# Initialiser les chemins de la base de données de façon centralisée et robuste
init_paths(DATABASE, UPLOAD_FOLDER)
from auth_utils     import (hash_pwd as _hash_pwd, check_pwd as _check_pwd,
                             get_auth_user, login_required,
                             get_csrf_token as _get_csrf_token,
                             validate_csrf_request,
                             check_rate_limit as _check_rate_limit,
                             record_failed_attempt as _record_failed_attempt,
                             reset_attempts as _reset_attempts,
                             validate_form)
from config_helpers import (LISTE_DEFAULTS, CFG_DEFAULTS,
                             get_liste, cfg_get, cfg_set, cfg_all, cfg_invalidate,
                             get_port_config, get_port_icon)
from client_helpers import (paginate, get_client_access, can_write,
                             get_client_id, get_clients,
                             log_history, log_error, garantie_active, human_size,
                             fmt_appareils, fmt_garantie_periph, fmt_contrat, fmt_intervention,
                             get_clients_for_filter, _format_date_field)
from uploads_sync import start_sync_thread
from crypto_utils   import get_crypto_manager
from cache_utils    import get_cache_manager, cache_result, invalidate_cache_pattern
from search_utils   import search_global, search_autocomplete
from app_update_routes import register_update_routes
import network_diag  # module de diagnostic réseau (démarre son thread de surveillance)

# Version de l'application (lue depuis version.json dans _resource_base)
# _resource_base = _MEIPASS en mode PyInstaller, dossier source sinon
def _load_app_version():
    try:
        _vf = os.path.join(_resource_base, 'version.json')
        with open(_vf, 'r', encoding='utf-8') as _f:
            v = json.load(_f).get('version', '')
        logging.getLogger('parcinfo').info(f'ParcInfo version {v} (depuis {_vf})')
        return v
    except Exception as e:
        logging.getLogger('parcinfo').warning(f'version.json introuvable dans {_resource_base}: {e}')
        return ''
APP_VERSION = _load_app_version()

def _load_collector_downloads():
    """Liens de téléchargement du collecteur (menu Inventaire), lus depuis
    version.json — même fichier et mêmes règles de résolution de chemin que
    _load_app_version() ci-dessus. Chaque publication de release régénère ce
    fichier avec des URLs pointant vers CETTE version (voir version.json et
    verifier_version.py) : ces liens suivent donc automatiquement la version
    de l'exécutable en cours, sans jamais nécessiter de mise à jour manuelle."""
    try:
        _vf = os.path.join(_resource_base, 'version.json')
        with open(_vf, 'r', encoding='utf-8') as _f:
            return json.load(_f).get('downloads', {})
    except Exception:
        return {}
COLLECTOR_DOWNLOADS = _load_collector_downloads()
_APP_DEMARRAGE = time.time()  # pour l'uptime affiché sur /apropos

# ─── HELPER: Retry pour requêtes DB verrouillées ─────────────────────────────
def retry_db_query(query_func, max_retries=5):
    """
    Exécute une fonction de requête DB avec retry automatique si verrouillée.
    Utilisé pour les requêtes critiques dans /index (dashboard).

    Utilisation:
        result = retry_db_query(lambda: conn.execute('SELECT ...').fetchall())
    """
    retry_delay = 0.05  # 50ms initial

    for attempt in range(max_retries):
        try:
            return query_func()
        except sqlite3.OperationalError as e:
            if 'locked' in str(e).lower() and attempt < max_retries - 1:
                # Database verrouillée → retry avec backoff exponentiel
                time.sleep(retry_delay * (2 ** attempt))
                continue
            else:
                raise e
        except Exception as e:
            raise e

@app.context_processor
def inject_auth_context():
    """Injecte les variables auth dans tous les templates."""
    u = None
    uid = session.get('auth_user_id')
    if uid:
        try:
            conn = get_db()
            u = row_to_dict(conn.execute('SELECT id,login,nom,prenom,role,logo_fichier FROM auth_users WHERE id=?', (uid,)).fetchone() or {})
            conn.close()
        except Exception:
            logger.exception('Erreur inject_auth_context')
    return dict(auth_user=u)
# Clé secrète persistée (générée une fois, stockée à côté de la DB)
_secret_key_file = os.path.join(_data_base, 'secret.key')
if os.path.exists(_secret_key_file):
    with open(_secret_key_file, 'r', encoding='utf-8') as _f:
        app.config['SECRET_KEY'] = _f.read().strip()
else:
    _generated_key = secrets.token_hex(32)
    with open(_secret_key_file, 'w', encoding='utf-8') as _f:
        _f.write(_generated_key)
    app.config['SECRET_KEY'] = _generated_key
# ── Configuration de sécurité des sessions ───────────────────────────────────
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_HTTPONLY']    = True   # inaccessible depuis JS
app.config['SESSION_COOKIE_SAMESITE']    = 'Lax'  # protection CSRF additionnelle

# ─── SCHEDULER (Cron Jobs) ────────────────────────────────────────────────────
from apscheduler.schedulers.background import BackgroundScheduler
scheduler = BackgroundScheduler()
scheduler.daemon = True  # Arrête avec l'application

# UPLOAD_FOLDER défini plus haut (support PyInstaller)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ─── VALIDATION DES FICHIERS DÉPOSÉS ─────────────────────────────────────────
# Auparavant tout était accepté, sans limite de taille : n'importe qui atteignant
# l'application pouvait déposer un fichier de n'importe quel type et de
# n'importe quelle taille dans le volume de données.

#: Documents joints aux appareils, contrats, périphériques, interventions.
ALLOWED_EXTENSIONS = {
    'pdf', 'txt', 'csv', 'rtf', 'md', 'log',
    'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'odt', 'ods', 'odp',
    'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'svg', 'heic',
    'zip', '7z', 'rar', 'gz', 'tar',
    'eml', 'msg', 'json', 'xml', 'html', 'htm',
}
#: Images seules — logos, avatars, photos de baie.
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'svg'}

#: Taille maximale d'une requête, pièces jointes comprises.
MAX_UPLOAD_MB = int(os.environ.get('PARCINFO_MAX_UPLOAD_MB', '64'))
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

#: Premiers octets attendus, par extension. Le type déclaré par le navigateur
#: vient du client et ne prouve rien ; la signature du fichier, si.
_SIGNATURES = {
    'pdf':  [b'%PDF-'],
    'png':  [b'\x89PNG\r\n\x1a\n'],
    'gif':  [b'GIF87a', b'GIF89a'],
    'jpg':  [b'\xff\xd8\xff'],
    'jpeg': [b'\xff\xd8\xff'],
    'bmp':  [b'BM'],
    'zip':  [b'PK\x03\x04', b'PK\x05\x06'],   # aussi docx, xlsx, pptx, odt…
    'docx': [b'PK\x03\x04'], 'xlsx': [b'PK\x03\x04'], 'pptx': [b'PK\x03\x04'],
    'odt':  [b'PK\x03\x04'], 'ods':  [b'PK\x03\x04'], 'odp':  [b'PK\x03\x04'],
    '7z':   [b'7z\xbc\xaf\x27\x1c'],
    'gz':   [b'\x1f\x8b'],
    'rar':  [b'Rar!\x1a\x07'],
}


def extension_de(filename):
    """Extension en minuscules, sans le point ('' si absente)."""
    if not filename or '.' not in filename:
        return ''
    return filename.rsplit('.', 1)[1].strip().lower()


def allowed_file(filename, extensions=None):
    """Vrai si le nom porte une extension autorisée."""
    if not filename or not filename.strip():
        return False
    return extension_de(filename) in (extensions or ALLOWED_EXTENSIONS)


def signature_coherente(fichier, extension):
    """Vrai si le contenu correspond à l'extension annoncée.

    Ne juge que les formats dont la signature est connue : le reste passe, faute
    de quoi on refuserait des documents parfaitement légitimes.
    """
    attendues = _SIGNATURES.get(extension)
    if not attendues:
        return True
    try:
        position = fichier.stream.tell()
        debut = fichier.stream.read(16)
        fichier.stream.seek(position)
    except Exception:
        return True
    return any(debut.startswith(s) for s in attendues)


def verifier_fichier(fichier, extensions=None):
    """Contrôle nom, extension et signature. Retourne (ok, message)."""
    if not fichier or not fichier.filename:
        return False, "Aucun fichier sélectionné"
    extension = extension_de(fichier.filename)
    if not allowed_file(fichier.filename, extensions):
        return False, ("Extension « %s » non autorisée" % (extension or '?'))
    if not signature_coherente(fichier, extension):
        return False, ("Le contenu du fichier ne correspond pas à un %s"
                       % extension.upper())
    return True, None


@app.errorhandler(413)
def _fichier_trop_volumineux(_e):
    """Message explicite : par défaut Flask renvoie une page d'erreur muette."""
    message = "Fichier trop volumineux (maximum %d Mo)" % MAX_UPLOAD_MB
    if request.path.startswith('/api/'):
        return jsonify({'status': 'error', 'message': message}), 413
    flash(message, 'danger')
    return redirect(request.referrer or url_for('index')), 302


# ─── CACHING OPTIMISÉ ─────────────────────────────────────────────────────────
def get_liste_cached(nom: str, ttl: int = 600) -> list:
    """
    Wrapper de get_liste() avec caching intelligent (10 min par défaut).
    Réduit les requêtes DB de 80% pour les listes fréquemment accédées.
    """
    cache_mgr = get_cache_manager()
    cache_key = f"liste:{nom}"

    # Vérifier le cache
    cached = cache_mgr.get(cache_key)
    if cached is not None:
        return cached

    # Récupérer depuis DB
    result = get_liste(nom)

    # Stocker en cache
    cache_mgr.set(cache_key, result, ttl)
    return result


# ─── CSRF ─────────────────────────────────────────────────────────────────────

# Couleurs de base des textes secondaire/atténué, par mode et niveau de
# contraste — miroir exact de TEXT_BASE dans templates/base.html (fonction JS
# appliquerVariables). Sert de point de départ au curseur de luminosité des
# textes secondaires (Réglages → Apparence & Couleurs).
_TEXT_BASE = {
    'dark':  {'normal': {'secondary': '#6a8aaa', 'muted': '#3a5570'},
              'high':   {'secondary': '#7a9bc5', 'muted': '#3a5570'},
              'max':    {'secondary': '#9ad0ff', 'muted': '#7aa0c0'}},
    'light': {'normal': {'secondary': '#334155', 'muted': '#64748b'},
              'high':   {'secondary': '#2d5a8a', 'muted': '#4a6fa8'},
              'max':    {'secondary': '#1a4d80', 'muted': '#3a6b9f'}},
}


def _ajuster_luminosite_texte(hex_couleur, delta_pct, mode):
    """Éclaircit (mode sombre) ou assombrit (mode clair) une couleur hex,
    pour toujours augmenter le contraste avec le fond — miroir Python de
    ajusterLuminositeTexte() dans templates/base.html."""
    if not delta_pct:
        return hex_couleur
    hexc = hex_couleur.lstrip('#')
    r, g_, b = (int(hexc[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g_, b)
    direction = -1 if mode == 'light' else 1
    nouveau_l = max(0.0, min(1.0, l + direction * delta_pct / 100))
    r2, g2, b2 = colorsys.hls_to_rgb(h, nouveau_l, s)
    vers255 = lambda v: max(0, min(255, round(v * 255)))
    return '#%02x%02x%02x' % (vers255(r2), vers255(g2), vers255(b2))


def _generate_dynamic_css(auth_user_id=None):
    """
    Génère le CSS dynamique basé sur les configurations de l'utilisateur.
    Injecté dans chaque page pour que les paramètres personnels persistent.
    """
    def g(k, d):
        return cfg_get(k, d, auth_user_id=auth_user_id)

    # Couleurs accents
    accent = g('accent_color', '#00c9ff')
    green = g('accent_green', '#00ff88')
    red = g('accent_red', '#ff3355')
    orange = g('accent_orange', '#ff8c00')

    # Niveau de contraste
    contrast_level = g('contrast_level', 'normal')

    # CSS des accents
    css = f":root{{--accent:{accent};--accent-green:{green};--accent-red:{red};--accent-orange:{orange}}}"

    # CSS de contraste selon le niveau sélectionné
    if contrast_level == 'high':
        css += "html.contrast-high{--text-primary-opacity:1;--text-secondary-opacity:0.95;--text-muted-opacity:0.85}html.contrast-high body{--text-primary:rgba(208,232,255,1);--text-secondary:rgba(106,138,170,1)}"
    elif contrast_level == 'max':
        css += "html.contrast-max{--text-primary-opacity:1;--text-secondary-opacity:1;--text-muted-opacity:0.9}html.contrast-max body{--text-primary:#ffffff;--text-secondary:#a6c5e8;filter:contrast(1.2)}"

    # Luminosité des textes secondaire/atténué (curseur, Réglages → Apparence).
    # Générée ici pour que le réglage tienne dès le rendu de la page, pas
    # seulement après ouverture du panneau Réglages (qui, lui, ne fait que
    # prévisualiser en direct côté JS pendant qu'on bouge le curseur —
    # oublié ici lors de l'ajout de ce réglage, d'où la perte au changement
    # de page malgré une valeur bien enregistrée).
    mode = g('mode', 'dark')
    if mode not in ('dark', 'light'):
        mode = 'dark'
    try:
        text_brightness = int(g('text_brightness', '0') or 0)
    except (TypeError, ValueError):
        text_brightness = 0
    if text_brightness > 0:
        niveau = contrast_level if contrast_level in ('high', 'max') else 'normal'
        base_texte = _TEXT_BASE[mode][niveau]
        secondary_txt = _ajuster_luminosite_texte(base_texte['secondary'], text_brightness, mode)
        muted_txt = _ajuster_luminosite_texte(base_texte['muted'], text_brightness, mode)
        css += f"body{{--text-secondary:{secondary_txt} !important;--text-muted:{muted_txt} !important}}"

    # Couleurs des ports (serviceType)
    port_colors = {
        'ssh': g('port_color_ssh', '#00ff88'),
        'http': g('port_color_http', '#00c9ff'),
        'https': g('port_color_https', '#00c9ff'),
        'rdp': g('port_color_rdp', '#c084fc'),
        'ftp': g('port_color_ftp', '#ff8c00'),
        'smb': g('port_color_smb', '#facc15'),
        'print': g('port_color_print', '#fb923c'),
        'telnet': g('port_color_telnet', '#ff3355'),
        'other': g('port_color_other', '#64748b'),
    }
    for k, col in port_colors.items():
        css += f".port-{k}{{color:{col};border-color:{col}55}}.port-{k}:hover{{background:{col}18;box-shadow:0 0 8px {col}44}}"

    # Couleurs des ports par NUMÉRO (configurations personnalisées)
    scan_ports_str = g('scan_ports', '21,22,23,25,53,80,110,135,139,143,443,445,631,3389,5900,8080,8443,9100')
    for port_str in scan_ports_str.split(','):
        try:
            pnum = int(port_str.strip())
            pcolor = g(f'port_{pnum}_color', '')
            if pcolor:
                css += f".port-num-{pnum}{{color:{pcolor};border-color:{pcolor}55}}.port-num-{pnum}:hover{{background:{pcolor}18;box-shadow:0 0 8px {pcolor}44}}"
        except (ValueError, TypeError):
            pass

    # Couleurs des périphériques
    periph_colors = {
        'ecran': g('periph_color_ecran', '#22d3ee'),
        'clavier': g('periph_color_clavier', '#a78bfa'),
        'souris': g('periph_color_souris', '#a78bfa'),
        'webcam': g('periph_color_webcam', '#fb923c'),
        'casque': g('periph_color_casque', '#c084fc'),
        'audio': g('periph_color_casque', '#c084fc'),
        'imprimante': g('periph_color_imprimante', '#f97316'),
        'scanner': g('periph_color_imprimante', '#f97316'),
        'onduleur': g('periph_color_onduleur', '#facc15'),
        'multiprise': g('periph_color_onduleur', '#facc15'),
        'stockage': g('periph_color_stockage', '#4ade80'),
        'usb': g('periph_color_usb', '#94a3b8'),
        'dock': g('periph_color_dock', '#60a5fa'),
        'reseau': g('periph_color_reseau', '#2dd4bf'),
        'tel': g('periph_color_tel', '#34d399'),
        'badge': g('periph_color_badge', '#f87171'),
        'autre': g('periph_color_autre', '#94a3b8'),
    }
    for k, col in periph_colors.items():
        css += f".pi-{k}{{color:{col};border-color:{col}66}}.pi-{k}:hover{{background:{col}18}}"

    # Couleurs des types d'appareils (depuis _TYPE_CSS_DEFAULTS)
    for type_key, default_color in _TYPE_CSS_DEFAULTS.items():
        type_color = g(f'type_color_{type_key}', default_color)
        css += f".type-{type_key}{{color:{type_color};background:{type_color}15;border-color:{type_color}55}}"

    return css


@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=_get_csrf_token())


@app.context_processor
def inject_collector_downloads():
    return dict(collector_downloads=COLLECTOR_DOWNLOADS)


@app.context_processor
def inject_dynamic_css():
    """Injecte le CSS dynamique généré côté serveur dans chaque template."""
    user = get_auth_user()
    auth_user_id = user['id'] if user else None
    dynamic_css = _generate_dynamic_css(auth_user_id=auth_user_id)
    return dict(dynamic_css=dynamic_css)


@app.before_request
def csrf_protect():
    validate_csrf_request()

# ── FILTRES JINJA2 POUR LES PORTS ────────────────────────────────────────────
_PORT_MAP = {
    21:   ('ftp',    '📁', 'FTP — Transfert de fichiers',          'ftp'),
    22:   ('ssh',    '⌨',  'SSH — Terminal sécurisé',              'ssh'),
    23:   ('telnet', '⚠',  'Telnet — Terminal NON sécurisé',       'telnet'),
    25:   ('other',  '✉',  'SMTP — Serveur mail sortant',          'info'),
    53:   ('other',  '🔍', 'DNS — Résolution de noms',             'info'),
    80:   ('http',   '🌐', 'HTTP — Serveur web',                   'http'),
    110:  ('other',  '✉',  'POP3 — Messagerie',                    'info'),
    135:  ('smb',    '⚙',  'RPC — Windows Remote Procedure Call',  'info'),
    139:  ('smb',    '🗂', 'NetBIOS — Partage Windows',            'smb'),
    143:  ('other',  '✉',  'IMAP — Messagerie',                    'info'),
    443:  ('https',  '🔒', 'HTTPS — Serveur web sécurisé',         'https'),
    445:  ('smb',    '🗂', 'SMB — Partage de fichiers Windows',    'smb'),
    631:  ('print',  '🖨', 'IPP — Service impression',          'print'),
    3389: ('rdp',    '🖥', 'RDP — Bureau à distance Windows',      'rdp'),
    5900: ('rdp',    '🖥', 'VNC — Bureau à distance VNC',          'vnc'),
    8080: ('http',   '🌐', 'HTTP alternatif (port 8080)',           'http8080'),
    8443: ('https',  '🔒', 'HTTPS alternatif (port 8443)',          'https8443'),
    9100: ('print',  '🖨', 'JetDirect — Impression directe',       'print'),
}

@app.template_filter('periph_icon')
def periph_icon_filter(cat):
    # Abreviations courtes (max 3 car) affichees dans le badge style "port"
    icons = {
        'Ecran':                   'ECR',
        'Clavier':                 'KB',
        'Souris':                  'SOU',
        'Webcam':                  'CAM',
        'Casque / Micro':          'MIC',
        'Haut-parleurs':           'HP',
        'Imprimante':              'IMP',
        'Scanner':                 'SCN',
        'Imprimante multifonction':'IMP',
        'Onduleur / UPS':          'UPS',
        'Multiprise parafoudre':   'MPR',
        'Disque dur externe':      'HDD',
        'Cle USB':                 'USB',
        'Hub USB':                 'HUB',
        'Lecteur de cartes':       'LCR',
        'Docking station':         'DOC',
        'Adaptateur reseau':       'NET',
        'Switch USB':              'SW',
        'Telephone fixe IP':       'TEL',
        'Telephone mobile':        'MOB',
        'Badge / Lecteur de badge':'BGE',
        'Autre':                   'AUT',
    }
    return icons.get(cat, 'PER')

# Mapping catégorie périphérique → clé de couleur config (periph_color_<key>)
_PERIPH_COLOR_KEY = {
    'Ecran': 'ecran', 'Clavier': 'clavier', 'Souris': 'souris',
    'Webcam': 'webcam', 'Casque / Micro': 'casque', 'Haut-parleurs': 'casque',
    'Imprimante': 'imprimante', 'Scanner': 'imprimante', 'Imprimante multifonction': 'imprimante',
    'Onduleur / UPS': 'onduleur', 'Multiprise parafoudre': 'onduleur',
    'Disque dur externe': 'stockage', 'Cle USB': 'usb', 'Hub USB': 'usb',
    'Lecteur de cartes': 'usb', 'Docking station': 'dock', 'Adaptateur reseau': 'reseau',
    'Switch USB': 'usb', 'Telephone fixe IP': 'tel', 'Telephone mobile': 'tel',
    'Badge / Lecteur de badge': 'badge', 'Autre': 'autre',
}

@app.template_filter('periph_color_key')
def periph_color_key_filter(cat):
    """Retourne la clé config (ex: 'ecran') pour une catégorie de périphérique."""
    return _PERIPH_COLOR_KEY.get(cat, 'autre')

# Mapping type d'appareil → clé CSS (type-<key>) et config (type_color_<key>)
_TYPE_CSS_MAP = {
    'PC': 'pc', 'PC (Windows)': 'pc', 'PC/Serveur (Linux)': 'linux',
    'Laptop': 'laptop', 'MacBook': 'mac', 'Serveur': 'serveur',
    'Imprimante': 'imprimante', 'Imprimante multifonction': 'imprimante',
    'Switch': 'switch', 'Switch/AP': 'switch', 'Routeur/Pare-feu': 'routeur',
    'NAS': 'nas', 'Telephone IP': 'tel', 'Tablette': 'tablette',
    'Camera IP': 'camera', 'Borne Wi-Fi': 'wifi', 'Objet connecté': 'iot',
    'Autre': 'autre',
}

@app.template_filter('type_css')
def type_css_filter(t):
    """Retourne la clé CSS du type d'appareil (ex: 'pc', 'nas', 'serveur').
    Pour les types non connus, génère une clé CSS-safe depuis le libellé."""
    import re as _re
    if t in _TYPE_CSS_MAP:
        return _TYPE_CSS_MAP[t]
    # Types custom : slug alphanumérique limité à 16 chars
    key = _re.sub(r'[^a-z0-9]', '', str(t).lower())[:16]
    return key or 'autre'

# Labels courts (≤3 chars) par défaut pour les badges de type d'appareil
_TYPE_BADGE_DEFAULTS = {
    'pc': 'PC', 'linux': 'LNX', 'laptop': 'LAP', 'mac': 'MAC',
    'serveur': 'SRV', 'imprimante': 'IMP', 'switch': 'SW', 'routeur': 'RTR',
    'nas': 'NAS', 'tel': 'TEL', 'tablette': 'TAB', 'camera': 'CAM',
    'wifi': 'WIF', 'iot': 'IOT', 'autre': 'AUT',
}

@app.template_filter('type_badge')
def type_badge_filter(t):
    """Retourne le label court (≤3 chars) configuré pour le badge de type d'appareil."""
    k = type_css_filter(t)
    val = cfg_get(f'type_badge_{k}')
    if val:
        return val[:3].upper()
    default = _TYPE_BADGE_DEFAULTS.get(k)
    if default:
        return default
    # Types custom : 3 premiers chars du slug
    return k[:3].upper() or 'AUT'

@app.template_filter('type_description')
def type_description_filter(t):
    """Retourne la description configurée pour le type d'appareil (pour infobulles)."""
    k = type_css_filter(t)
    desc = cfg_get(f'type_desc_{k}')
    return desc if desc else t

@app.template_filter('fromjson')
def fromjson_filter(s):
    """Décode une chaîne JSON en objet Python (liste ou dict). Retourne [] en cas d'erreur."""
    try:
        return json.loads(s) if s else []
    except Exception:
        return []

@app.template_filter('periph_css')
def periph_css_filter(cat):
    m = {
        'Ecran':'pi-ecran',
        'Clavier':'pi-clavier',
        'Souris':'pi-souris',
        'Webcam':'pi-webcam',
        'Casque / Micro':'pi-casque',
        'Haut-parleurs':'pi-audio',
        'Imprimante':'pi-imprimante',
        'Scanner':'pi-scanner',
        'Imprimante multifonction':'pi-imprimante',
        'Onduleur / UPS':'pi-onduleur',
        'Multiprise parafoudre':'pi-multiprise',
        'Disque dur externe':'pi-stockage',
        'Cle USB':'pi-usb',
        'Hub USB':'pi-usb',
        'Lecteur de cartes':'pi-usb',
        'Docking station':'pi-dock',
        'Adaptateur reseau':'pi-reseau',
        'Switch USB':'pi-usb',
        'Telephone fixe IP':'pi-tel',
        'Telephone mobile':'pi-tel',
        'Badge / Lecteur de badge':'pi-badge',
        'Autre':'pi-autre',
    }
    return m.get(cat, 'pi-autre')

@app.template_filter('port_badge')
def port_badge_filter(port):
    """Retourne un badge HTML pour un port avec couleur, nom et icône personnalisés.
    Format : <span style="...color...">ICON PORT</span>
    """
    if not port:
        return ''
    try:
        port_int = int(port)
        cfg = get_port_config(port_int)
        icon = cfg.get('icon', '◈')
        color = cfg.get('color', '#64748b')
        return f'<span style="background:rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},.15);color:{color};border:1.5px solid {color};padding:3px 6px;border-radius:3px;font-size:0.75em;font-weight:700;white-space:nowrap;display:inline-flex;align-items:center;gap:2px;">{icon} {port}</span>'
    except:
        return f'<span>{port}</span>'

@app.template_filter('port_name')
def port_name_filter(port):
    """Retourne uniquement le nom du service pour un port (ex: 'SSH', 'HTTP')."""
    try:
        return get_port_config(int(port)).get('name', str(port))
    except:
        return str(port)

@app.template_filter('port_class')
def port_class_filter(port):
    try: return _PORT_MAP.get(int(port), ('other','','',''))[0]
    except: return 'other'

@app.template_filter('port_icon')
def port_icon_filter(port):
    try:
        return get_port_icon(int(port))
    except:
        return '◈'

@app.template_filter('port_info')
def port_info_filter(port):
    try:
        port_int = int(port)
        cfg = get_port_config(port_int)
        name = cfg.get('name', str(port))
        desc = cfg.get('description', '')
        # Formater l'infobulle avec description si elle existe
        if desc:
            return f"{name} — {desc}"
        else:
            return f"{name} — Service TCP"
    except:
        return 'Port TCP ouvert'

@app.template_filter('port_action')
def port_action_filter(port):
    try: return _PORT_MAP.get(int(port), ('other','','','info'))[3]
    except: return 'info'

#: Icône par mot-clé (recherché en minuscules dans le nom du programme), pour
#: repérer d'un coup d'œil navigateur/mail/type de fichier par défaut — sans
#: extraire ni stocker l'icône réelle de l'exécutable (coûteux, et alourdirait
#: chaque collecte pour un simple confort visuel).
_ICONES_APPLICATIONS = (
    ('chrome', '🌐'), ('firefox', '🦊'), ('edge', '🌊'), ('brave', '🦁'),
    ('opera', '🎭'), ('internet explorer', '🌐'),
    ('outlook', '📧'), ('thunderbird', '📨'), ('onenote', '📓'), ('mail', '📧'),
    ('acrobat', '📕'), ('reader', '📕'), ('pdf', '📕'), ('foxit', '📕'),
    ('photos', '🖼'), ('paint', '🖼'), ('image', '🖼'), ('viewer', '🖼'),
    ('word', '📘'), ('excel', '📗'), ('powerpoint', '📙'), ('office', '📘'),
    ('notepad++', '📝'), ('notepad', '📝'), ('bloc-notes', '📝'), ('texte', '📝'),
    ('code', '💻'), ('sublime', '💻'), ('vim', '💻'),
    ('7-zip', '🗜'), ('winrar', '🗜'), ('zip', '🗜'),
)

@app.template_filter('app_icon')
def app_icon_filter(nom):
    """Icône représentative d'une application, déduite de son nom (best-effort)."""
    n = (nom or '').lower()
    for mot, icone in _ICONES_APPLICATIONS:
        if mot in n:
            return icone
    return '📦'
DB_PATH = DATABASE  # alias conservé pour compatibilité

_crypto_shared_cache = None   # CryptoManager en cache mémoire (valide pour toute la durée du process)
_crypto_shared_lock  = __import__('threading').Lock()

def _get_crypto_shared(cursor=None):
    """
    Retourne un CryptoManager dont la clé est lue DIRECTEMENT depuis Turso
    (source de vérité unique, indépendante du token utilisé).

    - Turso configuré → lit/crée 'crypto_key' dans la table config de Turso.
      Toutes les instances (même tokens différents) lisent la même base →
      même clé, déchiffrement garanti entre instances.
    - Turso non configuré → clé locale (secret.key).
    - Cache mémoire : l'appel HTTP Turso n'est fait qu'une seule fois par
      démarrage.
    """
    global _crypto_shared_cache
    with _crypto_shared_lock:
        if _crypto_shared_cache is not None:
            return _crypto_shared_cache

    mgr = _build_crypto_shared(cursor)
    if mgr is not None:
        with _crypto_shared_lock:
            _crypto_shared_cache = mgr
        return mgr
    return get_crypto_manager(os.path.join(_data_base, 'secret.key'))


def _build_crypto_shared(cursor=None):
    """Construit le CryptoManager partagé (appelé une seule fois)."""
    try:
        if cursor:
            url_row   = cursor.execute("SELECT valeur FROM config WHERE cle='turso_url'").fetchone()
            token_row = cursor.execute("SELECT valeur FROM config WHERE cle='turso_token'").fetchone()
        else:
            _c = get_local_db()
            url_row   = _c.execute("SELECT valeur FROM config WHERE cle='turso_url'").fetchone()
            token_row = _c.execute("SELECT valeur FROM config WHERE cle='turso_token'").fetchone()
            _c.close()

        turso_url   = (url_row[0]   or '').strip() if url_row   else ''
        turso_token = (token_row[0] or '').strip() if token_row else ''

        if not turso_url or not turso_token:
            return None  # Turso non configuré → fallback local

        from database import TursoConnection
        turso = TursoConnection(turso_url, turso_token)

        row = turso.execute("SELECT valeur FROM config WHERE cle='crypto_key'").fetchone()
        if row and row[0]:
            key = row[0]
        else:
            # Première instance → générer la clé et l'écrire dans Turso
            key = secrets.token_hex(32)
            now = _utcnow().isoformat()
            turso.execute(
                "INSERT OR REPLACE INTO config (cle, valeur, date_maj) VALUES (?, ?, ?)",
                ('crypto_key', key, now)
            )
            logger.info('🔑 Clé crypto générée et stockée dans Turso')

        logger.info('🔑 Clé crypto chargée depuis Turso (partagée entre instances)')
        return get_crypto_manager(shared_key=key)

    except Exception as e:
        logger.warning(f'_build_crypto_shared Turso inaccessible ({e}), fallback local')
        return None


def _invalidate_crypto_cache():
    """Vide le cache crypto (à appeler si les credentials Turso changent)."""
    global _crypto_shared_cache
    with _crypto_shared_lock:
        _crypto_shared_cache = None

def init_db():
    conn = get_db(); c = conn.cursor()

    # TABLE CLIENTS
    c.execute('''CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL DEFAULT '',
        contact TEXT DEFAULT '',
        telephone TEXT DEFAULT '',
        email TEXT DEFAULT '',
        adresse TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        couleur TEXT DEFAULT '#00c9ff',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '')''')

    # TABLE PARC (lié à un client)
    c.execute('''CREATE TABLE IF NOT EXISTS parc_general (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        nom_site TEXT DEFAULT '', adresse TEXT DEFAULT '',
        type_connexion TEXT DEFAULT '', debit_montant TEXT DEFAULT '', debit_descendant TEXT DEFAULT '',
        fournisseur_internet TEXT DEFAULT '', ip_publique TEXT DEFAULT '',
        plage_ip_locale TEXT DEFAULT '192.168.1.0/24', nb_machines INTEGER DEFAULT 0,
        nb_utilisateurs INTEGER DEFAULT 0, domaine TEXT DEFAULT '', serveur_dns TEXT DEFAULT '',
        passerelle TEXT DEFAULT '', baie_marque TEXT DEFAULT '', baie_nb_u INTEGER DEFAULT 0,
        switch_marque TEXT DEFAULT '', switch_nb_ports INTEGER DEFAULT 0, switch_nb_unites INTEGER DEFAULT 0,
        routeur_marque TEXT DEFAULT '', serveur_marque TEXT DEFAULT '', serveur_modele TEXT DEFAULT '',
        ups_marque TEXT DEFAULT '', ups_capacite TEXT DEFAULT '', autres_equipements TEXT DEFAULT '',
        logiciels_metier TEXT DEFAULT '', antivirus TEXT DEFAULT '', os_principal TEXT DEFAULT '',
        suite_bureautique TEXT DEFAULT '', notes TEXT DEFAULT '', date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # Ajouter "Wi-Fi" à la liste des catégories d'identifiants si absente
    try:
        nb = c.execute("SELECT COUNT(*) FROM config_listes WHERE nom_liste='categories_identifiants' AND valeur='Wi-Fi'").fetchone()[0]
        if nb == 0:
            # Ne forcer que si la liste a déjà été personnalisée
            nb_total = c.execute("SELECT COUNT(*) FROM config_listes WHERE nom_liste='categories_identifiants'").fetchone()[0]
            if nb_total > 0:
                ordre = c.execute("SELECT COALESCE(MAX(ordre),0)+1 FROM config_listes WHERE nom_liste='categories_identifiants'").fetchone()[0]
                c.execute("INSERT OR IGNORE INTO config_listes (nom_liste,valeur,ordre) VALUES ('categories_identifiants','Wi-Fi',?)", (ordre,))
    except: pass

    # Colonnes WiFi dans parc_general (migration)
    for col, defval in [
        ('wifi_ssid',       "''"),
        ('wifi_password',   "''"),
        ('wifi_securite',   "'WPA2'"),
        ('wifi_ssid2',      "''"),
        ('wifi_password2',  "''"),
        ('wifi_securite2',  "'WPA2'"),
        ('wifi_notes',      "''"),
    ]:
        try:
            c.execute(f"ALTER TABLE parc_general ADD COLUMN {col} TEXT DEFAULT {defval}")
        except: pass  # colonne déjà existante

    # TABLE APPAREILS (lié à un client)
    c.execute('''CREATE TABLE IF NOT EXISTS appareils (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        nom_machine TEXT DEFAULT '', type_appareil TEXT DEFAULT '',
        marque TEXT DEFAULT '', modele TEXT DEFAULT '', numero_serie TEXT DEFAULT '',
        adresse_ip TEXT DEFAULT '', adresse_mac TEXT DEFAULT '', nom_dns TEXT DEFAULT '',
        utilisateur TEXT DEFAULT '', service TEXT DEFAULT '', localisation TEXT DEFAULT '',
        date_achat TEXT DEFAULT '', duree_garantie INTEGER DEFAULT 0, date_fin_garantie TEXT DEFAULT '',
        fournisseur TEXT DEFAULT '', prix_achat REAL, numero_commande TEXT DEFAULT '',
        os TEXT DEFAULT '', version_os TEXT DEFAULT '', ram TEXT DEFAULT '', cpu TEXT DEFAULT '',
        stockage TEXT DEFAULT '', statut TEXT DEFAULT 'actif', dernier_ping TEXT DEFAULT '',
        en_ligne INTEGER DEFAULT 0, decouvert_scan INTEGER DEFAULT 0, ports_ouverts TEXT DEFAULT '',
        notes TEXT DEFAULT '', date_creation TEXT DEFAULT '', date_maj TEXT DEFAULT '',
        user_login TEXT DEFAULT '', user_password TEXT DEFAULT '',
        admin_login TEXT DEFAULT '', admin_password TEXT DEFAULT '',
        anydesk_id TEXT DEFAULT '', anydesk_password TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE IDENTIFIANTS GLOBAUX
    c.execute('''CREATE TABLE IF NOT EXISTS identifiants (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        categorie TEXT DEFAULT '',
        nom TEXT DEFAULT '',
        login TEXT DEFAULT '',
        mot_de_passe TEXT DEFAULT '',
        url TEXT DEFAULT '',
        description TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        date_expiration TEXT DEFAULT '',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # Colonnes WiFi dans identifiants (migration)
    for col, defval in [
        ('wifi_ssid',     "''"),
        ('wifi_securite', "'WPA2'"),
    ]:
        try:
            c.execute(f"ALTER TABLE identifiants ADD COLUMN {col} TEXT DEFAULT {defval}")
        except: pass

    # Lien optionnel vers un appareil/périphérique/utilisateur (migration) —
    # audit du 2026-08-24 : jusqu'ici un identifiant n'était jamais rattaché
    # à une entité précise (juste un nom/description en texte libre), alors
    # que la fiche appareil a ses propres champs user_login/user_password/
    # admin_login/admin_password. Deux silos de credentials pour une même
    # machine, sans le moindre rapprochement. utilisateur_id ajouté dans un
    # second temps (audit du même jour) : un identifiant nominatif (boîte
    # mail personnelle, VPN nominatif) n'appartient à aucune machine, juste
    # à une personne. NULL par défaut : ne change rien aux identifiants déjà
    # saisis.
    for col in ('appareil_id', 'peripherique_id', 'utilisateur_id'):
        try:
            c.execute(f"ALTER TABLE identifiants ADD COLUMN {col} INTEGER")
        except: pass

    # TABLE SERVICES
    c.execute('''CREATE TABLE IF NOT EXISTS services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        nom TEXT NOT NULL DEFAULT '',
        description TEXT DEFAULT '',
        responsable TEXT DEFAULT '',
        couleur TEXT DEFAULT '#6a8aaa',
        ordre INTEGER DEFAULT 0,
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE UTILISATEURS
    c.execute('''CREATE TABLE IF NOT EXISTS utilisateurs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        service_id INTEGER,
        prenom TEXT DEFAULT '',
        nom TEXT DEFAULT '',
        poste TEXT DEFAULT '',
        email TEXT DEFAULT '',
        telephone TEXT DEFAULT '',
        login_windows TEXT DEFAULT '',
        login_mail TEXT DEFAULT '',
        statut TEXT DEFAULT 'actif',
        notes TEXT DEFAULT '',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY(service_id) REFERENCES services(id) ON DELETE SET NULL)''')

    # Lien optionnel vers service/utilisateur sur appareils (migration) —
    # audit du 2026-08-24 : appareils.service et appareils.utilisateur
    # restaient du texte libre, sans le moindre lien vers les tables
    # services/utilisateurs (qui ont, elles, une vraie fiche structurée),
    # contrairement à peripheriques.utilisateur_id qui pointe déjà
    # proprement vers utilisateurs. Colonnes ajoutées en plus du texte
    # existant (rien ne casse), rattachées une fois ici par rapprochement de
    # nom (comme peripheriques_appareils l'a fait en son temps pour
    # appareil_id), puis tenues à jour à chaque sauvegarde de la fiche
    # appareil (voir editer_appareil/nouvel_appareil).
    for col in ('service_id', 'utilisateur_id'):
        try:
            c.execute(f"ALTER TABLE appareils ADD COLUMN {col} INTEGER")
        except: pass
    try:
        a_rattacher = c.execute(
            "SELECT id, client_id, service, utilisateur FROM appareils WHERE "
            "(service_id IS NULL AND service IS NOT NULL AND service!='') OR "
            "(utilisateur_id IS NULL AND utilisateur IS NOT NULL AND utilisateur!='')"
        ).fetchall()
        for aid, aclient, service_txt, user_txt in a_rattacher:
            svc_id = _resolve_service_id(conn, aclient, service_txt) if service_txt else None
            usr_id = _resolve_utilisateur_id(conn, aclient, user_txt) if user_txt else None
            if svc_id or usr_id:
                c.execute(
                    'UPDATE appareils SET service_id=COALESCE(service_id,?), '
                    'utilisateur_id=COALESCE(utilisateur_id,?) WHERE id=?',
                    (svc_id, usr_id, aid))
        if a_rattacher:
            conn.commit()
    except Exception:
        pass

    # TABLE TYPES_DROITS (référentiel configurable par client)
    c.execute('''CREATE TABLE IF NOT EXISTS types_droits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        categorie TEXT DEFAULT '',
        nom TEXT NOT NULL DEFAULT '',
        description TEXT DEFAULT '',
        icone TEXT DEFAULT '🔑',
        ordre INTEGER DEFAULT 0,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE DROITS UTILISATEURS (pivot users <-> types_droits)
    c.execute('''CREATE TABLE IF NOT EXISTS droits_utilisateurs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        utilisateur_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        categorie TEXT DEFAULT '',
        type_droit_id INTEGER,
        nom_droit TEXT DEFAULT '',
        valeur TEXT DEFAULT '',
        niveau TEXT DEFAULT 'lecture',
        notes TEXT DEFAULT '',
        date_attribution TEXT DEFAULT '',
        FOREIGN KEY(utilisateur_id) REFERENCES utilisateurs(id) ON DELETE CASCADE,
        FOREIGN KEY(type_droit_id) REFERENCES types_droits(id) ON DELETE SET NULL)''')

    # TABLE PERIPHERIQUES
    c.execute('''CREATE TABLE IF NOT EXISTS peripheriques (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        appareil_id INTEGER,
        utilisateur_id INTEGER,
        categorie TEXT DEFAULT '',
        marque TEXT DEFAULT '',
        modele TEXT DEFAULT '',
        numero_serie TEXT DEFAULT '',
        description TEXT DEFAULT '',
        localisation TEXT DEFAULT '',
        statut TEXT DEFAULT 'actif',
        date_achat TEXT DEFAULT '',
        duree_garantie INTEGER DEFAULT 0,
        date_fin_garantie TEXT DEFAULT '',
        fournisseur TEXT DEFAULT '',
        prix_achat REAL,
        numero_commande TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE SET NULL,
        FOREIGN KEY(utilisateur_id) REFERENCES utilisateurs(id) ON DELETE SET NULL)''')

    # TABLE LISTES PERSONNALISABLES
    c.execute('''CREATE TABLE IF NOT EXISTS config_listes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom_liste TEXT NOT NULL,
        valeur TEXT NOT NULL,
        ordre INTEGER DEFAULT 0,
        UNIQUE(nom_liste, valeur))''')

    # TABLE CONFIGURATION GLOBALE
    c.execute('''CREATE TABLE IF NOT EXISTS config (
        cle TEXT PRIMARY KEY,
        valeur TEXT DEFAULT '',
        date_maj TEXT DEFAULT '')''')

    # CLÉS DE RÉCUPÉRATION BITLOCKER — chiffrées, comme les mots de passe des
    # identifiants. Table dédiée plutôt qu'une colonne du rapport : le rapport
    # est repris tel quel dans le PDF joint à l'appareil, et un secret n'a rien
    # à faire dans une pièce jointe.
    c.execute('''CREATE TABLE IF NOT EXISTS cles_recuperation (
        cle           TEXT PRIMARY KEY,
        appareil_id   INTEGER NOT NULL,
        client_id     INTEGER NOT NULL,
        volume        TEXT DEFAULT '',
        identifiant   TEXT DEFAULT '',
        protection    TEXT DEFAULT '',
        chiffrement   TEXT DEFAULT '',
        valeur        TEXT DEFAULT '',
        date_maj      TEXT DEFAULT '')''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_cles_recuperation_appareil '
              'ON cles_recuperation(appareil_id)')

    # HISTORIQUE DES COLLECTES — un relevé par passage du collecteur.
    # Chaque collecte écrasait la précédente : on avait une photo, jamais une
    # trajectoire. Seules les grandeurs qui servent à comparer sont conservées,
    # pas le rapport entier — l'historique reste léger et synchronisable.
    # Clé TEXTE pour la même raison que journal_maj : chaque instance numérote
    # dans sa propre base, un identifiant auto-incrémenté entrerait en collision.
    c.execute('''CREATE TABLE IF NOT EXISTS collectes (
        cle               TEXT PRIMARY KEY,
        appareil_id       INTEGER NOT NULL,
        client_id         INTEGER NOT NULL,
        horodatage        TEXT NOT NULL,
        disque_total_go   REAL,
        disque_utilise_go REAL,
        disque_libre_go   REAL,
        ram_go            REAL,
        nb_logiciels      INTEGER,
        logiciels         TEXT DEFAULT '',
        os_version        TEXT DEFAULT '',
        cpu               TEXT DEFAULT '',
        numero_serie      TEXT DEFAULT '',
        nb_maj_attente    INTEGER,
        nb_peripheriques_erreur INTEGER,
        date_maj          TEXT DEFAULT '')''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_collectes_appareil '
              'ON collectes(appareil_id, horodatage DESC)')

    # JOURNAL DES MISES À JOUR DE L'APPLICATION — synchronisé entre instances,
    # pour qu'un poste voie ce qui a été installé sur les autres.
    # Clé TEXTE et non un id auto-incrémenté : chaque instance écrit dans sa
    # propre base, et deux machines qui se mettent à jour le même jour
    # produiraient le même id — la synchronisation écraserait l'une par l'autre.
    c.execute('''CREATE TABLE IF NOT EXISTS journal_maj (
        cle           TEXT PRIMARY KEY,
        horodatage    TEXT NOT NULL,
        machine       TEXT DEFAULT '',
        mode          TEXT DEFAULT '',
        version_avant TEXT DEFAULT '',
        version_apres TEXT DEFAULT '',
        statut        TEXT DEFAULT 'succes',
        detail        TEXT DEFAULT '',
        date_maj      TEXT DEFAULT '')''')

    # TABLE PRÉFÉRENCES UTILISATEUR (personnalisation par utilisateur)
    c.execute('''CREATE TABLE IF NOT EXISTS user_preferences (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        auth_user_id INTEGER NOT NULL,
        cle TEXT NOT NULL,
        valeur TEXT DEFAULT '',
        date_maj TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(auth_user_id, cle),
        FOREIGN KEY(auth_user_id) REFERENCES auth_users(id) ON DELETE CASCADE)''')

    # TABLE OUTILS (indépendant du client)
    c.execute('''CREATE TABLE IF NOT EXISTS outils (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        url TEXT NOT NULL,
        description TEXT DEFAULT '',
        categorie TEXT DEFAULT 'Général',
        icone TEXT DEFAULT '🔧',
        ordre INTEGER DEFAULT 0,
        actif INTEGER DEFAULT 1)''')

    # Insérer les outils par défaut si la table est vide
    nb_outils = c.execute('SELECT COUNT(*) FROM outils').fetchone()[0]
    if nb_outils == 0:
        defaults = [
            ('Test de débit',       'https://www.nperf.com/',              'Test vitesse download/upload/ping',   'Réseau',   '⚡', 0),
            ('Fast.com',            'https://fast.com/',                   'Test de débit Netflix',               'Réseau',   '⚡', 1),
            ('Test DNS Cloudflare', 'https://1.1.1.1/',                    'DNS Cloudflare & test connectivité',  'Réseau',   '🔒', 2),
            ('DNS Check Tools',     'https://dnschecker.org/',             'Vérification propagation DNS',        'DNS',      '🔍', 3),
            ('MXToolbox',           'https://mxtoolbox.com/',              'Outils DNS, blacklist, SMTP',         'DNS',      '📧', 4),
            ('What is my IP',       'https://www.whatismyip.com/',         'IP publique et géolocalisation',      'Réseau',   '🌐', 5),
            ('Cloudflare RADAR',    'https://radar.cloudflare.com/',       'Statistiques et état internet',       'Réseau',   '📡', 6),
            ('Test AdBlock d3ward', 'https://d3ward.github.io/toolz/adblock.html', 'Test efficacité bloqueur pubs', 'Sécurité','🛡', 7),
            ('SSL Labs',            'https://www.ssllabs.com/ssltest/',    'Analyse certificat SSL/TLS',          'Sécurité', '🔐', 8),
            ('Shodan',              'https://www.shodan.io/',              'Moteur de recherche IoT/sécurité',    'Sécurité', '🕵', 9),
            ('VirusTotal',          'https://www.virustotal.com/',         'Analyse fichiers et URLs',            'Sécurité', '🦠', 10),
            ('PingTools',           'https://ping.eu/',                    'Ping, traceroute, whois en ligne',    'Diagnostic','🏓', 11),
            ('Down For Everyone',   'https://downforeveryoneorjustme.com/','Site down ou problème local ?',       'Diagnostic','❓', 12),
            ('IPvFoo / IP Info',    'https://www.ipaddress.com/',          'Infos complètes sur une IP',          'DNS',      'ℹ', 13),
        ]
        for nom, url, desc, cat, ico, ordre in defaults:
            c.execute('INSERT INTO outils (nom,url,description,categorie,icone,ordre) VALUES (?,?,?,?,?,?)',
                      (nom, url, desc, cat, ico, ordre))

    # TABLE CONTRATS
    c.execute('''CREATE TABLE IF NOT EXISTS contrats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        titre TEXT DEFAULT '',
        type_contrat TEXT DEFAULT '',
        fournisseur TEXT DEFAULT '',
        contact_fournisseur TEXT DEFAULT '',
        email_fournisseur TEXT DEFAULT '',
        telephone_fournisseur TEXT DEFAULT '',
        numero_contrat TEXT DEFAULT '',
        date_debut TEXT DEFAULT '',
        date_fin TEXT DEFAULT '',
        reconduction_auto INTEGER DEFAULT 0,
        preavis_jours INTEGER DEFAULT 30,
        montant_ht REAL,
        periodicite TEXT DEFAULT 'annuel',
        description TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        statut TEXT DEFAULT 'actif',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE CONTRATS <-> APPAREILS (pivot)
    c.execute('''CREATE TABLE IF NOT EXISTS contrats_appareils (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contrat_id INTEGER NOT NULL,
        appareil_id INTEGER NOT NULL,
        FOREIGN KEY(contrat_id) REFERENCES contrats(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE CASCADE)''')

    # TABLE CONTRATS <-> PERIPHERIQUES (pivot)
    c.execute('''CREATE TABLE IF NOT EXISTS contrats_peripheriques (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contrat_id INTEGER NOT NULL,
        peripherique_id INTEGER NOT NULL,
        FOREIGN KEY(contrat_id) REFERENCES contrats(id) ON DELETE CASCADE,
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id) ON DELETE CASCADE)''')

    # TABLE MAINTENANCES
    c.execute('''CREATE TABLE IF NOT EXISTS maintenances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        appareil_id INTEGER,
        peripherique_id INTEGER,
        contrat_id INTEGER,
        type_maintenance TEXT NOT NULL,
        description TEXT DEFAULT '',
        date_planifiee TEXT NOT NULL,
        date_realisee TEXT,
        heure_debut TEXT DEFAULT '',
        heure_fin TEXT DEFAULT '',
        responsable TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        statut TEXT DEFAULT 'programmee',
        recurrence TEXT,
        date_fin_recurrence TEXT,
        parent_id INTEGER,
        created_by INTEGER,
        updated_by INTEGER,
        date_creation TEXT DEFAULT CURRENT_TIMESTAMP,
        date_maj TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE SET NULL,
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id) ON DELETE SET NULL,
        FOREIGN KEY(contrat_id) REFERENCES contrats(id) ON DELETE SET NULL,
        FOREIGN KEY(created_by) REFERENCES auth_users(id) ON DELETE SET NULL,
        FOREIGN KEY(updated_by) REFERENCES auth_users(id) ON DELETE SET NULL,
        FOREIGN KEY(parent_id) REFERENCES maintenances(id) ON DELETE CASCADE)''')

    # Ajouter colonne contrat_id si elle n'existe pas (migration)
    try:
        c.execute("ALTER TABLE maintenances ADD COLUMN contrat_id INTEGER")
    except: pass

    # ═══════════════════════════════════════════════════════════════════════════
    # CRÉATION DES INDICES (Optimisation Performance)
    # ═══════════════════════════════════════════════════════════════════════════
    try:
        # ── MAINTENANCES ──────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_client ON maintenances(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_appareil ON maintenances(appareil_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_peripherique ON maintenances(peripherique_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_date ON maintenances(date_planifiee)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_contrat ON maintenances(contrat_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_statut ON maintenances(statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_type ON maintenances(type_maintenance)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_client_statut ON maintenances(client_id, statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_maintenances_client_date ON maintenances(client_id, date_planifiee)')

        # ── APPAREILS ─────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_client ON appareils(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_statut ON appareils(statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_en_ligne ON appareils(en_ligne)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_type ON appareils(type_appareil)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_date_maj ON appareils(date_maj DESC)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_nom_machine ON appareils(nom_machine)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_client_statut ON appareils(client_id, statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_client_en_ligne ON appareils(client_id, en_ligne)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_date_fin_garantie ON appareils(date_fin_garantie)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_appareils_av_date_fin ON appareils(av_date_fin)')

        # ── PÉRIPHÉRIQUES ─────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_client ON peripheriques(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_statut ON peripheriques(statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_categorie ON peripheriques(categorie)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_date_creation ON peripheriques(date_creation)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_client_statut ON peripheriques(client_id, statut)')

        # ── CONTRATS ──────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_client ON contrats(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_statut ON contrats(statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_date_fin ON contrats(date_fin)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_client_statut ON contrats(client_id, statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_client_date_fin ON contrats(client_id, date_fin)')

        # ── INTERVENTIONS ─────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_client ON interventions(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_date ON interventions(date_intervention)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_statut ON interventions(statut)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_client_date ON interventions(client_id, date_intervention DESC)')

        # ── UTILISATEURS ──────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_utilisateurs_client ON utilisateurs(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_utilisateurs_prenom ON utilisateurs(prenom)')

        # ── SERVICES ──────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_services_client ON services(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_services_client_ordre ON services(client_id, ordre)')

        # ── IDENTIFIANTS ──────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_identifiants_client ON identifiants(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_identifiants_categorie ON identifiants(categorie)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_identifiants_client_categorie ON identifiants(client_id, categorie)')

        # ── AUTH_USERS ────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_auth_users_login ON auth_users(login)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_auth_users_actif ON auth_users(actif)')

        # ── CLIENT_PARTAGES ───────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_client_partages_client ON client_partages(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_client_partages_user ON client_partages(auth_user_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_client_partages_client_user ON client_partages(client_id, auth_user_id)')

        # ── HISTORIQUE ────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_historique_client ON historique(client_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_historique_date ON historique(date_action)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_historique_client_date ON historique(client_id, date_action DESC)')

        # ── DOCUMENTS ─────────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_documents_appareils_appareil ON documents_appareils(appareil_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_documents_contrats_contrat ON documents_contrats(contrat_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_documents_peripheriques_periph ON documents_peripheriques(peripherique_id)')

        # ── TABLES PIVOT ──────────────────────────────────────────────────────
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_appareils_contrat ON contrats_appareils(contrat_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_contrats_appareils_appareil ON contrats_appareils(appareil_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_appareils_intervention ON interventions_appareils(intervention_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_interventions_appareils_appareil ON interventions_appareils(appareil_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_appareils_periph ON peripheriques_appareils(peripherique_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_peripheriques_appareils_appareil ON peripheriques_appareils(appareil_id)')

        logger.info('✅ Indices de performance créés avec succès')
    except Exception as e:
        logger.warning(f'⚠️ Erreur lors de la création des indices: {e}')

    # ═══════════════════════════════════════════════════════════════════════════
    # CLÉ DE CHIFFREMENT PARTAGÉE (synchronisée via Turso)
    # ═══════════════════════════════════════════════════════════════════════════
    # MIGRATION: Chiffrer les identifiants existants en clair
    # ═══════════════════════════════════════════════════════════════════════════
    try:
        crypto = _get_crypto_shared(c)

        # Récupérer tous les identifiants avec mot de passe non chiffré
        not_encrypted = c.execute('''
            SELECT id, mot_de_passe FROM identifiants
            WHERE mot_de_passe IS NOT NULL
            AND mot_de_passe != ''
            AND mot_de_passe NOT LIKE 'gAAAAAB%'
        ''').fetchall()

        if not_encrypted:
            logger.info(f'🔐 Migration: chiffrement de {len(not_encrypted)} identifiants existants...')
            for ident_id, mdp_clair in not_encrypted:
                mdp_chiffre = crypto.encrypt(mdp_clair)
                c.execute('UPDATE identifiants SET mot_de_passe=? WHERE id=?', (mdp_chiffre, ident_id))
                logger.debug(f'  ✅ ID {ident_id} chiffré')
            conn.commit()
            logger.info(f'✅ Migration terminée: {len(not_encrypted)} identifiants chiffrés')
        else:
            logger.info('✅ Tous les identifiants sont déjà chiffrés')
    except Exception as e:
        logger.warning(f'⚠️ Erreur lors de la migration des identifiants: {e}')

    # TABLE MAINTENANCE_NOTIFICATIONS (tracking notifications envoyées)
    c.execute('''CREATE TABLE IF NOT EXISTS maintenance_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        maintenance_id INTEGER NOT NULL,
        notification_date TEXT NOT NULL,
        date_creation TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(maintenance_id) REFERENCES maintenances(id) ON DELETE CASCADE)''')

    # TABLE DOCUMENTS CONTRATS
    c.execute('''CREATE TABLE IF NOT EXISTS documents_contrats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contrat_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        nom TEXT DEFAULT '',
        description TEXT DEFAULT '',
        type_doc TEXT DEFAULT '',
        nom_fichier TEXT DEFAULT '',
        taille INTEGER DEFAULT 0,
        date_upload TEXT DEFAULT '',
        contenu_blob BLOB,
        sync_status TEXT DEFAULT 'local',
        date_sync TEXT DEFAULT '',
        FOREIGN KEY(contrat_id) REFERENCES contrats(id) ON DELETE CASCADE,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE DOCUMENTS APPAREILS
    c.execute('''CREATE TABLE IF NOT EXISTS documents_appareils (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        appareil_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        nom TEXT DEFAULT '',
        description TEXT DEFAULT '',
        type_doc TEXT DEFAULT '',
        nom_fichier TEXT DEFAULT '',
        taille INTEGER DEFAULT 0,
        date_upload TEXT DEFAULT '',
        contenu_blob BLOB,
        sync_status TEXT DEFAULT 'local',
        date_sync TEXT DEFAULT '',
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE CASCADE,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE HISTORIQUE
    # ── AUTH UTILISATEURS ────────────────────────────────────────────────────
    c.execute('''CREATE TABLE IF NOT EXISTS auth_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        login TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        nom TEXT NOT NULL,
        prenom TEXT DEFAULT '',
        email TEXT DEFAULT '',
        role TEXT DEFAULT 'user',
        logo_fichier TEXT DEFAULT '',
        actif INTEGER DEFAULT 1,
        date_creation TEXT,
        date_maj TEXT)''')

    c.execute('''CREATE TABLE IF NOT EXISTS client_partages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        auth_user_id INTEGER NOT NULL,
        niveau TEXT DEFAULT 'lecture',
        date_partage TEXT,
        UNIQUE(client_id, auth_user_id))''')

    # Migration : ajouter auth_user_id aux clients si absent
    cols_clients = [r[1] for r in c.execute('PRAGMA table_info(clients)').fetchall()]
    if 'auth_user_id' not in cols_clients:
        c.execute('ALTER TABLE clients ADD COLUMN auth_user_id INTEGER DEFAULT NULL')

    # Migration : jeton collecteur dédié à ce client (optionnel — voir
    # jeton_collecteur_valide()). Vide par défaut : aucun changement de
    # comportement tant qu'un admin ne l'a pas explicitement renseigné.
    if 'collecteur_token' not in cols_clients:
        c.execute("ALTER TABLE clients ADD COLUMN collecteur_token TEXT DEFAULT ''")

    # Migration : ajouter must_change_password si absent
    cols_auth = [r[1] for r in c.execute('PRAGMA table_info(auth_users)').fetchall()]
    if 'must_change_password' not in cols_auth:
        c.execute('ALTER TABLE auth_users ADD COLUMN must_change_password INTEGER DEFAULT 0')

    # Compte admin par defaut + rattachement des clients existants
    admin = c.execute("SELECT id FROM auth_users WHERE login='admin'").fetchone()
    if not admin:
        pwd_hash = _hash_pwd('admin')
        now2 = _utcnow().isoformat()
        c.execute("INSERT INTO auth_users (login,password_hash,nom,prenom,role,actif,must_change_password,date_creation,date_maj) VALUES (?,?,?,?,?,1,1,?,?)",
                  ('admin', pwd_hash, 'Administrateur', '', 'admin', now2, now2))
    c.execute("UPDATE clients SET auth_user_id=(SELECT id FROM auth_users WHERE login='admin') WHERE auth_user_id IS NULL")

    c.execute('''CREATE TABLE IF NOT EXISTS kb_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        icone TEXT DEFAULT '📋',
        ordre INTEGER DEFAULT 0)''')
    c.execute('''CREATE TABLE IF NOT EXISTS kb_articles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        categorie_id INTEGER NOT NULL,
        titre TEXT NOT NULL,
        contenu TEXT NOT NULL,
        tags TEXT DEFAULT '',
        date_creation TEXT,
        date_maj TEXT,
        FOREIGN KEY (categorie_id) REFERENCES kb_categories(id))''')

    c.execute('''CREATE TABLE IF NOT EXISTS historique (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        entite TEXT NOT NULL,
        entite_id INTEGER NOT NULL,
        entite_nom TEXT DEFAULT '',
        action TEXT NOT NULL,
        date_action TEXT NOT NULL,
        details TEXT DEFAULT '')''')

    # TABLE DOCUMENTS PÉRIPHÉRIQUES
    c.execute('''CREATE TABLE IF NOT EXISTS documents_peripheriques (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        peripherique_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        nom TEXT DEFAULT '',
        description TEXT DEFAULT '',
        type_doc TEXT DEFAULT '',
        nom_fichier TEXT DEFAULT '',
        taille INTEGER DEFAULT 0,
        date_upload TEXT DEFAULT '',
        contenu_blob BLOB,
        sync_status TEXT DEFAULT 'local',
        date_sync TEXT DEFAULT '',
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id) ON DELETE CASCADE,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE PIVOT PÉRIPHÉRIQUES <-> APPAREILS (N:N)
    c.execute('''CREATE TABLE IF NOT EXISTS peripheriques_appareils (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        peripherique_id INTEGER NOT NULL,
        appareil_id     INTEGER NOT NULL,
        UNIQUE(peripherique_id, appareil_id),
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id)     REFERENCES appareils(id)     ON DELETE CASCADE)''')

    # Migration : si appareil_id non-NULL dans peripheriques, copier dans la table pivot
    try:
        migrated = conn.execute(
            "SELECT id, appareil_id FROM peripheriques WHERE appareil_id IS NOT NULL").fetchall()
        for row in migrated:
            conn.execute(
                "INSERT OR IGNORE INTO peripheriques_appareils (peripherique_id, appareil_id) VALUES (?,?)",
                (row[0], row[1]))
        if migrated:
            conn.commit()
    except Exception:
        pass

    # ════════════════════════════════════════════════════════════════════════════
    # TABLE INTERVENTIONS
    # ════════════════════════════════════════════════════════════════════════════
    c.execute('''CREATE TABLE IF NOT EXISTS interventions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        titre TEXT DEFAULT '',
        type_intervention TEXT DEFAULT '',
        description TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        date_intervention TEXT NOT NULL,
        heure_debut TEXT DEFAULT '',
        heure_fin TEXT DEFAULT '',
        duree_minutes INTEGER DEFAULT 0,
        technicien_nom TEXT DEFAULT '',
        technicien_email TEXT DEFAULT '',
        statut TEXT DEFAULT 'completee',
        contrat_id INTEGER DEFAULT NULL,
        cout_ht REAL DEFAULT 0,
        devise TEXT DEFAULT 'EUR',
        date_creation TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        auth_user_id INTEGER,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY(contrat_id) REFERENCES contrats(id) ON DELETE SET NULL,
        FOREIGN KEY(auth_user_id) REFERENCES auth_users(id) ON DELETE SET NULL)''')

    # TABLE PIVOT: INTERVENTIONS <-> APPAREILS
    c.execute('''CREATE TABLE IF NOT EXISTS interventions_appareils (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL,
        appareil_id INTEGER NOT NULL,
        FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE CASCADE)''')

    # TABLE PIVOT: INTERVENTIONS <-> PERIPHERIQUES
    c.execute('''CREATE TABLE IF NOT EXISTS interventions_peripheriques (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL,
        peripherique_id INTEGER NOT NULL,
        FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE,
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id) ON DELETE CASCADE)''')

    # TABLE DOCUMENTS INTERVENTIONS
    c.execute('''CREATE TABLE IF NOT EXISTS documents_interventions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        nom TEXT DEFAULT '',
        description TEXT DEFAULT '',
        type_doc TEXT DEFAULT '',
        nom_fichier TEXT DEFAULT '',
        taille INTEGER DEFAULT 0,
        date_upload TEXT DEFAULT '',
        FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # Migration : colonnes antivirus / EDR / RMM sur appareils
    for _col, _def in [('av_marque', "TEXT DEFAULT ''"), ('av_nom', "TEXT DEFAULT ''"),
                        ('av_date_debut', "TEXT DEFAULT ''"), ('av_date_fin', "TEXT DEFAULT ''"),
                        ('av_contrat_id', 'INTEGER'),
                        ('edr_marque', "TEXT DEFAULT ''"), ('edr_nom', "TEXT DEFAULT ''"),
                        ('edr_date_fin', "TEXT DEFAULT ''"), ('edr_contrat_id', 'INTEGER'),
                        ('rmm_marque', "TEXT DEFAULT ''"), ('rmm_nom', "TEXT DEFAULT ''"),
                        ('rmm_agent_id', "TEXT DEFAULT ''"), ('rmm_date_fin', "TEXT DEFAULT ''"),
                        ('rmm_contrat_id', 'INTEGER')]:
        try:
            c.execute(f"ALTER TABLE appareils ADD COLUMN {_col} {_def}")
        except Exception:
            pass

    # Migration : colonne logiciels sur appareils (JSON array)
    try:
        c.execute("ALTER TABLE appareils ADD COLUMN logiciels TEXT DEFAULT '[]'")
    except Exception:
        pass

    # Migration : IP publique + opérateur constatés par le collecteur, par
    # APPAREIL — distinct de parc_general.ip_publique/fournisseur_internet
    # (un par site, saisi à la main) : utile pour un poste itinérant dont
    # l'IP publique varie selon le lieu de connexion.
    for _col, _def in [('adresse_ip_publique', "TEXT DEFAULT ''"),
                        ('operateur_ip_publique', "TEXT DEFAULT ''")]:
        try:
            c.execute(f"ALTER TABLE appareils ADD COLUMN {_col} {_def}")
        except Exception:
            pass

    # Migration : date_maj sur outils (pour sync bidirectionnelle) — baie_slots
    # n'existe pas encore à ce stade sur une installation neuve, sa colonne
    # date_maj est ajoutée juste après sa CREATE TABLE, plus bas.
    try:
        c.execute("ALTER TABLE outils ADD COLUMN date_maj TEXT DEFAULT ''")
    except Exception:
        pass

    # TABLE BAIE DE BRASSAGE — SLOTS
    c.execute('''CREATE TABLE IF NOT EXISTS baie_slots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        position INTEGER NOT NULL,
        col_index INTEGER DEFAULT 0,
        hauteur_u INTEGER DEFAULT 1,
        appareil_id INTEGER,
        nom_custom TEXT DEFAULT '',
        type_equipement TEXT DEFAULT '',
        couleur TEXT DEFAULT '#1e3a5f',
        description TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE SET NULL)''')

    # TABLE PHOTOS BAIE
    c.execute('''CREATE TABLE IF NOT EXISTS baie_photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        nom TEXT DEFAULT '',
        description TEXT DEFAULT '',
        nom_fichier TEXT DEFAULT '',
        taille INTEGER DEFAULT 0,
        date_upload TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE PLANS D'ÉTAGE
    c.execute('''CREATE TABLE IF NOT EXISTS plans (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id     INTEGER NOT NULL,
        nom           TEXT DEFAULT '',
        description   TEXT DEFAULT '',
        contenu       TEXT DEFAULT '{"elements":[]}',
        date_creation TEXT DEFAULT '',
        date_maj      TEXT DEFAULT '',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')

    # TABLE LICENCES LOGICIELS (par appareil)
    c.execute('''CREATE TABLE IF NOT EXISTS licences_appareils (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        appareil_id    INTEGER NOT NULL,
        client_id      INTEGER NOT NULL,
        editeur        TEXT DEFAULT '',
        produit        TEXT DEFAULT '',
        cle_licence    TEXT DEFAULT '',
        contrat_id     INTEGER,
        date_creation  TEXT DEFAULT '',
        FOREIGN KEY(appareil_id) REFERENCES appareils(id) ON DELETE CASCADE,
        FOREIGN KEY(client_id)   REFERENCES clients(id)   ON DELETE CASCADE,
        FOREIGN KEY(contrat_id)  REFERENCES contrats(id)  ON DELETE SET NULL)''')

    # TABLE PRESTATAIRES (commun à tous les clients)

    conn.execute('PRAGMA foreign_keys = ON')

    # Migration : si ancienne table parc_general sans client_id, migrer
    cols_parc = [r[1] for r in conn.execute('PRAGMA table_info(parc_general)').fetchall()]
    cols_app  = [r[1] for r in conn.execute('PRAGMA table_info(appareils)').fetchall()]

    if 'client_id' not in cols_parc:
        # Créer client par défaut et migrer les données
        now = _utcnow().isoformat()
        c.execute("INSERT INTO clients (nom, date_creation, date_maj) VALUES ('Client par défaut', ?, ?)", (now, now))
        default_cid = c.lastrowid
        c.execute(f'ALTER TABLE parc_general ADD COLUMN client_id INTEGER DEFAULT {default_cid}')
        c.execute('UPDATE parc_general SET client_id=?', (default_cid,))

    if 'client_id' not in cols_app:
        now = _utcnow().isoformat()
        cid = conn.execute('SELECT id FROM clients ORDER BY id LIMIT 1').fetchone()
        if cid:
            c.execute(f"ALTER TABLE appareils ADD COLUMN client_id INTEGER DEFAULT {cid['id']}")
            c.execute('UPDATE appareils SET client_id=?', (cid['id'],))

    # Migration : col_index + baie_nom + date_maj (sync bidirectionnelle) dans baie_slots
    for col_add, defval in [('col_index','0'), ('baie_nom',"'Baie principale'"), ('date_maj', "''")]:
        try:
            c.execute(f"ALTER TABLE baie_slots ADD COLUMN {col_add} TEXT DEFAULT {defval}")
        except: pass

    # Lien optionnel vers un périphérique (migration) — audit du 2026-08-24 :
    # un emplacement de baie ne pouvait être associé qu'à un appareil ; un
    # onduleur/UPS ou un panneau de brassage — pourtant des catégories
    # périphérique existantes, typiquement montés en baie — ne pouvaient être
    # qu'une étiquette texte libre (nom_custom/type_equipement), jamais liés
    # à leur vraie fiche.
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN peripherique_id INTEGER")
    except: pass

    # Ports (RJ45) d'un élément réseau de la baie — demandé : compter et
    # numéroter les ports d'un switch/patch panel, lier chacun à un appareil/
    # périphérique existant ou à un usage libre (téléphonie, alarme...).
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN nb_ports INTEGER DEFAULT 0")
    except: pass

    # Disposition des ports RJ d'un switch — 'ligne' (comportement
    # historique, une seule rangée) ou 'deux_lignes' (façon vrai switch/
    # patch panel : impairs en haut, pairs en dessous — 1 3 5 7.../2 4 6
    # 8...), centrées dans les 80% gauche de l'élément — demandé.
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN ports_disposition TEXT DEFAULT 'ligne'")
    except: pass

    # Ports SFP (fibre) d'un switch — distincts des ports RJ, centrés dans
    # les 20% de largeur restants (voir _ports_avec_details/le rendu côté
    # client). Numérotés dans un espace SÉPARÉ des ports RJ (voir
    # SFP_NUMERO_OFFSET plus bas) plutôt qu'une colonne type_port
    # supplémentaire sur baie_slot_ports : les deux plages ne se
    # chevauchent jamais (RJ 1-48, SFP 1001-1008), pas besoin de toucher à
    # sa contrainte UNIQUE(slot_id, numero).
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN nb_ports_sfp INTEGER DEFAULT 0")
    except: pass

    # Ports WAN d'un routeur/pare-feu — troisième groupe de connecteurs
    # générique, dans un espace de numérotation encore séparé (voir
    # WAN_NUMERO_OFFSET plus bas). Un switch ou un onduleur/PDU n'utilisent
    # jamais cette colonne (0 par défaut, jamais modifiée pour ces types).
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN nb_ports_wan INTEGER DEFAULT 0")
    except: pass

    # Orientation d'un élément — 'horizontal' (comportement historique) ou
    # 'vertical' (PDU/onduleur monté sur le rail latéral, courant en baie
    # réelle) — demandé. Reste dans la MÊME grille position+col_index+
    # hauteur_u+largeur_u (aucune géométrie séparée à gérer) : seul le rendu
    # interne (nom, ports/prises) pivote côté client, voir cell-vertical.
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN orientation TEXT DEFAULT 'horizontal'")
    except: pass

    # Puissance nominale d'un onduleur, en VA — demandé, purement informatif
    # (affiché sur l'élément et dans son infobulle), aucun calcul de charge
    # cumulée des appareils branchés dessus. 0 = non renseignée. Seul le
    # type 'UPS' expose ce champ côté formulaire (voir appliquerReglesType),
    # mais la colonne reste générique comme le reste du schéma baie_slots.
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN puissance_va INTEGER DEFAULT 0")
    except: pass

    # TABLE PORTS DE BAIE — un port par numéro, par slot. FOREIGN KEY sur
    # slot_id sans ON DELETE CASCADE (SQLite ne l'applique de toute façon que
    # si PRAGMA foreign_keys=ON est actif sur LA connexion courante, pas
    # garanti partout dans ce fichier) : le nettoyage à la suppression d'un
    # slot est fait explicitement (voir api_baie_slot, branche DELETE), même
    # pattern que pour appareils/peripheriques/utilisateurs ailleurs dans
    # init_db().
    c.execute('''CREATE TABLE IF NOT EXISTS baie_slot_ports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slot_id INTEGER NOT NULL,
        numero INTEGER NOT NULL,
        appareil_id INTEGER,
        peripherique_id INTEGER,
        usage_libre TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(slot_id) REFERENCES baie_slots(id),
        FOREIGN KEY(appareil_id) REFERENCES appareils(id),
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id),
        UNIQUE(slot_id, numero))''')

    # Lien port-à-port (câblage physique switch<->routeur, etc.) — demandé
    # séparément de la liaison appareil/périphérique/usage : deux ports se
    # référencent mutuellement (lie_slot_id+lie_port_numero sur CHACUN des
    # deux, pas une seule ligne de jonction) — reste dans le même schéma
    # d'adressage (slot_id, numero) que tout le reste de baie_slot_ports.
    for col_add, defval in [('lie_slot_id', 'INTEGER'), ('lie_port_numero', 'INTEGER')]:
        try:
            c.execute(f"ALTER TABLE baie_slot_ports ADD COLUMN {col_add} {defval}")
        except: pass

    # Pièce du site desservie par ce port (prise murale), et étiquette de
    # câble (couleur/longueur) pour un lien port-à-port — deux métadonnées
    # DÉLIBÉRÉMENT INDÉPENDANTES de appareil_id/peripherique_id/usage_libre/
    # lie_slot_id (jamais effacées quand l'un de ces champs change, et
    # n'effacent jamais rien d'autre) : un port de bandeau RJ dessert une
    # pièce ET peut ensuite être câblé à un port de switch en même temps —
    # avant cette colonne, `piece` réutilisait usage_libre, effacé dès que
    # le port était câblé à un autre (perte de l'info "quelle pièce" au
    # moment où elle compte le plus). Sert la traçabilité de bout en bout
    # (pièce -> port bandeau -> câble -> port switch -> appareil).
    for col_add in ('piece', 'cable_couleur', 'cable_longueur'):
        try:
            c.execute(f"ALTER TABLE baie_slot_ports ADD COLUMN {col_add} TEXT DEFAULT ''")
        except: pass

    # PRISES MURALES — demandé : sur un bandeau RJ, séparer la prise murale
    # (côté local — pièce desservie, appareil branché dans le bureau) du
    # port RJ (côté baie — sert désormais UNIQUEMENT à interconnecter avec
    # un autre élément de la baie via "🔗 Lier des ports", comme dans un
    # vrai système de brassage structuré). Avant cette table, les deux
    # étaient conflées sur la même ligne baie_slot_ports (voir commentaire
    # piece/cable_couleur/cable_longueur ci-dessus) : un port de bandeau
    # portait À LA FOIS son appareil_id/peripherique_id/usage_libre/piece
    # ET, éventuellement, un lien vers un switch.
    # Correspondance AUTOMATIQUE par numéro de port avec baie_slot_ports —
    # pas de lien à créer manuellement, la prise murale #12 est
    # structurellement celle du port RJ #12 du même bandeau (voir
    # _reconcilier_prises_murales, en miroir de _reconcilier_ports).
    # cable_couleur/cable_longueur ici documentent le câble FIXE mur ->
    # bandeau, distinct du cordon de brassage (bandeau -> switch, resté sur
    # baie_slot_ports.cable_couleur/cable_longueur).
    c.execute('''CREATE TABLE IF NOT EXISTS baie_prises_murales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slot_id INTEGER NOT NULL,
        numero INTEGER NOT NULL,
        piece TEXT DEFAULT '',
        identification TEXT DEFAULT '',
        appareil_id INTEGER,
        peripherique_id INTEGER,
        usage_libre TEXT DEFAULT '',
        cable_couleur TEXT DEFAULT '',
        cable_longueur TEXT DEFAULT '',
        date_maj TEXT DEFAULT '',
        FOREIGN KEY(slot_id) REFERENCES baie_slots(id),
        FOREIGN KEY(appareil_id) REFERENCES appareils(id),
        FOREIGN KEY(peripherique_id) REFERENCES peripheriques(id),
        UNIQUE(slot_id, numero))''')
    # identification (repère physique de la prise, ex. "RJ 3.12" imprimé sur
    # la plaque murale — distinct de `piece`, qui nomme la pièce desservie,
    # pas la prise elle-même) : ajoutée après coup (v2.18.75), migration
    # nécessaire pour toute base ayant déjà créé la table ci-dessus (2.18.74).
    try:
        c.execute("ALTER TABLE baie_prises_murales ADD COLUMN identification TEXT DEFAULT ''")
    except: pass

    # Migration : reprend les ports de bandeau RJ déjà en base (créés avant
    # cette table) — copie appareil_id/peripherique_id/usage_libre/piece
    # vers la prise murale de même numéro, puis les vide côté port RJ (le
    # lien port-à-port éventuel — lie_slot_id/lie_port_numero/cable_couleur/
    # cable_longueur — n'est PAS touché, c'est le cordon vers le switch,
    # toujours du ressort du port). Idempotente par construction, sans
    # marqueur séparé : une fois migrée, une ligne n'a plus aucun de ces 4
    # champs renseigné, donc plus rien à reprendre au prochain démarrage.
    a_migrer = conn.execute(
        "SELECT bp.slot_id, bp.numero, bp.appareil_id, bp.peripherique_id, bp.usage_libre, bp.piece "
        "FROM baie_slot_ports bp JOIN baie_slots s ON s.id=bp.slot_id "
        "WHERE s.type_equipement='Bandeau RJ' AND ("
        "  bp.appareil_id IS NOT NULL OR bp.peripherique_id IS NOT NULL OR "
        "  (bp.usage_libre IS NOT NULL AND bp.usage_libre!='') OR "
        "  (bp.piece IS NOT NULL AND bp.piece!=''))"
    ).fetchall()
    if a_migrer:
        _now_migration = _utcnow().isoformat()
        for _slot_id, _numero, _aid, _pid, _usage, _piece in a_migrer:
            c.execute('''INSERT OR IGNORE INTO baie_prises_murales
                (slot_id, numero, piece, appareil_id, peripherique_id, usage_libre, date_maj)
                VALUES (?,?,?,?,?,?,?)''',
                (_slot_id, _numero, _piece or '', _aid, _pid, _usage or '', _now_migration))
            c.execute('''UPDATE baie_slot_ports SET appareil_id=NULL, peripherique_id=NULL,
                usage_libre='', piece='' WHERE slot_id=? AND numero=?''', (_slot_id, _numero))
        logger.info('Migration prises murales : %d port(s) de bandeau RJ repris', len(a_migrer))

    # Largeur d'un élément de baie, en dixièmes de la largeur du rack (1-10) —
    # demandé : redimensionner un élément en largeur à la souris, sur une
    # grille de 10 positions. NULL (pas juste 10) est le défaut délibéré :
    # ça signifie "jamais redimensionné, garder le partage égal automatique
    # entre éléments côte à côte" (comportement historique, calculé par
    # renderRack() côté client) — mettre 10 par défaut casserait l'affichage
    # existant de deux/trois éléments déjà placés côte à côte via col_index
    # (10+10 > 10 en tenants littéraux). Une valeur n'apparaît ici qu'après
    # un glisser explicite de la poignée de largeur (voir
    # redimensionnerLargeurSlot() dans baie_brassage.html).
    try:
        c.execute("ALTER TABLE baie_slots ADD COLUMN largeur_u INTEGER")
    except: pass

    # (ancienne migration col_index conservée pour compatibilité)
    cols_baie = [r[1] for r in conn.execute('PRAGMA table_info(baie_slots)').fetchall()]
    if 'col_index' not in cols_baie:
        c.execute("ALTER TABLE baie_slots ADD COLUMN col_index INTEGER DEFAULT 0")

    # Migration : colonnes identifiants appareils + carte graphique
    cols_app2 = [r[1] for r in conn.execute('PRAGMA table_info(appareils)').fetchall()]
    for col in ['user_login','user_password','admin_login','admin_password','anydesk_id','anydesk_password','carte_graphique']:
        if col not in cols_app2:
            c.execute(f"ALTER TABLE appareils ADD COLUMN {col} TEXT DEFAULT ''")
    if 'garantie_alerte_ignoree' not in cols_app2:
        c.execute("ALTER TABLE appareils ADD COLUMN garantie_alerte_ignoree INTEGER DEFAULT 0")

    # Legacy, plus alimentée (voir plus bas où les triggers _trg_del_* sont
    # retirés) : mécanisme d'avant _sync_journal, jamais relu par la sync
    # actuelle. CREATE conservé pour ne rien casser sur une base existante qui
    # porte encore ses anciennes lignes ; rien ne les purge plus, sans
    # conséquence (elles ne grossissent plus et ne gênent personne).
    c.execute('''CREATE TABLE IF NOT EXISTS _sync_deletions (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        tbl        TEXT    NOT NULL,
        record_id  INTEGER NOT NULL,
        deleted_at TEXT    NOT NULL,
        UNIQUE(tbl, record_id) ON CONFLICT REPLACE)''')

    # TABLE DE CHANGE-TRACKING : enregistre INSERT/UPDATE/DELETE pour optimiser la sync
    # record_id est TEXT (pas INTEGER) pour supporter aussi les tables à clé texte (ex: config.cle)
    c.execute('''CREATE TABLE IF NOT EXISTS _sync_journal (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        tbl       TEXT    NOT NULL,
        record_id TEXT    NOT NULL,
        action    TEXT    NOT NULL,  -- 'INSERT', 'UPDATE', 'DELETE'
        timestamp TEXT    NOT NULL,
        UNIQUE(tbl, record_id, action) ON CONFLICT REPLACE)''')

    # Garde anti-rebouclage : quand la sync applique une donnée reçue de Turso
    # (pull) dans les tables locales, ces écritures NE DOIVENT PAS re-déclencher
    # les triggers _trg_journal_* ci-dessous (sinon la donnée qu'on vient de
    # recevoir est immédiatement re-marquée comme "modification locale" et
    # repoussée au cycle suivant — avec un vieil état si la ligne existait déjà,
    # ce qui écrase silencieusement les changements faits par une autre instance).
    # database.py insère une ligne ici le temps d'appliquer le pull, puis la retire
    # dans un `finally` — sauf si le PROCESS entier disparaît entre les deux (kill,
    # coupure de courant, crash) : la ligne reste alors sur disque, survit au
    # redémarrage, et bloquerait silencieusement TOUS les triggers ci-dessous pour
    # toujours (WHEN NOT EXISTS (SELECT 1 FROM _sync_applying) ne serait plus
    # jamais vrai) — plus aucune modification locale ne serait journalisée, donc
    # plus jamais poussée vers les autres instances, sans la moindre erreur pour
    # le signaler. Un redémarrage de l'application est la seule preuve possible
    # qu'aucun pull n'est en cours : la vider ici est donc toujours sûr.
    c.execute('''CREATE TABLE IF NOT EXISTS _sync_applying (id INTEGER PRIMARY KEY)''')
    c.execute('DELETE FROM _sync_applying')

    # Triggers : enregistre automatiquement chaque modification dans _sync_journal
    # Toutes les tables de données sont couvertes (auth_users, client_partages, config, etc.
    # inclus — leur absence empêchait la sync entre instances de ces données critiques).
    # Valeur = nom de la colonne clé primaire ('id' pour la quasi-totalité, 'cle' pour config).
    _TRACKED_JOURNAL = {
        'appareils': 'id', 'peripheriques': 'id', 'identifiants': 'id', 'contrats': 'id',
        'utilisateurs': 'id', 'services': 'id', 'clients': 'id', 'baie_slots': 'id',
        'outils': 'id', 'kb_articles': 'id', 'kb_categories': 'id',
        'documents_appareils': 'id', 'documents_contrats': 'id',
        'documents_peripheriques': 'id', 'baie_photos': 'id',
        'types_droits': 'id', 'droits_utilisateurs': 'id',
        'contrats_appareils': 'id', 'contrats_peripheriques': 'id',
        'peripheriques_appareils': 'id', 'parc_general': 'id', 'historique': 'id', 'plans': 'id',
        'maintenances': 'id', 'interventions': 'id', 'licences_appareils': 'id',
        # Tables précédemment oubliées (jamais synchronisées entre instances) :
        'auth_users': 'id', 'client_partages': 'id',
        'config_listes': 'id', 'user_preferences': 'id',
        'documents_interventions': 'id', 'interventions_appareils': 'id',
        'interventions_peripheriques': 'id', 'maintenance_notifications': 'id',
        # baie_slot_ports : signalé en usage réel — le positionnement des
        # éléments de la baie (table baie_slots, suivie ci-dessus) se
        # synchronisait bien entre instances, mais pas leurs ports/câblage
        # (numéro, appareil/périphérique associé, usage libre, pièce, lien
        # port-à-port lie_slot_id/lie_port_numero) : cette table entière
        # avait été oubliée de _TRACKED_JOURNAL depuis sa création. Voir
        # rattraper_sync_baie_slot_ports() plus bas pour le rattrapage des
        # ports déjà créés avant ce correctif (ajouter le trigger ici ne
        # journalise que les écritures FUTURES).
        'baie_slot_ports': 'id',
        # baie_prises_murales : nouvelle table (voir migration ci-dessus),
        # suit le même besoin que baie_slot_ports — ajoutée au trigger dès
        # sa création, pas de rattrapage nécessaire pour les écritures
        # futures. Les lignes issues de la migration ci-dessus (données déjà
        # existantes) ont, elles, besoin d'un rattrapage séparé : voir
        # rattraper_sync_baie_prises_murales() plus bas.
        'baie_prises_murales': 'id',
        # Tables à clé texte (pas de colonne 'id')
        'config': 'cle', 'journal_maj': 'cle', 'collectes': 'cle',
        'cles_recuperation': 'cle',
    }
    # DROP + CREATE (pas IF NOT EXISTS) : garantit que la définition du trigger
    # correspond toujours au code, même après une mise à jour de cette liste sur
    # une base déjà initialisée (CREATE TRIGGER IF NOT EXISTS ne mettrait pas à
    # jour un trigger déjà présent avec une ancienne définition).
    for _t, _pk in _TRACKED_JOURNAL.items():
        c.execute(f"DROP TRIGGER IF EXISTS _trg_journal_ins_{_t}")
        c.execute(f"DROP TRIGGER IF EXISTS _trg_journal_upd_{_t}")
        c.execute(f"DROP TRIGGER IF EXISTS _trg_journal_del_{_t}")
        # DELETE puis INSERT plutôt qu'un seul INSERT OR REPLACE : signalé en
        # production, « UNIQUE constraint failed » sur _sync_journal côté
        # Turso alors que la contrainte est UNIQUE(...) ON CONFLICT REPLACE
        # ET que le trigger dit bien OR REPLACE — la contrainte est donc
        # correctement détectée, mais sa résolution REPLACE ne semble pas
        # s'appliquer de façon fiable une fois le trigger exécuté à distance
        # par Turso (écart de compatibilité SQLite/libSQL plausible plutôt
        # qu'une erreur de schéma : la table _sync_journal locale porte cette
        # clause ON CONFLICT REPLACE depuis sa toute première version). Deux
        # instructions simples plutôt qu'une résolution de conflit évite ce
        # risque, quelle qu'en soit la cause exacte.
        # INSERT trigger
        c.execute(f"""CREATE TRIGGER _trg_journal_ins_{_t}
            AFTER INSERT ON {_t}
            WHEN NOT EXISTS (SELECT 1 FROM _sync_applying) BEGIN
                DELETE FROM _sync_journal WHERE tbl='{_t}' AND record_id=NEW.{_pk} AND action='INSERT';
                INSERT INTO _sync_journal (tbl, record_id, action, timestamp)
                VALUES ('{_t}', NEW.{_pk}, 'INSERT', datetime('now'));
            END""")
        # UPDATE trigger
        c.execute(f"""CREATE TRIGGER _trg_journal_upd_{_t}
            AFTER UPDATE ON {_t}
            WHEN NOT EXISTS (SELECT 1 FROM _sync_applying) BEGIN
                DELETE FROM _sync_journal WHERE tbl='{_t}' AND record_id=NEW.{_pk} AND action='UPDATE';
                INSERT INTO _sync_journal (tbl, record_id, action, timestamp)
                VALUES ('{_t}', NEW.{_pk}, 'UPDATE', datetime('now'));
            END""")
        # DELETE trigger
        c.execute(f"""CREATE TRIGGER _trg_journal_del_{_t}
            AFTER DELETE ON {_t}
            WHEN NOT EXISTS (SELECT 1 FROM _sync_applying) BEGIN
                DELETE FROM _sync_journal WHERE tbl='{_t}' AND record_id=OLD.{_pk} AND action='DELETE';
                INSERT INTO _sync_journal (tbl, record_id, action, timestamp)
                VALUES ('{_t}', OLD.{_pk}, 'DELETE', datetime('now'));
            END""")

    # _sync_deletions / _trg_del_* : mécanisme legacy d'avant _sync_journal,
    # jamais relu par la synchronisation actuelle (voir _sync_using_journal
    # dans database.py, entièrement basée sur _sync_journal). Trouvé lors
    # d'un contrôle du système de synchronisation : ces triggers écrivaient
    # encore sur chaque suppression, sur 23 tables, pour une table que plus
    # rien ne lisait — et une liste restée figée à l'ancienne, jamais mise à
    # jour avec les tables ajoutées depuis (contrairement à _TRACKED_JOURNAL
    # ci-dessus). Supprimés plutôt que recréés ; DROP explicite (pas juste
    # l'absence de CREATE) pour retirer aussi ceux déjà posés sur les bases
    # existantes. La table _sync_deletions elle-même est laissée en place
    # (vide désormais) : un DROP TABLE n'apporterait rien et risquerait une
    # erreur sur un code qui la référencerait encore ailleurs sans qu'on l'ait vu.
    for _t in ('appareils', 'peripheriques', 'identifiants', 'contrats',
               'utilisateurs', 'services', 'clients', 'baie_slots',
               'outils', 'kb_articles', 'kb_categories',
               'documents_appareils', 'documents_contrats',
               'documents_peripheriques', 'baie_photos',
               'types_droits', 'droits_utilisateurs',
               'contrats_appareils', 'contrats_peripheriques',
               'peripheriques_appareils', 'parc_general', 'historique', 'plans'):
        c.execute(f"DROP TRIGGER IF EXISTS _trg_del_{_t}")

    # Migration : ajouter colonnes BLOB + sync si n'existent pas
    for table in ['documents_appareils', 'documents_contrats', 'documents_peripheriques',
                  'documents_interventions', 'baie_photos']:
        try:
            c.execute(f'ALTER TABLE {table} ADD COLUMN contenu_blob BLOB')
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        try:
            c.execute(f"ALTER TABLE {table} ADD COLUMN sync_status TEXT DEFAULT 'local'")
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        try:
            c.execute(f"ALTER TABLE {table} ADD COLUMN date_sync TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà

    # Anti-collision d'ID multi-machines (sync bidirectionnelle Turso) : deux
    # installations indépendantes qui créent chacune un nouvel enregistrement
    # démarrent TOUTES LES DEUX leur compteur AUTOINCREMENT à 1. La sync applique
    # ensuite un UPSERT sur cet id partagé par coïncidence : la seconde machine à
    # synchroniser écrase silencieusement l'enregistrement de la première (confirmé :
    # perte de données à la fois dans Turso et localement sur la première machine
    # dès son cycle de sync suivant). On attribue donc à chaque machine un point de
    # départ aléatoire (espace ~2^48) pour chaque table suivie, une seule fois tant
    # que la table est encore vide - n'affecte jamais une table déjà peuplée.
    try:
        c.execute("SELECT valeur FROM config WHERE cle='_sync_id_offset'")
        _offset_row = c.fetchone()
        if _offset_row and _offset_row[0]:
            _id_offset = int(_offset_row[0])
        else:
            import secrets as _secrets
            _id_offset = _secrets.randbits(48)
            c.execute("INSERT OR REPLACE INTO config (cle, valeur, date_maj) VALUES ('_sync_id_offset', ?, ?)",
                      (str(_id_offset), _utcnow().isoformat()))

        for _tbl, _pk in _TRACKED_JOURNAL.items():
            if _pk != 'id':
                continue  # seules les tables à clé entière 'id' sont concernées (pas 'config', clé 'cle')
            try:
                _count = c.execute(f"SELECT COUNT(*) FROM [{_tbl}]").fetchone()[0]
                _has_seq = c.execute("SELECT 1 FROM sqlite_sequence WHERE name=?", (_tbl,)).fetchone()
                if _count == 0 and not _has_seq:
                    c.execute("INSERT INTO sqlite_sequence (name, seq) VALUES (?, ?)", (_tbl, _id_offset))
            except Exception:
                pass
    except Exception:
        logger.exception("Anti-collision ID (sqlite_sequence) - échec non bloquant")

    # Migration : colonnes pour système auto-remplissage (v2.6.24)
    # rapport_systeme_json (v2.6.30) stocke le snapshot complet du collecteur :
    # sans lui, 90 % des données collectées n'existaient que dans le PDF joint
    cols_app3 = [r[1] for r in conn.execute('PRAGMA table_info(appareils)').fetchall()]
    for col, defval in [
        ('antivirus', "TEXT DEFAULT ''"),
        ('logiciels_installes_json', "TEXT DEFAULT '[]'"),
        ('derniere_synchro', "TEXT DEFAULT NULL"),
        ('rapport_systeme_json', "TEXT DEFAULT ''"),
    ]:
        if col not in cols_app3:
            try:
                c.execute(f"ALTER TABLE appareils ADD COLUMN {col} {defval}")
            except Exception:
                pass

    # Migration : identité USB des périphériques créés par le collecteur.
    # Sans cette colonne, chaque collecte recréerait les mêmes périphériques.
    cols_periph = [r[1] for r in conn.execute('PRAGMA table_info(peripheriques)').fetchall()]
    if 'source_usb_id' not in cols_periph:
        try:
            c.execute("ALTER TABLE peripheriques ADD COLUMN source_usb_id TEXT DEFAULT ''")
        except Exception:
            pass

    # TABLES DIAGNOSTIC RÉSEAU (module network_diag)
    c.execute('''CREATE TABLE IF NOT EXISTS diag_reseau_evenements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        horodatage TEXT DEFAULT '',
        gravite TEXT DEFAULT 'info',
        categorie TEXT DEFAULT '',
        titre TEXT DEFAULT '',
        details_json TEXT DEFAULT '{}',
        source TEXT DEFAULT 'actif',
        signature TEXT DEFAULT '',
        appareil_id INTEGER,
        resolu INTEGER DEFAULT 0,
        date_resolu TEXT,
        premiere_occurrence TEXT DEFAULT '',
        derniere_occurrence TEXT DEFAULT '',
        nb_occurrences INTEGER DEFAULT 1,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_diag_evt
        ON diag_reseau_evenements(client_id, resolu, horodatage)''')
    if 'appareil_id' not in [r[1] for r in c.execute('PRAGMA table_info(diag_reseau_evenements)').fetchall()]:
        try:
            c.execute('ALTER TABLE diag_reseau_evenements ADD COLUMN appareil_id INTEGER')
        except Exception:
            pass
    c.execute('''CREATE TABLE IF NOT EXISTS diag_reseau_runs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        debut TEXT DEFAULT '', fin TEXT DEFAULT '', duree_s INTEGER DEFAULT 0,
        mode TEXT DEFAULT 'snapshot', plage TEXT DEFAULT '',
        capture_utilisee INTEGER DEFAULT 0,
        resume_json TEXT DEFAULT '{}',
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')
    # Palier 3 — relevés SNMP par port (compteurs cumulatifs : deux relevés
    # sont nécessaires pour calculer un taux / un delta).
    c.execute('''CREATE TABLE IF NOT EXISTS diag_snmp_releves (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL,
        appareil_id INTEGER,
        equipement_ip TEXT DEFAULT '',
        port_index INTEGER DEFAULT 0,
        port_nom TEXT DEFAULT '',
        horodatage TEXT DEFAULT '',
        epoch REAL DEFAULT 0,
        compteurs_json TEXT DEFAULT '{}',
        duplex INTEGER DEFAULT 0,
        speed_mbps INTEGER DEFAULT 0,
        oper_status INTEGER DEFAULT 0,
        FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE)''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_snmp_releves
        ON diag_snmp_releves(client_id, equipement_ip, port_index, epoch)''')

    # Client par défaut si aucun
    if not c.execute('SELECT id FROM clients').fetchone():
        now = _utcnow().isoformat()
        c.execute("INSERT INTO clients (nom, date_creation, date_maj) VALUES ('Mon Client', ?, ?)", (now, now))
        cid = c.lastrowid
        c.execute("INSERT INTO parc_general (client_id, nom_site, date_maj) VALUES (?, 'Mon Parc Informatique', ?)", (cid, now))

    conn.commit(); conn.close()

# init_db() est appelé ici en mode dev (python app.py)
# En mode launcher/PyInstaller, le launcher le gère après avoir surchargé DATABASE
# En mode normal (python app.py ou import dev), appeler init_db()
# En mode launcher, le launcher le fait après avoir surchargé DATABASE
import inspect as _inspect
_called_from_launcher = any('launcher' in f.filename for f in _inspect.stack())
if not _called_from_launcher:
    init_db()


# ════════════════════════════════════════════════════════════════════════════
# SAUVEGARDE AUTOMATIQUE DE LA BASE
# ════════════════════════════════════════════════════════════════════════════

BACKUP_DIR = os.path.join(_data_base, 'backups')
# Nombre de sauvegardes conservées. Au-delà, les plus anciennes sont supprimées.
BACKUP_KEEP = 3
# Intervalle entre deux sauvegardes automatiques.
BACKUP_INTERVAL_HOURS = 24
_backup_lock = threading.Lock()


def _backup_files():
    """Sauvegardes existantes, de la plus récente à la plus ancienne."""
    try:
        fichiers = [
            os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR)
            if f.startswith('parc_info_') and f.endswith('.db')
        ]
    except OSError:
        return []
    return sorted(fichiers, key=lambda f: os.path.getmtime(f), reverse=True)


def creer_sauvegarde(raison='automatique'):
    """Copie cohérente de la base, puis rotation des anciennes.

    La copie passe par l'API `backup` de SQLite et non par un copier-coller de
    fichier : la base est en mode WAL et une copie brute pendant une écriture
    donnerait une sauvegarde tronquée, inutilisable au moment où elle servirait.

    Retourne (chemin, erreur) — l'un des deux vaut None.
    """
    # Une seule sauvegarde à la fois : le déclencheur périodique et une demande
    # manuelle peuvent se présenter en même temps.
    if not _backup_lock.acquire(blocking=False):
        return (None, 'une sauvegarde est déjà en cours')
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        horodatage = _utcnow().strftime('%Y%m%d_%H%M%S')
        destination = os.path.join(BACKUP_DIR, f'parc_info_{horodatage}.db')
        # L'horodatage est à la seconde : deux sauvegardes rapprochées portaient
        # le même nom et la seconde écrasait la première. C'était grave au moment
        # de restaurer — le filet de sécurité pris juste avant remplaçait la
        # sauvegarde qu'on s'apprêtait à recharger.
        suffixe = 1
        while os.path.exists(destination):
            destination = os.path.join(BACKUP_DIR, f'parc_info_{horodatage}_{suffixe}.db')
            suffixe += 1

        source = sqlite3.connect(DATABASE)
        try:
            cible = sqlite3.connect(destination)
            try:
                source.backup(cible)
            finally:
                cible.close()
        finally:
            source.close()

        taille = os.path.getsize(destination)
        # Rotation : ne garder que les BACKUP_KEEP plus récentes.
        supprimees = []
        for ancienne in _backup_files()[BACKUP_KEEP:]:
            try:
                os.remove(ancienne)
                supprimees.append(os.path.basename(ancienne))
            except OSError:
                logger.warning('Sauvegarde non supprimée : %s', ancienne)

        logger.info('Sauvegarde %s : %s (%s)%s', raison, os.path.basename(destination),
                    human_size(taille),
                    ' — supprimé %s' % ', '.join(supprimees) if supprimees else '')
        try:
            from database import log_sync_event
            log_sync_event('sauvegarde', 'ok',
                           'Sauvegarde %s : %s' % (raison, human_size(taille)),
                           {'fichier': os.path.basename(destination),
                            'supprimees': supprimees})
        except Exception:
            # Le journal est un confort : son absence ne doit pas faire échouer
            # une sauvegarde qui, elle, a réussi.
            pass
        return (destination, None)
    except Exception as exc:
        logger.exception('Échec de la sauvegarde automatique')
        return (None, str(exc))
    finally:
        _backup_lock.release()


def _boucle_sauvegarde():
    """Déclenche une sauvegarde au démarrage puis à intervalle régulier."""
    # Au démarrage : une sauvegarde n'est faite que si la dernière date de plus
    # d'un intervalle, pour ne pas en créer une à chaque redémarrage et faire
    # tourner la rotation jusqu'à perdre les sauvegardes utiles.
    while True:
        try:
            recentes = _backup_files()
            derniere = os.path.getmtime(recentes[0]) if recentes else 0
            age_heures = (time.time() - derniere) / 3600
            if age_heures >= BACKUP_INTERVAL_HOURS:
                creer_sauvegarde('automatique')
        except Exception:
            logger.exception('Boucle de sauvegarde')
        time.sleep(3600)


def demarrer_sauvegardes():
    """Lance le thread de sauvegarde, sauf si la fonction est désactivée."""
    if str(os.environ.get('PARCINFO_BACKUP', '1')).lower() in ('0', 'false', 'no'):
        logger.info('Sauvegarde automatique désactivée (PARCINFO_BACKUP)')
        return
    fil = threading.Thread(target=_boucle_sauvegarde, daemon=True,
                           name='SauvegardeBase')
    fil.start()


# ─── CONFIGURATION GLOBALE ───────────────────────────────────────────────────

_TYPE_CSS_DEFAULTS = {
    'pc':'#00c9ff','linux':'#4ade80','laptop':'#60a5fa','mac':'#e2e8f0',
    'serveur':'#c084fc','imprimante':'#f97316','switch':'#facc15',
    'routeur':'#ff3355','nas':'#4ade80','tel':'#34d399',
    'tablette':'#a78bfa','camera':'#fb923c','wifi':'#2dd4bf','autre':'#94a3b8',
}

@app.context_processor
def inject_cfg():
    types = get_liste_cached('types_appareils')
    user = get_auth_user()
    auth_user_id = user['id'] if user else None
    # Fusionner config globale + préférences personnelles de l'utilisateur
    cfg = cfg_all(auth_user_id=auth_user_id)
    # Calcule les labels courts pour les badges de types (pour les templates)
    type_badges = {}
    for t in types:
        k = type_css_filter(t)
        if k not in type_badges:
            val = cfg.get(f'type_badge_{k}', '')
            if val:
                type_badges[k] = val[:3].upper()
            else:
                type_badges[k] = _TYPE_BADGE_DEFAULTS.get(k, k[:3].upper() or 'AUT')
    return {
        'cfg': cfg,
        'types_appareils_ctx': types,
        'type_css_defaults': _TYPE_CSS_DEFAULTS,
        'type_badge_defaults': _TYPE_BADGE_DEFAULTS,
        'type_badges': type_badges,
        'app_version': APP_VERSION,
    }

# ─── REGISTER UPDATE NOTIFICATION ROUTES ──────────────────────────────────────
register_update_routes(app)

@app.route('/api/config', methods=['GET'])
@login_required
def api_config_get():
    user = get_auth_user()
    auth_user_id = user['id'] if user else None
    # Retourne config globale + préférences personnelles de l'utilisateur
    return jsonify(cfg_all(auth_user_id=auth_user_id))

@app.route('/api/config', methods=['POST'])
@login_required
def api_config_save():
    from config_helpers import cfg_set_batch
    user = get_auth_user()
    auth_user_id = user['id'] if user else None

    data = request.json or {}
    old_db_type = cfg_get('db_type', auth_user_id=auth_user_id)  # Récupérer avant les modifications

    # Filtrer et valider les clés à sauvegarder
    valid_config = {}
    for k, v in data.items():
        if (k in CFG_DEFAULTS
                or k.startswith('port_color_')  # Anciennes clés (par serviceType): port_color_ssh, port_color_http, etc.
                or k.startswith('port_icon_')   # Anciennes clés (par serviceType): port_icon_ssh, port_icon_http, etc.
                or (k.startswith('port_') and k.endswith(('_name', '_description', '_color', '_icon')))  # Nouvelles clés (par numéro de port): port_22_color, port_5000_icon, etc.
                or k.startswith('periph_color_')
                or k.startswith('type_color_')
                or k.startswith('type_badge_')
                or k.startswith('type_desc_')  # Nouveau: descriptions de types
                or k == 'mode'):
            # Special validation for dashboard_widgets_size (Phase 9)
            if k == 'dashboard_widgets_size':
                # Ensure it's valid JSON before saving
                try:
                    sizes = json.loads(str(v))
                    # Validate all sizes are valid
                    for widget_id, size in sizes.items():
                        if size not in ('small', 'medium', 'large'):
                            logger.warning(f"Invalid widget size: {widget_id}={size}")
                            return jsonify({'error': 'Invalid widget size value'}), 400
                    valid_config[k] = json.dumps(sizes)  # Re-serialize to ensure valid JSON
                except (json.JSONDecodeError, ValueError) as e:
                    logger.warning(f"Invalid dashboard_widgets_size JSON: {str(e)}")
                    return jsonify({'error': 'Invalid JSON in dashboard_widgets_size'}), 400
            # Special validation for dashboard_widgets_height (5 levels: xs, s, m, l, xl)
            elif k == 'dashboard_widgets_height':
                # Ensure it's valid JSON before saving
                try:
                    heights = json.loads(str(v))
                    # Valid height levels (5 levels for more precision)
                    valid_heights = {'xs', 's', 'm', 'l', 'xl', 'compact', 'normal', 'tall'}  # Include legacy names
                    # Map legacy names to new ones for backwards compatibility
                    legacy_map = {'compact': 's', 'normal': 'm', 'tall': 'l'}

                    # Validate all heights are valid
                    for widget_id, height in heights.items():
                        if height not in valid_heights:
                            logger.warning(f"Invalid widget height: {widget_id}={height}")
                            return jsonify({'error': 'Invalid widget height value'}), 400
                        # Convert legacy names to new ones
                        if height in legacy_map:
                            heights[widget_id] = legacy_map[height]

                    valid_config[k] = json.dumps(heights)  # Re-serialize to ensure valid JSON
                except (json.JSONDecodeError, ValueError) as e:
                    logger.warning(f"Invalid dashboard_widgets_height JSON: {str(e)}")
                    return jsonify({'error': 'Invalid JSON in dashboard_widgets_height'}), 400
            else:
                valid_config[k] = str(v)

    # ✓ Sauvegarder tout en UNE SEULE transaction (beaucoup plus rapide)
    # Les clés personnelles iront dans user_preferences, les autres dans config
    if valid_config:
        cfg_set_batch(valid_config, auth_user_id=auth_user_id)

    cfg_invalidate()

    # Invalider le cache crypto si les credentials Turso ont changé
    if any(k in valid_config for k in ('turso_url', 'turso_token', 'db_type')):
        _invalidate_crypto_cache()

    # ✓ Optimisation: ne reconfigurer la sync que si db_type a changé
    new_db_type = cfg_get('db_type', auth_user_id=auth_user_id)
    if old_db_type != new_db_type:
        _handle_sync_config()

    return jsonify({'ok': True})

@app.route('/api/config/reset', methods=['POST'])
@login_required
def api_config_reset():
    conn = get_db()
    conn.execute('DELETE FROM config')
    conn.commit(); conn.close()
    cfg_invalidate()
    return jsonify({'ok': True})


@app.route('/api/db/test', methods=['POST'])
@login_required
def api_db_test():
    data  = request.get_json() or {}
    url   = (data.get('url',   '') or '').strip()
    token = (data.get('token', '') or '').strip()
    if not url or not token:
        return jsonify({'ok': False, 'message': 'URL et token requis'})
    from database import test_turso
    ok, msg = test_turso(url, token)
    return jsonify({'ok': ok, 'message': msg})


@app.route('/api/db/transfer', methods=['POST'])
@login_required
def api_db_transfer():
    if not can_write():
        return jsonify({'ok': False, 'error': 'Accès en lecture seule'})
    data      = request.get_json() or {}
    direction = data.get('direction', '')
    if direction not in ('local_to_turso', 'turso_to_local'):
        return jsonify({'ok': False, 'error': 'Direction invalide'})
    url   = cfg_get('turso_url',   '').strip()
    token = cfg_get('turso_token', '').strip()
    if not url or not token:
        return jsonify({'ok': False, 'error': 'Turso non configuré'})
    from database import TursoConnection, test_turso, migrate_db, get_local_db
    ok, msg = test_turso(url, token)
    if not ok:
        return jsonify({'ok': False, 'error': f'Connexion impossible: {msg}'})
    local_conn = get_local_db()
    turso_conn = TursoConnection(url, token)
    try:
        source = local_conn if direction == 'local_to_turso' else turso_conn
        target = turso_conn if direction == 'local_to_turso' else local_conn
        ok, stats, error = migrate_db(source, target)
        local_conn.close()
        if ok:
            total    = sum(stats.values())
            n_tables = sum(1 for v in stats.values() if v > 0)
            return jsonify({'ok': True,
                            'summary': f'{total} enregistrements sur {n_tables} tables',
                            'stats': stats})
        return jsonify({'ok': False, 'error': (error or '')[:500]})
    except Exception as e:
        local_conn.close()
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/db/sauvegarde', methods=['GET', 'POST'])
@login_required
def api_db_sauvegarde():
    """Liste les sauvegardes (GET) ou en déclenche une (POST, administrateur)."""
    user = get_auth_user()
    if request.method == 'GET':
        sauvegardes = []
        for chemin in _backup_files():
            try:
                sauvegardes.append({
                    'fichier': os.path.basename(chemin),
                    'taille': human_size(os.path.getsize(chemin)),
                    'date': datetime.utcfromtimestamp(
                        os.path.getmtime(chemin)).isoformat(timespec='seconds'),
                })
            except OSError:
                continue
        return jsonify({'sauvegardes': sauvegardes, 'conservees': BACKUP_KEEP,
                        'intervalle_heures': BACKUP_INTERVAL_HOURS})

    # Une sauvegarde copie l'intégralité des données de tous les clients :
    # la déclencher reste une opération d'administration.
    if not user or user.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Réservé aux administrateurs'}), 403

    chemin, erreur = creer_sauvegarde('manuelle')
    if erreur:
        return jsonify({'ok': False, 'error': erreur}), 500
    return jsonify({'ok': True, 'fichier': os.path.basename(chemin),
                    'taille': human_size(os.path.getsize(chemin))})


@app.route('/api/db/sauvegarde/restaurer', methods=['POST'])
@login_required
def api_db_restaurer():
    """Remet la base dans l'état d'une sauvegarde (administrateur).

    Restaurer efface les données actuelles : une sauvegarde de sécurité est
    prise d'abord, sans quoi une erreur de choix serait sans retour.
    """
    user = get_auth_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Réservé aux administrateurs'}), 403

    demande = (request.json or {}).get('fichier') or request.form.get('fichier') or ''
    # Seul le nom de fichier est accepté, et il doit désigner une sauvegarde
    # existante : sans ce filtre, un chemin relatif permettrait de charger
    # n'importe quel fichier de la machine à la place de la base.
    nom = os.path.basename(demande)
    source = os.path.join(BACKUP_DIR, nom)
    if not nom or source not in _backup_files():
        return jsonify({'ok': False, 'error': 'Sauvegarde introuvable'}), 404

    filet, erreur_filet = creer_sauvegarde('avant restauration')
    if erreur_filet:
        return jsonify({'ok': False,
                        'error': "Restauration annulée : la sauvegarde de "
                                 "sécurité a échoué (%s)" % erreur_filet}), 500

    try:
        # L'API backup de SQLite écrit dans la base en place, verrou compris :
        # remplacer le fichier pendant que l'application tourne laisserait les
        # connexions ouvertes sur l'ancien inode, et la base en mode WAL.
        origine = sqlite3.connect(source)
        try:
            cible = sqlite3.connect(DATABASE)
            try:
                origine.backup(cible)
            finally:
                cible.close()
        finally:
            origine.close()
    except Exception as exc:
        logger.exception('Restauration impossible depuis %s', nom)
        return jsonify({'ok': False, 'error': str(exc)}), 500

    cfg_invalidate()
    logger.warning('Base restaurée depuis %s par %s', nom, user.get('login'))
    try:
        from database import log_sync_event
        log_sync_event('restauration', 'ok', 'Base restaurée depuis %s' % nom,
                       {'filet': os.path.basename(filet) if filet else None,
                        'par': user.get('login')})
    except Exception:
        pass

    return jsonify({'ok': True, 'fichier': nom,
                    'filet': os.path.basename(filet) if filet else None})


@app.route('/api/db/sync', methods=['GET', 'POST'])
@login_required
def api_db_sync():
    from database import sync_once, get_sync_state
    if request.method == 'GET':
        return jsonify(get_sync_state())
    if not can_write():
        return jsonify({'ok': False, 'error': 'Accès en lecture seule'})
    ok, stats, error = sync_once()
    _log_sync_errors(error)
    state = get_sync_state()
    state['ok']    = ok
    state['error'] = error
    return jsonify(state)


@app.route('/journal-synchronisation')
@login_required
def journal_synchronisation():
    """Journal : mises à jour de l'application, puis cycles de synchronisation.

    Les mises à jour sont lues dans une table synchronisée : ce poste voit donc
    aussi celles appliquées sur les autres installations. Les cycles de
    synchronisation, eux, restent locaux — ils décrivent ce que CETTE instance a
    échangé, et n'auraient pas de sens répliqués.
    """
    from database import (get_local_db, get_sync_state,
                          creer_journal_synchronisation, creer_journal_maj)
    conn = get_local_db()
    try:
        # Définitions tenues par database.py : les dupliquer ici laissait deux
        # copies libres de diverger sans que rien ne le signale.
        creer_journal_synchronisation(conn)
        creer_journal_maj(conn)
        conn.commit()
        entries = [row_to_dict(r) for r in conn.execute(
            "SELECT * FROM journal_synchronisation ORDER BY id DESC LIMIT 200"
        ).fetchall()]
        majs = [row_to_dict(r) for r in conn.execute(
            "SELECT * FROM journal_maj ORDER BY horodatage DESC LIMIT 100"
        ).fetchall()]
    finally:
        conn.close()

    for e in entries:
        try:
            e['details_parsed'] = json.loads(e['details']) if e.get('details') else None
        except Exception:
            e['details_parsed'] = None

    machine_locale = socket.gethostname()
    for m in majs:
        m['locale'] = (m.get('machine') == machine_locale)

    return render_template('journal_synchronisation.html', entries=entries, majs=majs,
                           machine_locale=machine_locale, sync_state=get_sync_state())


# ─── THREAD DE SYNCHRONISATION TURSO ─────────────────────────────────────────

_sync_thread: threading.Thread | None = None
_sync_stop   = threading.Event()

_last_sync_error_logged: str | None = None

def _log_sync_errors(error_msg: str | None):
    """Enregistre une erreur de sync dans le journal (client_id=0 = entrée système).
    Réinitialise la déduplication si la sync réussit (error_msg is None).
    """
    global _last_sync_error_logged
    if not error_msg:
        _last_sync_error_logged = None
        return
    if error_msg == _last_sync_error_logged:
        return
    _last_sync_error_logged = error_msg
    try:
        import json as _j
        from database import get_local_db
        from client_helpers import log_history as _lh
        conn = get_local_db()
        details = _j.dumps({'message': error_msg[:800]}, ensure_ascii=False)
        _lh(conn, 0, 'système', 0, 'Synchronisation Turso', 'Erreur', details)
        conn.commit()
        conn.close()
    except Exception:
        pass


def _bg_sync_loop():
    """Boucle de synchronisation bidirectionnelle en arrière-plan."""
    from database import sync_once
    while not _sync_stop.is_set():
        try:
            if cfg_get('db_type') == 'sync':
                ok, stats, error = sync_once()
                _log_sync_errors(error)
            else:
                break   # Mode sync désactivé : on arrête le thread
        except Exception:
            pass
        try:
            interval = int(cfg_get('db_sync_interval', '30'))
        except Exception:
            interval = 30
        _sync_stop.wait(timeout=max(5, interval))


def _start_sync_thread():
    global _sync_thread, _sync_stop
    if _sync_thread and _sync_thread.is_alive():
        return
    _sync_stop.clear()
    _sync_thread = threading.Thread(target=_bg_sync_loop, daemon=True, name='turso-sync')
    _sync_thread.start()
    logger.info('Thread de synchronisation Turso démarré (intervalle=%ss)',
                cfg_get('db_sync_interval', '30'))


def _stop_sync_thread():
    global _sync_thread
    _sync_stop.set()
    if _sync_thread:
        _sync_thread.join(timeout=3)
        _sync_thread = None
    logger.info('Thread de synchronisation Turso arrêté')


def _handle_sync_config():
    """Démarre ou arrête le thread de sync selon la config db_type."""
    # Vérifier si la sync Turso est explicitement désactivée
    if os.environ.get('DISABLE_TURSO_SYNC', '0') == '1':
        logger.info('Sync Turso désactivée (DISABLE_TURSO_SYNC=1)')
        _stop_sync_thread()
        return
    if cfg_get('db_type') == 'sync':
        _stop_sync_thread()   # redémarre avec le nouvel intervalle éventuel
        _start_sync_thread()
    else:
        _stop_sync_thread()


_sync_init_done = False

@app.before_request
def _auto_start_sync():
    """Démarre le thread de sync au premier appel si le mode sync est actif."""
    global _sync_init_done
    if not _sync_init_done:
        _sync_init_done = True
        # Vérifier si la sync Turso est explicitement désactivée
        if os.environ.get('DISABLE_TURSO_SYNC', '0') == '1':
            logger.info('Sync Turso désactivée (DISABLE_TURSO_SYNC=1)')
            return
        if cfg_get('db_type') == 'sync':
            _start_sync_thread()


# ─── OUTILS ──────────────────────────────────────────────────────────────────

@app.route('/outils')
@login_required
def page_outils():
    conn = get_db()
    outils = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM outils ORDER BY categorie, ordre, nom").fetchall()]
    # Group by categorie
    cats = {}
    for o in outils:
        c = o['categorie'] or 'Général'
        cats.setdefault(c, []).append(o)
    conn.close()
    cid = get_client_id()
    return render_template('outils.html', outils=outils, cats=cats,
                           clients=get_clients(), client_actif_id=cid)

@app.route('/api/outils', methods=['GET'])
@login_required
def api_outils_get():
    conn = get_db()
    rows = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM outils ORDER BY categorie, ordre, nom").fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/outils/ajouter', methods=['POST'])
@login_required
def api_outils_ajouter():
    d = request.json or {}
    nom  = d.get('nom','').strip()
    url  = d.get('url','').strip()
    if not nom or not url: return jsonify({'error':'Nom et URL requis'}), 400
    if not url.startswith('http'): url = 'https://' + url
    conn = get_db()
    ordre = conn.execute('SELECT COALESCE(MAX(ordre),0)+1 FROM outils').fetchone()[0]
    conn.execute('INSERT INTO outils (nom,url,description,categorie,icone,ordre,date_maj) VALUES (?,?,?,?,?,?,?)',
        (nom, url, d.get('description',''), d.get('categorie','Général'),
         d.get('icone','🔧'), ordre, _utcnow().isoformat()))
    conn.commit()
    outils = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM outils ORDER BY categorie, ordre, nom").fetchall()]
    conn.close()
    return jsonify({'ok': True, 'outils': outils})

@app.route('/api/outils/<int:id>/supprimer', methods=['POST'])
@login_required
def api_outils_supprimer(id):
    conn = get_db()
    conn.execute('DELETE FROM outils WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/outils/<int:id>/toggle', methods=['POST'])
@login_required
def api_outils_toggle(id):
    conn = get_db()
    conn.execute('UPDATE outils SET actif = 1 - actif WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ─── GESTION CLIENTS ─────────────────────────────────────────────────────────

@app.route('/clients')
@login_required
def liste_clients():
    clients = get_clients()   # filtre par user + champ acces
    conn = get_db()
    for cl in clients:
        cl['nb_appareils'] = conn.execute('SELECT COUNT(*) FROM appareils WHERE client_id=?', (cl['id'],)).fetchone()[0]
        cl['nb_actifs']    = conn.execute('SELECT COUNT(*) FROM appareils WHERE client_id=? AND en_ligne=1', (cl['id'],)).fetchone()[0]
    conn.close()
    return render_template('clients.html', clients=clients, client_actif_id=get_client_id())

@app.route('/client/nouveau', methods=['GET','POST'])
@login_required
def nouveau_client():
    if request.method == 'POST':
        f = request.form
        now = _utcnow().isoformat()
        uid = session.get('auth_user_id')
        conn = get_db()
        c = conn.execute(
            "INSERT INTO clients (nom,contact,telephone,email,adresse,notes,couleur,auth_user_id,collecteur_token,date_creation,date_maj) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (f.get('nom','Nouveau client'), f.get('contact',''), f.get('telephone',''),
             f.get('email',''), f.get('adresse',''), f.get('notes',''),
             f.get('couleur','#00c9ff'), uid, f.get('collecteur_token','').strip(), now, now))
        cid = c.lastrowid
        conn.execute("INSERT INTO parc_general (client_id, nom_site, date_maj) VALUES (?,?,?)",
                     (cid, f.get('nom','Nouveau client'), now))
        conn.commit(); conn.close()
        session['client_id'] = cid
        flash(f"Client « {f.get('nom')} » créé avec succès", 'success')
        return redirect(url_for('index'))
    return render_template('form_client.html', client=None, action='Nouveau')

@app.route('/client/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_client(id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE clients SET nom=?,contact=?,telephone=?,email=?,adresse=?,notes=?,couleur=?,collecteur_token=?,date_maj=?
            WHERE id=?''', (f.get('nom',''), f.get('contact',''), f.get('telephone',''),
             f.get('email',''), f.get('adresse',''), f.get('notes',''),
             f.get('couleur','#00c9ff'), f.get('collecteur_token','').strip(),
             _utcnow().isoformat(), id))
        conn.commit(); conn.close()
        flash('Client mis à jour', 'success')
        return redirect(url_for('liste_clients'))
    cl = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (id,)).fetchone())
    conn.close()
    return render_template('form_client.html', client=cl, action='Modifier')

@app.route('/client/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_client(id):
    conn = get_db()
    conn.execute('PRAGMA foreign_keys = ON')
    nom = row_to_dict(conn.execute('SELECT nom FROM clients WHERE id=?', (id,)).fetchone() or {}).get('nom','')
    conn.execute('DELETE FROM clients WHERE id=?', (id,))
    conn.commit(); conn.close()
    if session.get('client_id') == id:
        session.pop('client_id', None)
    flash(f"Client « {nom} » supprimé avec toutes ses données", 'info')
    return redirect(url_for('liste_clients'))

@app.route('/client/<int:id>/selectionner')
@login_required
def selectionner_client(id):
    """
    Sélectionne un client et redirige vers son dashboard.
    Fonctionne depuis n'importe quelle page (user_dashboard ou dashboard d'un autre client).
    """
    session['client_id'] = id
    # Toujours rediriger vers le dashboard du client sélectionné
    # pour une UX cohérente (quand on sélectionne un client, on voit son dashboard)
    return redirect(url_for('client_dashboard_view', cid=id))

# ─── ROUTES PRINCIPALES ──────────────────────────────────────────────────────

def _compute_client_dashboard_stats(conn, cid, today):
    """
    Calcule les statistiques du dashboard pour un client spécifique.
    Retourne un dictionnaire avec tous les compteurs et données agrégées.
    """
    appareils = fmt_appareils([row_to_dict(r) for r in retry_db_query(lambda: conn.execute(
        'SELECT * FROM appareils WHERE client_id=? ORDER BY adresse_ip', (cid,)).fetchall())])

    # Compteurs principaux
    nb_en_ligne   = sum(1 for a in appareils if a.get('en_ligne'))
    nb_hors_ligne = sum(1 for a in appareils if not a.get('en_ligne') and a.get('statut') == 'actif')
    nb_garantie   = sum(1 for a in appareils if a.get('garantie_active'))
    nb_periph     = retry_db_query(lambda: conn.execute('SELECT COUNT(*) FROM peripheriques WHERE client_id=?', (cid,)).fetchone()[0])
    nb_contrats   = retry_db_query(lambda: conn.execute("SELECT COUNT(*) FROM contrats WHERE client_id=? AND statut='actif'", (cid,)).fetchone()[0])
    nb_identifiants = retry_db_query(lambda: conn.execute('SELECT COUNT(*) FROM identifiants WHERE client_id=?', (cid,)).fetchone()[0])

    # Répartition par type
    repartition_rows = retry_db_query(lambda: conn.execute(
        "SELECT type_appareil, COUNT(*) as nb FROM appareils WHERE client_id=? GROUP BY type_appareil ORDER BY nb DESC",
        (cid,)).fetchall())
    repartition = [{'type': r[0] or 'Autre', 'nb': r[1]} for r in repartition_rows]

    # Statistiques périphériques
    periph_stats = {
        'actif': conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='actif'", (cid,)).fetchone()[0],
        'stock': conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='stock'", (cid,)).fetchone()[0],
        'hs':    conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='hors_service'", (cid,)).fetchone()[0],
    }

    # Calculs dérivés
    nb_app_total = len(appareils)
    taux_dispo   = round(nb_en_ligne / nb_app_total * 100) if nb_app_total else 0

    # Montant annuel des contrats
    montant_annuel = 0.0
    _period_map = {'mensuel':12,'trimestriel':4,'semestriel':2,'annuel':1,'pluriannuel':0.5,'unique':0}
    for ct in conn.execute("SELECT montant_ht,periodicite FROM contrats WHERE client_id=? AND statut='actif'", (cid,)).fetchall():
        try:
            if ct[0]: montant_annuel += float(ct[0]) * _period_map.get(ct[1] or 'annuel', 1)
        except: pass

    # Graphiques
    types_chart = [(r['type'], r['nb']) for r in repartition]
    _p_cats = conn.execute(
        'SELECT categorie, COUNT(*) as n FROM peripheriques WHERE client_id=? GROUP BY categorie ORDER BY n DESC',
        (cid,)).fetchall()
    periph_chart = [(r[0], r[1]) for r in _p_cats]

    # Appareils récents
    recents = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM appareils WHERE client_id=? ORDER BY date_maj DESC LIMIT 5", (cid,)).fetchall()]

    # Derniers périphériques et contrats
    derniers_periph = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM peripheriques WHERE client_id=? ORDER BY date_creation DESC LIMIT 3', (cid,)).fetchall()]
    derniers_contrats = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM contrats WHERE client_id=? ORDER BY date_creation DESC LIMIT 3', (cid,)).fetchall()]

    # Nombre d'utilisateurs
    nb_users = conn.execute('SELECT COUNT(*) FROM utilisateurs WHERE client_id=?', (cid,)).fetchone()[0]

    # Nombre hors service
    nb_hors_service = conn.execute("SELECT COUNT(*) FROM appareils WHERE client_id=? AND statut='hors_service'", (cid,)).fetchone()[0]

    # Activités récentes
    hist_recent = []
    try:
        hist_recent = [row_to_dict(r) for r in conn.execute(
            "SELECT entite, entite_nom, action, date_action FROM historique "
            "WHERE client_id=? ORDER BY id DESC LIMIT 6", (cid,)).fetchall()]
    except Exception:
        pass

    return {
        'appareils': appareils,
        'nb_en_ligne': nb_en_ligne,
        'nb_hors_ligne': nb_hors_ligne,
        'nb_garantie': nb_garantie,
        'nb_periph': nb_periph,
        'nb_contrats': nb_contrats,
        'nb_identifiants': nb_identifiants,
        'repartition': repartition,
        'periph_stats': periph_stats,
        'nb_app_total': nb_app_total,
        'taux_dispo': taux_dispo,
        'montant_annuel': montant_annuel,
        'types_chart': types_chart,
        'periph_chart': periph_chart,
        'recents': recents,
        'derniers_periph': derniers_periph,
        'derniers_contrats': derniers_contrats,
        'nb_users': nb_users,
        'nb_hors_service': nb_hors_service,
        'hist_recent': hist_recent,
    }


def _compute_alerts_for_client(conn, cid, today):
    """
    Calcule les alertes (contrats, garanties, antivirus) pour un client.
    Retourne un dictionnaire avec les différentes listes d'alertes.
    """
    alerte_jours = int(cfg_get('garantie_alerte_jours', '90'))

    # Alertes contrats
    contrats_alertes = []
    for row in retry_db_query(lambda: conn.execute("SELECT * FROM contrats WHERE client_id=? AND statut='actif' AND date_fin!='' ORDER BY date_fin", (cid,)).fetchall()):
        ct = row_to_dict(row)
        if not ct.get('date_fin'): continue
        try:
            df = date.fromisoformat(ct['date_fin'])
            delta = (df - today).days
            ct['jours_restants'] = delta
            ct['date_fin_fmt']   = df.strftime('%d/%m/%Y')
            preavis = ct.get('preavis_jours') or 30
            if delta < 0 or delta <= preavis:
                ct['expire_depasse'] = delta < 0
                contrats_alertes.append(ct)
        except: pass

    # Alertes garanties
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM appareils WHERE client_id=? ORDER BY adresse_ip', (cid,)).fetchall()]
    appareils = fmt_appareils(appareils)

    garanties_alertes = []
    for a in appareils:
        if not a.get('date_fin_garantie'): continue
        if a.get('garantie_alerte_ignoree'): continue
        try:
            df = date.fromisoformat(a['date_fin_garantie'])
            delta = (df - today).days
            if delta < 0 or delta <= alerte_jours:
                a['garantie_jours'] = delta
                a['garantie_fin_fmt'] = df.strftime('%d/%m/%Y')
                garanties_alertes.append(a)
        except: pass
    garanties_alertes.sort(key=lambda x: x.get('garantie_jours', 9999))

    # Prochains contrats à renouveler
    prochains_renouvellements = []
    for row in conn.execute(
        "SELECT * FROM contrats WHERE client_id=? AND statut='actif' AND date_fin!='' ORDER BY date_fin LIMIT 5",
        (cid,)).fetchall():
        ct = row_to_dict(row)
        try:
            df = date.fromisoformat(ct['date_fin'])
            ct['jours_restants'] = (df - today).days
            ct['date_fin_fmt']   = df.strftime('%d/%m/%Y')
            prochains_renouvellements.append(ct)
        except: pass

    # Antivirus urgents (30 jours)
    today_iso    = date.today().isoformat()
    seuil_av_iso = (date.today() + timedelta(days=30)).isoformat()
    av_urgents = []
    for _r in conn.execute(
        "SELECT nom_machine, av_nom, av_marque, av_date_fin FROM appareils "
        "WHERE client_id=? AND av_date_fin!='' AND av_date_fin IS NOT NULL AND av_date_fin<=? "
        "ORDER BY av_date_fin LIMIT 5", (cid, seuil_av_iso)).fetchall():
        _item = row_to_dict(_r)
        _item['expire_depasse'] = _item.get('av_date_fin', '') < today_iso
        av_urgents.append(_item)

    # Diagnostic réseau — évènements actifs
    diag_alertes = {'total': 0, 'critiques': 0, 'items': []}
    try:
        for _r in conn.execute(
            "SELECT titre, gravite, categorie FROM diag_reseau_evenements "
            "WHERE client_id=? AND resolu=0 "
            "ORDER BY CASE gravite WHEN 'critique' THEN 0 WHEN 'avertissement' THEN 1 ELSE 2 END, "
            "derniere_occurrence DESC", (cid,)).fetchall():
            it = row_to_dict(_r)
            diag_alertes['total'] += 1
            if it['gravite'] == 'critique':
                diag_alertes['critiques'] += 1
            if len(diag_alertes['items']) < 5:
                diag_alertes['items'].append(it)
    except Exception:
        pass

    return {
        'contrats_alertes': contrats_alertes[:5],
        'garanties_alertes': garanties_alertes[:5],
        'prochains_renouvellements': prochains_renouvellements,
        'av_urgents': av_urgents,
        'diag_alertes': diag_alertes,
    }


# ═══════════════════════════════════════════════════════════════════════════════════════
# PHASE 8: WIDGET DATA COMPUTATION HELPERS
# These functions compute data for individual dashboard widgets
# ═══════════════════════════════════════════════════════════════════════════════════════

def _compute_critical_alerts(conn, cid, today):
    """
    Consolidates all critical issues for the client:
    - Expired warranties
    - Expired contracts
    - Offline devices
    - Expiring licenses (AV/RMM/EDR)
    Returns list sorted by urgency.
    """
    alerts = []

    # Expired/expiring warranties — même délai configuré (garantie_alerte_jours)
    # et même drapeau « ignorer cette alerte » que le widget garanties du
    # dashboard client (_compute_alerts_for_client) : un appareil explicitement
    # mis en sourdine ne doit pas réapparaître ici.
    alerte_jours_garantie = int(cfg_get('garantie_alerte_jours', '90'))
    for row in conn.execute(
        "SELECT id, nom_machine, date_fin_garantie FROM appareils "
        "WHERE client_id=? AND date_fin_garantie!='' AND date_fin_garantie<=? "
        "AND (garantie_alerte_ignoree IS NULL OR garantie_alerte_ignoree=0)",
        (cid, (today + timedelta(days=alerte_jours_garantie)).isoformat())).fetchall():
        a = row_to_dict(row)
        try:
            df = date.fromisoformat(a['date_fin_garantie'])
            if df < today:
                alerts.append({'type': 'warranty_expired', 'device': a['nom_machine'], 'date': a['date_fin_garantie'], 'severity': 'critical'})
            else:
                alerts.append({'type': 'warranty_expiring', 'device': a['nom_machine'], 'date': a['date_fin_garantie'], 'severity': 'warning'})
        except: pass

    # Expired/expiring contracts
    for row in conn.execute(
        "SELECT id, description, date_fin FROM contrats "
        "WHERE client_id=? AND statut='actif' AND date_fin!='' AND date_fin<=?",
        (cid, (today + timedelta(days=30)).isoformat())).fetchall():
        c = row_to_dict(row)
        try:
            df = date.fromisoformat(c['date_fin'])
            if df < today:
                alerts.append({'type': 'contract_expired', 'contract': c['description'], 'date': c['date_fin'], 'severity': 'critical'})
            else:
                alerts.append({'type': 'contract_expiring', 'contract': c['description'], 'date': c['date_fin'], 'severity': 'warning'})
        except: pass

    # Offline devices (no recent ping)
    for row in conn.execute(
        "SELECT id, nom_machine FROM appareils WHERE client_id=? AND en_ligne=0",
        (cid,)).fetchall():
        a = row_to_dict(row)
        alerts.append({'type': 'device_offline', 'device': a['nom_machine'], 'severity': 'warning'})

    # Expiring AV/RMM/EDR licenses
    for row in conn.execute(
        "SELECT nom_machine, av_date_fin FROM appareils "
        "WHERE client_id=? AND av_date_fin!='' AND av_date_fin<=?",
        (cid, (today + timedelta(days=30)).isoformat())).fetchall():
        a = row_to_dict(row)
        try:
            df = date.fromisoformat(a['av_date_fin'])
            severity = 'critical' if df < today else 'warning'
            alerts.append({'type': 'av_expiring', 'device': a['nom_machine'], 'date': a['av_date_fin'], 'severity': severity})
        except: pass

    # Évènements de diagnostic réseau non résolus
    try:
        for row in conn.execute(
            "SELECT titre, gravite, nb_occurrences FROM diag_reseau_evenements "
            "WHERE client_id=? AND resolu=0", (cid,)).fetchall():
            d = row_to_dict(row)
            alerts.append({
                'type': 'diag_reseau',
                'device': d['titre'],
                'severity': 'critical' if d['gravite'] == 'critique' else 'warning',
                'link': '/diag-reseau',
            })
    except Exception:
        pass  # table absente sur une très vieille base non encore migrée

    # Sort by severity (critical first) then by date
    severity_order = {'critical': 0, 'warning': 1}
    alerts.sort(key=lambda x: (severity_order.get(x.get('severity'), 2), x.get('date', '')))

    return {'alerts': alerts, 'count': len(alerts)}


def _compute_kpi_cards(stats, alerts, today):
    """Returns data for the 6 main KPI cards."""
    return {
        'nb_app_total': stats['nb_app_total'],
        'nb_en_ligne': stats['nb_en_ligne'],
        'taux_dispo': stats['taux_dispo'],
        'nb_garantie': stats['nb_garantie'],
        'nb_contrats': stats['nb_contrats'],
        'montant_annuel': stats['montant_annuel'],
        'nb_alertes': len(alerts['contrats_alertes']) + len(alerts['garanties_alertes']) + len(alerts['av_urgents']),
    }


def _compute_av_status(conn, cid):
    """
    Returns AV/RMM/EDR license health status across all devices.
    Counts by status: active, expiring soon (30 days), expired.
    """
    today = date.today()
    seuil_futur = (today + timedelta(days=30)).isoformat()
    today_iso = today.isoformat()

    devices = conn.execute(
        "SELECT nom_machine, av_marque, av_date_fin, rmm_marque, rmm_date_fin, edr_marque, edr_date_fin FROM appareils WHERE client_id=?",
        (cid,)).fetchall()

    av_status = {'active': 0, 'expiring': 0, 'expired': 0}
    rmm_status = {'active': 0, 'expiring': 0, 'expired': 0}
    edr_status = {'active': 0, 'expiring': 0, 'expired': 0}

    for row in devices:
        d = row_to_dict(row)

        # AV
        if d.get('av_date_fin'):
            if d['av_date_fin'] < today_iso:
                av_status['expired'] += 1
            elif d['av_date_fin'] <= seuil_futur:
                av_status['expiring'] += 1
            else:
                av_status['active'] += 1

        # RMM
        if d.get('rmm_date_fin'):
            if d['rmm_date_fin'] < today_iso:
                rmm_status['expired'] += 1
            elif d['rmm_date_fin'] <= seuil_futur:
                rmm_status['expiring'] += 1
            else:
                rmm_status['active'] += 1

        # EDR
        if d.get('edr_date_fin'):
            if d['edr_date_fin'] < today_iso:
                edr_status['expired'] += 1
            elif d['edr_date_fin'] <= seuil_futur:
                edr_status['expiring'] += 1
            else:
                edr_status['active'] += 1

    return {
        'av': av_status,
        'rmm': rmm_status,
        'edr': edr_status,
    }


def _compute_network_status(stats):
    """Returns device online/offline status summary."""
    return {
        'nb_en_ligne': stats['nb_en_ligne'],
        'nb_hors_ligne': stats['nb_hors_ligne'],
        'taux_dispo': stats['taux_dispo'],
        'devices': stats['appareils'][:10],  # Top 10 devices
    }


def _compute_device_types(stats):
    """Returns device type distribution."""
    return {
        'repartition': stats['repartition'],
        'types_chart': stats['types_chart'],
    }


def _compute_peripherals_distribution(stats):
    """Returns peripheral category distribution."""
    return {
        'periph_chart': stats['periph_chart'],
        'periph_stats': stats['periph_stats'],
    }


def _compute_device_age(conn, cid, today):
    """
    Groups devices by acquisition date into age buckets:
    - 0-1 year
    - 1-3 years
    - 3-5 years
    - 5+ years
    """
    devices = conn.execute(
        "SELECT nom_machine, date_achat FROM appareils WHERE client_id=? AND date_achat!='' AND date_achat IS NOT NULL",
        (cid,)).fetchall()

    age_groups = {
        '0-1_year': [],
        '1-3_years': [],
        '3-5_years': [],
        '5plus_years': [],
        'unknown': []
    }

    for row in devices:
        d = row_to_dict(row)
        if not d.get('date_achat'):
            age_groups['unknown'].append(d)
            continue

        try:
            acq_date = date.fromisoformat(d['date_achat'])
            age_days = (today - acq_date).days
            age_years = age_days / 365.25

            if age_years < 1:
                age_groups['0-1_year'].append(d)
            elif age_years < 3:
                age_groups['1-3_years'].append(d)
            elif age_years < 5:
                age_groups['3-5_years'].append(d)
            else:
                age_groups['5plus_years'].append(d)
        except:
            age_groups['unknown'].append(d)

    return {
        'age_groups': age_groups,
        '0-1_year_count': len(age_groups['0-1_year']),
        '1-3_years_count': len(age_groups['1-3_years']),
        '3-5_years_count': len(age_groups['3-5_years']),
        '5plus_years_count': len(age_groups['5plus_years']),
    }


def _compute_contracts_timeline(conn, cid, today):
    """Returns upcoming contract renewals/expirations timeline."""
    contracts = []
    for row in conn.execute(
        "SELECT id, titre, date_fin, montant_ht FROM contrats "
        "WHERE client_id=? AND statut='actif' AND date_fin!='' "
        "ORDER BY date_fin LIMIT 10",
        (cid,)).fetchall():
        c = row_to_dict(row)
        try:
            df = date.fromisoformat(c['date_fin'])
            days_left = (df - today).days
            c['jours_restants'] = days_left
            c['date_fin_fmt'] = df.strftime('%d/%m/%Y')
            c['urgence'] = 'expired' if days_left < 0 else ('urgent' if days_left < 30 else 'ok')
            contracts.append(c)
        except: pass

    return {
        'contracts': contracts,
        'total_count': len(contracts),
    }


def _compute_recent_activity(stats):
    """Returns recent activity/modifications."""
    return {
        'recents': stats['recents'],
        'hist_recent': stats['hist_recent'],
    }


def _compute_interventions_summary(recent_interventions):
    """Returns recent interventions summary."""
    return {
        'interventions': recent_interventions,
        'count': len(recent_interventions),
    }


def _compute_business_software(logiciels, stats):
    """Returns business software deployments."""
    return {
        'logiciels': logiciels,
        'count': len(logiciels),
        'appareils_count': stats['nb_app_total'],
    }


def _compute_network_info(parc):
    """Returns network configuration information."""
    return {
        'nom_site': parc.get('nom_site', 'N/A'),
        'type_connexion': parc.get('type_connexion', 'N/A'),
        'debit_montant': parc.get('debit_montant', 'N/A'),
        'debit_descendant': parc.get('debit_descendant', 'N/A'),
        'fournisseur_internet': parc.get('fournisseur_internet', 'N/A'),
        'ip_publique': parc.get('ip_publique', 'N/A'),
        'plage_ip_locale': parc.get('plage_ip_locale', 'N/A'),
        'domaine': parc.get('domaine', 'N/A'),
        'serveur_dns': parc.get('serveur_dns', 'N/A'),
        'passerelle': parc.get('passerelle', 'N/A'),
    }


def single_client_dashboard(cid):
    """
    Affiche le dashboard pour un seul client (vue classique).
    """
    conn = get_db()
    today = date.today()
    user = get_auth_user()

    try:
        # Fetch parc and client info
        parc    = row_to_dict(retry_db_query(lambda: conn.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {}))
        client  = row_to_dict(retry_db_query(lambda: conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {}))

        # Get dashboard stats
        stats = _compute_client_dashboard_stats(conn, cid, today)

        # Get alerts
        alerts = _compute_alerts_for_client(conn, cid, today)

        # Recent interventions
        recent_interventions = [fmt_intervention(row_to_dict(r)) for r in conn.execute(
            "SELECT * FROM interventions WHERE client_id=? AND statut != ? ORDER BY date_intervention DESC LIMIT 5",
            (cid, 'archivee')).fetchall()]

        # Logiciels & antivirus (depuis parc_general)
        logiciels = [l.strip() for l in (parc.get('logiciels_metier') or '').splitlines() if l.strip()]

        # Calcul valeur parc
        valeur_parc = sum(a.get('prix_achat') or 0 for a in stats['appareils'])

        # ═══════════════════════════════════════════════════════════════════
        # PHASE 8.4: Fetch and parse user's widget preferences
        # ═══════════════════════════════════════════════════════════════════
        user_id = user['id'] if user else None
        default_enabled = 'critical-alerts,kpi,av-status,network-status,device-types,peripherals,device-age,contracts-timeline,recent-activity,interventions,business-software,network-info'
        default_order = 'critical-alerts,kpi,av-status,network-status,device-types,peripherals,device-age,contracts-timeline,recent-activity,interventions,business-software,network-info'

        enabled_widgets_str = cfg_get('dashboard_widgets_enabled', default_enabled, user_id)
        widget_order_str = cfg_get('dashboard_widgets_order', default_order, user_id)

        # CRITICAL: Handle empty strings by using defaults
        # This prevents widgets from disappearing if saved as empty
        if not enabled_widgets_str or not enabled_widgets_str.strip():
            enabled_widgets_str = default_enabled
        if not widget_order_str or not widget_order_str.strip():
            widget_order_str = default_order

        enabled_widgets = [w.strip() for w in enabled_widgets_str.split(',') if w.strip()]
        widget_order = [w.strip() for w in widget_order_str.split(',') if w.strip()]

        # CRITICAL: Ensure all widgets from default list are included (for new widgets added later)
        # If a widget is in the default list but not in the saved config, add it
        default_enabled_list = [w.strip() for w in default_enabled.split(',') if w.strip()]
        default_order_list = [w.strip() for w in default_order.split(',') if w.strip()]

        for widget_id in default_enabled_list:
            if widget_id not in enabled_widgets:
                enabled_widgets.append(widget_id)

        for widget_id in default_order_list:
            if widget_id not in widget_order:
                widget_order.append(widget_id)

        # ═══════════════════════════════════════════════════════════════════
        # PHASE 9: Parse widget sizes from user preferences (Phase 9)
        # ═══════════════════════════════════════════════════════════════════
        # Default widget sizes per widget_id
        WIDGET_DEFAULT_SIZES = {
            'critical-alerts': 'large',
            'kpi': 'large',
            'av-status': 'medium',
            'network-status': 'medium',  # Changed from 'large' to match device-types for visual cohesion
            'device-types': 'medium',
            'peripherals': 'medium',
            'device-age': 'small',
            'contracts-timeline': 'large',
            'recent-activity': 'large',
            'interventions': 'small',
            'business-software': 'medium',
            'network-info': 'medium',
        }

        # Parse user's widget size preferences (JSON)
        widget_sizes_str = cfg_get('dashboard_widgets_size', '{}', user_id)
        try:
            widget_sizes_json = json.loads(widget_sizes_str)
        except (json.JSONDecodeError, ValueError):
            widget_sizes_json = {}

        # Build final sizes dict with defaults
        widget_sizes = {}
        for widget_id in enabled_widgets + widget_order:
            if widget_id in WIDGET_DEFAULT_SIZES:
                # Use user's size if specified, otherwise use default
                widget_sizes[widget_id] = widget_sizes_json.get(widget_id, WIDGET_DEFAULT_SIZES[widget_id])

        # ═══════════════════════════════════════════════════════════════════
        # PHASE 10: Parse widget heights from user preferences (NEW)
        # ═══════════════════════════════════════════════════════════════════
        # Default widget heights per widget_id (5 LEVEL SYSTEM: xs, s, m, l, xl)
        # Tous les widgets utilisent 'm' (280px) pour une présentation cohérente et harmonieuse
        WIDGET_DEFAULT_HEIGHTS = {
            'critical-alerts': 'm',      # Medium (280px)
            'kpi': 'm',
            'av-status': 'm',
            'network-status': 'm',       # Changed from 'compact' to 'm' for consistency
            'device-types': 'm',
            'peripherals': 'm',
            'device-age': 'm',
            'contracts-timeline': 'm',
            'recent-activity': 'm',
            'interventions': 'm',        # Changed from 'compact' to 'm' for consistency
            'business-software': 'm',    # Changed from 'compact' to 'm' for consistency
            'network-info': 'm',         # Changed from 'compact' to 'm' for consistency
        }

        # Parse user's widget height preferences (JSON)
        widget_heights_str = cfg_get('dashboard_widgets_height', '{}', user_id)
        try:
            widget_heights_json = json.loads(widget_heights_str)
        except (json.JSONDecodeError, ValueError):
            widget_heights_json = {}

        # Build final heights dict with defaults
        widget_heights = {}
        for widget_id in enabled_widgets + widget_order:
            if widget_id in WIDGET_DEFAULT_HEIGHTS:
                # Use user's height if specified, otherwise use default
                widget_heights[widget_id] = widget_heights_json.get(widget_id, WIDGET_DEFAULT_HEIGHTS[widget_id])

        # Build complete widget_data dict with all widget calculations
        # (even disabled widgets may be re-enabled later without reload)
        try:
            widget_data = {
                'critical_alerts': _compute_critical_alerts(conn, cid, today),
                'kpi': _compute_kpi_cards(stats, alerts, today),
                'av_status': _compute_av_status(conn, cid),
                'network_status': _compute_network_status(stats),
                'device_types': _compute_device_types(stats),
                'peripherals': _compute_peripherals_distribution(stats),
                'device_age': _compute_device_age(conn, cid, today),
                'contracts_timeline': _compute_contracts_timeline(conn, cid, today),
                'recent_activity': _compute_recent_activity(stats),
                'interventions': _compute_interventions_summary(recent_interventions),
                'business_software': _compute_business_software(logiciels, stats),
                'network_info': _compute_network_info(parc),
            }
        except Exception as e:
            logger.exception("Error computing widget data for client %d: %s", cid, str(e))
            # Fallback: empty widget data
            widget_data = {
                'critical_alerts': {},
                'kpi': {},
                'av_status': {},
                'network_status': {},
                'device_types': {},
                'peripherals': {},
                'device_age': {},
                'contracts_timeline': {},
                'recent_activity': {},
                'interventions': {},
                'business_software': {},
                'network_info': {},
            }

        # Combine all data for template
        template_data = {
            'parc': parc,
            'client': client,
            'appareils': stats['appareils'],
            'nb_en_ligne': stats['nb_en_ligne'],
            'nb_hors_ligne': stats['nb_hors_ligne'],
            'nb_actifs': stats['nb_en_ligne'],
            'nb_garantie': stats['nb_garantie'],
            'nb_app_total': stats['nb_app_total'],
            'taux_dispo': stats['taux_dispo'],
            'valeur_parc': valeur_parc,
            'nb_periph': stats['nb_periph'],
            'nb_contrats': stats['nb_contrats'],
            'nb_identifiants': stats['nb_identifiants'],
            'repartition': stats['repartition'],
            'types_chart': stats['types_chart'],
            'periph_chart': stats['periph_chart'],
            'contrats_alertes': alerts['contrats_alertes'],
            'garanties_alertes': alerts['garanties_alertes'],
            'prochains_renouvellements': alerts['prochains_renouvellements'],
            'periph_stats': stats['periph_stats'],
            'montant_annuel': stats['montant_annuel'],
            'recents': stats['recents'],
            'derniers_periph': stats['derniers_periph'],
            'derniers_contrats': stats['derniers_contrats'],
            'recent_interventions': recent_interventions,
            'logiciels': logiciels,
            'av_urgents': alerts['av_urgents'],
            'hist_recent': stats['hist_recent'],
            'nb_users': stats['nb_users'],
            'nb_hors_service': stats['nb_hors_service'],
            'clients': get_clients(),
            'client_actif_id': cid,
            # Widget preferences (Phase 8.4)
            'enabled_widgets': enabled_widgets,
            'widget_order': widget_order,
            'widget_data': widget_data,
            # Widget sizes (Phase 9)
            'widget_sizes': widget_sizes,
            # Widget heights (Phase 10)
            'widget_heights': widget_heights,
        }

        return render_template('client_dashboard.html', **template_data)
    finally:
        conn.close()


def user_dashboard():
    """
    Affiche le dashboard utilisateur avec vue d'ensemble de tous les clients accessibles.
    Agrège les statistiques et les alertes par client.
    """
    user = get_auth_user()
    clients = get_clients()
    conn = get_db()
    today = date.today()

    try:
        # Agrégation des données par client
        clients_data = []
        all_alerts = []

        for client in clients:
            cid = client['id']

            # Récupérer les stats du client
            stats = _compute_client_dashboard_stats(conn, cid, today)
            alerts = _compute_alerts_for_client(conn, cid, today)

            # Compter les alertes pour ce client
            alert_count = (
                len(alerts['contrats_alertes']) +
                len(alerts['garanties_alertes']) +
                len(alerts['av_urgents'])
            )

            # Données du client pour le template
            client_summary = {
                'client': client,
                'stats': stats,
                'alerts': alerts,
                'alert_count': alert_count,
            }
            clients_data.append(client_summary)

            # Ajouter les alertes à la liste consolidée avec contexte du client
            for contract_alert in alerts['contrats_alertes']:
                all_alerts.append({
                    'type': 'contract',
                    'client_id': cid,
                    'client_nom': client['nom'],
                    'description': contract_alert.get('description', f"Contrat: {contract_alert.get('numero_contrat', 'N/A')}"),
                    'days_remaining': contract_alert.get('jours_restants', 999),
                    'date': contract_alert.get('date_fin', ''),
                    'expired': contract_alert.get('expire_depasse', False),
                    'object': contract_alert,
                })

            for warranty_alert in alerts['garanties_alertes']:
                all_alerts.append({
                    'type': 'warranty',
                    'client_id': cid,
                    'client_nom': client['nom'],
                    'description': warranty_alert.get('nom_machine', 'Appareil'),
                    'days_remaining': warranty_alert.get('garantie_jours', 999),
                    'date': warranty_alert.get('date_fin_garantie', ''),
                    'expired': warranty_alert.get('garantie_jours', 0) < 0,
                    'object': warranty_alert,
                })

            for av_alert in alerts['av_urgents']:
                all_alerts.append({
                    'type': 'antivirus',
                    'client_id': cid,
                    'client_nom': client['nom'],
                    'description': av_alert.get('nom_machine', 'Appareil'),
                    'days_remaining': 0,
                    'date': av_alert.get('av_date_fin', ''),
                    'expired': av_alert.get('expire_depasse', False),
                    'object': av_alert,
                })

        # Trier les alertes par urgence (jours restants croissant)
        all_alerts.sort(key=lambda x: (x['expired'] == False, x['days_remaining']))

        # Calculs globaux
        total_devices = sum(s['stats']['nb_app_total'] for s in clients_data)
        total_online = sum(s['stats']['nb_en_ligne'] for s in clients_data)
        total_contracts = sum(s['stats']['nb_contrats'] for s in clients_data)
        total_peripherals = sum(s['stats']['nb_periph'] for s in clients_data)
        total_alerts_count = len(all_alerts)

        # Récupérer le client actif de la session (si existe)
        # pour afficher le bon nom dans la topbar du dashboard utilisateur
        active_client_id = session.get('client_id')
        active_client = None
        if active_client_id:
            active_client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (active_client_id,)).fetchone() or {})

        # Récupérer les préférences de vue du dashboard utilisateur
        user_id = user['id'] if user else None
        user_view_mode = cfg_get('user_dashboard_view_mode', 'grid', user_id)
        user_detail_level = cfg_get('user_dashboard_detail_level', 'standard', user_id)

        return render_template('user_dashboard.html',
                             user=user,
                             clients_data=clients_data,
                             all_alerts=all_alerts[:20],  # Top 20 most urgent alerts
                             total_devices=total_devices,
                             total_online=total_online,
                             total_contracts=total_contracts,
                             total_peripherals=total_peripherals,
                             total_alerts_count=total_alerts_count,
                             clients=clients,
                             client=active_client,  # Pass active client from session to topbar
                             client_actif_id=active_client_id,  # Pass active client ID to dropdown
                             user_dashboard_view_mode=user_view_mode,
                             user_dashboard_detail_level=user_detail_level)
    finally:
        conn.close()


@app.route('/')
@login_required
def index():
    """
    Route dashboard intelligente qui détecte le nombre de clients accessibles.
    - Zéro clients: redirection vers création
    - Un client: affiche le dashboard single-client classique
    - Plusieurs clients: affiche le dashboard multi-client avec vue d'ensemble
    """
    user = get_auth_user()
    clients = get_clients()  # Récupère tous les clients accessibles (respects ACL)

    # Cas 1: Pas de clients accessibles
    if not clients:
        return redirect(url_for('nouveau_client'))

    # Cas 2: Un seul client accessible -> affiche le dashboard classique
    if len(clients) == 1:
        cid = clients[0]['id']
        session['client_id'] = cid
        return single_client_dashboard(cid)

    # Cas 3: Plusieurs clients accessibles -> affiche le dashboard utilisateur
    return user_dashboard()


@app.route('/client/<int:cid>/dashboard')
@login_required
def client_dashboard_view(cid):
    """
    Affiche le dashboard pour un client spécifique.
    Vérifie d'abord que l'utilisateur a accès au client.
    """
    # Vérifier l'accès
    if not get_client_access(cid):
        flash('Accès refusé à ce client', 'danger')
        return redirect(url_for('index'))

    # Définir le client actif dans la session
    session['client_id'] = cid

    # Afficher le dashboard du client
    return single_client_dashboard(cid)

def _marque_modele_combos(rows):
    """Combine marque+modèle en une seule chaîne par ligne, dédoublonnée et
    triée — pour peupler un datalist de suggestions depuis l'inventaire réel
    (appareils/périphériques) plutôt qu'une liste de marques figée."""
    vals = set()
    for marque, modele in rows:
        combo = ' '.join(p for p in (marque or '', modele or '') if p).strip()
        if combo:
            vals.add(combo)
    return sorted(vals)


@app.route('/parc', methods=['GET','POST'])
@login_required
def parc_general():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    parc = row_to_dict(conn.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    if request.method == 'POST':
        if not can_write():
            flash('Accès en lecture seule — modification non autorisée', 'danger')
            return redirect(url_for('index'))
        f = request.form
        if parc.get('id'):
            conn.execute('''UPDATE parc_general SET nom_site=?,adresse=?,type_connexion=?,debit_montant=?,
                debit_descendant=?,fournisseur_internet=?,ip_publique=?,plage_ip_locale=?,nb_machines=?,
                nb_utilisateurs=?,domaine=?,serveur_dns=?,passerelle=?,baie_marque=?,baie_nb_u=?,
                switch_marque=?,switch_nb_ports=?,switch_nb_unites=?,routeur_marque=?,serveur_marque=?,
                serveur_modele=?,ups_marque=?,ups_capacite=?,autres_equipements=?,logiciels_metier=?,
                antivirus=?,os_principal=?,suite_bureautique=?,notes=?,
                wifi_ssid=?,wifi_password=?,wifi_securite=?,
                wifi_ssid2=?,wifi_password2=?,wifi_securite2=?,wifi_notes=?,
                date_maj=? WHERE client_id=?''', (
                f.get('nom_site',''), f.get('adresse',''), f.get('type_connexion',''),
                f.get('debit_montant',''), f.get('debit_descendant',''), f.get('fournisseur_internet',''),
                f.get('ip_publique',''), f.get('plage_ip_locale','192.168.1.0/24'),
                int(f.get('nb_machines') or 0), int(f.get('nb_utilisateurs') or 0),
                f.get('domaine',''), f.get('serveur_dns',''), f.get('passerelle',''),
                f.get('baie_marque',''), int(f.get('baie_nb_u') or 0), f.get('switch_marque',''),
                int(f.get('switch_nb_ports') or 0), int(f.get('switch_nb_unites') or 0),
                f.get('routeur_marque',''), f.get('serveur_marque',''), f.get('serveur_modele',''),
                f.get('ups_marque',''), f.get('ups_capacite',''), f.get('autres_equipements',''),
                f.get('logiciels_metier',''), f.get('antivirus',''), f.get('os_principal',''),
                f.get('suite_bureautique',''), f.get('notes',''),
                f.get('wifi_ssid',''), f.get('wifi_password',''), f.get('wifi_securite','WPA2'),
                f.get('wifi_ssid2',''), f.get('wifi_password2',''), f.get('wifi_securite2','WPA2'),
                f.get('wifi_notes',''),
                _utcnow().isoformat(), cid))
        else:
            conn.execute('''INSERT INTO parc_general (client_id,nom_site,plage_ip_locale,date_maj) VALUES (?,?,?,?)''',
                         (cid, f.get('nom_site',''), f.get('plage_ip_locale','192.168.1.0/24'), _utcnow().isoformat()))
        conn.commit(); conn.close()
        flash('Informations du parc sauvegardées', 'success')
        return redirect(url_for('parc_general'))
    # Lien croisé avec la baie de brassage : parc_general (résumé
    # switch/routeur/UPS) et baie_slots (position physique par U) décrivent
    # le même matériel sans jamais se recouper. Un simple comptage suffit à
    # signaler l'écart le plus fréquent (ex. switch/routeur renseignés ici,
    # jamais positionnés dans la baie) sans dupliquer la donnée.
    baie_nb_slots_occupes = conn.execute(
        "SELECT COUNT(*) FROM baie_slots WHERE client_id=?", (cid,)).fetchone()[0]

    # Suggestions issues de l'inventaire réel (appareils/périphériques déjà
    # saisis pour ce client) : proposées dans un datalist à côté de chaque
    # champ concerné, sans jamais empêcher la saisie manuelle.
    switch_suggestions = _marque_modele_combos(conn.execute(
        "SELECT marque, modele FROM appareils WHERE client_id=? AND type_appareil IN ('Switch','Switch/AP')",
        (cid,)).fetchall())
    routeur_suggestions = _marque_modele_combos(conn.execute(
        "SELECT marque, modele FROM appareils WHERE client_id=? AND type_appareil='Routeur/Pare-feu'",
        (cid,)).fetchall())
    ups_suggestions = _marque_modele_combos(conn.execute(
        "SELECT marque, modele FROM peripheriques WHERE client_id=? AND categorie='Onduleur / UPS'",
        (cid,)).fetchall())
    serveur_marques = sorted({r[0] for r in conn.execute(
        "SELECT DISTINCT marque FROM appareils WHERE client_id=? AND type_appareil='Serveur' AND marque!=''",
        (cid,)).fetchall()})
    serveur_modeles = sorted({r[0] for r in conn.execute(
        "SELECT DISTINCT modele FROM appareils WHERE client_id=? AND type_appareil='Serveur' AND modele!=''",
        (cid,)).fetchall()})
    conn.close()
    return render_template('parc_general.html', parc=parc, client=client,
                           baie_nb_slots_occupes=baie_nb_slots_occupes,
                           switch_suggestions=switch_suggestions,
                           routeur_suggestions=routeur_suggestions,
                           ups_suggestions=ups_suggestions,
                           serveur_marques=serveur_marques,
                           serveur_modeles=serveur_modeles,
                           clients=get_clients(), client_actif_id=cid)

# ─── ROUTES API PRESTATAIRES ─────────────────────────────────────────────────────

# ─── ROUTES API POUR CHARGER LES ENTITÉS D'UN CLIENT ──────────────────────────

@app.route('/api/client/<int:client_id>/appareils', methods=['GET'])
@login_required
def api_get_client_appareils(client_id):
    """Lister les appareils d'un client"""
    if not get_client_access(client_id):
        return jsonify({'error': 'Forbidden'}), 403

    conn = get_db()
    try:
        rows = conn.execute(
            'SELECT id, nom_machine as nom FROM appareils WHERE client_id=? ORDER BY nom_machine ASC',
            (client_id,)
        ).fetchall()
        conn.close()
        return jsonify([{'id': r[0], 'nom': r[1]} for r in rows])
    except Exception as e:
        conn.close()
        logger.exception(f'Erreur lecture appareils client {client_id}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/client/<int:client_id>/contrats', methods=['GET'])
@login_required
def api_get_client_contrats(client_id):
    """Lister les contrats d'un client"""
    if not get_client_access(client_id):
        return jsonify({'error': 'Forbidden'}), 403

    conn = get_db()
    try:
        rows = conn.execute(
            'SELECT id, description FROM contrats WHERE client_id=? ORDER BY description ASC',
            (client_id,)
        ).fetchall()
        conn.close()
        return jsonify([{'id': r[0], 'nom': r[1]} for r in rows])
    except Exception as e:
        conn.close()
        logger.exception(f'Erreur lecture contrats client {client_id}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/client/<int:client_id>/peripheriques', methods=['GET'])
@login_required
def api_get_client_peripheriques(client_id):
    """Lister les périphériques d'un client"""
    if not get_client_access(client_id):
        return jsonify({'error': 'Forbidden'}), 403

    conn = get_db()
    try:
        rows = conn.execute(
            'SELECT id, (marque || " " || modele) as nom FROM peripheriques WHERE client_id=? ORDER BY marque, modele ASC',
            (client_id,)
        ).fetchall()
        conn.close()
        return jsonify([{'id': r[0], 'nom': r[1]} for r in rows])
    except Exception as e:
        conn.close()
        logger.exception(f'Erreur lecture périphériques client {client_id}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/client/<int:client_id>/services', methods=['GET'])
@login_required
def api_get_client_services(client_id):
    """Lister les services d'un client"""
    if not get_client_access(client_id):
        return jsonify({'error': 'Forbidden'}), 403

    conn = get_db()
    try:
        rows = conn.execute(
            'SELECT id, nom FROM services WHERE client_id=? ORDER BY nom ASC',
            (client_id,)
        ).fetchall()
        conn.close()
        return jsonify([{'id': r[0], 'nom': r[1]} for r in rows])
    except Exception as e:
        conn.close()
        logger.exception(f'Erreur lecture services client {client_id}')
        return jsonify({'error': str(e)}), 500

#: Colonnes de `appareils` jamais affichées par la LISTE d'inventaire (seule
#: la fiche détail les utilise) mais pouvant peser plusieurs centaines de Ko
#: chacune une fois remplies par une vraie collecte — jusqu'à 1 Mo pour
#: rapport_systeme_json (voir sa limite dans /api/device-info). `SELECT a.*`
#: les ramenait sur CHAQUE page de la liste, pour CHAQUE appareil affiché :
#: signalé en usage réel comme un peu lent avec une soixantaine d'appareils,
#: exactement le genre de ralentissement qu'un tel sur-chargement produit.
_APP_COLONNES_LOURDES = frozenset({'rapport_systeme_json', 'logiciels_installes_json'})
_app_colonnes_liste_cache = None

def _colonnes_appareils_liste():
    """Colonnes de `appareils` pour la liste d'inventaire, sans les blobs
    lourds ci-dessus. Construite depuis PRAGMA (mise en cache : le schéma ne
    change pas en cours de vie du process) plutôt qu'à la main — `appareils`
    a beaucoup grandi et continue de grandir (voir claude.md), une liste
    figée se démoderait silencieusement à la prochaine colonne ajoutée
    ailleurs dans le code.
    """
    global _app_colonnes_liste_cache
    if _app_colonnes_liste_cache is None:
        conn = get_db()
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(appareils)").fetchall()
                    if r[1] not in _APP_COLONNES_LOURDES]
        finally:
            conn.close()
        _app_colonnes_liste_cache = ', '.join(f'a.{c}' for c in cols)
    return _app_colonnes_liste_cache


# Colonnes triables pour l'inventaire appareils
_APP_SORT_COLS = {
    'nom':      'a.nom_machine',
    'type':     'a.type_appareil, a.nom_machine',
    'ip':       'ip_sort_key(a.adresse_ip), a.nom_machine',
    'user':     'a.utilisateur, a.nom_machine',
    'garantie': "CASE WHEN a.date_fin_garantie='' OR a.date_fin_garantie IS NULL THEN '9999-99-99' ELSE a.date_fin_garantie END, a.nom_machine",
    'statut':   'a.statut, a.nom_machine',
    'marque':   'a.marque, a.modele, a.nom_machine',
    'os':       'a.os, a.nom_machine',
}

@app.route('/appareils')
@login_required
def liste_appareils():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    page      = request.args.get('page', 1, type=int)
    sort_col  = request.args.get('sort', 'ip')
    sort_dir  = request.args.get('dir',  'asc')
    f_types   = request.args.getlist('type')
    f_statut  = request.args.get('statut', '')
    f_av      = request.args.get('av', '')
    f_service = request.args.get('service', '')

    order_expr = _APP_SORT_COLS.get(sort_col, 'ip_sort_key(a.adresse_ip)')
    direction  = 'DESC' if sort_dir == 'desc' else 'ASC'

    q = f'''SELECT {_colonnes_appareils_liste()},
            (SELECT COUNT(*) FROM documents_appareils d WHERE d.appareil_id=a.id) as nb_docs,
            (SELECT COUNT(*) FROM contrats_appareils ca JOIN contrats ct ON ca.contrat_id=ct.id
             WHERE ca.appareil_id=a.id AND ct.client_id=a.client_id) as nb_contrats
            FROM appareils a WHERE a.client_id=?'''
    params = [cid]

    if f_types:
        placeholders = ','.join('?' * len(f_types))
        q += f' AND a.type_appareil IN ({placeholders})'
        params.extend(f_types)

    if f_statut:
        q += ' AND a.statut=?'
        params.append(f_statut)

    if f_service:
        q += ' AND a.service_id=?'
        params.append(int(f_service))

    # Filtre antivirus (reconstruit la logique de fmt_appareils)
    if f_av == 'none':
        q += " AND (a.av_nom='' OR a.av_nom IS NULL) AND (a.av_marque='' OR a.av_marque IS NULL)"
    elif f_av == 'expired':
        q += " AND (a.av_nom!='' AND a.av_nom IS NOT NULL) AND a.av_date_fin!='' AND a.av_date_fin IS NOT NULL AND date(a.av_date_fin)<date('now')"
    elif f_av == 'expiring':
        q += " AND (a.av_nom!='' AND a.av_nom IS NOT NULL) AND a.av_date_fin!='' AND a.av_date_fin IS NOT NULL AND date(a.av_date_fin)>=date('now') AND date(a.av_date_fin)<=date('now','+30 days')"
    elif f_av == 'active':
        q += " AND (a.av_nom!='' AND a.av_nom IS NOT NULL) AND (a.av_date_fin='' OR a.av_date_fin IS NULL OR date(a.av_date_fin)>date('now','+30 days'))"

    q += f' ORDER BY {order_expr} {direction}'

    rows, pagination = paginate(q, tuple(params), page)
    appareils = fmt_appareils([row_to_dict(r) for r in rows])
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    periph_rows = conn.execute(
        '''SELECT pa.appareil_id, p.id, p.categorie, p.marque, p.modele,
                  p.description, p.statut, p.numero_serie
           FROM peripheriques p
           JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id
           WHERE p.client_id=? ORDER BY p.categorie''',
        (cid,)).fetchall()
    conn.close()
    periph_by_app = {}
    for r in periph_rows:
        aid = r[0]
        periph_by_app.setdefault(aid, []).append({
            'id': r[1], 'categorie': r[2], 'marque': r[3],
            'modele': r[4], 'description': r[5], 'statut': r[6], 'numero_serie': r[7]
        })
    for a in appareils:
        a['peripheriques'] = periph_by_app.get(a['id'], [])
    # Badge UPS (demandé) : un appareil branché sur une prise d'onduleur —
    # directement, ou via un PDU dont l'entrée remonte, en cascade, jusqu'à
    # un onduleur (voir _appareils_secourus_par_onduleur) — toutes les
    # baies du client, pas seulement une en particulier, l'inventaire étant
    # une vue globale du parc.
    conn3 = get_db()
    ups_par_appareil = _appareils_secourus_par_onduleur(conn3, cid)
    conn3.close()
    for a in appareils:
        a['ups_nom'] = ups_par_appareil.get(a['id'])
    service_filtre_nom = None
    if f_service:
        conn2 = get_db()
        row = conn2.execute('SELECT nom FROM services WHERE id=? AND client_id=?', (f_service, cid)).fetchone()
        conn2.close()
        service_filtre_nom = row[0] if row else None
    return render_template('liste_appareils.html', appareils=appareils, client=client,
                           clients=get_clients(), client_actif_id=cid, pagination=pagination,
                           sort_col=sort_col, sort_dir=sort_dir,
                           f_types=f_types, f_statut=f_statut, f_av=f_av,
                           f_service=f_service, service_filtre_nom=service_filtre_nom)

def _save_licences(conn, appareil_id, cid, form):
    """Supprime puis réinsère les licences d'un appareil depuis les données du formulaire."""
    conn.execute('DELETE FROM licences_appareils WHERE appareil_id=?', (appareil_id,))
    editeurs    = form.getlist('lic_editeur')
    produits    = form.getlist('lic_produit')
    cles        = form.getlist('lic_cle')
    contrat_ids = form.getlist('lic_contrat_id')
    now = _utcnow().isoformat()
    for i, editeur in enumerate(editeurs):
        editeur  = editeur.strip()
        produit  = produits[i].strip()  if i < len(produits)    else ''
        cle      = cles[i].strip()      if i < len(cles)         else ''
        cid_lic  = contrat_ids[i]       if i < len(contrat_ids)  else ''
        if not editeur and not produit and not cle:
            continue
        contrat_id_val = None
        try: contrat_id_val = int(cid_lic) if cid_lic else None
        except: pass
        conn.execute(
            '''INSERT INTO licences_appareils
               (appareil_id,client_id,editeur,produit,cle_licence,contrat_id,date_creation)
               VALUES (?,?,?,?,?,?,?)''',
            (appareil_id, cid, editeur, produit, cle, contrat_id_val, now))


def _get_logiciels_metier_list(conn, cid):
    """Retourne la liste des logiciels métier depuis parc_general pour le client donné."""
    parc_row = conn.execute('SELECT logiciels_metier FROM parc_general WHERE client_id=?', (cid,)).fetchone()
    raw = (parc_row[0] if parc_row else '') or ''
    return [l.strip() for l in re.split(r'[,\n]', raw) if l.strip()]


@app.route('/appareil/nouveau', methods=['GET','POST'])
@login_required
def nouvel_appareil():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    contrats = [row_to_dict(r) for r in conn.execute(
        "SELECT id,titre,fournisseur,statut FROM contrats WHERE client_id=? ORDER BY titre", (cid,)).fetchall()]
    lm_list = _get_logiciels_metier_list(conn, cid)
    utilisateurs_noms, utilisateurs_variantes = _utilisateurs_pour_formulaire(conn, cid)
    services_noms = _services_pour_formulaire(conn, cid)
    conn.close()
    if request.method == 'POST':
        if not can_write():
            flash('Accès en lecture seule — modification non autorisée', 'danger')
            return redirect(url_for('liste_appareils'))
        errs = validate_form([
            ('nom_machine',  'str',   True),
            ('adresse_ip',   'ip',    False),
            ('adresse_mac',  'mac',   False),
            ('date_achat',   'date',  False),
            ('date_fin_garantie', 'date', False),
        ], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        now = _utcnow().isoformat()
        vals = (cid,) + _extract_form(request.form) + (now, now)
        conn = get_db()
        svc_id = _resolve_service_id(conn, cid, request.form.get('service', ''))
        usr_id = _resolve_utilisateur_id(conn, cid, request.form.get('utilisateur', ''))
        conn.execute('''INSERT INTO appareils (client_id,nom_machine,type_appareil,marque,modele,numero_serie,
            adresse_ip,adresse_mac,nom_dns,utilisateur,service,localisation,date_achat,duree_garantie,
            date_fin_garantie,fournisseur,prix_achat,numero_commande,os,version_os,ram,cpu,stockage,carte_graphique,
            statut,notes,user_login,user_password,admin_login,admin_password,anydesk_id,anydesk_password,
            av_marque,av_nom,av_date_debut,av_date_fin,av_contrat_id,
            edr_marque,edr_nom,edr_date_fin,edr_contrat_id,
            rmm_marque,rmm_nom,rmm_agent_id,rmm_date_fin,rmm_contrat_id,
            logiciels,garantie_alerte_ignoree,date_creation,date_maj,service_id,utilisateur_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            vals + (svc_id, usr_id))
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        _save_licences(conn, new_id, cid, request.form)
        log_history(conn, cid, 'appareil', new_id, request.form.get('nom_machine','') or 'Nouvel appareil', 'Création')
        _sync_appareil_to_periph(conn, new_id, cid)
        _propager_utilisateur_aux_peripheriques(
            conn, new_id, cid, '', request.form.get('utilisateur', ''))
        conn.commit(); conn.close()
        flash('Appareil ajouté avec succès', 'success')
        return redirect(url_for('liste_appareils'))
    return render_template('form_appareil.html', appareil=None, action='Ajouter',
                           types_appareils=get_liste_cached('types_appareils'),
                           marques_av=get_liste('marques_antivirus'),
                           noms_av=get_liste('noms_antivirus'),
                           marques_edr=get_liste('marques_edr'),
                           noms_edr=get_liste('noms_edr'),
                           marques_rmm=get_liste('marques_rmm'),
                           noms_rmm=get_liste('noms_rmm'),
                           contrats=contrats,
                           sw_courants_groups=SW_COURANTS_GROUPS,
                           sw_courants_all=list(SW_COURANTS_ALL),
                           logiciels_metier_list=lm_list,
                           sw_sel=[],
                           sw_custom_sel=[],
                           licences=[],
                           utilisateurs_noms=utilisateurs_noms,
                           utilisateurs_variantes=utilisateurs_variantes,
                           services_noms=services_noms,
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/appareil/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_appareil(id):
    cid = get_client_id()
    conn = get_db()
    # Vérification d'appartenance AVANT toute lecture/écriture — sans elle, un
    # utilisateur avec un accès écriture sur un seul client pouvait éditer/
    # supprimer les appareils de n'importe quel autre client en changeant
    # juste l'id dans l'URL (aucune des requêtes ci-dessous ne filtrait par
    # client_id). Voir le même correctif sur supprimer_appareil ci-dessous.
    if not conn.execute('SELECT 1 FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone():
        conn.close()
        flash('Appareil introuvable', 'danger')
        return redirect(url_for('liste_appareils'))
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    if request.method == 'POST':
        if not can_write():
            flash('Accès en lecture seule — modification non autorisée', 'danger')
            return redirect(url_for('liste_appareils'))
        errs = validate_form([
            ('nom_machine',  'str',   True),
            ('adresse_ip',   'ip',    False),
            ('adresse_mac',  'mac',   False),
            ('date_achat',   'date',  False),
            ('date_fin_garantie', 'date', False),
        ], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        now = _utcnow().isoformat()
        _old = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        svc_id = _resolve_service_id(conn, cid, request.form.get('service', ''))
        usr_id = _resolve_utilisateur_id(conn, cid, request.form.get('utilisateur', ''))
        vals = _extract_form(request.form) + (now, svc_id, usr_id, id, cid)
        conn.execute('''UPDATE appareils SET nom_machine=?,type_appareil=?,marque=?,modele=?,numero_serie=?,
            adresse_ip=?,adresse_mac=?,nom_dns=?,utilisateur=?,service=?,localisation=?,date_achat=?,
            duree_garantie=?,date_fin_garantie=?,fournisseur=?,prix_achat=?,numero_commande=?,os=?,
            version_os=?,ram=?,cpu=?,stockage=?,carte_graphique=?,statut=?,notes=?,
            user_login=?,user_password=?,admin_login=?,admin_password=?,anydesk_id=?,anydesk_password=?,
            av_marque=?,av_nom=?,av_date_debut=?,av_date_fin=?,av_contrat_id=?,
            edr_marque=?,edr_nom=?,edr_date_fin=?,edr_contrat_id=?,
            rmm_marque=?,rmm_nom=?,rmm_agent_id=?,rmm_date_fin=?,rmm_contrat_id=?,
            logiciels=?,garantie_alerte_ignoree=?,date_maj=?,service_id=?,utilisateur_id=?
            WHERE id=? AND client_id=?''', vals)
        nom = request.form.get('nom_machine','') or f'Appareil #{id}'
        _cols_a = _ENTITE_COLS['appareil']
        _details_a = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_a},
                                 {k: str(request.form.get(k,'') or '') for k in _cols_a})
        _save_licences(conn, id, cid, request.form)
        log_history(conn, cid, 'appareil', id, nom, 'Modification', _details_a)
        _sync_appareil_to_periph(conn, id, cid)
        _propager_utilisateur_aux_peripheriques(
            conn, id, cid, _old.get('utilisateur', ''), request.form.get('utilisateur', ''))
        conn.commit(); conn.close()
        flash('Appareil mis à jour', 'success')
        return redirect(url_for('liste_appareils'))
    a = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, appareil_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_appareils WHERE appareil_id=? ORDER BY date_upload DESC', (id,)).fetchall()]
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
    contrats = [row_to_dict(r) for r in conn.execute(
        "SELECT id,titre,fournisseur,statut FROM contrats WHERE client_id=? ORDER BY titre", (cid,)).fetchall()]
    utilisateurs_noms, utilisateurs_variantes = _utilisateurs_pour_formulaire(conn, cid)
    services_noms = _services_pour_formulaire(conn, cid)
    licences = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM licences_appareils WHERE appareil_id=? ORDER BY id', (id,)).fetchall()]
    lm_list = _get_logiciels_metier_list(conn, cid)
    # Clés de récupération BitLocker : seules les métadonnées voyagent jusqu'à
    # la page. La valeur reste chiffrée en base et ne part qu'à la demande,
    # comme pour les mots de passe des identifiants.
    cles_bitlocker = [row_to_dict(r) for r in conn.execute(
        'SELECT volume, identifiant, protection, chiffrement, date_maj '
        'FROM cles_recuperation WHERE appareil_id=? AND client_id=? ORDER BY volume',
        (id, cid)).fetchall()]
    # Demandé : retrouver depuis la fiche appareil les ports de baie qui le
    # référencent (un appareil peut être câblé sur plusieurs ports — deux
    # cartes réseau, une carte de gestion...).
    # UNION avec baie_prises_murales : un appareil câblé via le système de
    # prise murale d'un bandeau RJ (voir _prises_murales_avec_details) n'a
    # plus son appareil_id sur le PORT lui-même mais sur sa prise murale —
    # sans cette union, il aurait silencieusement disparu de cette section
    # dès la migration prises murales (voir init_db()).
    ports_baie = [row_to_dict(r) for r in conn.execute(
        '''SELECT bp.numero, bs.id AS slot_id, bs.baie_nom, bs.position,
                  bs.nom_custom, bs.type_equipement, 'port' AS origine
           FROM baie_slot_ports bp JOIN baie_slots bs ON bp.slot_id=bs.id
           WHERE bp.appareil_id=? AND bs.client_id=?
           UNION ALL
           SELECT pm.numero, bs.id AS slot_id, bs.baie_nom, bs.position,
                  bs.nom_custom, bs.type_equipement, 'prise_murale' AS origine
           FROM baie_prises_murales pm JOIN baie_slots bs ON pm.slot_id=bs.id
           WHERE pm.appareil_id=? AND bs.client_id=?
           ORDER BY baie_nom, position, numero''', (id, cid, id, cid)).fetchall()]
    conn.close()
    try:
        sw_sel = json.loads(a.get('logiciels') or '[]')
        if not isinstance(sw_sel, list): sw_sel = []
    except Exception:
        sw_sel = []
    lm_set = set(lm_list)
    sw_custom_sel = [sw for sw in sw_sel if sw not in SW_COURANTS_ALL and sw not in lm_set]

    return render_template('form_appareil.html', appareil=a, documents=docs, action='Modifier',
                           cles_bitlocker=cles_bitlocker,
                           ports_baie=ports_baie,
                           utilisateurs_noms=utilisateurs_noms,
                           utilisateurs_variantes=utilisateurs_variantes,
                           services_noms=services_noms,
                           types_appareils=get_liste_cached('types_appareils'),
                           marques_av=get_liste('marques_antivirus'),
                           noms_av=get_liste('noms_antivirus'),
                           marques_edr=get_liste('marques_edr'),
                           noms_edr=get_liste('noms_edr'),
                           marques_rmm=get_liste('marques_rmm'),
                           noms_rmm=get_liste('noms_rmm'),
                           contrats=contrats,
                           sw_courants_groups=SW_COURANTS_GROUPS,
                           sw_courants_all=list(SW_COURANTS_ALL),
                           logiciels_metier_list=lm_list,
                           sw_sel=sw_sel,
                           sw_custom_sel=sw_custom_sel,
                           licences=licences, can_write_flag=can_write(),
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/appareil/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_appareil(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_appareils'))
    cid = get_client_id()
    conn = get_db()
    conn.execute('PRAGMA foreign_keys = ON')
    # id seul ne suffit pas : sans le filtre client_id, un utilisateur avec un
    # accès écriture sur un seul client pouvait supprimer l'appareil de
    # n'importe quel autre client en changeant l'id dans l'URL.
    a = row_to_dict(conn.execute('SELECT nom_machine FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if not a:
        conn.close()
        flash('Appareil introuvable', 'danger')
        return redirect(url_for('liste_appareils'))
    log_history(conn, cid, 'appareil', id, a.get('nom_machine','?'), 'Suppression')
    # Tables sans FK déclarée (ajoutées via ALTER TABLE) — non couvertes par le
    # PRAGMA ci-dessus, nettoyage manuel nécessaire pour éviter les orphelins.
    conn.execute('DELETE FROM cles_recuperation WHERE appareil_id=? AND client_id=?', (id, cid))
    conn.execute('DELETE FROM collectes WHERE appareil_id=? AND client_id=?', (id, cid))
    conn.execute('UPDATE identifiants SET appareil_id=NULL WHERE appareil_id=? AND client_id=?', (id, cid))
    conn.execute('''UPDATE baie_slot_ports SET appareil_id=NULL WHERE appareil_id=? AND slot_id IN
        (SELECT id FROM baie_slots WHERE client_id=?)''', (id, cid))
    conn.execute('''UPDATE baie_prises_murales SET appareil_id=NULL WHERE appareil_id=? AND slot_id IN
        (SELECT id FROM baie_slots WHERE client_id=?)''', (id, cid))
    conn.execute('DELETE FROM appareils WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Appareil supprimé', 'info')
    return redirect(url_for('liste_appareils'))

@app.route('/appareil/<int:id>/rdp')
@login_required
def telecharger_rdp(id):
    """Génère et télécharge un fichier .rdp valide pour lancer une session RDP"""
    cid = get_client_id()
    conn = get_db()
    appareil = row_to_dict(conn.execute(
        'SELECT id, nom_machine, adresse_ip FROM appareils WHERE id=? AND client_id=?',
        (id, cid)
    ).fetchone() or {})
    conn.close()

    if not appareil or not appareil.get('adresse_ip'):
        flash('Appareil introuvable ou sans adresse IP', 'danger')
        return redirect(url_for('liste_appareils'))

    ip = appareil['adresse_ip'].strip()
    nom = (appareil.get('nom_machine') or 'rdp').replace(' ', '_').replace('/', '_')

    # Contenu fichier RDP - version minimale qui fonctionne
    rdp_lines = [
        'full address:s:' + ip,
        'prompt for credentials:i:1',
        'username:s:',
        'domain:s:',
        'desktopwidth:i:1920',
        'desktopheight:i:1080',
        'session bpp:i:32',
        'compression:i:1',
        'keyboardhook:i:2',
        'audiocapturemode:i:0',
        'videoplaybackmode:i:1',
        'connection type:i:7',
        'networkautodetect:i:1',
        'bandwidthautodetect:i:1',
        'displayconnectionbar:i:1',
        'redirectclipboard:i:1',
    ]

    # Joindre les lignes avec des sauts de ligne Windows (CRLF)
    rdp_content = '\r\n'.join(rdp_lines) + '\r\n'

    response = make_response(rdp_content)
    response.headers['Content-Type'] = 'application/x-rdp'
    # RFC 6266 : un en-tête HTTP ne transporte que de l'ASCII. Un nom de machine
    # accentué (« Bureau-Réception ») produisait un nom de fichier corrompu, car
    # Werkzeug encode l'en-tête en latin-1. On fournit donc une version ASCII de
    # repli et la version UTF-8 percent-encodée que lisent les navigateurs.
    import unicodedata as _ud
    from urllib.parse import quote as _quote
    _nom_fichier = f'{nom}_{ip}.rdp'
    _ascii = _ud.normalize('NFKD', _nom_fichier).encode('ascii', 'ignore').decode() or 'connexion.rdp'
    _ascii = _ascii.replace('"', '')
    response.headers['Content-Disposition'] = (
        "attachment; filename=\"%s\"; filename*=UTF-8''%s"
        % (_ascii, _quote(_nom_fichier, safe=''))
    )

    return response

@app.route('/plans')
@login_required
def liste_plans():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    plans = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom,description,date_creation,date_maj FROM plans WHERE client_id=? ORDER BY date_maj DESC',
        (cid,)).fetchall()]
    conn.close()
    return render_template('liste_plans.html', plans=plans, client=client,
                           clients=get_clients(), client_actif_id=cid)


@app.route('/plan/nouveau', methods=['POST'])
@login_required
def nouveau_plan():
    if not can_write():
        flash('Accès en lecture seule', 'danger')
        return redirect(url_for('liste_plans'))
    cid = get_client_id()
    nom = (request.form.get('nom') or '').strip() or 'Nouveau plan'
    desc = request.form.get('description', '')
    now = _utcnow().isoformat()
    conn = get_db()
    conn.execute(
        "INSERT INTO plans (client_id,nom,description,contenu,date_creation,date_maj) VALUES (?,?,?,?,?,?)",
        (cid, nom, desc, '{"elements":[]}', now, now))
    plan_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.commit(); conn.close()
    return redirect(url_for('editer_plan', id=plan_id))


@app.route('/plan/<int:id>')
@login_required
def editer_plan(id):
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    plan = row_to_dict(conn.execute(
        'SELECT * FROM plans WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if not plan:
        conn.close()
        flash('Plan introuvable', 'danger')
        return redirect(url_for('liste_plans'))
    appareils = [row_to_dict(r) for r in conn.execute(
        "SELECT id,nom_machine,type_appareil,statut,en_ligne FROM appareils WHERE client_id=? ORDER BY nom_machine",
        (cid,)).fetchall()]
    for a in appareils:
        k = type_css_filter(a.get('type_appareil', ''))
        a['color'] = cfg_get(f'type_color_{k}') or '#2563eb'
    conn.close()
    # Désérialiser le contenu JSON stocké en base pour éviter le double-encodage en template
    try:
        plan['contenu'] = json.loads(plan.get('contenu') or '{"elements":[]}')
    except Exception:
        plan['contenu'] = {'elements': []}
    return render_template('plan_editeur.html', plan=plan, appareils=appareils,
                           client=client, clients=get_clients(), client_actif_id=cid)


@app.route('/api/plan/<int:id>/sauvegarder', methods=['POST'])
@login_required
def api_plan_save(id):
    if not can_write():
        return jsonify({'ok': False, 'error': 'read-only'}), 403
    cid = get_client_id()
    data = request.get_json(force=True, silent=True) or {}
    contenu = json.dumps(data.get('contenu', {'elements': []}), ensure_ascii=False)
    nom = (data.get('nom') or '').strip()
    now = _utcnow().isoformat()
    conn = get_db()
    if nom:
        conn.execute('UPDATE plans SET contenu=?,nom=?,date_maj=? WHERE id=? AND client_id=?',
                     (contenu, nom, now, id, cid))
    else:
        conn.execute('UPDATE plans SET contenu=?,date_maj=? WHERE id=? AND client_id=?',
                     (contenu, now, id, cid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route('/plan/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_plan(id):
    if not can_write():
        flash('Accès en lecture seule', 'danger')
        return redirect(url_for('liste_plans'))
    cid = get_client_id()
    conn = get_db()
    conn.execute('DELETE FROM plans WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Plan supprimé', 'info')
    return redirect(url_for('liste_plans'))


@app.route('/scan')
@login_required
def page_scan():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    parc   = row_to_dict(conn.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,adresse_ip,type_appareil,marque,modele FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    conn.close()
    appareils_ips = [a['adresse_ip'] for a in appareils if a.get('adresse_ip')]
    # Déduire une plage par défaut
    parc_plage = '192.168.1.0/24'
    if appareils_ips:
        try:
            import ipaddress as _ipa
            parc_plage = str(_ipa.ip_interface(appareils_ips[0] + '/24').network)
        except: pass
    # Statut base OUI
    from app import _OUI_FULL, _OUI
    if _OUI_FULL is None:
        _oui_load_full()
    oui_loaded = bool(_OUI_FULL)
    oui_count  = len(_OUI_FULL) if oui_loaded else len(_OUI)
    oui_maj    = cfg_get('oui_derniere_maj', '')

    return render_template('scan_reseau.html', parc=parc, client=client,
                           appareils=appareils, appareils_ips=appareils_ips,
                           parc_plage=parc_plage, oui_loaded=oui_loaded, oui_count=oui_count,
                           oui_maj=oui_maj,
                           clients=get_clients(), client_actif_id=cid)

@app.route('/api/oui/telecharger', methods=['POST'])
@login_required
def api_oui_telecharger():
    """Télécharge/rafraîchit la base OUI IEEE complète à la demande (bouton
    "Mettre à jour" de la page Scan) — voir _oui_telecharger()."""
    resultat = _oui_telecharger(force=True)
    resultat['oui_count'] = len(_OUI_FULL or {})
    return jsonify(resultat), (200 if resultat.get('ok') else 502)

# ─── SERVICES ────────────────────────────────────────────────────────────────

@app.route('/services')
@login_required
def liste_services():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    services = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()]
    for s in services:
        s['nb_users'] = conn.execute(
            'SELECT COUNT(*) FROM utilisateurs WHERE service_id=? AND statut="actif"', (s['id'],)).fetchone()[0]
        s['nb_appareils'] = conn.execute(
            'SELECT COUNT(*) FROM appareils WHERE service_id=?', (s['id'],)).fetchone()[0]
    conn.close()
    return render_template('services.html', services=services, client=client,
                           clients=get_clients(), client_actif_id=cid)

@app.route('/service/nouveau', methods=['GET','POST'])
@login_required
def nouveau_service():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('index'))
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    conn.close()
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        conn = get_db()
        conn.execute('INSERT INTO services (client_id,nom,description,responsable,couleur,ordre,date_creation,date_maj) VALUES (?,?,?,?,?,?,?,?)',
            (cid, f.get('nom',''), f.get('description',''), f.get('responsable',''),
             f.get('couleur','#6a8aaa'), int(f.get('ordre',0) or 0), now, now))
        conn.commit(); conn.close()
        flash('Service créé', 'success')
        return redirect(url_for('liste_services'))
    return render_template('form_service.html', service=None, action='Nouveau',
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/service/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_service(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('index'))
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        conn.execute('UPDATE services SET nom=?,description=?,responsable=?,couleur=?,ordre=?,date_maj=? WHERE id=? AND client_id=?',
            (f.get('nom',''), f.get('description',''), f.get('responsable',''),
             f.get('couleur','#6a8aaa'), int(f.get('ordre',0) or 0), now, id, cid))
        conn.commit(); conn.close()
        flash('Service mis à jour', 'success')
        return redirect(url_for('liste_services'))
    svc = row_to_dict(conn.execute('SELECT * FROM services WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    return render_template('form_service.html', service=svc, action='Modifier',
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/service/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_service(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('index'))
    cid = get_client_id()
    conn = get_db()
    # appareils.service_id : colonne ajoutée via ALTER TABLE, sans FK
    # déclarée — nettoyage manuel pour éviter les orphelins.
    conn.execute('UPDATE appareils SET service_id=NULL WHERE service_id=? AND client_id=?', (id, cid))
    conn.execute('DELETE FROM services WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Service supprimé', 'info')
    return redirect(url_for('liste_services'))

# ─── TYPES DE DROITS ─────────────────────────────────────────────────────────

CATEGORIES_DROITS = ['Dossiers réseau', 'Logiciels', 'Messagerie', 'Applications web',
                     'Accès physique', 'Administration', 'Autre']

def get_types_droits(cid):
    conn = get_db()
    types = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM types_droits WHERE client_id=? ORDER BY categorie,ordre,nom', (cid,)).fetchall()]
    conn.close()
    return types

@app.route('/types-droits')
@login_required
def liste_types_droits():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    conn.close()
    types = get_types_droits(cid)
    return render_template('types_droits.html', types=types, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           categories_droits=CATEGORIES_DROITS)

@app.route('/api/type-droit', methods=['POST'])
@login_required
def api_creer_type_droit():
    cid = get_client_id()
    f = request.json or {}
    now = _utcnow().isoformat()
    conn = get_db()
    c = conn.execute('INSERT INTO types_droits (client_id,categorie,nom,description,icone,ordre) VALUES (?,?,?,?,?,?)',
        (cid, f.get('categorie','Autre'), f.get('nom',''), f.get('description',''),
         f.get('icone','🔑'), int(f.get('ordre',0) or 0)))
    tid = c.lastrowid
    conn.commit()
    row = row_to_dict(conn.execute('SELECT * FROM types_droits WHERE id=?', (tid,)).fetchone())
    conn.close()
    return jsonify(row)

@app.route('/api/type-droit/<int:id>', methods=['PUT','DELETE'])
@login_required
def api_type_droit(id):
    cid = get_client_id()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute('DELETE FROM types_droits WHERE id=? AND client_id=?', (id, cid))
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    f = request.json or {}
    conn.execute('UPDATE types_droits SET categorie=?,nom=?,description=?,icone=?,ordre=? WHERE id=? AND client_id=?',
        (f.get('categorie',''), f.get('nom',''), f.get('description',''),
         f.get('icone','🔑'), int(f.get('ordre',0) or 0), id, cid))
    conn.commit()
    row = row_to_dict(conn.execute('SELECT * FROM types_droits WHERE id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(row)

# ─── UTILISATEURS ────────────────────────────────────────────────────────────

@app.route('/utilisateurs')
@login_required
def liste_utilisateurs():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    filtre_svc = request.args.get('service', '')
    filtre_statut = request.args.get('statut', 'actif')
    q = 'SELECT u.*, s.nom as service_nom, s.couleur as service_couleur FROM utilisateurs u LEFT JOIN services s ON u.service_id=s.id WHERE u.client_id=?'
    params = [cid]
    if filtre_svc:
        q += ' AND u.service_id=?'; params.append(int(filtre_svc))
    if filtre_statut:
        q += ' AND u.statut=?'; params.append(filtre_statut)
    q += ' ORDER BY s.nom, u.nom, u.prenom'
    users = [row_to_dict(r) for r in conn.execute(q, params).fetchall()]
    services = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()]
    conn.close()
    return render_template('utilisateurs.html', utilisateurs=users, services=services,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           filtre_svc=filtre_svc, filtre_statut=filtre_statut)

@app.route('/utilisateur/nouveau', methods=['GET','POST'])
@login_required
def nouvel_utilisateur():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_utilisateurs'))
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    services = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()]
    conn.close()
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        svc_id = int(f.get('service_id') or 0) or None
        conn = get_db()
        c = conn.execute('''INSERT INTO utilisateurs (client_id,service_id,prenom,nom,poste,email,
            telephone,login_windows,login_mail,statut,notes,date_creation,date_maj)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (cid, svc_id, f.get('prenom',''), f.get('nom',''), f.get('poste',''),
             f.get('email',''), f.get('telephone',''), f.get('login_windows',''),
             f.get('login_mail',''), f.get('statut','actif'), f.get('notes',''), now, now))
        uid = c.lastrowid
        nom_u = (f.get('prenom','') + ' ' + f.get('nom','')).strip() or 'Nouvel utilisateur'
        log_history(conn, cid, 'utilisateur', uid, nom_u, 'Création')
        conn.commit(); conn.close()
        flash('Utilisateur créé', 'success')
        return redirect(url_for('droits_utilisateur', id=uid))
    return render_template('form_utilisateur.html', utilisateur=None, action='Nouveau',
                           services=services, client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/utilisateur/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_utilisateur(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_utilisateurs'))
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    services = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()]
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        _old = row_to_dict(conn.execute('SELECT * FROM utilisateurs WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        svc_id = int(f.get('service_id') or 0) or None
        conn.execute('''UPDATE utilisateurs SET service_id=?,prenom=?,nom=?,poste=?,email=?,
            telephone=?,login_windows=?,login_mail=?,statut=?,notes=?,date_maj=? WHERE id=? AND client_id=?''',
            (svc_id, f.get('prenom',''), f.get('nom',''), f.get('poste',''),
             f.get('email',''), f.get('telephone',''), f.get('login_windows',''),
             f.get('login_mail',''), f.get('statut','actif'), f.get('notes',''), now, id, cid))
        nom = (request.form.get('prenom','') + ' ' + request.form.get('nom','')).strip() or f'Utilisateur #{id}'
        _cols_u = _ENTITE_COLS['utilisateur']
        _details_u = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_u},
                                  {k: str(f.get(k,'') or '') for k in _cols_u})
        log_history(conn, cid, 'utilisateur', id, nom, 'Modification', _details_u)
        conn.commit(); conn.close()
        flash('Utilisateur mis à jour', 'success')
        return redirect(url_for('liste_utilisateurs'))
    u = row_to_dict(conn.execute('SELECT * FROM utilisateurs WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    return render_template('form_utilisateur.html', utilisateur=u, action='Modifier',
                           services=services, client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/utilisateur/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_utilisateur(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_utilisateurs'))
    cid = get_client_id()
    conn = get_db()
    u = row_to_dict(conn.execute('SELECT prenom,nom FROM utilisateurs WHERE id=?',(id,)).fetchone() or {})
    nom = (u.get('prenom','') + ' ' + u.get('nom','')).strip() or '?'
    log_history(conn, cid, 'utilisateur', id, nom, 'Suppression')
    # appareils.utilisateur_id / identifiants.utilisateur_id : colonnes
    # ajoutées via ALTER TABLE, sans FK déclarée — nettoyage manuel pour
    # éviter les orphelins (même pattern que supprimer_appareil/
    # supprimer_peripherique ci-dessus). peripheriques.utilisateur_id a une
    # FK déclarée à l'origine mais son ON DELETE SET NULL ne s'applique que
    # si PRAGMA foreign_keys=ON — jamais activé sur cette connexion jusqu'ici
    # — donc traité manuellement aussi, par prudence.
    conn.execute('UPDATE appareils SET utilisateur_id=NULL WHERE utilisateur_id=? AND client_id=?', (id, cid))
    conn.execute('UPDATE identifiants SET utilisateur_id=NULL WHERE utilisateur_id=? AND client_id=?', (id, cid))
    conn.execute('UPDATE peripheriques SET utilisateur_id=NULL WHERE utilisateur_id=? AND client_id=?', (id, cid))
    conn.execute('DELETE FROM utilisateurs WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Utilisateur supprimé', 'info')
    return redirect(url_for('liste_utilisateurs'))

@app.route('/utilisateur/<int:id>/droits')
@login_required
def droits_utilisateur(id):
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    u = row_to_dict(conn.execute(
        'SELECT u.*, s.nom as service_nom, s.couleur as service_couleur FROM utilisateurs u LEFT JOIN services s ON u.service_id=s.id WHERE u.id=? AND u.client_id=?',
        (id, cid)).fetchone() or {})
    droits = [row_to_dict(r) for r in conn.execute(
        'SELECT d.*, t.icone, t.categorie as t_categorie FROM droits_utilisateurs d LEFT JOIN types_droits t ON d.type_droit_id=t.id WHERE d.utilisateur_id=? ORDER BY d.categorie, d.nom_droit',
        (id,)).fetchall()]
    types = get_types_droits(cid)
    appareils_affectes = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine, type_appareil FROM appareils WHERE utilisateur_id=? AND client_id=? ORDER BY nom_machine',
        (id, cid)).fetchall()]
    identifiants_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom, categorie, login FROM identifiants WHERE utilisateur_id=? AND client_id=? ORDER BY nom',
        (id, cid)).fetchall()]
    conn.close()
    # Grouper par catégorie
    cats = {}
    for d in droits:
        cat = d.get('categorie') or 'Autre'
        cats.setdefault(cat, []).append(d)
    nb_droits_total = sum(len(v) for v in cats.values())
    return render_template('droits_utilisateur.html', utilisateur=u, droits_par_cat=cats,
                           nb_droits_total=nb_droits_total, appareils_affectes=appareils_affectes,
                           identifiants_lies=identifiants_lies,
                           types=types, categories_droits=CATEGORIES_DROITS,
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/api/droit', methods=['POST'])
@login_required
def api_ajouter_droit():
    cid = get_client_id()
    f = request.json or {}
    uid = f.get('utilisateur_id')
    now = _utcnow().isoformat()
    conn = get_db()
    c = conn.execute('''INSERT INTO droits_utilisateurs
        (utilisateur_id, client_id, categorie, type_droit_id, nom_droit, valeur, niveau, notes, date_attribution)
        VALUES (?,?,?,?,?,?,?,?,?)''',
        (uid, cid, f.get('categorie',''), f.get('type_droit_id') or None,
         f.get('nom_droit',''), f.get('valeur',''), f.get('niveau','lecture'),
         f.get('notes',''), now))
    did = c.lastrowid
    conn.commit()
    row = row_to_dict(conn.execute('SELECT * FROM droits_utilisateurs WHERE id=?', (did,)).fetchone())
    conn.close()
    return jsonify(row)

@app.route('/api/droit/<int:id>', methods=['PUT','DELETE'])
@login_required
def api_droit(id):
    cid = get_client_id()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute('DELETE FROM droits_utilisateurs WHERE id=?', (id,))
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    f = request.json or {}
    conn.execute('UPDATE droits_utilisateurs SET categorie=?,nom_droit=?,valeur=?,niveau=?,notes=? WHERE id=?',
        (f.get('categorie',''), f.get('nom_droit',''), f.get('valeur',''),
         f.get('niveau','lecture'), f.get('notes',''), id))
    conn.commit()
    row = row_to_dict(conn.execute('SELECT * FROM droits_utilisateurs WHERE id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(row)

@app.route('/api/utilisateurs')
@login_required
def api_utilisateurs():
    cid = get_client_id()
    conn = get_db()
    users = [row_to_dict(r) for r in conn.execute(
        'SELECT id, prenom, nom, service_id FROM utilisateurs WHERE client_id=? AND statut="actif" ORDER BY nom', (cid,)).fetchall()]
    conn.close()
    return jsonify(users)

# ─── IDENTIFIANTS ────────────────────────────────────────────────────────────

@app.route('/identifiants')
@login_required
def liste_identifiants():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    page = request.args.get('page', 1, type=int)
    filtre_cat = request.args.get('cat', '')
    _q_base = ('SELECT i.*, a.nom_machine AS lie_appareil_nom, '
               'p.categorie AS lie_periph_categorie, p.marque AS lie_periph_marque, p.modele AS lie_periph_modele, '
               'u.prenom AS lie_util_prenom, u.nom AS lie_util_nom '
               'FROM identifiants i '
               'LEFT JOIN appareils a ON a.id = i.appareil_id '
               'LEFT JOIN peripheriques p ON p.id = i.peripherique_id '
               'LEFT JOIN utilisateurs u ON u.id = i.utilisateur_id '
               'WHERE i.client_id=?')
    if filtre_cat:
        q, params = _q_base + ' AND i.categorie=? ORDER BY i.categorie, i.nom', (cid, filtre_cat)
    else:
        q, params = _q_base + ' ORDER BY i.categorie, i.nom', (cid,)
    rows, pagination = paginate(q, params, page)
    ids_ = [row_to_dict(r) for r in rows]
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    cats = [r[0] for r in conn.execute(
        'SELECT DISTINCT categorie FROM identifiants WHERE client_id=? ORDER BY categorie', (cid,)).fetchall()]

    # Générer les statistiques par catégorie
    stats = {'total': 0}
    total_result = conn.execute('SELECT COUNT(*) FROM identifiants WHERE client_id=?', (cid,)).fetchone()
    stats['total'] = total_result[0] if total_result else 0
    for cat in cats:
        count_result = conn.execute('SELECT COUNT(*) FROM identifiants WHERE client_id=? AND categorie=?', (cid, cat)).fetchone()
        stats[cat] = count_result[0] if count_result else 0

    conn.close()
    crypto = _get_crypto_shared()
    for i in ids_:
        if i.get('mot_de_passe'):
            i['mot_de_passe'] = crypto.decrypt(i['mot_de_passe']) or i['mot_de_passe']
        if i.get('date_expiration'):
            try:
                d = date.fromisoformat(i['date_expiration'])
                i['expire_bientot'] = (d - date.today()).days <= 30
                i['expire_depasse'] = d < date.today()
                i['date_expiration_fmt'] = d.strftime('%d/%m/%Y')
            except: i['expire_bientot'] = i['expire_depasse'] = False; i['date_expiration_fmt'] = ''
        else: i['expire_bientot'] = i['expire_depasse'] = False; i['date_expiration_fmt'] = ''
    # Récupérer les identifiants WiFi du parc général (en lecture seule)
    conn2 = get_db()
    parc = row_to_dict(conn2.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    conn2.close()
    wifi_parc = []
    if parc.get('wifi_ssid'):
        wifi_parc.append({
            'id': None, 'from_parc': True,
            'nom': parc['wifi_ssid'] + ' (Parc général)',
            'categorie': 'Wi-Fi',
            'login': parc.get('wifi_ssid',''),
            'mot_de_passe': parc.get('wifi_password',''),
            'wifi_ssid': parc.get('wifi_ssid',''),
            'wifi_securite': parc.get('wifi_securite','WPA2'),
            'description': 'Réseau principal — depuis le Parc général',
            'url': '', 'notes': parc.get('wifi_notes',''),
            'expire_bientot': False, 'expire_depasse': False, 'date_expiration_fmt': '',
        })
    if parc.get('wifi_ssid2'):
        wifi_parc.append({
            'id': None, 'from_parc': True,
            'nom': parc['wifi_ssid2'] + ' (Parc général)',
            'categorie': 'Wi-Fi',
            'login': parc.get('wifi_ssid2',''),
            'mot_de_passe': parc.get('wifi_password2',''),
            'wifi_ssid': parc.get('wifi_ssid2',''),
            'wifi_securite': parc.get('wifi_securite2','WPA2'),
            'description': 'Réseau invités — depuis le Parc général',
            'url': '', 'notes': '',
            'expire_bientot': False, 'expire_depasse': False, 'date_expiration_fmt': '',
        })
    return render_template('identifiants.html', identifiants=ids_, wifi_parc=wifi_parc, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_identifiants'), cats_utilisees=cats,
                           filtre_cat=filtre_cat, pagination=pagination, stats=stats)

@app.route('/identifiant/nouveau', methods=['GET','POST'])
@login_required
def nouvel_identifiant():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_identifiants'))
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils_liste = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    peripheriques_liste = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie, marque', (cid,)).fetchall()]
    utilisateurs_liste = [row_to_dict(r) for r in conn.execute(
        "SELECT id, prenom, nom FROM utilisateurs WHERE client_id=? AND statut='actif' ORDER BY nom", (cid,)).fetchall()]
    conn.close()
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        errs = validate_form([
            ('nom',            'str',   True),
            ('url',            'url',   False),
            ('date_expiration','date',  False),
        ], f)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        conn = get_db()
        # ✅ Chiffrer le mot de passe avant stockage
        crypto = _get_crypto_shared()
        mdp_chiffre = crypto.encrypt(f.get('mot_de_passe','')) if f.get('mot_de_passe') else ''
        appareil_id = int(f['appareil_id']) if f.get('appareil_id') else None
        peripherique_id = int(f['peripherique_id']) if f.get('peripherique_id') else None
        utilisateur_id = int(f['utilisateur_id']) if f.get('utilisateur_id') else None
        conn.execute('''INSERT INTO identifiants (client_id,categorie,nom,login,mot_de_passe,url,
            description,notes,date_expiration,wifi_ssid,wifi_securite,appareil_id,peripherique_id,
            utilisateur_id,date_creation,date_maj)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (cid, f.get('categorie',''), f.get('nom',''), f.get('login',''), mdp_chiffre,
             f.get('url',''), f.get('description',''), f.get('notes',''),
             f.get('date_expiration',''),
             f.get('wifi_ssid','') if f.get('categorie') == 'Wi-Fi' else '',
             f.get('wifi_securite','WPA2') if f.get('categorie') == 'Wi-Fi' else '',
             appareil_id, peripherique_id, utilisateur_id,
             now, now))
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        nom = request.form.get('nom','') or 'Nouvel identifiant'
        log_history(conn, cid, 'identifiant', new_id, nom, 'Création')
        conn.commit(); conn.close()
        flash('Identifiant ajouté', 'success')
        return redirect(url_for('liste_identifiants'))
    return render_template('form_identifiant.html', identifiant=None, action='Ajouter',
                           appareils_liste=appareils_liste, peripheriques_liste=peripheriques_liste,
                           utilisateurs_liste=utilisateurs_liste,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_identifiants'))

@app.route('/identifiant/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_identifiant(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_identifiants'))
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    if request.method == 'POST':
        f = request.form; now = _utcnow().isoformat()
        errs = validate_form([
            ('nom',            'str',   True),
            ('url',            'url',   False),
            ('date_expiration','date',  False),
        ], f)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        _old = row_to_dict(conn.execute('SELECT * FROM identifiants WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        # ✅ Chiffrer le mot de passe avant mise à jour
        crypto = _get_crypto_shared()
        mdp_chiffre = crypto.encrypt(f.get('mot_de_passe','')) if f.get('mot_de_passe') else ''
        appareil_id = int(f['appareil_id']) if f.get('appareil_id') else None
        peripherique_id = int(f['peripherique_id']) if f.get('peripherique_id') else None
        utilisateur_id = int(f['utilisateur_id']) if f.get('utilisateur_id') else None
        conn.execute('''UPDATE identifiants SET categorie=?,nom=?,login=?,mot_de_passe=?,url=?,
            description=?,notes=?,date_expiration=?,wifi_ssid=?,wifi_securite=?,
            appareil_id=?,peripherique_id=?,utilisateur_id=?,date_maj=?
            WHERE id=? AND client_id=?''',
            (f.get('categorie',''), f.get('nom',''), f.get('login',''), mdp_chiffre,
             f.get('url',''), f.get('description',''), f.get('notes',''),
             f.get('date_expiration',''),
             f.get('wifi_ssid','') if f.get('categorie') == 'Wi-Fi' else '',
             f.get('wifi_securite','WPA2') if f.get('categorie') == 'Wi-Fi' else '',
             appareil_id, peripherique_id, utilisateur_id,
             now, id, cid))
        nom = request.form.get('nom','') or f'Identifiant #{id}'
        _cols_i = _ENTITE_COLS['identifiant']
        _details_i = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_i},
                                  {k: str(f.get(k,'') or '') for k in _cols_i})
        log_history(conn, cid, 'identifiant', id, nom, 'Modification', _details_i)
        conn.commit(); conn.close()
        flash('Identifiant mis à jour', 'success')
        return redirect(url_for('liste_identifiants'))
    ident = row_to_dict(conn.execute('SELECT * FROM identifiants WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    appareils_liste = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    peripheriques_liste = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie, marque', (cid,)).fetchall()]
    utilisateurs_liste = [row_to_dict(r) for r in conn.execute(
        "SELECT id, prenom, nom FROM utilisateurs WHERE client_id=? AND statut='actif' ORDER BY nom", (cid,)).fetchall()]
    conn.close()
    # ✅ Déchiffrer le mot de passe pour l'affichage
    if ident and ident.get('mot_de_passe'):
        crypto = _get_crypto_shared()
        ident['mot_de_passe'] = crypto.decrypt(ident['mot_de_passe']) or ident['mot_de_passe']
    return render_template('form_identifiant.html', identifiant=ident, action='Modifier',
                           appareils_liste=appareils_liste, peripheriques_liste=peripheriques_liste,
                           utilisateurs_liste=utilisateurs_liste,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_identifiants'))

@app.route('/identifiant/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_identifiant(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_identifiants'))
    cid = get_client_id()
    conn = get_db()
    # Sans le filtre client_id, cette lecture pouvait révéler le nom d'un
    # identifiant appartenant à un autre client dans l'historique du client
    # actif (le DELETE ci-dessous, lui, était déjà correctement scopé — rien
    # n'était donc réellement supprimé, seul le nom fuitait dans le journal).
    idn = row_to_dict(conn.execute('SELECT nom FROM identifiants WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    log_history(conn, cid, 'identifiant', id, idn.get('nom','?'), 'Suppression')
    conn.execute('DELETE FROM identifiants WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Identifiant supprimé', 'info')
    return redirect(url_for('liste_identifiants'))

@app.route('/api/appareil/<int:id>/cle-bitlocker')
@login_required
def api_cle_bitlocker(id):
    """Déchiffre une clé de récupération BitLocker, à la demande.

    Chaque consultation est inscrite à l'historique : une clé qui déverrouille
    un disque mérite qu'on sache qui l'a lue et quand, ce qui n'a pas de coût
    ici puisque l'historique existe déjà.
    """
    cid = get_client_id()
    volume = (request.args.get('volume') or '').strip()

    conn = get_db()
    try:
        ligne = conn.execute(
            'SELECT valeur, volume FROM cles_recuperation '
            'WHERE appareil_id=? AND client_id=? AND volume=?',
            (id, cid, volume)).fetchone()
        if not ligne:
            return jsonify({'error': 'not found'}), 404

        crypto = _get_crypto_shared()
        claire = crypto.decrypt(ligne[0]) if ligne[0] else ''

        user = get_auth_user()
        appareil = conn.execute(
            'SELECT nom_machine FROM appareils WHERE id=? AND client_id=?',
            (id, cid)).fetchone()
        log_history(conn, cid, 'appareil', id,
                    appareil[0] if appareil else str(id),
                    'Consultation d\'une clé de récupération BitLocker',
                    {'volume': ligne[1], 'par': (user or {}).get('login')})
        conn.commit()
    finally:
        conn.close()

    return jsonify({'cle': claire, 'volume': ligne[1]})


@app.route('/api/identifiant/<int:id>/mdp')
@login_required
def api_get_mdp(id):
    cid = get_client_id()
    conn = get_db()
    row = conn.execute('SELECT mot_de_passe FROM identifiants WHERE id=? AND client_id=?', (id, cid)).fetchone()
    conn.close()
    if not row: return jsonify({'error': 'not found'}), 404
    # ✅ Déchiffrer le mot de passe
    crypto = _get_crypto_shared()
    mdp_dechiffre = crypto.decrypt(row[0]) if row[0] else ''
    return jsonify({'mdp': mdp_dechiffre})

# ─── DOCUMENTS APPAREILS ─────────────────────────────────────────────────────

def _fiche_systeme_disques(rapport):
    """Décompose les volumes logiques pour un affichage en barres.

    `disk_drives` est une liste de lignes déjà formatées par le collecteur, dont
    le format diffère selon l'OS. `_parse_drive` sait lire celui de Windows ;
    quand la ligne n'est pas reconnue (macOS, Linux), on la restitue telle
    quelle plutôt que d'inventer un pourcentage.
    """
    from collector_core import _disk_level, _parse_drive

    volumes = []
    for ligne in rapport.get('disk_drives') or []:
        analyse = _parse_drive(str(ligne))
        if analyse and analyse[1] and analyse[2] is not None:
            libelle, total, utilise, libre = analyse
            pct = round(utilise / total * 100) if total else 0
            volumes.append({
                'libelle': libelle, 'total': total, 'utilise': utilise,
                'libre': libre if libre is not None else total - utilise,
                'pct': pct, 'level': _disk_level(pct), 'brut': None,
            })
        else:
            volumes.append({'libelle': '', 'brut': str(ligne)})
    return volumes


def _fiche_systeme_disques_physiques(rapport):
    """Décompose les disques physiques pour un tableau aligné.

    Le collecteur assemble ces lignes à partir de champs distincts ; les
    afficher telles quelles obligeait à lire la phrase entière pour retrouver
    la capacité ou l'état SMART. Comme pour les volumes logiques, une ligne au
    format non reconnu est restituée brute.
    """
    from collector_core import parse_physical_disk

    disques = []
    for ligne in rapport.get('physical_disks') or []:
        disques.append(parse_physical_disk(str(ligne))
                       or {'nom': '', 'brut': str(ligne)})
    return disques


#: Couleur fixe par type de partition, la même sur tous les disques de la
#: fiche système : contrairement au rouge/orange/vert « niveau de
#: remplissage » (déjà porté par le pourcentage affiché sur le segment), ce
#: code n'a pas vocation à changer avec l'usage — une partition EFI est
#: toujours sarcelle, qu'elle soit pleine à 10 % ou 90 %. Clé = type traduit
#: en minuscules (voir `_TYPE_PARTITION` dans collector_core.py) ; « données »
#: couvre aussi bien Windows (Basic/IFS) que macOS/Linux (type non détecté,
#: mais une lettre/un point de montage prouve qu'il s'agit de données).
_COULEUR_PARTITION = {
    'données': '#00c9ff',
    'réservé (msr)': '#a78bfa',
    'récupération': '#fbbf24',
    'système efi': '#2dd4bf',
    'non attribué': '#64748b',
}
_COULEUR_PARTITION_INCONNU = '#475569'


def _fiche_systeme_disk_layout(rapport):
    """Regroupe les partitions par disque physique pour la vue « un disque,
    ses partitions dedans » de la fiche système.

    `disk_layout` vient du collecteur (3.1+, `_win_extras`/`_unix_disks` dans
    collector_core.py) : une fiche collectée avant cet ajout n'a pas ce champ,
    et la vue retombe alors sur l'ancienne carte à plat (gérée directement
    dans le template).

    Windows expose désormais TOUTES les partitions, pas seulement celles avec
    une lettre de lecteur : réservée système (MSR), EFI et récupération en
    ont, et n'en restent pas moins des partitions réelles à montrer — un
    disque qui n'affichait qu'une seule partition sur cinq laissait croire à
    de l'espace non identifié là où il n'y en avait pas. macOS/Linux n'ont pas
    cette granularité (`df` ne voit que les systèmes de fichiers montés) : le
    type y reste vide et la partition s'affiche par son seul nom d'appareil.
    """
    from collector_core import _SANTE_DISQUE, _TYPE_PARTITION

    def _taille_lisible(go):
        if go is None:
            return None
        if go < 1:
            return '%d Mo' % round(go * 1024)
        return '%.1f GB' % go if go < 10 else '%d GB' % round(go)

    def _couleur_partition(non_attribue, lettre, type_libelle):
        if non_attribue:
            cle = 'non attribué'
        elif lettre is not None:
            # Une partition montée/lettrée est par définition des données,
            # même si le type n'a pas pu être traduit (macOS/Linux).
            cle = (type_libelle or 'données').lower()
        else:
            cle = (type_libelle or '').lower()
        return _COULEUR_PARTITION.get(cle, _COULEUR_PARTITION_INCONNU)

    disques = []
    for d in rapport.get('disk_layout') or []:
        partitions = []
        total_partitionne = 0.0
        for p in d.get('partitions') or []:
            total = p.get('total')
            if total is None:
                continue
            pct = p.get('pct')
            lettre = p.get('letter') or None
            type_brut = (p.get('type') or '').strip()
            type_libelle = _TYPE_PARTITION.get(type_brut.lower(), type_brut or None) if type_brut else None
            partitions.append({
                'letter': lettre, 'type': type_libelle, 'total': total,
                'taille_txt': _taille_lisible(total),
                'used': p.get('used'), 'free': p.get('free'), 'pct': pct,
                'non_attribue': False,
                'couleur': _couleur_partition(False, lettre, type_libelle),
            })
            total_partitionne += total

        taille_disque = d.get('size_gb') or round(total_partitionne, 1)
        # Résiduel rare (arrondi, espace jamais partitionné en fin de disque) :
        # les partitions système sans lettre sont désormais listées ci-dessus,
        # ce segment ne comble plus que ce qu'aucune partition ne couvre.
        non_attribue = round(taille_disque - total_partitionne, 2)
        if non_attribue > 0.05:
            partitions.append({
                'letter': None, 'type': None, 'total': non_attribue,
                'taille_txt': _taille_lisible(non_attribue), 'used': None,
                'free': non_attribue, 'pct': None,
                'non_attribue': True, 'couleur': _couleur_partition(True, None, None),
            })

        brut_sante = (d.get('health') or '').strip()
        if brut_sante:
            sante, sante_niveau = _SANTE_DISQUE.get(brut_sante.lower(), (brut_sante, 'muted'))
        else:
            sante, sante_niveau = None, None

        op_status = (d.get('op_status') or '').strip()

        disques.append({
            'number': d.get('number'), 'model': (d.get('model') or '').strip(),
            'media_type': (d.get('media_type') or '').strip(),
            'sante': sante, 'sante_niveau': sante_niveau,
            'etat': op_status if op_status and op_status != 'OK' else None,
            'size_gb': taille_disque, 'partitions': partitions,
        })
    return disques


def _fiche_systeme_kpis(rapport):
    """Vignettes chiffrées de la fiche système : valeur, barre et criticité.

    Les seuils et le calcul sont ceux de `collector_core`, pour que la page et
    le rapport PDF portent exactement le même jugement sur une machine.
    """
    from collector_core import (
        BATTERY_DANGER_PCT, BATTERY_WARN_PCT, _battery_pct, _disk_level, _num,
    )

    kpis = []

    total = _num(rapport.get('disk_total_gb'))
    used = _num(rapport.get('disk_used_gb'))
    if total and used is not None and total > 0:
        pct = round(used / total * 100)
        kpis.append({'label': 'Stockage', 'value': pct, 'unit': '%', 'pct': pct,
                     'level': _disk_level(pct),
                     'sub': '%g GB utilisés sur %g GB' % (used, total)})

    ram = _num(rapport.get('ram_gb'))
    if ram:
        libre = _num(rapport.get('ram_free_gb'))
        if libre is not None and ram > 0:
            pct = round((ram - libre) / ram * 100)
            level = 'danger' if pct >= 90 else 'warn' if pct >= 75 else 'ok'
            sub = '%.1f GB libres sur %g GB' % (libre, ram)
        else:
            # Sans mesure d'occupation, la barre situe la machine sur une
            # échelle 0–64 Go plutôt que d'inventer un pourcentage.
            pct = min(ram / 64 * 100, 100)
            level = 'ok' if ram >= 8 else 'warn'
            sub = 'Confortable' if ram >= 16 else 'Correct' if ram >= 8 else 'Juste'
        kpis.append({'label': 'Mémoire vive', 'value': ('%g' % ram), 'unit': 'GB',
                     'pct': pct, 'level': level, 'sub': sub})

    charge = _battery_pct(rapport)
    if charge is not None:
        level = ('danger' if charge <= BATTERY_DANGER_PCT
                 else 'warn' if charge <= BATTERY_WARN_PCT else 'ok')
        sante = _num(rapport.get('battery_health_percent'))
        sub = ('Santé %g %% de la capacité d\'origine' % sante) if sante is not None else 'Charge restante'
        kpis.append({'label': 'Batterie', 'value': charge, 'unit': '%', 'pct': charge,
                     'level': level, 'sub': sub})

    uptime = _num(rapport.get('uptime_hours'))
    if uptime is not None:
        jours = uptime / 24
        kpis.append({'label': 'Sans redémarrage', 'value': '%.0f' % jours, 'unit': 'j',
                     'pct': min(jours / 30 * 100, 100),
                     'level': 'warn' if jours > 30 else 'ok',
                     'sub': '%.0f heures d\'uptime' % uptime})

    return kpis


@app.route('/appareil/<int:id>/fiche-systeme')
@login_required
def fiche_systeme_appareil(id):
    """Affiche le snapshot complet remonté par le collecteur système.

    Le collecteur remonte bien plus que les colonnes dédiées de `appareils`
    (mémoire par slot, licences, usure SSD, écrans, correctifs…) : tout cela
    est stocké tel quel dans rapport_systeme_json et rendu ici.
    """
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    a = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()

    if not a:
        abort(404)

    def _parse(raw, fallback):
        if not raw:
            return fallback
        try:
            return json.loads(raw)
        except (ValueError, TypeError):
            return fallback

    rapport = _parse(a.get('rapport_systeme_json'), {})
    logiciels = _parse(a.get('logiciels_installes_json'), [])
    # Les collectes antérieures à la v3 envoyaient une liste de chaînes
    logiciels = [s if isinstance(s, dict) else {'name': str(s)} for s in logiciels]

    # Analyse et mise en forme reprises de collector_core : la page porte ainsi
    # le même jugement que le rapport PDF sur une machine donnée.
    alertes, kpis, disques, ports_cartes, ports_masques = [], [], [], [], 0
    disques_physiques = []
    disque_layout = []
    age_materiel = None
    if rapport:
        try:
            from collector_core import (
                build_alerts, describe_listening_port, notable_ports,
            )
            alertes = build_alerts(rapport)
            kpis = _fiche_systeme_kpis(rapport)
            disques = _fiche_systeme_disques(rapport)
            disques_physiques = _fiche_systeme_disques_physiques(rapport)
            disque_layout = _fiche_systeme_disk_layout(rapport)
            from collector_core import hardware_age_years
            age_materiel = hardware_age_years(rapport.get('bios_release_date'))
            ports = [describe_listening_port(p)
                     for p in (rapport.get('listening_ports') or [])
                     if isinstance(p, dict)]
            ports_cartes = notable_ports(ports)
            ports_masques = len(ports) - len(ports_cartes)
        except Exception:
            # Une analyse qui échoue ne doit pas priver l'utilisateur de la
            # fiche : les données brutes restent affichées.
            logger.exception("Analyse de la fiche système")

    # Historique des collectes : la fiche ne montre que le dernier relevé, mais
    # c'est la comparaison avec les précédents qui révèle un disque qui se
    # remplit, un logiciel apparu ou une pièce remplacée.
    conn = get_db()
    try:
        historique = historique_appareil(conn, cid, id)
    except Exception:
        logger.exception("Historique des collectes de l'appareil %s", id)
        historique = None
    finally:
        conn.close()

    return render_template('fiche_systeme.html', appareil=a, rapport=rapport,
                           logiciels=logiciels, client=client, clients=get_clients(),
                           client_actif_id=cid, alertes=alertes, kpis=kpis,
                           disques=disques, disques_physiques=disques_physiques,
                           disque_layout=disque_layout,
                           ports_cartes=ports_cartes,
                           ports_masques=ports_masques, age_materiel=age_materiel,
                           historique=historique)


@app.route('/appareil/<int:id>/documents')
@login_required
def documents_appareil(id):
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    a = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, appareil_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_appareils WHERE appareil_id=? ORDER BY date_upload DESC', (id,)).fetchall()]

    # Fetch related interventions
    interventions = [fmt_intervention(row_to_dict(r)) for r in conn.execute(
        'SELECT i.* FROM interventions i JOIN interventions_appareils ia ON i.id=ia.intervention_id '
        'WHERE ia.appareil_id=? AND i.statut != ? ORDER BY i.date_intervention DESC LIMIT 10',
        (id, 'archivee')).fetchall()]

    conn.close()
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
    return render_template('documents_appareil.html', appareil=a, documents=docs, interventions=interventions,
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/appareil/<int:id>/documents/upload', methods=['POST'])
@login_required
def upload_document(id):
    cid = get_client_id()
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('documents_appareil', id=id))
    f = request.files['fichier']
    ok, motif = verifier_fichier(f)
    if not ok:
        flash(motif, 'danger')
        return redirect(url_for('documents_appareil', id=id))
    safe = secure_filename(f.filename)
    if not safe:
        flash('Nom de fichier invalide', 'danger')
        return redirect(url_for('documents_appareil', id=id))
    unique = f"app{id}_{int(time.time())}_{safe}"
    save_path = os.path.join(UPLOAD_FOLDER, unique)
    logger.info(f"Upload document appareil #{id}: saving to {save_path}")
    try:
        f.save(save_path)
        taille = os.path.getsize(save_path)
        logger.info(f"Upload document appareil #{id}: saved {taille} bytes")
    except Exception as e:
        logger.exception(f"Upload document appareil #{id}: save FAILED → {save_path}")
        flash(f'Erreur lors de la sauvegarde du fichier : {e}', 'danger')
        return redirect(url_for('documents_appareil', id=id))

    nom = request.form.get('nom', '') or f.filename
    desc = request.form.get('description', '')
    type_doc = request.form.get('type_doc', '')
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute('''INSERT INTO documents_appareils
        (appareil_id,client_id,nom,description,type_doc,nom_fichier,taille,date_upload,contenu_blob,sync_status,date_sync)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (id, cid, nom, desc, type_doc, unique, taille, now, None, 'local', ''))

    # Log document upload
    app_title = conn.execute('SELECT nom_machine FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone()
    app_name = app_title[0] if app_title else f'Appareil #{id}'
    log_history(conn, cid, 'appareil', id, app_name, 'Ajout de document',
                _diff_json({}, {'nom': nom, 'fichier': unique, 'type_doc': type_doc}))

    conn.commit(); conn.close()
    flash(f'Document « {nom} » uploadé avec succès', 'success')
    next_url = request.form.get('next') or url_for('editer_appareil', id=id)
    return redirect(next_url)

@app.route('/document/<int:id>/telecharger')
@login_required
def telecharger_document(id):
    cid = get_client_id()
    if not cid:
        flash('Vous devez être connecté pour télécharger un document', 'danger')
        return redirect(url_for('page_login'))

    conn = get_db()
    try:
        doc = row_to_dict(conn.execute('SELECT id, appareil_id, client_id, nom, nom_fichier, contenu_blob FROM documents_appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        if not doc:
            flash('Document introuvable', 'danger')
            return redirect(url_for('liste_appareils'))

        # Préférer servir depuis BLOB si disponible (synced)
        if doc.get('contenu_blob'):
            return send_file(
                io.BytesIO(doc['contenu_blob']),
                as_attachment=True,
                download_name=doc.get('nom_fichier') or doc['nom']
            )

        # Fallback: servir depuis fichier local
        if doc.get('nom_fichier'):
            fichier_path = os.path.join(UPLOAD_FOLDER, doc['nom_fichier'])
            if os.path.exists(fichier_path):
                return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=True, download_name=doc['nom'])

        flash('Le fichier n\'existe plus', 'danger')
        return redirect(url_for('liste_appareils'))
    finally:
        conn.close()

@app.route('/document/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_document(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_appareils'))
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute('SELECT id, appareil_id, client_id, nom, nom_fichier, contenu_blob FROM documents_appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    appareil_id = doc.get('appareil_id', 0)
    if doc:
        conn.execute('DELETE FROM documents_appareils WHERE id=?', (id,))

        # Log document deletion
        app_title = conn.execute('SELECT nom_machine FROM appareils WHERE id=? AND client_id=?', (appareil_id, cid)).fetchone()
        app_name = app_title[0] if app_title else f'Appareil #{appareil_id}'
        log_history(conn, cid, 'appareil', appareil_id, app_name, 'Suppression de document',
                    _diff_json({'nom': doc.get('nom', ''), 'fichier': doc.get('nom_fichier', '')}, {}))

        conn.commit()
        try:
            os.remove(os.path.join(UPLOAD_FOLDER, doc['nom_fichier']))
        except:
            pass
    conn.close()
    next_url = request.args.get('next') or url_for('editer_appareil', id=appareil_id)
    flash('Document supprimé', 'info')
    return redirect(next_url)

@app.route('/document/<int:id>/apercu')
@login_required
def apercu_document(id):
    cid = get_client_id()
    if not cid:
        return 'Unauthorized', 403

    conn = get_db()
    try:
        doc = row_to_dict(conn.execute('SELECT id, appareil_id, client_id, nom, nom_fichier, contenu_blob FROM documents_appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        if not doc:
            return 'Not found', 404

        # Préférer servir depuis BLOB si disponible (synced)
        if doc.get('contenu_blob'):
            return send_file(
                io.BytesIO(doc['contenu_blob']),
                as_attachment=False,
                download_name=doc.get('nom_fichier') or doc.get('nom', 'document')
            )

        # Fallback: servir depuis fichier local
        if doc.get('nom_fichier'):
            fichier_path = os.path.join(UPLOAD_FOLDER, doc['nom_fichier'])
            if os.path.exists(fichier_path):
                return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=False)

        return f"Fichier introuvable : {doc.get('nom_fichier', '?')}", 404
    finally:
        conn.close()


# ─── API APPAREIL : IGNORER ALERTE GARANTIE ──────────────────────────────────

@app.route('/api/appareil/<int:id>/garantie-ignorer', methods=['POST'])
@login_required
def api_garantie_ignorer(id):
    """Active ou désactive le flag 'ignorer alerte garantie' sur un appareil."""
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    ignorer = (request.json or {}).get('ignorer', True)
    now = _utcnow().isoformat()
    conn = get_db()
    conn.execute('UPDATE appareils SET garantie_alerte_ignoree=?, date_maj=? WHERE id=?',
                 (1 if ignorer else 0, now, id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'garantie_alerte_ignoree': bool(ignorer)})


# ─── API DOCUMENTS MODALE ────────────────────────────────────────────────────

@app.route('/api/appareil/<int:id>/documents')
@login_required
def api_docs_appareil(id):
    cid = get_client_id()
    conn = get_db()
    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, appareil_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_appareils WHERE appareil_id=? AND client_id=? ORDER BY date_upload DESC',
        (id, cid)).fetchall()]
    conn.close()
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
        ext = (d.get('nom_fichier','').rsplit('.',1)[-1] or '').lower()
        d['is_img'] = ext in ('png','jpg','jpeg','gif','webp')
        d['is_pdf'] = ext == 'pdf'
    return jsonify(docs)

@app.route('/api/peripherique/<int:id>/documents')
@login_required
def api_docs_peripherique(id):
    cid = get_client_id()
    conn = get_db()
    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, peripherique_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_peripheriques WHERE peripherique_id=? AND client_id=? ORDER BY date_upload DESC',
        (id, cid)).fetchall()]
    conn.close()
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
        ext = (d.get('nom_fichier','').rsplit('.',1)[-1] or '').lower()
        d['is_img'] = ext in ('png','jpg','jpeg','gif','webp')
        d['is_pdf'] = ext == 'pdf'
    return jsonify(docs)

# ─── DOCUMENTS PÉRIPHÉRIQUES ───────────────────────────────────────

@app.route('/peripherique/<int:id>/documents/upload', methods=['POST'])
@login_required
def upload_doc_peripherique(id):
    cid = get_client_id()
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('editer_peripherique', id=id))
    f = request.files['fichier']
    ok, motif = verifier_fichier(f)
    if not ok:
        flash(motif, 'danger')
        return redirect(url_for('editer_peripherique', id=id))
    safe = secure_filename(f.filename)
    if not safe:
        flash('Nom de fichier invalide', 'danger')
        return redirect(url_for('editer_peripherique', id=id))
    unique = f"per{id}_{int(time.time())}_{safe}"
    save_path = os.path.join(UPLOAD_FOLDER, unique)
    logger.info(f"Upload document périphérique #{id}: saving to {save_path}")
    try:
        f.save(save_path)
        taille = os.path.getsize(save_path)
        logger.info(f"Upload document périphérique #{id}: saved {taille} bytes")
    except Exception as e:
        logger.exception(f"Upload document périphérique #{id}: save FAILED → {save_path}")
        flash(f'Erreur lors de la sauvegarde du fichier : {e}', 'danger')
        return redirect(url_for('editer_peripherique', id=id))

    nom = request.form.get('nom', '') or f.filename
    desc = request.form.get('description', '')
    type_doc = request.form.get('type_doc', '')
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute('''INSERT INTO documents_peripheriques
        (peripherique_id,client_id,nom,description,type_doc,nom_fichier,taille,date_upload,contenu_blob,sync_status,date_sync)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (id, cid, nom, desc, type_doc, unique, taille, now, None, 'local', ''))

    # Log document upload
    per_title = conn.execute('SELECT CONCAT(marque, \' \', modele) FROM peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone()
    per_name = per_title[0] if per_title else f'Périphérique #{id}'
    log_history(conn, cid, 'peripherique', id, per_name, 'Ajout de document',
                _diff_json({}, {'nom': nom, 'fichier': unique, 'type_doc': type_doc}))

    conn.commit(); conn.close()
    flash(f'Document « {nom} » uploadé', 'success')
    return redirect(url_for('editer_peripherique', id=id))

@app.route('/doc-peripherique/<int:id>/telecharger')
@login_required
def telecharger_doc_peripherique(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, peripherique_id, client_id, nom, nom_fichier, contenu_blob FROM documents_peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not doc: return 'Not found', 404

    # Préférer servir depuis BLOB si disponible (synced)
    if doc.get('contenu_blob'):
        return send_file(
            io.BytesIO(doc['contenu_blob']),
            as_attachment=True,
            download_name=doc['nom']
        )

    # Fallback: servir depuis fichier local
    return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=True, download_name=doc['nom'])

@app.route('/doc-peripherique/<int:id>/apercu')
@login_required
def apercu_doc_peripherique(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, peripherique_id, client_id, nom, nom_fichier, contenu_blob FROM documents_peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not doc: return 'Not found', 404
    if doc.get('contenu_blob'):
        return send_file(io.BytesIO(doc['contenu_blob']), as_attachment=False,
                         download_name=doc.get('nom_fichier', 'document'))
    try:
        return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=False)
    except Exception:
        return f"Fichier introuvable : {doc.get('nom_fichier', '?')}", 404

@app.route('/doc-peripherique/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_doc_peripherique(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_peripheriques'))
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, peripherique_id, client_id, nom, nom_fichier, contenu_blob FROM documents_peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    periph_id = doc.get('peripherique_id', 0)
    if doc:
        conn.execute('DELETE FROM documents_peripheriques WHERE id=?', (id,))

        # Log document deletion
        per_title = conn.execute('SELECT CONCAT(marque, \' \', modele) FROM peripheriques WHERE id=? AND client_id=?', (periph_id, cid)).fetchone()
        per_name = per_title[0] if per_title else f'Périphérique #{periph_id}'
        log_history(conn, cid, 'peripherique', periph_id, per_name, 'Suppression de document',
                    _diff_json({'nom': doc.get('nom', ''), 'fichier': doc.get('nom_fichier', '')}, {}))

        conn.commit()
        try: os.remove(os.path.join(UPLOAD_FOLDER, doc['nom_fichier']))
        except: pass
    conn.close()
    flash('Document supprimé', 'info')
    return redirect(url_for('editer_peripherique', id=periph_id))

# ─── BAIE DE BRASSAGE ────────────────────────────────────────────────────────

@app.route('/baie')
@login_required
def baie_brassage():
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    parc = row_to_dict(conn.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    nb_u = parc.get('baie_nb_u', 12) or 12
    # Récupérer les slots existants
    slots_db = [row_to_dict(r) for r in conn.execute(
        '''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.marque, a.modele, a.en_ligne,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id
           WHERE s.client_id=? ORDER BY s.position''', (cid,)).fetchall()]
    # Appareils disponibles pour association
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,type_appareil,adresse_ip,marque,modele FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    # Périphériques disponibles pour association (onduleurs, panneaux de
    # brassage… typiquement montés en baie mais jusqu'ici jamais liables)
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id,categorie,marque,modele FROM peripheriques WHERE client_id=? ORDER BY categorie,marque',
        (cid,)).fetchall()]
    photos = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM baie_photos WHERE client_id=? ORDER BY date_upload DESC', (cid,)).fetchall()]
    # Suggestions de pièces pour l'affectation des ports d'un bandeau RJ (prises
    # murales) — mêmes valeurs de localisation déjà saisies sur les appareils/
    # périphériques du client, pas une nouvelle liste séparée à maintenir.
    localisations = sorted({r[0].strip() for r in conn.execute(
        "SELECT localisation FROM appareils WHERE client_id=? AND localisation!='' "
        "UNION SELECT localisation FROM peripheriques WHERE client_id=? AND localisation!=''",
        (cid, cid)).fetchall() if r[0] and r[0].strip()})
    conn.close()
    # Construire la grille : dict (position, col_index) -> slot
    slots_map = {}
    for s in slots_db:
        key = (s['position'], s.get('col_index', 0))
        slots_map[key] = s
    return render_template('baie_brassage.html', parc=parc, client=client, nb_u=nb_u,
                           slots_map=slots_map, slots_db=slots_db, appareils=appareils,
                           peripheriques=peripheriques, localisations=localisations,
                           photos=photos, clients=get_clients(), client_actif_id=cid)

def _couleur_port(appareil_type=None, periph_categorie=None, usage_libre=None, est_lien=False):
    """Couleur d'un port selon ce qu'il vaut : même palette configurable que
    les badges de type d'appareil/catégorie de périphérique ailleurs dans
    l'app (type_color_*/periph_color_*, réglages utilisateur) — un port lié
    à un serveur prend la même couleur que partout ailleurs où ce type est
    représenté, pas une couleur inventée pour l'occasion. Un lien port-à-port
    (câblage vers un autre port, sans appareil/périphérique propre) a sa
    propre couleur distincte de l'usage libre (ambre) et du port non affecté
    (gris)."""
    if appareil_type:
        return cfg_get('type_color_' + type_css_filter(appareil_type), '#94a3b8')
    if periph_categorie:
        return cfg_get('periph_color_' + periph_color_key_filter(periph_categorie), '#94a3b8')
    if est_lien:
        return '#818cf8'
    if usage_libre:
        return '#f59e0b'
    return '#334155'

def _prises_murales_avec_details(conn, slot_id):
    """Prises murales d'un bandeau RJ, enrichies du nom et de la couleur de
    leur cible — même logique de résolution qu'un port directement associé
    (voir _ports_avec_details), mais côté prise murale : c'est ELLE qui
    porte désormais l'appareil/périphérique/usage branché dans la pièce,
    le port RJ correspondant ne servant plus qu'à interconnecter avec un
    autre élément de la baie. Retourne un dict {numero: détails}, vide pour
    un slot qui n'est pas (ou plus) un bandeau RJ."""
    rows = conn.execute(
        '''SELECT pm.slot_id, pm.numero, pm.piece, pm.identification, pm.appareil_id, pm.peripherique_id, pm.usage_libre,
                  pm.cable_couleur, pm.cable_longueur,
                  a.nom_machine, a.type_appareil, a.en_ligne, a.dernier_ping,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_prises_murales pm
           LEFT JOIN appareils a ON pm.appareil_id=a.id
           LEFT JOIN peripheriques p ON pm.peripherique_id=p.id
           WHERE pm.slot_id=? ORDER BY pm.numero''', (slot_id,)).fetchall()
    par_numero = {}
    for r in rows:
        d = row_to_dict(r)
        if d['appareil_id']:
            nom = d['nom_machine'] or ('Appareil #%d' % d['appareil_id'])
        elif d['peripherique_id']:
            nom = ' '.join(filter(None, [d['p_categorie'], d['p_marque'], d['p_modele']])) or ('Périphérique #%d' % d['peripherique_id'])
        else:
            nom = d['usage_libre'] or ''
        d['nom_cible'] = nom
        d['couleur'] = _couleur_port(d.get('type_appareil'), d.get('p_categorie'), d.get('usage_libre'), False)
        par_numero[d['numero']] = d
    return par_numero

def _ports_avec_details(conn, slot_id):
    """Ports d'un slot, enrichis du nom et de la couleur de leur cible —
    y compris, pour un port relié à un autre port (câblage physique switch
    <-> bandeau RJ, routeur, etc.), le nom de l'élément en face ET, si un
    appareil/périphérique est joignable au bout de ce câble, son nom et son
    statut (cible_finale/cible_hors_ligne/lie_appareil_id — sert aussi à
    colorer le port en direct côté client, et à le pinger). Trois façons
    DISTINCTES pour l'élément en face de porter cette association, toutes
    trois résolues ici :
    1. Sur le PORT EN FACE lui-même (bp2.appareil_id/peripherique_id) —
       un élément d'interconnexion (Patch Panel générique, PDU...) dont le
       port sert directement un appareil/périphérique, sans passer par le
       système de prise murale (réservé au bandeau RJ, voir ci-dessous).
    2. Sur le SLOT en face tout entier (baie_slots.appareil_id) — cas d'un
       appareil RACK-MONTÉ (ex. routeur), "Associer à un appareil" dans le
       panneau "+ Placer" du slot lui-même, un de ses ports relié au switch
       sans que CE port porte l'association (elle vit au niveau du slot).
       C'est le mécanisme d'origine, conservé tel quel.
    3. Sur la PRISE MURALE en face, quand l'élément en face est un bandeau
       RJ (voir _prises_murales_avec_details) — cas standard du brassage
       structuré : prise murale d'un bureau câblée en fixe vers ce port du
       bandeau, lui-même relié par cordon à un port de switch/routeur. La
       prise murale porte l'appareil, le port RJ ne sert qu'à interconnecter.
    Le port en face ne peut par construction cumuler les cas 1/2 avec être
    lui-même un lien (voir api_baie_lien_port : un port n'a qu'une seule
    cible à la fois) — priorité au port, repli sur le slot, repli sur sa
    prise murale. `piece`/appareil/périphérique/usage d'un port de bandeau
    RJ lui-même sont, depuis la migration prises murales (voir init_db()),
    normalement toujours vides — la prise murale de même numéro
    (d['prise_murale'], via _prises_murales_avec_details) les porte à sa
    place."""
    type_slot = conn.execute('SELECT type_equipement FROM baie_slots WHERE id=?', (slot_id,)).fetchone()
    est_bandeau = bool(type_slot and type_slot[0] == 'Bandeau RJ')
    prises_par_numero = _prises_murales_avec_details(conn, slot_id) if est_bandeau else {}
    rows = conn.execute(
        '''SELECT bp.slot_id, bp.numero, bp.appareil_id, bp.peripherique_id, bp.usage_libre,
                  bp.lie_slot_id, bp.lie_port_numero, bp.piece, bp.cable_couleur, bp.cable_longueur,
                  a.nom_machine, a.type_appareil, a.en_ligne, a.dernier_ping,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slot_ports bp
           LEFT JOIN appareils a ON bp.appareil_id=a.id
           LEFT JOIN peripheriques p ON bp.peripherique_id=p.id
           WHERE bp.slot_id=? ORDER BY bp.numero''', (slot_id,)).fetchall()
    ports = []
    for r in rows:
        d = row_to_dict(r)
        d['cible_finale'] = ''
        d['cible_hors_ligne'] = None
        d['lie_appareil_id'] = None
        lie_type_appareil = None
        lie_p_categorie = None
        if d['appareil_id']:
            nom = d['nom_machine'] or ('Appareil #%d' % d['appareil_id'])
        elif d['peripherique_id']:
            nom = ' '.join(filter(None, [d['p_categorie'], d['p_marque'], d['p_modele']])) or ('Périphérique #%d' % d['peripherique_id'])
        elif d['lie_slot_id']:
            cible = conn.execute(
                'SELECT nom_custom, type_equipement, appareil_id FROM baie_slots WHERE id=?',
                (d['lie_slot_id'],)).fetchone()
            nom_elem = (cible[0] or cible[1] if cible else None) or ('Élément #%d' % d['lie_slot_id'])
            # Libellé du port en face — voir _libelle_port_pour_type
            # (miroir de libellePortAffiche côté client) : numéro AFFICHÉ,
            # sans l'offset de sa plage (1001+/2001+/9001), et distingue
            # SFP/Fibre/prise ondulée-protégée/entrée selon le TYPE de
            # l'élément en face — pas juste "WAN %d"/"SFP %d" génériques
            # (bug trouvé en auditant : un lien vers l'entrée d'un PDU,
            # numéro brut 9001, s'affichait à tort comme "WAN 7001").
            far_num = d['lie_port_numero']
            far_label = _libelle_port_pour_type(cible[1] if cible else None, far_num)
            nom = '%s — %s' % (far_label, nom_elem)
            far_port = conn.execute(
                '''SELECT bp2.appareil_id, bp2.peripherique_id, a2.nom_machine, a2.en_ligne, a2.type_appareil,
                          p2.categorie, p2.marque, p2.modele, a2.dernier_ping
                   FROM baie_slot_ports bp2
                   LEFT JOIN appareils a2 ON bp2.appareil_id=a2.id
                   LEFT JOIN peripheriques p2 ON bp2.peripherique_id=p2.id
                   WHERE bp2.slot_id=? AND bp2.numero=?''',
                (d['lie_slot_id'], d['lie_port_numero'])).fetchone()
            # Cas STANDARD du brassage structuré (voir cas 3 de la docstring) :
            # le port en face est lui-même un port RJ de bandeau, dont
            # l'appareil/périphérique éventuel vit désormais sur sa PRISE
            # MURALE de même numéro, jamais sur le port lui-même (far_port
            # ci-dessus reste vide pour un bandeau) — requête seulement si
            # far_port n'a lui-même rien porté, pour ne pas interroger cette
            # table pour rien sur un lien vers un élément non-bandeau.
            far_pm = None
            if cible and cible[1] == 'Bandeau RJ' and not (far_port and (far_port[0] or far_port[1])):
                far_pm = conn.execute(
                    '''SELECT pm.appareil_id, pm.peripherique_id, a3.nom_machine, a3.en_ligne, a3.type_appareil,
                              p3.categorie, p3.marque, p3.modele, a3.dernier_ping
                       FROM baie_prises_murales pm
                       LEFT JOIN appareils a3 ON pm.appareil_id=a3.id
                       LEFT JOIN peripheriques p3 ON pm.peripherique_id=p3.id
                       WHERE pm.slot_id=? AND pm.numero=?''',
                    (d['lie_slot_id'], d['lie_port_numero'])).fetchone()
            if far_port and far_port[0]:
                d['cible_finale'] = far_port[2] or ('Appareil #%d' % far_port[0])
                # en_ligne vaut 0 par défaut en base, indiscernable d'un vrai
                # échec sans dernier_ping (même garde que côté client pour
                # le voyant d'un port, voir portHTML()) — sans lui, un
                # appareil simplement JAMAIS pingé s'affichait à tort comme
                # "⚠️ hors ligne" dans la bulle du port relié.
                d['cible_hors_ligne'] = bool(far_port[8]) and (far_port[3] == 0)
                d['lie_appareil_id'] = far_port[0]
                lie_type_appareil = far_port[4]
            elif far_port and far_port[1]:
                d['cible_finale'] = ' '.join(filter(None, far_port[5:8])) or ('Périphérique #%d' % far_port[1])
                lie_p_categorie = far_port[5]
            elif far_pm and far_pm[0]:
                d['cible_finale'] = far_pm[2] or ('Appareil #%d' % far_pm[0])
                d['cible_hors_ligne'] = bool(far_pm[8]) and (far_pm[3] == 0)
                d['lie_appareil_id'] = far_pm[0]
                lie_type_appareil = far_pm[4]
            elif far_pm and far_pm[1]:
                d['cible_finale'] = ' '.join(filter(None, far_pm[5:8])) or ('Périphérique #%d' % far_pm[1])
                lie_p_categorie = far_pm[5]
            elif cible and cible[2]:
                # Repli ultime : le SLOT en face porte l'association
                # (appareil rack-monté, voir docstring) — jamais atteint si
                # far_pm a déjà résolu quelque chose ci-dessus.
                far_app = conn.execute(
                    'SELECT nom_machine, en_ligne, type_appareil, dernier_ping FROM appareils WHERE id=?', (cible[2],)).fetchone()
                if far_app:
                    d['cible_finale'] = far_app[0] or ('Appareil #%d' % cible[2])
                    d['cible_hors_ligne'] = bool(far_app[3]) and (far_app[1] == 0)
                    d['lie_appareil_id'] = cible[2]
                    lie_type_appareil = far_app[2]
        else:
            nom = d['usage_libre'] or ''
        d['nom_cible'] = nom
        if d['lie_slot_id'] and (lie_type_appareil or lie_p_categorie):
            # Cible résolue au bout du câble : sa propre couleur de type,
            # pas l'indigo générique "lien" — c'est elle qui doit se voir.
            d['couleur'] = _couleur_port(lie_type_appareil, lie_p_categorie, None, False)
        else:
            d['couleur'] = _couleur_port(d.get('type_appareil'), d.get('p_categorie'), d.get('usage_libre'), bool(d['lie_slot_id']))
        d['prise_murale'] = prises_par_numero.get(d['numero'])
        ports.append(d)
    return ports

def _detacher_liens_vers(conn, slot_id, numeros=None):
    """Efface le lien port-à-port de tout port (n'importe où, y compris dans
    une autre baie) qui pointait vers slot_id — éventuellement restreint à
    certains numéros. À appeler AVANT de supprimer/réduire des ports ou un
    slot entier, pour ne jamais laisser un lien à sens unique pointant vers
    un port qui n'existe plus (même défaut que peripherique_id/appareil_id
    ailleurs dans ce fichier — colonnes ajoutées après coup, sans FK
    ON DELETE, nettoyage manuel obligatoire)."""
    if numeros:
        placeholders = ','.join('?' * len(numeros))
        conn.execute(
            f"UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
            f"cable_couleur='', cable_longueur='' "
            f'WHERE lie_slot_id=? AND lie_port_numero IN ({placeholders})',
            (slot_id, *numeros))
    else:
        conn.execute("UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
                     "cable_couleur='', cable_longueur='' WHERE lie_slot_id=?", (slot_id,))

def _plafond_nb_ports(type_equipement):
    """Nombre de ports maximum d'un élément de baie. Un bandeau RJ est
    plafonné à 24 (et non 48 comme les autres éléments) : demandé pour
    garantir la place d'empiler la rangée des prises murales AU-DESSUS de
    la rangée des ports RJ (voir _prises_murales_avec_details / rendu côté
    client) sans avoir à compacter les cellules ni agrandir la hauteur du
    bandeau — un bandeau RJ réel fait d'ailleurs déjà typiquement 24 ports
    (voir le raccourci "+ Bandeau RJ" côté client)."""
    return 24 if type_equipement == 'Bandeau RJ' else 48

# Ports SFP (fibre) d'un switch — espace de numérotation séparé des ports
# RJ (voir la migration nb_ports_sfp dans init_db()) : 1001, 1002... plutôt
# que 1, 2... pour ne jamais entrer en collision avec un numéro de port RJ
# (plafonné à 48) sur la même ligne baie_slot_ports (UNIQUE(slot_id,
# numero)). Un switch réel a rarement plus de 8 ports SFP/SFP+ uplink. Le
# MÊME espace de numérotation (mêmes colonne et offset) sert de "Fibre"
# pour un routeur/pare-feu (voir appliquerReglesType côté client) — un
# routeur n'a jamais besoin des deux groupes RJ ET SFP d'un switch à la
# fois, la plage peut donc être réutilisée telle quelle sous un autre nom.
SFP_NUMERO_OFFSET = 1000
PLAFOND_SFP = 8

# Ports WAN d'un routeur/pare-feu — troisième et dernier espace de
# numérotation générique (après la plage "de base" nb_ports/1-48 et la
# plage "haute" nb_ports_sfp/1001+ ci-dessus), pour un équipement qui a
# BESOIN des trois groupes en même temps (WAN + LAN + Fibre) — un switch
# ou un onduleur/PDU n'en ont jamais besoin, seul un routeur/pare-feu
# active ce troisième groupe (voir la migration nb_ports_wan dans
# init_db()). Un routeur réel a rarement plus de 4 liens WAN.
WAN_NUMERO_OFFSET = 2000
PLAFOND_WAN = 4

# Port "Entrée" d'un PDU — son propre cordon d'alimentation (typiquement
# branché sur une prise d'onduleur, éventuellement sur un AUTRE PDU en
# cascade), distinct des prises de SORTIE (nb_ports) qu'il redistribue —
# demandé, pour propager le statut "alimenté par un onduleur" (voir
# liste_appareils()) aux appareils branchés sur ce PDU quand son entrée
# remonte, directement ou en cascade, jusqu'à un vrai onduleur. Plage la
# plus haute de toutes (après WAN_NUMERO_OFFSET) : TOUJOURS EXACTEMENT 1
# port, jamais configurable (pas de "nombre d'entrées" — un PDU n'a
# physiquement qu'un seul cordon), créé/retiré automatiquement selon
# type_equipement (voir _reconcilier_port_entree), jamais depuis un champ
# de formulaire. Se lie avec le MÊME mécanisme générique que tout autre
# port (api_baie_lien_port) — un onduleur/switch/routeur n'a jamais ce
# port.
ENTREE_NUMERO_OFFSET = 9000
NUMERO_ENTREE_PDU = ENTREE_NUMERO_OFFSET + 1

def _numero_logique_port(numero):
    """Numéro AFFICHÉ d'un port — retire l'offset de sa plage (voir
    SFP_NUMERO_OFFSET/WAN_NUMERO_OFFSET/ENTREE_NUMERO_OFFSET) pour ne
    jamais montrer le numéro BRUT stocké en base (ex. 1001 pour un port
    SFP) là où l'utilisateur attend le numéro logique (1) qu'il voit
    partout ailleurs dans l'app. Équivalent serveur de numeroAffiche()
    côté client (baie_brassage.html) — DOIT rester en miroir exact."""
    if numero > ENTREE_NUMERO_OFFSET:
        return numero
    if numero > WAN_NUMERO_OFFSET:
        return numero - WAN_NUMERO_OFFSET
    if numero > SFP_NUMERO_OFFSET:
        return numero - SFP_NUMERO_OFFSET
    return numero

def _libelle_port_pour_type(type_equipement, numero):
    """Équivalent serveur de libellePortAffiche() côté client (voir
    baie_brassage.html) — DOIT rester en miroir exact : même logique de
    plage/type des deux côtés. Utilisé partout où un numéro de port doit
    être présenté à l'utilisateur côté serveur (nom_cible d'un lien dans
    _ports_avec_details, fiche de câblage imprimable/CSV) — sans cette
    fonction centralisée, ces deux endroits affichaient le numéro BRUT
    (avec offset, ex. 1001/2001/9001) au lieu du libellé logique (bugs
    trouvés en auditant tout le mécanisme de la baie de brassage)."""
    if numero > ENTREE_NUMERO_OFFSET:
        return 'Entrée'
    if type_equipement == 'UPS':
        return ('Prise protégée ' if SFP_NUMERO_OFFSET < numero <= WAN_NUMERO_OFFSET else 'Prise ondulée ') + str(_numero_logique_port(numero))
    if type_equipement == 'PDU':
        return 'Prise %d' % numero
    if numero > WAN_NUMERO_OFFSET:
        return 'Port WAN %d' % (numero - WAN_NUMERO_OFFSET)
    if SFP_NUMERO_OFFSET < numero <= WAN_NUMERO_OFFSET:
        nom = 'Fibre' if type_equipement == 'Routeur/Pare-feu' else 'SFP'
        return 'Port %s %d' % (nom, numero - SFP_NUMERO_OFFSET)
    if type_equipement == 'Routeur/Pare-feu':
        return 'Port LAN %d' % numero
    return 'Port %d' % numero

def _reconcilier_ports(conn, slot_id, nb_ports, ports_existants=None, type_equipement=None):
    """Ajuste les lignes baie_slot_ports d'un slot à nb_ports (0-48, 0-24
    pour un bandeau RJ — voir _plafond_nb_ports) : crée les numéros
    manquants, retire ceux au-delà (en détachant d'abord tout lien
    port-à-port pointant vers un numéro sur le point de disparaître).
    ports_existants (dict numero -> (appareil_id, peripherique_id,
    usage_libre, lie_slot_id, lie_port_numero, piece, cable_couleur,
    cable_longueur)), s'il est fourni, permet de reporter les liaisons d'un
    ancien slot remplacé au même emplacement
    (voir api_baie_ajouter_slot) au lieu de les perdre — l'appelant doit
    ensuite repointer vers le NOUVEAU slot_id les ports ailleurs qui
    visaient l'ancien (UPDATE ... SET lie_slot_id=nouveau WHERE
    lie_slot_id=ancien), ce que cette fonction ne peut pas faire elle-même
    puisqu'elle ne connaît pas l'ancien id."""
    nb_ports = min(_plafond_nb_ports(type_equipement), max(0, int(nb_ports or 0)))
    now = _utcnow().isoformat()
    if ports_existants is None:
        supprimes = [r[0] for r in conn.execute(
            'SELECT numero FROM baie_slot_ports WHERE slot_id=? AND numero>?', (slot_id, nb_ports)).fetchall()]
        if supprimes:
            _detacher_liens_vers(conn, slot_id, supprimes)
        conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=? AND numero>?', (slot_id, nb_ports))
        deja = {r[0] for r in conn.execute('SELECT numero FROM baie_slot_ports WHERE slot_id=?', (slot_id,)).fetchall()}
        for numero in range(1, nb_ports + 1):
            if numero not in deja:
                conn.execute('INSERT INTO baie_slot_ports (slot_id,numero,date_maj) VALUES (?,?,?)', (slot_id, numero, now))
    else:
        for numero in range(1, nb_ports + 1):
            ap, pe, us, lsid, lnum, piece, cc, cl = ports_existants.get(
                numero, (None, None, '', None, None, '', '', ''))
            conn.execute('''INSERT INTO baie_slot_ports
                (slot_id,numero,appareil_id,peripherique_id,usage_libre,lie_slot_id,lie_port_numero,
                 piece,cable_couleur,cable_longueur,date_maj)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''', (slot_id, numero, ap, pe, us, lsid, lnum, piece, cc, cl, now))
    return nb_ports

def _reconcilier_ports_sfp(conn, slot_id, nb_ports_sfp, ports_existants=None):
    """Miroir de _reconcilier_ports pour les ports SFP/Fibre (0-8, voir
    PLAFOND_SFP) — même mécanique, sur la plage de numéros dédiée
    SFP_NUMERO_OFFSET+1..+nb_ports_sfp, jamais partagée avec les numéros
    de port RJ/LAN (plafonnés à 48) : évite de toucher à la contrainte
    UNIQUE(slot_id, numero) de baie_slot_ports pour distinguer les deux
    types. ports_existants, s'il est fourni, est le MÊME dict complet
    (numero -> tuple, sans filtre de plage) que celui passé à
    _reconcilier_ports — la clé (numero) suffit à isoler les entrées SFP
    sans requête séparée.
    Bornée en haut par WAN_NUMERO_OFFSET : un routeur/pare-feu peut avoir
    À LA FOIS des ports Fibre (cette plage) ET des ports WAN (plage encore
    plus haute, voir _reconcilier_ports_wan) sur le MÊME slot — sans cette
    borne, réduire nb_ports_sfp aurait aussi supprimé/détaché à tort tous
    les ports WAN déjà configurés (numero > plafond_numero de cette
    fonction, quel que soit ce plafond)."""
    nb_ports_sfp = min(PLAFOND_SFP, max(0, int(nb_ports_sfp or 0)))
    plafond_numero = SFP_NUMERO_OFFSET + nb_ports_sfp
    now = _utcnow().isoformat()
    if ports_existants is None:
        supprimes = [r[0] for r in conn.execute(
            'SELECT numero FROM baie_slot_ports WHERE slot_id=? AND numero>? AND numero<?',
            (slot_id, plafond_numero, WAN_NUMERO_OFFSET)).fetchall()]
        if supprimes:
            _detacher_liens_vers(conn, slot_id, supprimes)
        conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=? AND numero>? AND numero<?',
                     (slot_id, plafond_numero, WAN_NUMERO_OFFSET))
        deja = {r[0] for r in conn.execute('SELECT numero FROM baie_slot_ports WHERE slot_id=?', (slot_id,)).fetchall()}
        for numero in range(SFP_NUMERO_OFFSET + 1, plafond_numero + 1):
            if numero not in deja:
                conn.execute('INSERT INTO baie_slot_ports (slot_id,numero,date_maj) VALUES (?,?,?)', (slot_id, numero, now))
    else:
        for numero in range(SFP_NUMERO_OFFSET + 1, plafond_numero + 1):
            ap, pe, us, lsid, lnum, piece, cc, cl = ports_existants.get(
                numero, (None, None, '', None, None, '', '', ''))
            conn.execute('''INSERT INTO baie_slot_ports
                (slot_id,numero,appareil_id,peripherique_id,usage_libre,lie_slot_id,lie_port_numero,
                 piece,cable_couleur,cable_longueur,date_maj)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''', (slot_id, numero, ap, pe, us, lsid, lnum, piece, cc, cl, now))
    return nb_ports_sfp

def _reconcilier_ports_wan(conn, slot_id, nb_ports_wan, ports_existants=None):
    """Miroir de _reconcilier_ports_sfp pour les ports WAN d'un routeur/
    pare-feu (0-4, voir PLAFOND_WAN), sur la plage WAN_NUMERO_OFFSET+1..+n.
    Bornée en haut par ENTREE_NUMERO_OFFSET (même raison que SFP bornée par
    WAN_NUMERO_OFFSET, voir _reconcilier_ports_sfp) : un slot peut avoir à
    la fois des ports WAN ET, s'il s'agit d'un PDU — cas normalement
    exclusif en pratique, mais la borne coûte rien et évite toute
    ambiguïté future — un port d'entrée (voir _reconcilier_port_entree)."""
    nb_ports_wan = min(PLAFOND_WAN, max(0, int(nb_ports_wan or 0)))
    plafond_numero = WAN_NUMERO_OFFSET + nb_ports_wan
    now = _utcnow().isoformat()
    if ports_existants is None:
        supprimes = [r[0] for r in conn.execute(
            'SELECT numero FROM baie_slot_ports WHERE slot_id=? AND numero>? AND numero<?',
            (slot_id, plafond_numero, ENTREE_NUMERO_OFFSET)).fetchall()]
        if supprimes:
            _detacher_liens_vers(conn, slot_id, supprimes)
        conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=? AND numero>? AND numero<?',
                     (slot_id, plafond_numero, ENTREE_NUMERO_OFFSET))
        deja = {r[0] for r in conn.execute('SELECT numero FROM baie_slot_ports WHERE slot_id=?', (slot_id,)).fetchall()}
        for numero in range(WAN_NUMERO_OFFSET + 1, plafond_numero + 1):
            if numero not in deja:
                conn.execute('INSERT INTO baie_slot_ports (slot_id,numero,date_maj) VALUES (?,?,?)', (slot_id, numero, now))
    else:
        for numero in range(WAN_NUMERO_OFFSET + 1, plafond_numero + 1):
            ap, pe, us, lsid, lnum, piece, cc, cl = ports_existants.get(
                numero, (None, None, '', None, None, '', '', ''))
            conn.execute('''INSERT INTO baie_slot_ports
                (slot_id,numero,appareil_id,peripherique_id,usage_libre,lie_slot_id,lie_port_numero,
                 piece,cable_couleur,cable_longueur,date_maj)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''', (slot_id, numero, ap, pe, us, lsid, lnum, piece, cc, cl, now))
    return nb_ports_wan

def _reconcilier_port_entree(conn, slot_id, type_equipement, ports_existants=None):
    """Crée ou retire le port d'entrée unique d'un PDU (numero=
    NUMERO_ENTREE_PDU, voir plus haut) selon que type_equipement vaut
    'PDU' ou non — même mécanique que les autres _reconcilier_ports*, mais
    TOUJOURS 0 ou 1 (jamais un compte fourni par l'appelant, contrairement
    à nb_ports/nb_ports_sfp/nb_ports_wan)."""
    doit_exister = (type_equipement == 'PDU')
    now = _utcnow().isoformat()
    if ports_existants is None:
        existe = conn.execute('SELECT 1 FROM baie_slot_ports WHERE slot_id=? AND numero=?',
                               (slot_id, NUMERO_ENTREE_PDU)).fetchone()
        if not doit_exister:
            if existe:
                _detacher_liens_vers(conn, slot_id, [NUMERO_ENTREE_PDU])
                conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=? AND numero=?', (slot_id, NUMERO_ENTREE_PDU))
        elif not existe:
            conn.execute('INSERT INTO baie_slot_ports (slot_id,numero,date_maj) VALUES (?,?,?)',
                         (slot_id, NUMERO_ENTREE_PDU, now))
    elif doit_exister:
        ap, pe, us, lsid, lnum, piece, cc, cl = ports_existants.get(
            NUMERO_ENTREE_PDU, (None, None, '', None, None, '', '', ''))
        conn.execute('''INSERT INTO baie_slot_ports
            (slot_id,numero,appareil_id,peripherique_id,usage_libre,lie_slot_id,lie_port_numero,
             piece,cable_couleur,cable_longueur,date_maj)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (slot_id, NUMERO_ENTREE_PDU, ap, pe, us, lsid, lnum, piece, cc, cl, now))
    # else : ports_existants fourni mais doit_exister faux — rien à
    # reporter, l'ancien port d'entrée (s'il existait) ne survit
    # simplement pas au remplacement, cohérent avec le reste du report.

def _appareils_secourus_par_onduleur(conn, cid):
    """appareil_id -> nom de l'élément qui l'alimente directement, pour
    tout appareil branché — même transitivement à travers une cascade de
    PDU — sur un vrai onduleur (demandé, voir badge_ups dans
    liste_appareils.html). Un PDU n'est secouru QUE si son port d'entrée
    (NUMERO_ENTREE_PDU, voir _reconcilier_port_entree) est lié, directement
    ou via un autre PDU déjà secouru, à un port d'un onduleur — jamais
    déduit automatiquement, toujours un lien explicite posé via
    "🔗 Lier des ports" comme n'importe quel autre câblage. Une prise
    ondulée ou protégée d'un onduleur compte pareil (toutes deux une
    connexion ÉLECTRIQUE via sa batterie/son parafoudre), le port d'entrée
    lui-même jamais (il n'alimente rien directement, appareil_id y est
    toujours NULL par construction).
    Propagation par point fixe (boucle jusqu'à stabilité) plutôt qu'une
    requête SQL récursive : le nombre de PDU en cascade dans une baie
    réelle est trivialement petit, une implémentation simple en Python
    reste largement assez rapide et bien plus lisible qu'un WITH RECURSIVE."""
    slots_rows = conn.execute(
        "SELECT id, type_equipement, nom_custom FROM baie_slots WHERE client_id=? AND type_equipement IN ('UPS','PDU')",
        (cid,)).fetchall()
    noms = {r[0]: (r[2] or '') for r in slots_rows}
    ups_ids = {r[0] for r in slots_rows if r[1] == 'UPS'}
    pdu_ids = {r[0] for r in slots_rows if r[1] == 'PDU'}
    if not ups_ids and not pdu_ids:
        return {}
    # Où pointe (slot_id en face) le port d'entrée de chaque PDU, s'il est
    # lié à quelque chose.
    entree_vers = {}
    if pdu_ids:
        placeholders = ','.join('?' * len(pdu_ids))
        rows = conn.execute(
            f'''SELECT slot_id, lie_slot_id FROM baie_slot_ports
                WHERE numero=? AND slot_id IN ({placeholders}) AND lie_slot_id IS NOT NULL''',
            (NUMERO_ENTREE_PDU, *pdu_ids)).fetchall()
        entree_vers = {r[0]: r[1] for r in rows}
    secourus = set(ups_ids)
    changement = True
    while changement:
        changement = False
        for pdu_id, cible_slot_id in entree_vers.items():
            if pdu_id not in secourus and cible_slot_id in secourus:
                secourus.add(pdu_id)
                changement = True
    if not secourus:
        return {}
    placeholders = ','.join('?' * len(secourus))
    rows = conn.execute(
        f'''SELECT bsp.appareil_id, bsp.slot_id FROM baie_slot_ports bsp
            WHERE bsp.slot_id IN ({placeholders}) AND bsp.appareil_id IS NOT NULL AND bsp.numero != ?''',
        (*secourus, NUMERO_ENTREE_PDU)).fetchall()
    # Texte complet du tooltip (voir badge_ups dans liste_appareils.html) —
    # distingue le cas direct (branché sur l'onduleur lui-même) du cas en
    # cascade (branché sur un PDU dont l'entrée remonte à un onduleur),
    # pour ne jamais dire à tort "alimenté par l'onduleur « PDU X »".
    resultat = {}
    for appareil_id, slot_id in rows:
        nom = noms[slot_id] or ('Onduleur' if slot_id in ups_ids else 'PDU')
        resultat[appareil_id] = (f"l'onduleur « {nom} »" if slot_id in ups_ids
                                  else f"le PDU « {nom} », lui-même alimenté par un onduleur")
    return resultat

def _reconcilier_prises_murales(conn, slot_id, nb_ports, type_equipement, prises_existantes=None):
    """Ajuste les lignes baie_prises_murales d'un bandeau RJ à nb_ports —
    même mécanique que _reconcilier_ports, en miroir côté prise murale.
    Un élément qui n'est PAS (ou plus) un bandeau RJ n'a par construction
    aucune prise murale : celles qui existeraient (changement de
    type_equipement après coup) sont supprimées.
    prises_existantes (dict numero -> (piece, identification, appareil_id,
    peripherique_id, usage_libre, cable_couleur, cable_longueur)), s'il est
    fourni, reporte les prises murales d'un ancien slot remplacé au même
    emplacement — même rôle que ports_existants pour _reconcilier_ports."""
    if type_equipement != 'Bandeau RJ':
        conn.execute('DELETE FROM baie_prises_murales WHERE slot_id=?', (slot_id,))
        return
    nb_ports = min(24, max(0, int(nb_ports or 0)))
    now = _utcnow().isoformat()
    if prises_existantes is None:
        conn.execute('DELETE FROM baie_prises_murales WHERE slot_id=? AND numero>?', (slot_id, nb_ports))
        deja = {r[0] for r in conn.execute('SELECT numero FROM baie_prises_murales WHERE slot_id=?', (slot_id,)).fetchall()}
        for numero in range(1, nb_ports + 1):
            if numero not in deja:
                conn.execute('INSERT INTO baie_prises_murales (slot_id,numero,date_maj) VALUES (?,?,?)', (slot_id, numero, now))
    else:
        for numero in range(1, nb_ports + 1):
            piece, ident, ap, pe, us, cc, cl = prises_existantes.get(numero, ('', '', None, None, '', '', ''))
            conn.execute('''INSERT INTO baie_prises_murales
                (slot_id,numero,piece,identification,appareil_id,peripherique_id,usage_libre,cable_couleur,cable_longueur,date_maj)
                VALUES (?,?,?,?,?,?,?,?,?,?)''', (slot_id, numero, piece, ident, ap, pe, us, cc, cl, now))

def _clamp_largeur_u(v, col=0):
    """1-10 (dixièmes de la largeur du rack, position dans la grille à 10
    colonnes du client — voir renderRack() dans baie_brassage.html). Valeur
    manquante/invalide -> 10 (pleine largeur) : chaque emplacement a
    désormais TOUJOURS une largeur explicite, plus de notion de "partage
    égal automatique" entre éléments d'une même rangée (l'ancien design,
    où largeur_u valait None tant que l'utilisateur n'avait jamais
    redimensionné, cassait dès qu'un élément à la fois partiel en largeur
    ET en hauteur (hauteur_u > 1) partageait sa rangée avec un autre —
    voir le commentaire de renderRack() côté client pour le détail).
    `col` (0-9, colonne de départ déjà bornée par l'appelant) plafonne en
    plus la largeur à 10-col : sans ça, un slot à col=9 pourrait se voir
    attribuer largeur_u=5, débordant de la grille à 10 colonnes (grid-column
    créerait alors des pistes implicites au-delà de la 10e, décalant tout
    ce qui suit dans la même rangée)."""
    if v in (None, '', 0, '0'):
        n = 10
    else:
        try:
            n = int(v)
        except (TypeError, ValueError):
            n = 10
    n = min(10, max(1, n))
    return min(n, max(1, 10 - col))

def _clamp_col_index(v):
    """0-9 : position de départ dans la grille à 10 colonnes du rack."""
    try:
        n = int(v or 0)
    except (TypeError, ValueError):
        return 0
    return min(9, max(0, n))

def _slots_en_collision(conn, cid, baie_nom, position, col_index, hauteur_u, largeur_u, exclude_id=None):
    """Retourne les slots de la baie dont le rectangle (rangées U × colonnes
    0-9) chevauche celui donné.

    Signalé en usage réel : placer, redimensionner ou déplacer un élément
    à la fois moins large que la baie ET haut de plusieurs U restait
    "toujours très compliqué" — la cause réelle n'était pas la précision du
    positionnement (déjà corrigée en 2.18.69) mais l'absence TOTALE de
    détection de chevauchement : aucune route ne vérifiait quoi que ce soit
    au-delà de la case d'origine exacte (position, col_index). Un élément
    haut de 3U et large de 4/10 pouvait donc silencieusement se superposer
    à un autre élément occupant une partie de ces mêmes rangées, sans le
    moindre avertissement — visuellement, les deux se chevauchaient/se
    battaient en z-index, strictement impossible à comprendre ou corriger
    depuis l'interface.
    """
    fin_u = position + hauteur_u - 1
    fin_col = col_index + largeur_u - 1
    # baie_nom IS NULL traité comme 'Baie principale' — mêmes emplacements
    # "historiques" que partout ailleurs dans ce fichier (voir
    # api_baie_supprimer/api_baie_slots) : sans cette équivalence, un
    # ancien slot jamais migré échapperait silencieusement au contrôle.
    sql = ('SELECT id, position, col_index, hauteur_u, largeur_u, nom_custom, type_equipement '
           'FROM baie_slots WHERE client_id=? AND (baie_nom=? OR (baie_nom IS NULL AND ?=\'Baie principale\'))')
    params = [cid, baie_nom, baie_nom]
    if exclude_id:
        sql += ' AND id!=?'
        params.append(exclude_id)
    collisions = []
    for r in conn.execute(sql, params).fetchall():
        s_id, s_pos, s_col, s_hu, s_lu, s_nom, s_type = r
        s_hu = s_hu or 1
        s_lu = s_lu or 10
        s_col = s_col or 0
        s_fin_u = s_pos + s_hu - 1
        s_fin_col = s_col + s_lu - 1
        if position <= s_fin_u and s_pos <= fin_u and col_index <= s_fin_col and s_col <= fin_col:
            collisions.append({
                'id': s_id, 'nom': s_nom or s_type or ('Emplacement #%d' % s_id),
                'position': s_pos, 'col_index': s_col, 'hauteur_u': s_hu, 'largeur_u': s_lu,
            })
    return collisions

def _msg_collision(collisions):
    """Message d'erreur listant nommément ce qui bloque le placement — pour
    que l'utilisateur comprenne IMMÉDIATEMENT pourquoi, plutôt qu'un
    "conflit" générique qui ne dit pas quoi déplacer ou redimensionner."""
    noms = ', '.join(
        '%s (U%d%s, col %d-%d)' % (
            c['nom'], c['position'],
            '-%d' % (c['position'] + c['hauteur_u'] - 1) if c['hauteur_u'] > 1 else '',
            c['col_index'], c['col_index'] + c['largeur_u'] - 1)
        for c in collisions)
    return 'Chevauche %s.' % noms

@app.route('/api/baie/slot', methods=['POST'])
@login_required
def api_baie_ajouter_slot():
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    f = request.json or {}
    conn = get_db()
    pos = f.get('position', 1)
    col = _clamp_col_index(f.get('col_index', 0))
    baie_nom = f.get('baie_nom', 'Baie principale')
    try:
        hauteur_u = max(1, int(f.get('hauteur_u') or 1))
    except (TypeError, ValueError):
        hauteur_u = 1
    largeur_u = _clamp_largeur_u(f.get('largeur_u'), col)
    ancien = conn.execute('SELECT id FROM baie_slots WHERE client_id=? AND position=? AND col_index=?', (cid, pos, col)).fetchone()
    # Chevauchement avec un AUTRE élément (voir _slots_en_collision) —
    # AVANT toute suppression : l'éventuel "ancien" à la case d'origine
    # EXACTE est exclu du contrôle (c'est lui qu'on remplace en place, pas
    # un chevauchement), mais tant qu'on n'a pas vérifié, rien n'est encore
    # supprimé — sans quoi un rejet ici aurait déjà perdu l'ancien élément
    # et ses ports pour rien.
    collisions = _slots_en_collision(conn, cid, baie_nom, pos, col, hauteur_u, largeur_u,
                                      exclude_id=(ancien[0] if ancien else None))
    if collisions:
        conn.close()
        return jsonify({'error': _msg_collision(collisions)}), 409
    # Supprimer l'ancien slot à cette position+col si existe — en conservant
    # ses liaisons de ports (placerEquip() ré-envoie systématiquement un
    # POST, y compris pour éditer un slot existant : sans ce report, chaque
    # modification via le panneau "Placer l'équipement" effacerait tous les
    # ports déjà affectés).
    anciens_ports = None
    anciennes_prises = None
    if ancien:
        anciens_ports = {r[0]: (r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]) for r in conn.execute(
            'SELECT numero, appareil_id, peripherique_id, usage_libre, lie_slot_id, lie_port_numero, '
            'piece, cable_couleur, cable_longueur '
            'FROM baie_slot_ports WHERE slot_id=?', (ancien[0],)).fetchall()}
        # Prises murales de l'ancien slot — même report que pour les ports
        # (placerEquip() ré-envoie systématiquement un POST, y compris pour
        # éditer un bandeau existant).
        anciennes_prises = {r[0]: (r[1], r[2], r[3], r[4], r[5], r[6], r[7]) for r in conn.execute(
            'SELECT numero, piece, identification, appareil_id, peripherique_id, usage_libre, cable_couleur, cable_longueur '
            'FROM baie_prises_murales WHERE slot_id=?', (ancien[0],)).fetchall()}
        # Numéros qui ne seront pas recréés (nb_ports/nb_ports_sfp/
        # nb_ports_wan réduits, ou entrée PDU disparue si le type change) :
        # détacher leur éventuel partenaire AVANT de perdre la trace de
        # l'ancien slot_id. Quatre plages distinctes (voir
        # SFP_NUMERO_OFFSET/WAN_NUMERO_OFFSET/ENTREE_NUMERO_OFFSET) : un
        # seuil unique aurait à tort marqué TOUS les ports d'une plage
        # haute comme "supprimés" dès que le plafond d'une AUTRE plage est
        # comparé à leur numéro.
        nb_ports_demande = min(_plafond_nb_ports(f.get('type_equipement', '')), max(0, int(f.get('nb_ports', 0) or 0)))
        nb_ports_sfp_demande = min(PLAFOND_SFP, max(0, int(f.get('nb_ports_sfp', 0) or 0)))
        nb_ports_wan_demande = min(PLAFOND_WAN, max(0, int(f.get('nb_ports_wan', 0) or 0)))
        garde_entree = f.get('type_equipement', '') == 'PDU'
        supprimes = ([n for n in anciens_ports if n <= SFP_NUMERO_OFFSET and n > nb_ports_demande]
                     + [n for n in anciens_ports if SFP_NUMERO_OFFSET < n <= WAN_NUMERO_OFFSET and n > SFP_NUMERO_OFFSET + nb_ports_sfp_demande]
                     + [n for n in anciens_ports if WAN_NUMERO_OFFSET < n <= ENTREE_NUMERO_OFFSET and n > WAN_NUMERO_OFFSET + nb_ports_wan_demande]
                     + ([n for n in anciens_ports if n > ENTREE_NUMERO_OFFSET] if not garde_entree else []))
        if supprimes:
            _detacher_liens_vers(conn, ancien[0], supprimes)
        conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=?', (ancien[0],))
        conn.execute('DELETE FROM baie_prises_murales WHERE slot_id=?', (ancien[0],))
    conn.execute('DELETE FROM baie_slots WHERE client_id=? AND position=? AND col_index=?', (cid, pos, col))
    ports_disposition = f.get('ports_disposition') if f.get('ports_disposition') in ('ligne', 'deux_lignes') else 'ligne'
    orientation = f.get('orientation') if f.get('orientation') in ('horizontal', 'vertical') else 'horizontal'
    puissance_va = max(0, int(f.get('puissance_va', 0) or 0))
    conn.execute('''INSERT INTO baie_slots
        (client_id,position,col_index,hauteur_u,appareil_id,peripherique_id,nom_custom,type_equipement,couleur,description,baie_nom,nb_ports,largeur_u,date_maj,ports_disposition,nb_ports_sfp,nb_ports_wan,orientation,puissance_va)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (cid, pos, col, hauteur_u,
         f.get('appareil_id') or None, f.get('peripherique_id') or None, f.get('nom_custom', ''),
         f.get('type_equipement', ''), f.get('couleur', '#1e3a5f'),
         f.get('description', ''), baie_nom, 0, largeur_u, _utcnow().isoformat(), ports_disposition, 0, 0,
         orientation, puissance_va))
    conn.commit()
    sid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    nb_ports = _reconcilier_ports(conn, sid, f.get('nb_ports', 0), anciens_ports, f.get('type_equipement', ''))
    nb_ports_sfp = _reconcilier_ports_sfp(conn, sid, f.get('nb_ports_sfp', 0), anciens_ports)
    nb_ports_wan = _reconcilier_ports_wan(conn, sid, f.get('nb_ports_wan', 0), anciens_ports)
    _reconcilier_port_entree(conn, sid, f.get('type_equipement', ''), anciens_ports)
    _reconcilier_prises_murales(conn, sid, nb_ports, f.get('type_equipement', ''), anciennes_prises)
    if ancien:
        # Le remplacement change l'id du slot (voir commentaire ci-dessus) :
        # tout port — sur ce slot ou un autre — qui pointait vers l'ANCIEN id
        # doit suivre vers le nouveau, sans quoi le câblage documenté casse à
        # chaque simple modification via le panneau "Placer l'équipement".
        conn.execute('UPDATE baie_slot_ports SET lie_slot_id=? WHERE lie_slot_id=?', (sid, ancien[0]))
    conn.execute('UPDATE baie_slots SET nb_ports=?, nb_ports_sfp=?, nb_ports_wan=? WHERE id=?', (nb_ports, nb_ports_sfp, nb_ports_wan, sid))
    conn.commit()
    slot = row_to_dict(conn.execute(
        '''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.marque, a.en_ligne,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id WHERE s.id=?''', (sid,)).fetchone() or {})
    slot['ports'] = _ports_avec_details(conn, sid)
    nom_slot = slot.get('nom_custom') or slot.get('type_equipement') or slot.get('nom_machine') or f'Emplacement #{sid}'
    detail_sfp = f", {nb_ports_sfp} SFP" if nb_ports_sfp else ''
    detail_wan = f", {nb_ports_wan} WAN" if nb_ports_wan else ''
    log_history(conn, cid, 'baie_slot', sid, nom_slot,
                'Modification (baie)' if ancien else 'Placement',
                f"{baie_nom} · U{pos}·C{col}, {slot.get('type_equipement') or 'sans type'}, {nb_ports} port(s){detail_sfp}{detail_wan}")
    conn.commit()
    conn.close()
    return jsonify(slot)

@app.route('/api/baie/slot/<int:id>', methods=['PUT','DELETE'])
@login_required
def api_baie_slot(id):
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    if request.method == 'DELETE':
        avant = row_to_dict(conn.execute(
            'SELECT nom_custom, type_equipement, position, col_index, baie_nom FROM baie_slots '
            'WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        _detacher_liens_vers(conn, id)
        conn.execute('DELETE FROM baie_slot_ports WHERE slot_id=?', (id,))
        conn.execute('DELETE FROM baie_prises_murales WHERE slot_id=?', (id,))
        conn.execute('DELETE FROM baie_slots WHERE id=? AND client_id=?', (id, cid))
        if avant:
            nom_slot = avant.get('nom_custom') or avant.get('type_equipement') or f'Emplacement #{id}'
            log_history(conn, cid, 'baie_slot', id, nom_slot, 'Retrait',
                        f"{avant.get('baie_nom') or 'Baie principale'} · U{avant.get('position')}·C{avant.get('col_index', 0)}")
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    f = request.json or {}
    nb_ports = min(_plafond_nb_ports(f.get('type_equipement', '')), max(0, int(f.get('nb_ports', 0) or 0)))
    nb_ports_sfp = min(PLAFOND_SFP, max(0, int(f.get('nb_ports_sfp', 0) or 0)))
    nb_ports_wan = min(PLAFOND_WAN, max(0, int(f.get('nb_ports_wan', 0) or 0)))
    ports_disposition = f.get('ports_disposition') if f.get('ports_disposition') in ('ligne', 'deux_lignes') else 'ligne'
    orientation = f.get('orientation') if f.get('orientation') in ('horizontal', 'vertical') else 'horizontal'
    puissance_va = max(0, int(f.get('puissance_va', 0) or 0))
    actuel = conn.execute(
        'SELECT col_index, baie_nom FROM baie_slots WHERE id=? AND client_id=?', (id, cid)).fetchone()
    if not actuel:
        conn.close()
        return jsonify({'error': 'Slot introuvable'}), 404
    # col_index est désormais modifiable par cette route (repli sur la
    # valeur actuelle si absente du payload, pour ne rien changer au geste
    # de redimensionnement qui n'en envoie jamais) — auparavant réservé à
    # /deplacer, ce qui empêchait de déplacer HORIZONTALEMENT un élément en
    # modifiant simplement le champ "Colonne" du panneau "+ Placer" : le
    # formulaire recréait alors un DOUBLON à la nouvelle position sans
    # jamais toucher l'ancien (placerEquip() ne postait que des créations).
    col_index = _clamp_col_index(f['col_index']) if 'col_index' in f else (actuel[0] or 0)
    baie_nom = actuel[1] or 'Baie principale'
    try:
        position = int(f.get('position') or 1)
    except (TypeError, ValueError):
        position = 1
    try:
        hauteur_u = max(1, int(f.get('hauteur_u') or 1))
    except (TypeError, ValueError):
        hauteur_u = 1
    largeur_u = _clamp_largeur_u(f.get('largeur_u'), col_index)
    # Chevauchement avec un AUTRE élément (voir _slots_en_collision),
    # celui-ci exclu de son propre contrôle — sinon impossible de
    # redimensionner/déplacer quoi que ce soit, un slot chevauche toujours
    # SA PROPRE position actuelle.
    collisions = _slots_en_collision(conn, cid, baie_nom, position, col_index, hauteur_u, largeur_u, exclude_id=id)
    if collisions:
        conn.close()
        return jsonify({'error': _msg_collision(collisions)}), 409
    conn.execute('''UPDATE baie_slots SET position=?,col_index=?,hauteur_u=?,appareil_id=?,peripherique_id=?,
        nom_custom=?,type_equipement=?,couleur=?,description=?,nb_ports=?,largeur_u=?,date_maj=?,
        ports_disposition=?,nb_ports_sfp=?,nb_ports_wan=?,orientation=?,puissance_va=? WHERE id=? AND client_id=?''',
        (position, col_index, hauteur_u, f.get('appareil_id') or None, f.get('peripherique_id') or None,
         f.get('nom_custom',''), f.get('type_equipement',''),
         f.get('couleur','#1e3a5f'), f.get('description',''), nb_ports,
         largeur_u,
         _utcnow().isoformat(), ports_disposition, nb_ports_sfp, nb_ports_wan,
         orientation, puissance_va, id, cid))
    _reconcilier_ports(conn, id, nb_ports, type_equipement=f.get('type_equipement', ''))
    nb_ports_sfp = _reconcilier_ports_sfp(conn, id, nb_ports_sfp)
    nb_ports_wan = _reconcilier_ports_wan(conn, id, nb_ports_wan)
    _reconcilier_port_entree(conn, id, f.get('type_equipement', ''))
    conn.execute('UPDATE baie_slots SET nb_ports_sfp=?, nb_ports_wan=? WHERE id=?', (nb_ports_sfp, nb_ports_wan, id))
    _reconcilier_prises_murales(conn, id, nb_ports, f.get('type_equipement', ''))
    conn.commit()
    slot = row_to_dict(conn.execute(
        '''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.marque, a.en_ligne,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id WHERE s.id=?''', (id,)).fetchone() or {})
    slot['ports'] = _ports_avec_details(conn, id)
    if slot:
        nom_slot = slot.get('nom_custom') or slot.get('type_equipement') or f'Emplacement #{id}'
        detail_sfp = f", {nb_ports_sfp} SFP" if nb_ports_sfp else ''
        detail_wan = f", {nb_ports_wan} WAN" if nb_ports_wan else ''
        log_history(conn, cid, 'baie_slot', id, nom_slot, 'Modification (baie)',
                    f"{slot.get('baie_nom') or 'Baie principale'} · U{slot.get('position')}·C{slot.get('col_index',0)}, "
                    f"{slot.get('hauteur_u',1)}U, {nb_ports} port(s){detail_sfp}{detail_wan}")
        conn.commit()
    conn.close()
    return jsonify(slot)

@app.route('/api/baie/slot/<int:slot_id>/port/<int:numero>', methods=['PUT'])
@login_required
def api_baie_port(slot_id, numero):
    """Lie un port à un appareil, un périphérique, ou un usage libre —
    mutuellement exclusifs, comme pour un slot entier (un port = une seule
    cible à la fois). Sert aussi de déliaison pour un lien port-à-port : si
    ce port était câblé à un autre (voir api_baie_lien_port), ce dernier est
    détaché des deux côtés — le champ "— Libre —" du sélecteur de port
    appelle cette même route sans rien renseigner."""
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    slot = conn.execute('SELECT id, nom_custom, type_equipement FROM baie_slots WHERE id=? AND client_id=?', (slot_id, cid)).fetchone()
    if not slot:
        conn.close()
        return jsonify({'error': 'Slot introuvable'}), 404
    f = request.json or {}
    now = _utcnow().isoformat()
    # `piece` a déménagé sur la prise murale (voir api_baie_prise_murale) —
    # un port RJ, bandeau ou non, ne la porte plus : cette route n'a donc
    # plus besoin de la distinguer d'un changement de cible.
    cible_fournie = any(k in f for k in ('appareil_id', 'peripherique_id', 'usage_libre'))
    if cible_fournie:
        appareil_id = f.get('appareil_id') or None
        peripherique_id = f.get('peripherique_id') or None
        usage_libre = f.get('usage_libre', '').strip() if not (appareil_id or peripherique_id) else ''
        ancien_lien = conn.execute(
            'SELECT lie_slot_id, lie_port_numero FROM baie_slot_ports WHERE slot_id=? AND numero=?',
            (slot_id, numero)).fetchone()
        if ancien_lien and ancien_lien[0]:
            conn.execute("UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
                         "cable_couleur='', cable_longueur='', date_maj=? "
                         'WHERE slot_id=? AND numero=?', (now, ancien_lien[0], ancien_lien[1]))
        conn.execute('''UPDATE baie_slot_ports SET appareil_id=?,peripherique_id=?,usage_libre=?,
            lie_slot_id=NULL,lie_port_numero=NULL,cable_couleur='',cable_longueur='',date_maj=?
            WHERE slot_id=? AND numero=?''',
            (appareil_id, peripherique_id, usage_libre, now, slot_id, numero))
    elif f.get('detacher_lien'):
        # Détache SEULEMENT le lien (bouton ✕ sur un port câblé, voir
        # delierPort() côté client) — sans toucher à une éventuelle
        # association directe sur CE port (appareil_id/peripherique_id/
        # usage_libre) : depuis que api_baie_lien_port ne les efface plus
        # à la création d'un lien (un port de bandeau RJ peut légitimement
        # porter les deux à la fois), les effacer ICI aussi aurait perdu
        # l'association directe pour un simple débranchement du cordon.
        ancien_lien = conn.execute(
            'SELECT lie_slot_id, lie_port_numero FROM baie_slot_ports WHERE slot_id=? AND numero=?',
            (slot_id, numero)).fetchone()
        if ancien_lien and ancien_lien[0]:
            conn.execute("UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
                         "cable_couleur='', cable_longueur='', date_maj=? "
                         'WHERE slot_id=? AND numero=?', (now, ancien_lien[0], ancien_lien[1]))
        conn.execute("UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
                     "cable_couleur='', cable_longueur='', date_maj=? WHERE slot_id=? AND numero=?",
                     (now, slot_id, numero))
    conn.commit()
    ports = _ports_avec_details(conn, slot_id)
    port = next((p for p in ports if p['numero'] == numero), None)
    detacher_lien = (not cible_fournie) and bool(f.get('detacher_lien'))
    if port and (cible_fournie or detacher_lien):
        nom_slot = slot[1] or slot[2] or f'Emplacement #{slot_id}'
        if cible_fournie:
            detail = f"Port {numero} -> {port.get('nom_cible') or '— Libre —'}"
        else:
            detail = f"Port {numero} : lien détaché"
        log_history(conn, cid, 'baie_slot', slot_id, nom_slot, 'Modification (port baie)', detail)
        conn.commit()
    conn.close()
    return jsonify(port or {'error': 'Port introuvable'}), (200 if port else 404)

@app.route('/api/baie/prise-murale/<int:slot_id>/<int:numero>', methods=['PUT'])
@login_required
def api_baie_prise_murale(slot_id, numero):
    """Édite UNE prise murale d'un bandeau RJ — la pièce desservie, son
    identification (repère physique de la prise, ex. "RJ 3.12" imprimé sur
    la plaque murale — distinct de la pièce) et l'appareil/périphérique/
    usage qui y est branché, plus l'étiquette du câble fixe mur -> bandeau.
    Chaque champ peut être fourni seul (payload partiel, voir les panneaux
    d'édition dédiés) ou tous ensemble (payload complet, voir la modale
    d'édition par double-clic) : chaque clé du payload ne touche que sa
    propre colonne, les absentes restent inchangées. Entité séparée du port
    RJ de même numéro (voir baie_prises_murales dans init_db()) : le port
    RJ, lui, ne sert plus qu'à interconnecter avec un autre élément de la
    baie (voir api_baie_lien_port) et n'est jamais modifié par cette
    route."""
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    slot = conn.execute(
        "SELECT id, nom_custom, type_equipement FROM baie_slots "
        "WHERE id=? AND client_id=? AND type_equipement='Bandeau RJ'", (slot_id, cid)).fetchone()
    if not slot:
        conn.close()
        return jsonify({'error': 'Bandeau RJ introuvable'}), 404
    existe = conn.execute(
        'SELECT 1 FROM baie_prises_murales WHERE slot_id=? AND numero=?', (slot_id, numero)).fetchone()
    if not existe:
        conn.close()
        return jsonify({'error': 'Prise murale introuvable'}), 404
    f = request.json or {}
    now = _utcnow().isoformat()
    cible_fournie = any(k in f for k in ('appareil_id', 'peripherique_id', 'usage_libre'))
    if cible_fournie:
        appareil_id = f.get('appareil_id') or None
        peripherique_id = f.get('peripherique_id') or None
        usage_libre = f.get('usage_libre', '').strip() if not (appareil_id or peripherique_id) else ''
        conn.execute('''UPDATE baie_prises_murales SET appareil_id=?,peripherique_id=?,usage_libre=?,date_maj=?
            WHERE slot_id=? AND numero=?''', (appareil_id, peripherique_id, usage_libre, now, slot_id, numero))
    piece_fournie = 'piece' in f
    if piece_fournie:
        conn.execute('UPDATE baie_prises_murales SET piece=?, date_maj=? WHERE slot_id=? AND numero=?',
                     (str(f.get('piece') or '').strip(), now, slot_id, numero))
    identification_fournie = 'identification' in f
    if identification_fournie:
        conn.execute('UPDATE baie_prises_murales SET identification=?, date_maj=? WHERE slot_id=? AND numero=?',
                     (str(f.get('identification') or '').strip(), now, slot_id, numero))
    cable_fourni = ('cable_couleur' in f) or ('cable_longueur' in f)
    if cable_fourni:
        conn.execute('UPDATE baie_prises_murales SET cable_couleur=?, cable_longueur=?, date_maj=? '
                     'WHERE slot_id=? AND numero=?',
                     (str(f.get('cable_couleur') or '').strip(), str(f.get('cable_longueur') or '').strip(),
                      now, slot_id, numero))
    conn.commit()
    prises = _prises_murales_avec_details(conn, slot_id)
    prise = prises.get(numero)
    if prise and (cible_fournie or piece_fournie or identification_fournie or cable_fourni):
        nom_slot = slot[1] or slot[2] or f'Emplacement #{slot_id}'
        if cible_fournie:
            detail = f"Prise murale {numero} -> {prise.get('nom_cible') or '— Libre —'}"
        elif piece_fournie and not identification_fournie:
            detail = f"Prise murale {numero}, pièce -> {prise.get('piece') or '(vide)'}"
        elif identification_fournie and not piece_fournie:
            detail = f"Prise murale {numero}, identification -> {prise.get('identification') or '(vide)'}"
        elif piece_fournie or identification_fournie:
            detail = f"Prise murale {numero} : infos mises à jour"
        else:
            detail = f"Prise murale {numero} : câble mis à jour"
        log_history(conn, cid, 'baie_slot', slot_id, nom_slot, 'Modification (prise murale)', detail)
        conn.commit()
    conn.close()
    return jsonify(prise or {'error': 'Prise murale introuvable'}), (200 if prise else 404)

@app.route('/api/baie/lien-port', methods=['POST'])
@login_required
def api_baie_lien_port():
    """Crée un lien port-à-port (câblage physique switch<->routeur, etc.),
    bidirectionnel : les deux ports se référencent mutuellement. Toute
    liaison précédente sur L'UN OU L'AUTRE port — y compris avec un
    troisième port différent — est détachée des deux côtés au passage, pour
    ne jamais laisser un lien à sens unique après un reclassement.

    N'efface PLUS appareil_id/peripherique_id/usage_libre du port : un port
    de bandeau RJ peut légitimement porter les DEUX à la fois — l'appareil
    qu'il dessert (câblage fixe vers la prise murale, "Associer à un
    appareil") ET un lien (cordon de brassage vers un switch) — ce n'est
    pas un conflit, ce sont deux informations physiquement distinctes sur
    le MÊME port. Les effacer empêchait de jamais représenter le cas
    standard du brassage structuré (bug corrigé, signalé en usage réel :
    lier un port de bandeau déjà associé à un appareil au switch effaçait
    silencieusement cette association). La résolution du port EN FACE
    (couleur/ping/tooltip, voir _ports_avec_details) donne toujours la
    priorité à l'association directe d'un port sur son propre lien."""
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    f = request.json or {}
    s1, n1 = f.get('slot1_id'), f.get('numero1')
    s2, n2 = f.get('slot2_id'), f.get('numero2')
    # Étiquette de câble optionnelle (couleur/longueur) — texte libre, purement
    # documentaire, indépendante de tout le reste (voir migration init_db()).
    cable_couleur = str(f.get('cable_couleur') or '').strip()
    cable_longueur = str(f.get('cable_longueur') or '').strip()
    if not (s1 and n1 and s2 and n2) or (s1 == s2 and n1 == n2):
        return jsonify({'error': 'Deux ports distincts sont requis'}), 400
    conn = get_db()
    slots_noms = {r[0]: (r[1] or r[2] or f'Emplacement #{r[0]}') for r in conn.execute(
        'SELECT id, nom_custom, type_equipement FROM baie_slots WHERE id IN (?,?) AND client_id=?',
        (s1, s2, cid)).fetchall()}
    # Comparer l'appartenance de s1 ET s2 individuellement, pas len(...) != 2 :
    # relier deux ports du MÊME équipement (boucle de test, brassage interne)
    # est un cas légitime où s1 == s2 — la requête IN (?,?) ne renvoie alors
    # qu'UNE ligne pour les deux, ce qui faisait échouer ce cas pourtant
    # valide avec un faux "Port introuvable" (bug trouvé en vérifiant tout le
    # mécanisme de liaison suite à un signalement d'affichage de liens).
    if s1 not in slots_noms or s2 not in slots_noms:
        conn.close()
        return jsonify({'error': 'Port introuvable'}), 404
    now = _utcnow().isoformat()
    for sid, num in ((s1, n1), (s2, n2)):
        ancien = conn.execute(
            'SELECT lie_slot_id, lie_port_numero FROM baie_slot_ports WHERE slot_id=? AND numero=?',
            (sid, num)).fetchone()
        if ancien and ancien[0]:
            conn.execute("UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL, "
                         "cable_couleur='', cable_longueur='', date_maj=? "
                         'WHERE slot_id=? AND numero=?', (now, ancien[0], ancien[1]))
    conn.execute('''UPDATE baie_slot_ports SET lie_slot_id=?,lie_port_numero=?,
        cable_couleur=?,cable_longueur=?,date_maj=?
        WHERE slot_id=? AND numero=?''',
        (s2, n2, cable_couleur, cable_longueur, now, s1, n1))
    conn.execute('''UPDATE baie_slot_ports SET lie_slot_id=?,lie_port_numero=?,
        cable_couleur=?,cable_longueur=?,date_maj=?
        WHERE slot_id=? AND numero=?''',
        (s1, n1, cable_couleur, cable_longueur, now, s2, n2))
    log_history(conn, cid, 'baie_slot', s1, slots_noms[s1], 'Câblage (baie)',
                f"Port {n1} relié à {slots_noms[s2]} port {n2}")
    conn.commit()
    ports1 = _ports_avec_details(conn, s1)
    ports2 = ports1 if s2 == s1 else _ports_avec_details(conn, s2)
    conn.close()
    port1 = next((p for p in ports1 if p['numero'] == n1), None)
    port2 = next((p for p in ports2 if p['numero'] == n2), None)
    return jsonify({'port1': port1, 'port2': port2})

@app.route('/api/baie/slot/<int:id>/deplacer', methods=['POST'])
@login_required
def api_baie_deplacer_slot(id):
    '''Drag & drop : déplace un slot vers une nouvelle position/col.'''
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    f = request.json or {}
    new_pos = f.get('position', 1)
    new_col = _clamp_col_index(f.get('col_index', 0))
    conn = get_db()
    # La largeur/hauteur du slot déplacé ne changent pas ici (seule sa
    # position bouge) — mais new_col doit rester compatible avec la
    # largeur, sans quoi le slot déborderait de la grille à 10 colonnes
    # (voir _clamp_largeur_u), et les deux sont nécessaires pour le
    # contrôle de chevauchement ci-dessous (rectangle complet, pas
    # seulement la case d'origine visée).
    actuel = conn.execute(
        'SELECT largeur_u, hauteur_u, baie_nom FROM baie_slots WHERE id=? AND client_id=?', (id, cid)).fetchone()
    if not actuel:
        conn.close()
        return jsonify({'error': 'Slot introuvable'}), 404
    largeur_actuelle = actuel[0] or 10
    hauteur_actuelle = actuel[1] or 1
    baie_nom = actuel[2] or 'Baie principale'
    new_col = min(new_col, max(0, 10 - largeur_actuelle))
    # Chevauchement avec un AUTRE élément (voir _slots_en_collision) — avant
    # (2.18.69), un dépôt écrasait silencieusement tout ce qui occupait
    # EXACTEMENT la case visée (jamais atteignable en pratique, on ne peut
    # déposer que sur une case vide) sans jamais vérifier les rangées
    # supplémentaires couvertes par un élément haut de plusieurs U : un
    # déplacement pouvait ainsi silencieusement chevaucher un élément
    # occupant une partie de ces rangées, sans le moindre avertissement.
    collisions = _slots_en_collision(conn, cid, baie_nom, new_pos, new_col, hauteur_actuelle, largeur_actuelle, exclude_id=id)
    if collisions:
        conn.close()
        return jsonify({'error': _msg_collision(collisions)}), 409
    conn.execute('UPDATE baie_slots SET position=?, col_index=? WHERE id=? AND client_id=?',
                 (new_pos, new_col, id, cid))
    conn.commit()
    slot = row_to_dict(conn.execute(
        '''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.marque, a.en_ligne,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id WHERE s.id=?''', (id,)).fetchone() or {})
    slot['ports'] = _ports_avec_details(conn, id)
    if slot:
        nom_slot = slot.get('nom_custom') or slot.get('type_equipement') or f'Emplacement #{id}'
        log_history(conn, cid, 'baie_slot', id, nom_slot, 'Déplacement',
                    f"{slot.get('baie_nom') or 'Baie principale'} · vers U{new_pos}·C{new_col}")
        conn.commit()
    conn.close()
    return jsonify(slot)

@app.route('/api/baie', methods=['DELETE'])
@login_required
def api_baie_supprimer():
    """Supprime une baie entière (tous ses emplacements) — pas un simple
    « vider » côté client (toutVider(), qui garde la baie mais efface son
    contenu) : ici la baie elle-même disparaît. Une baie n'existe
    qu'implicitement, comme valeur distincte de baie_slots.baie_nom (voir
    api_baie_slots ci-dessous) — la supprimer revient donc à supprimer tous
    les emplacements qui portent ce nom, y compris les emplacements
    "historiques" sans baie_nom explicite (NULL, traité comme 'Baie
    principale' partout ailleurs dans ce fichier)."""
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    baie_nom = request.args.get('baie', 'Baie principale')
    conn = get_db()
    if baie_nom == 'Baie principale':
        cond, params = "(baie_nom=? OR baie_nom IS NULL)", (cid, baie_nom)
    else:
        cond, params = "baie_nom=?", (cid, baie_nom)
    slot_ids = [r[0] for r in conn.execute(
        f'SELECT id FROM baie_slots WHERE client_id=? AND {cond}', params).fetchall()]
    if slot_ids:
        # Un port d'une AUTRE baie peut avoir été câblé vers un slot d'ici —
        # sans ce nettoyage, il resterait à pointer vers un slot supprimé.
        placeholders = ','.join('?' * len(slot_ids))
        conn.execute(f'UPDATE baie_slot_ports SET lie_slot_id=NULL, lie_port_numero=NULL '
                     f'WHERE lie_slot_id IN ({placeholders})', slot_ids)
        conn.execute(f'DELETE FROM baie_slot_ports WHERE slot_id IN ({placeholders})', slot_ids)
        conn.execute(f'DELETE FROM baie_prises_murales WHERE slot_id IN ({placeholders})', slot_ids)
    conn.execute(f'DELETE FROM baie_slots WHERE client_id=? AND {cond}', params)
    log_history(conn, cid, 'baie', 0, baie_nom, 'Suppression (baie)',
                f"{len(slot_ids)} emplacement(s) supprimé(s)")
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/baie/slots')
@login_required
def api_baie_slots():
    cid = get_client_id()
    baie_nom = request.args.get('baie', 'Baie principale')
    conn = get_db()
    slots = [row_to_dict(r) for r in conn.execute(
        '''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.marque, a.modele, a.en_ligne, a.ports_ouverts,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id
           WHERE s.client_id=? AND (s.baie_nom=? OR (s.baie_nom IS NULL AND ?='Baie principale'))
           ORDER BY s.position, s.col_index''', (cid, baie_nom, baie_nom)).fetchall()]
    for s in slots:
        if s.get('nb_ports'):
            s['ports'] = _ports_avec_details(conn, s['id'])
    parc = row_to_dict(conn.execute('SELECT baie_nb_u, switch_nb_unites FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    # Liste des baies existantes
    baies = [r[0] for r in conn.execute(
        "SELECT DISTINCT COALESCE(baie_nom,'Baie principale') FROM baie_slots WHERE client_id=? ORDER BY 1",
        (cid,)).fetchall()]
    if not baies: baies = ['Baie principale']
    if 'Baie principale' not in baies: baies.insert(0, 'Baie principale')
    conn.close()
    return jsonify({'slots': slots, 'nb_u': parc.get('baie_nb_u', 12) or 12, 'baies': baies})

def _liste_cablage(conn, cid):
    """Chaque lien port-à-port du client, une seule fois (pas les deux sens),
    avec le nom des deux éléments — sert à la fois à la page imprimable et à
    l'export CSV. La condition d'ordre (slot_id, numero) élimine le doublon
    inhérent au stockage bidirectionnel de baie_slot_ports (chaque port du
    lien porte sa propre ligne, référençant l'autre).
    Exclut désormais les liens dont un bout est un port de bandeau RJ — ces
    liens-là (bandeau -> switch/routeur, le cas standard du brassage
    structuré) sont documentés en détail, avec la prise murale d'en face,
    par _fiche_prises_murales() : les reprendre ici ferait doublon avec la
    fiche de brassage. Ne restent que les interconnexions "backbone" entre
    deux éléments ni l'un ni l'autre bandeau (switch<->routeur,
    switch<->switch, PDU...)."""
    rows = conn.execute('''
        SELECT COALESCE(s1.baie_nom,'Baie principale') AS baie,
               s1.position AS pos1, s1.nom_custom AS nom1_custom, s1.type_equipement AS type1,
               bp.numero AS numero1, bp.cable_couleur, bp.cable_longueur,
               s2.position AS pos2, s2.nom_custom AS nom2_custom, s2.type_equipement AS type2,
               bp.lie_port_numero AS numero2
        FROM baie_slot_ports bp
        JOIN baie_slots s1 ON bp.slot_id = s1.id
        JOIN baie_slots s2 ON bp.lie_slot_id = s2.id
        WHERE s1.client_id=? AND s2.client_id=? AND bp.lie_slot_id IS NOT NULL
          AND s1.type_equipement != 'Bandeau RJ' AND s2.type_equipement != 'Bandeau RJ'
          AND (bp.slot_id < bp.lie_slot_id
               OR (bp.slot_id = bp.lie_slot_id AND bp.numero < bp.lie_port_numero))
        ORDER BY baie, pos1, bp.numero
    ''', (cid, cid)).fetchall()
    liens = []
    for r in rows:
        d = row_to_dict(r)
        d['nom1'] = d['nom1_custom'] or d['type1'] or f"U{d['pos1']}"
        d['nom2'] = d['nom2_custom'] or d['type2'] or f"U{d['pos2']}"
        # Libellé AFFICHÉ du port (voir _libelle_port_pour_type), pas le
        # numéro BRUT stocké — bug trouvé en auditant : la fiche imprimable
        # et l'export CSV montraient jusqu'ici "Port 1001"/"Port 9001" au
        # lieu de "Port SFP 1"/"Entrée" pour un lien SFP/WAN/entrée de PDU.
        d['numero1_label'] = _libelle_port_pour_type(d['type1'], d['numero1'])
        d['numero2_label'] = _libelle_port_pour_type(d['type2'], d['numero2'])
        liens.append(d)
    return liens

def _fiche_prises_murales(conn, cid):
    """Une ligne par port de CHAQUE bandeau RJ du client (les 24 au complet,
    affectés ou non — une fiche de brassage documente tout le panneau, pas
    seulement les prises déjà câblées, pour savoir d'un coup d'œil ce qui
    reste disponible) : pièce desservie, identification de la prise,
    appareil/périphérique/usage qui y est branché, et — si le port RJ
    correspondant est câblé vers un autre élément — la chaîne complète
    jusqu'à la cible finale. Réutilise telle quelle la résolution déjà
    faite par _ports_avec_details() (même donnée que l'écran de la baie,
    y compris la couleur/le statut), pas de requête SQL séparée à
    maintenir en double."""
    bandeaux = conn.execute(
        "SELECT id, baie_nom, position, nom_custom FROM baie_slots "
        "WHERE client_id=? AND type_equipement='Bandeau RJ' "
        "ORDER BY COALESCE(baie_nom,'Baie principale'), position", (cid,)).fetchall()
    lignes = []
    for slot_id, baie_nom, position, nom_custom in bandeaux:
        nom_bandeau = nom_custom or f'U{position}'
        for p in _ports_avec_details(conn, slot_id):
            pm = p.get('prise_murale') or {}
            lignes.append({
                'baie': baie_nom or 'Baie principale',
                'bandeau': nom_bandeau,
                'numero': p['numero'],
                'piece': pm.get('piece') or '',
                'identification': pm.get('identification') or '',
                'prise_cible': pm.get('nom_cible') or '',
                # Deux câbles PHYSIQUEMENT distincts (voir baie_prises_murales
                # dans init_db()) : le câble mural fixe (mur -> bandeau, sur
                # la prise murale elle-même) et le cordon de brassage
                # (bandeau -> élément relié, sur le port RJ, baie_slot_ports)
                # — les confondre ferait perdre exactement l'info que cette
                # table sépare depuis la 2.18.74.
                'cable_mural_couleur': pm.get('cable_couleur') or '',
                'cable_mural_longueur': pm.get('cable_longueur') or '',
                'lien_cible': p.get('nom_cible') if p.get('lie_slot_id') else '',
                'cible_finale': p.get('cible_finale') or '',
                'cordon_couleur': p.get('cable_couleur') or '',
                'cordon_longueur': p.get('cable_longueur') or '',
            })
    return lignes

@app.route('/baie/cablage')
@login_required
def baie_cablage():
    """Fiche de brassage imprimable — documentation complète à laisser sur
    site ou pour un audit, en deux parties : chaque prise murale d'un
    bandeau RJ (pièce, identification, appareil branché, chaîne jusqu'à sa
    cible finale — voir _fiche_prises_murales) puis les interconnexions
    "backbone" entre éléments non-bandeau (switch<->routeur, etc. — voir
    _liste_cablage), sans naviguer élément par élément dans l'éditeur."""
    cid = get_client_id()
    if not get_client_access(cid):
        flash('Accès refusé', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    prises = _fiche_prises_murales(conn, cid)
    liens = _liste_cablage(conn, cid)
    conn.close()
    return render_template('baie_cablage.html', client=client, liens=liens, prises=prises,
                           clients=get_clients(), client_actif_id=cid,
                           date_export=_utcnow().strftime('%d/%m/%Y %H:%M'))

@app.route('/baie/cablage.csv')
@login_required
def baie_cablage_csv():
    cid = get_client_id()
    if not get_client_access(cid):
        return jsonify({'error': 'Accès refusé'}), 403
    conn = get_db()
    prises = _fiche_prises_murales(conn, cid)
    liens = _liste_cablage(conn, cid)
    conn.close()
    buf = io.StringIO()
    buf.write('﻿')  # BOM — Excel ouvre le CSV en UTF-8 sans le corrompre
    w = csv.writer(buf, delimiter=';')
    w.writerow(['PRISES MURALES'])
    w.writerow(['Baie', 'Bandeau', 'Port', 'Pièce', 'Identification', 'Prise branchée à',
                'Câble mural', 'Relié (port RJ) à', 'Cordon', 'Cible finale'])
    for p in prises:
        w.writerow([p['baie'], p['bandeau'], p['numero'], p['piece'], p['identification'],
                    p['prise_cible'],
                    ' · '.join(filter(None, [p['cable_mural_couleur'], p['cable_mural_longueur']])),
                    p['lien_cible'],
                    ' · '.join(filter(None, [p['cordon_couleur'], p['cordon_longueur']])),
                    p['cible_finale']])
    w.writerow([])
    w.writerow(['INTERCONNEXIONS'])
    w.writerow(['Baie', 'Élément A', 'Port A', 'Élément B', 'Port B', 'Couleur câble', 'Longueur câble'])
    for l in liens:
        w.writerow([l['baie'], l['nom1'], l['numero1_label'], l['nom2'], l['numero2_label'],
                    l['cable_couleur'], l['cable_longueur']])
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename="fiche_brassage.csv"'
    return resp

# ─── PHOTOS BAIE ─────────────────────────────────────────────────────────────

@app.route('/baie/photo/upload', methods=['POST'])
@login_required
def upload_photo_baie():
    cid = get_client_id()
    if not can_write():
        flash('Accès en lecture seule', 'danger')
        return redirect(url_for('baie_brassage'))
    if 'fichier' not in request.files:
        return redirect(url_for('baie_brassage'))
    f = request.files['fichier']
    # Une photo de baie est une image : rien d'autre n'a de sens ici.
    ok, motif = verifier_fichier(f, ALLOWED_IMAGE_EXTENSIONS)
    if not ok:
        flash(motif, 'danger')
        return redirect(url_for('baie_brassage'))
    safe = secure_filename(f.filename)
    unique = f"baie{cid}_{int(time.time())}_{safe}"
    save_path = os.path.join(UPLOAD_FOLDER, unique)
    f.save(save_path)
    taille = os.path.getsize(save_path)
    nom = request.form.get('nom', '') or f.filename
    desc = request.form.get('description', '')
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute('INSERT INTO baie_photos (client_id,nom,description,nom_fichier,taille,date_upload) VALUES (?,?,?,?,?,?)',
                 (cid, nom, desc, unique, taille, now))
    conn.commit(); conn.close()
    flash(f'Photo « {nom} » ajoutée', 'success')
    return redirect(url_for('baie_brassage'))

@app.route('/baie/photo/<int:id>', methods=['PUT'])
@login_required
def modifier_photo_baie(id):
    """Édite le nom/la description d'une photo déjà uploadée — jusqu'ici
    seuls l'upload (nom/description figés une fois pour toutes, voir
    upload_photo_baie ci-dessus) et la suppression existaient, aucun moyen
    de corriger une photo mal nommée sans la supprimer puis la
    re-uploader."""
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    photo = row_to_dict(conn.execute('SELECT id FROM baie_photos WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if not photo:
        conn.close()
        return jsonify({'error': 'Photo introuvable'}), 404
    f = request.json or {}
    nom = str(f.get('nom') or '').strip()
    if not nom:
        conn.close()
        return jsonify({'error': 'Nom requis'}), 400
    description = str(f.get('description') or '').strip()
    conn.execute('UPDATE baie_photos SET nom=?, description=? WHERE id=? AND client_id=?',
                 (nom, description, id, cid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'nom': nom, 'description': description})

@app.route('/baie/photo/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_photo_baie(id):
    cid = get_client_id()
    if not can_write():
        flash('Accès en lecture seule', 'danger')
        return redirect(url_for('baie_brassage'))
    conn = get_db()
    photo = row_to_dict(conn.execute('SELECT * FROM baie_photos WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if photo:
        conn.execute('DELETE FROM baie_photos WHERE id=?', (id,))
        conn.commit()
        try: os.remove(os.path.join(UPLOAD_FOLDER, photo['nom_fichier']))
        except: pass
    conn.close()
    return redirect(url_for('baie_brassage'))

@app.route('/baie/photo/<int:id>/apercu')
@login_required
def apercu_photo_baie(id):
    cid = get_client_id()
    conn = get_db()
    photo = row_to_dict(conn.execute('SELECT * FROM baie_photos WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not photo: return 'Not found', 404

    # Préférer servir depuis BLOB si disponible (synced depuis une autre machine)
    if photo.get('contenu_blob'):
        return send_file(io.BytesIO(photo['contenu_blob']), as_attachment=False, download_name=photo.get('nom_fichier', 'photo'))

    fichier_path = os.path.join(UPLOAD_FOLDER, photo['nom_fichier'])
    if not os.path.exists(fichier_path):
        return f"Fichier introuvable : {photo.get('nom_fichier', '?')}", 404
    return send_from_directory(UPLOAD_FOLDER, photo['nom_fichier'], as_attachment=False)

@app.route('/api/baie/nb_u', methods=['POST'])
@login_required
def api_baie_nb_u():
    cid = get_client_id()
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule'}), 403
    data = request.json or {}
    nb = max(6, min(48, int(data.get('nb_u', 12))))
    conn = get_db()
    conn.execute('UPDATE parc_general SET baie_nb_u=? WHERE client_id=?', (nb, cid))
    conn.commit(); conn.close()
    return jsonify({'nb_u': nb})

# ─── MOTEUR DE SCAN MULTI-THREAD ─────────────────────────────────────────────

scan_status = {"running": False, "progress": 0, "message": "", "results": [], "errors": []}
scan_lock = threading.Lock()
IS_WINDOWS = platform.system() == "Windows"

# ── OUI VENDOR LOOKUP (top fabricants embarqués) ────────────────────────────
_OUI = {
    # ── VMware / Virtualisation ───────────────────────────────────────────────
    "00:50:56":"VMware","00:0c:29":"VMware","00:05:69":"VMware","00:1c:14":"VMware",
    "08:00:27":"VirtualBox","0a:00:27":"VirtualBox",
    "52:54:00":"QEMU/KVM","00:16:3e":"Xen",
    # ── Raspberry Pi Foundation ───────────────────────────────────────────────
    "b8:27:eb":"Raspberry Pi","dc:a6:32":"Raspberry Pi","e4:5f:01":"Raspberry Pi",
    "d8:3a:dd":"Raspberry Pi","28:cd:c1":"Raspberry Pi",
    # ── Apple ────────────────────────────────────────────────────────────────
    "00:03:93":"Apple","00:05:02":"Apple","00:0a:27":"Apple","00:0a:95":"Apple",
    "00:11:24":"Apple","00:14:51":"Apple","00:16:cb":"Apple","00:17:f2":"Apple",
    "00:19:e3":"Apple","00:1b:63":"Apple","00:1c:b3":"Apple","00:1d:4f":"Apple",
    "00:1e:52":"Apple","00:1e:c2":"Apple","00:1f:5b":"Apple","00:1f:f3":"Apple",
    "00:21:e9":"Apple","00:22:41":"Apple","00:23:12":"Apple","00:23:32":"Apple",
    "00:23:6c":"Apple","00:24:36":"Apple","00:25:00":"Apple","00:25:4b":"Apple",
    "00:25:bc":"Apple","00:26:08":"Apple","00:26:b0":"Apple","00:26:bb":"Apple",
    "3c:07:54":"Apple","3c:15:c2":"Apple","3c:22:fb":"Apple","40:6c:8f":"Apple",
    "40:a6:d9":"Apple","4c:8d:79":"Apple","60:fb:42":"Apple","64:a3:cb":"Apple",
    "70:11:24":"Apple","70:cd:60":"Apple","78:7b:8a":"Apple","7c:f0:5f":"Apple",
    "88:19:08":"Apple","88:e8:7f":"Apple","8c:00:6d":"Apple","8c:58:77":"Apple",
    "90:72:40":"Apple","98:01:a7":"Apple","98:03:d8":"Apple","a4:b1:97":"Apple",
    "a4:c3:f0":"Apple","a8:20:66":"Apple","ac:bc:32":"Apple","b4:18:d1":"Apple",
    "b8:09:8a":"Apple","b8:8d:12":"Apple","b8:c7:5d":"Apple","bc:52:b7":"Apple",
    "c8:69:cd":"Apple","c8:b5:b7":"Apple","cc:29:f5":"Apple","d0:23:db":"Apple",
    "d4:61:9d":"Apple","d8:bb:c1":"Apple","dc:2b:2a":"Apple","e0:ac:cb":"Apple",
    "e4:ce:8f":"Apple","e8:04:0b":"Apple","f0:b4:79":"Apple","f0:d1:a9":"Apple",
    "f4:0f:24":"Apple","f8:1e:df":"Apple","f8:27:93":"Apple","fc:25:3f":"Apple",
    # ── Microsoft ────────────────────────────────────────────────────────────
    "00:03:ff":"Microsoft","00:12:5a":"Microsoft","00:15:5d":"Microsoft (Hyper-V)",
    "00:17:fa":"Microsoft","00:50:f2":"Microsoft","28:18:78":"Microsoft",
    "48:b0:2d":"Microsoft","50:1a:c5":"Microsoft","7c:1e:52":"Microsoft",
    # ── Dell ─────────────────────────────────────────────────────────────────
    "00:06:5b":"Dell","00:08:74":"Dell","00:0b:db":"Dell","00:0d:56":"Dell",
    "00:0f:1f":"Dell","00:11:43":"Dell","00:12:3f":"Dell","00:13:72":"Dell",
    "00:14:22":"Dell","00:15:c5":"Dell","00:16:f0":"Dell","00:18:8b":"Dell",
    "00:19:b9":"Dell","00:1a:a0":"Dell","00:1c:23":"Dell","00:1d:09":"Dell",
    "00:1e:4f":"Dell","00:21:70":"Dell","00:22:19":"Dell","00:23:ae":"Dell",
    "00:24:e8":"Dell","00:25:64":"Dell","00:26:b9":"Dell","08:00:37":"Dell",
    "0c:c4:7a":"Dell","10:98:36":"Dell","14:18:77":"Dell","14:fe:b5":"Dell",
    "18:03:73":"Dell","18:66:da":"Dell","18:a9:9b":"Dell","1c:40:24":"Dell",
    "20:47:47":"Dell","24:b6:fd":"Dell","28:f1:0e":"Dell","2c:76:8a":"Dell",
    "34:17:eb":"Dell","34:e6:d7":"Dell","38:63:bb":"Dell","3c:a8:2a":"Dell",
    "44:a8:42":"Dell","48:4d:7e":"Dell","4c:d9:8f":"Dell","50:9a:4c":"Dell",
    "54:9f:13":"Dell","58:8a:5a":"Dell","5c:f9:dd":"Dell","60:57:18":"Dell",
    "6c:2b:59":"Dell","70:10:6f":"Dell","74:86:e2":"Dell","78:45:c4":"Dell",
    "7c:b1:1c":"Dell","84:8f:69":"Dell","8c:04:ba":"Dell","90:b1:1c":"Dell",
    "98:90:96":"Dell","9c:eb:e8":"Dell","a0:36:9f":"Dell","a4:1f:72":"Dell",
    "a4:ba:db":"Dell","a8:9d:21":"Dell","b0:83:fe":"Dell","b4:96:91":"Dell",
    "b8:ca:3a":"Dell","bc:30:5b":"Dell","c8:1f:66":"Dell","d4:ae:52":"Dell",
    "d4:be:d9":"Dell","d8:9e:f3":"Dell","e0:db:55":"Dell","e4:43:4b":"Dell",
    "e8:b0:c3":"Dell","ec:f4:bb":"Dell","f0:1f:af":"Dell","f4:02:70":"Dell",
    "f8:db:88":"Dell","f8:bc:12":"Dell","fc:aa:14":"Dell",
    # ── HP / HPE ──────────────────────────────────────────────────────────────
    "00:01:e6":"HP","00:04:ea":"HP","00:08:02":"HP","00:0b:cd":"HP",
    "00:0e:7f":"HP","00:10:83":"HP","00:11:0a":"HP","00:12:79":"HP",
    "00:13:21":"HP","00:14:38":"HP","00:15:60":"HP","00:16:35":"HP",
    "00:17:08":"HP","00:17:a4":"HP","00:18:71":"HP","00:19:bb":"HP",
    "00:1a:4b":"HP","00:1b:78":"HP","00:1c:c4":"HP","00:1d:b3":"HP",
    "00:1e:0b":"HP","00:1f:28":"HP","00:1f:29":"HP","00:21:5a":"HP",
    "00:22:64":"HP","00:23:7d":"HP","00:24:81":"HP","00:25:b3":"HP",
    "00:26:55":"HP","00:30:6e":"HP","3c:4a:92":"HP","3c:d9:2b":"HP",
    "40:a8:f0":"HP","40:b0:34":"HP","58:20:b1":"HP","5c:b9:01":"HP",
    "6c:c2:17":"HP","6c:c2:6b":"HP","70:10:6f":"HP","70:5a:0f":"HP",
    "78:ac:c0":"HP","80:c1:6e":"HP","84:34:97":"HP","9c:8e:99":"HP",
    "a0:b3:cc":"HP","a4:5d:36":"HP","b8:af:67":"HP","bc:ea:fa":"HP",
    "c4:34:6b":"HP","c8:d3:ff":"HP","d4:85:64":"HP","d8:9d:67":"HP",
    "dc:4a:3e":"HP","e4:11:5b":"HP","e8:39:35":"HP","ec:b1:d7":"HP",
    "f0:92:1c":"HP","f4:ce:46":"HP","f8:b1:56":"HP","fc:15:b4":"HP",
    # ── Lenovo ────────────────────────────────────────────────────────────────
    "00:1a:6b":"Lenovo","18:56:80":"Lenovo","28:d2:44":"Lenovo","40:8d:5c":"Lenovo",
    "40:f0:2f":"Lenovo","48:0f:cf":"Lenovo","4c:79:6e":"Lenovo","50:7b:9d":"Lenovo",
    "54:ee:75":"Lenovo","58:8f:c7":"Lenovo","5c:f3:70":"Lenovo","60:67:20":"Lenovo",
    "70:72:3c":"Lenovo","70:f3:95":"Lenovo","74:df:bf":"Lenovo","78:92:9c":"Lenovo",
    "80:5e:c0":"Lenovo","84:2b:2b":"Lenovo","88:70:8c":"Lenovo","8c:d5:d9":"Lenovo",
    "8c:ec:4b":"Lenovo","90:2b:34":"Lenovo","98:41:5c":"Lenovo","9c:93:4e":"Lenovo",
    "a4:4c:c8":"Lenovo","ac:b3:13":"Lenovo","b8:ae:ed":"Lenovo","c4:65:16":"Lenovo",
    "c8:5b:76":"Lenovo","cc:f9:54":"Lenovo","d0:37:45":"Lenovo","d4:81:d7":"Lenovo",
    "d4:c9:ef":"Lenovo","e8:6a:64":"Lenovo","ec:f4:bb":"Lenovo","f8:16:54":"Lenovo",
    # ── Cisco ─────────────────────────────────────────────────────────────────
    "00:00:0c":"Cisco","00:00:7f":"Cisco","00:01:42":"Cisco","00:01:43":"Cisco",
    "00:01:63":"Cisco","00:01:64":"Cisco","00:01:96":"Cisco","00:01:97":"Cisco",
    "00:02:16":"Cisco","00:02:17":"Cisco","00:03:6b":"Cisco","00:03:e3":"Cisco",
    "00:04:6d":"Cisco","00:0a:8a":"Cisco","00:0b:be":"Cisco","00:0b:fd":"Cisco",
    "00:0c:ce":"Cisco","00:0d:28":"Cisco","00:0d:29":"Cisco","00:0e:38":"Cisco",
    "00:0e:83":"Cisco","00:0e:84":"Cisco","00:0f:23":"Cisco","00:0f:24":"Cisco",
    "00:0f:8f":"Cisco","00:0f:90":"Cisco","00:1a:2f":"Cisco","00:1a:30":"Cisco",
    "00:1b:2a":"Cisco","00:1b:2b":"Cisco","00:1c:57":"Cisco","00:1c:58":"Cisco",
    "00:1d:45":"Cisco","00:1d:46":"Cisco","00:1e:13":"Cisco","00:1e:14":"Cisco",
    "00:1f:6c":"Cisco","00:1f:6d":"Cisco","00:21:55":"Cisco","00:21:56":"Cisco",
    "00:22:55":"Cisco","00:22:56":"Cisco","00:23:33":"Cisco","00:23:34":"Cisco",
    "00:24:13":"Cisco","00:24:14":"Cisco","00:25:83":"Cisco","00:25:84":"Cisco",
    "00:26:0b":"Cisco","00:26:0c":"Cisco","00:30:f2":"Cisco","00:50:0f":"Cisco",
    "00:60:70":"Cisco","00:90:21":"Cisco","00:90:6d":"Cisco","00:90:86":"Cisco",
    "04:62:73":"Cisco","08:96:ad":"Cisco","10:bd:18":"Cisco","14:f1:08":"Cisco",
    "18:33:9d":"Cisco","1c:de:a7":"Cisco","20:37:06":"Cisco","24:e9:b3":"Cisco",
    "28:94:0f":"Cisco","2c:54:91":"Cisco","30:37:a6":"Cisco","34:a8:4e":"Cisco",
    "38:ed:18":"Cisco","3c:08:f6":"Cisco","40:f4:ec":"Cisco","44:d3:ca":"Cisco",
    "48:39:50":"Cisco","4c:4e:35":"Cisco","50:61:84":"Cisco","54:75:d0":"Cisco",
    "58:97:bd":"Cisco","5c:71:0d":"Cisco","60:73:5c":"Cisco","64:14:13":"Cisco",
    "68:86:a7":"Cisco","6c:20:56":"Cisco","70:69:5a":"Cisco","74:26:ac":"Cisco",
    "78:ba:f9":"Cisco","7c:69:f6":"Cisco","80:e0:1d":"Cisco","84:b8:02":"Cisco",
    "88:75:56":"Cisco","8c:60:4f":"Cisco","90:e2:ba":"Cisco","94:d4:69":"Cisco",
    "98:90:96":"Cisco","9c:57:ad":"Cisco","a0:55:4f":"Cisco","a4:4c:11":"Cisco",
    "a8:b4:56":"Cisco","ac:17:c8":"Cisco","b0:aa:77":"Cisco","b4:a4:e3":"Cisco",
    "b8:38:61":"Cisco","bc:16:f5":"Cisco","c0:62:6b":"Cisco","c4:72:95":"Cisco",
    "c8:9c:1d":"Cisco","cc:d8:c1":"Cisco","d0:72:dc":"Cisco","d4:8c:b5":"Cisco",
    "d8:24:bd":"Cisco","dc:7b:94":"Cisco","e0:5f:b9":"Cisco","e4:aa:5d":"Cisco",
    "e8:ba:70":"Cisco","ec:1d:8b":"Cisco","f0:25:72":"Cisco","f4:cf:e2":"Cisco",
    "f8:7b:20":"Cisco","fc:5b:39":"Cisco","fc:fb:fb":"Cisco",
    # ── Intel ─────────────────────────────────────────────────────────────────
    "00:02:b3":"Intel","00:03:47":"Intel","00:04:23":"Intel","00:07:e9":"Intel",
    "00:0e:35":"Intel","00:12:f0":"Intel","00:13:02":"Intel","00:13:20":"Intel",
    "00:15:00":"Intel","00:16:ea":"Intel","00:16:eb":"Intel","00:18:de":"Intel",
    "00:19:d1":"Intel","00:1b:21":"Intel","00:1c:bf":"Intel","00:1e:64":"Intel",
    "00:1e:65":"Intel","00:1e:67":"Intel","00:1f:3b":"Intel","00:1f:3c":"Intel",
    "00:21:6a":"Intel","00:22:fa":"Intel","00:23:14":"Intel","00:24:d7":"Intel",
    "00:27:10":"Intel","04:0e:3c":"Intel","08:11:96":"Intel","0c:8b:fd":"Intel",
    "10:02:b5":"Intel","10:f0:05":"Intel","18:67:b0":"Intel","1c:69:7a":"Intel",
    "24:77:03":"Intel","28:d2:44":"Intel","2c:41:38":"Intel","34:13:e8":"Intel",
    "34:de:1a":"Intel","38:2c:4a":"Intel","40:a5:ef":"Intel","44:85:00":"Intel",
    "48:51:b7":"Intel","4c:eb:42":"Intel","54:27:1e":"Intel","5c:51:4f":"Intel",
    "60:67:20":"Intel","60:f6:77":"Intel","64:00:6a":"Intel","68:05:ca":"Intel",
    "6c:29:95":"Intel","70:1a:04":"Intel","74:d4:35":"Intel","78:92:9c":"Intel",
    "7c:5c:f8":"Intel","80:19:34":"Intel","80:86:f2":"Intel","84:3a:4b":"Intel",
    "88:53:2e":"Intel","8c:8d:28":"Intel","8c:ec:4b":"Intel","90:e2:ba":"Intel",
    "94:65:9c":"Intel","98:4b:e1":"Intel","9c:eb:e8":"Intel","a0:88:b4":"Intel",
    "a4:c3:f0":"Intel","a8:6b:ad":"Intel","ac:72:89":"Intel","b4:96:91":"Intel",
    "b8:08:cf":"Intel","bc:ee:7b":"Intel","c0:3f:d5":"Intel","c4:d9:87":"Intel",
    "c8:d9:d2":"Intel","cc:3d:82":"Intel","d0:50:99":"Intel","d4:3d:7e":"Intel",
    "d8:fc:93":"Intel","dc:53:60":"Intel","e0:d5:5e":"Intel","e4:b3:18":"Intel",
    "e8:b4:70":"Intel","ec:b1:d7":"Intel","f0:4d:a2":"Intel","f4:8e:38":"Intel",
    # ── TP-Link ───────────────────────────────────────────────────────────────
    "00:23:cd":"TP-Link","08:57:00":"TP-Link","10:fe:ed":"TP-Link","14:cc:20":"TP-Link",
    "18:a6:f7":"TP-Link","18:d6:c7":"TP-Link","1c:61:b4":"TP-Link","20:dc:e6":"TP-Link",
    "24:69:68":"TP-Link","28:2c:b2":"TP-Link","2c:54:91":"TP-Link","30:b5:c2":"TP-Link",
    "34:60:f9":"TP-Link","38:94:ed":"TP-Link","3c:52:82":"TP-Link","40:16:9f":"TP-Link",
    "44:94:fc":"TP-Link","48:8f:5a":"TP-Link","4c:e1:73":"TP-Link","50:c7:bf":"TP-Link",
    "54:a7:03":"TP-Link","5c:89:9a":"TP-Link","60:a4:b7":"TP-Link","64:70:02":"TP-Link",
    "68:ff:7b":"TP-Link","6c:5a:b0":"TP-Link","70:4f:57":"TP-Link","74:da:38":"TP-Link",
    "78:44:fd":"TP-Link","7c:8b:ca":"TP-Link","80:8f:1d":"TP-Link","84:16:f9":"TP-Link",
    "88:d7:f6":"TP-Link","8c:21:0a":"TP-Link","90:f6:52":"TP-Link","94:d9:b3":"TP-Link",
    "98:da:c4":"TP-Link","9c:a6:15":"TP-Link","a0:f3:c1":"TP-Link","a4:2b:b0":"TP-Link",
    "a8:57:4e":"TP-Link","ac:84:c6":"TP-Link","b0:4e:26":"TP-Link","b4:b0:24":"TP-Link",
    "b8:a3:86":"TP-Link","bc:46:99":"TP-Link","c0:06:c3":"TP-Link","c4:e9:84":"TP-Link",
    "c8:3a:35":"TP-Link","cc:32:e5":"TP-Link","d4:6e:5c":"TP-Link","d8:07:b6":"TP-Link",
    "dc:fe:18":"TP-Link","e0:28:6d":"TP-Link","e4:c3:2a":"TP-Link","e8:48:b8":"TP-Link",
    "ec:08:6b":"TP-Link","f0:a7:31":"TP-Link","f4:f2:6d":"TP-Link","f8:1a:67":"TP-Link",
    "fc:d7:33":"TP-Link",
    # ── Ubiquiti ──────────────────────────────────────────────────────────────
    "00:15:6d":"Ubiquiti","00:27:22":"Ubiquiti","04:18:d6":"Ubiquiti","0a:27:22":"Ubiquiti",
    "18:e8:29":"Ubiquiti","24:a4:3c":"Ubiquiti","2c:27:d7":"Ubiquiti","34:1a:35":"Ubiquiti",
    "44:d9:e7":"Ubiquiti","4c:e9:e4":"Ubiquiti","60:22:32":"Ubiquiti","68:72:51":"Ubiquiti",
    "6e:27:d3":"Ubiquiti","70:a7:41":"Ubiquiti","74:83:c8":"Ubiquiti","78:8a:20":"Ubiquiti",
    "78:d2:94":"Ubiquiti","7c:dd:90":"Ubiquiti","80:2a:a8":"Ubiquiti","b4:fb:e4":"Ubiquiti",
    "b6:fb:e4":"Ubiquiti","d8:21:e8":"Ubiquiti","d8:b3:70":"Ubiquiti","dc:9f:db":"Ubiquiti",
    "e0:63:da":"Ubiquiti","e4:38:83":"Ubiquiti","e6:38:83":"Ubiquiti","e8:48:b8":"Ubiquiti",
    "f0:9f:c2":"Ubiquiti","f4:92:bf":"Ubiquiti","f4:e2:c6":"Ubiquiti","fc:ec:da":"Ubiquiti",
    "a4:4c:11":"Ubiquiti",
    # ── Netgear ───────────────────────────────────────────────────────────────
    "00:09:5b":"Netgear","00:0f:b5":"Netgear","00:14:6c":"Netgear","00:18:4d":"Netgear",
    "00:1b:2f":"Netgear","00:1e:2a":"Netgear","00:1f:33":"Netgear","00:22:3f":"Netgear",
    "00:24:b2":"Netgear","00:26:f2":"Netgear","10:0d:7f":"Netgear","20:0c:c8":"Netgear",
    "20:4e:7f":"Netgear","28:80:23":"Netgear","2c:30:33":"Netgear","30:46:9a":"Netgear",
    "3c:37:86":"Netgear","44:94:fc":"Netgear","4c:60:de":"Netgear","6c:b0:ce":"Netgear",
    "74:44:01":"Netgear","7c:b7:33":"Netgear","80:37:73":"Netgear","84:1b:5e":"Netgear",
    "9c:3d:cf":"Netgear","9c:d3:6d":"Netgear","a0:21:b7":"Netgear","a0:40:a0":"Netgear",
    "a4:2b:8c":"Netgear","b0:7f:b9":"Netgear","c0:3f:0e":"Netgear","c4:04:15":"Netgear",
    "c4:3d:c7":"Netgear","c8:d7:19":"Netgear","cc:40:d0":"Netgear","e0:46:9a":"Netgear",
    "e0:91:f5":"Netgear","e4:f4:c6":"Netgear","e8:fc:af":"Netgear","f8:1a:67":"Netgear",
    # ── D-Link ────────────────────────────────────────────────────────────────
    "00:05:5d":"D-Link","00:0d:88":"D-Link","00:0f:3d":"D-Link","00:11:95":"D-Link",
    "00:13:46":"D-Link","00:15:e9":"D-Link","00:17:9a":"D-Link","00:19:5b":"D-Link",
    "00:1b:11":"D-Link","00:1c:f0":"D-Link","00:1e:58":"D-Link","00:21:91":"D-Link",
    "00:22:b0":"D-Link","00:24:01":"D-Link","00:26:5a":"D-Link","1c:7e:e5":"D-Link",
    "28:10:7b":"D-Link","2c:b0:5d":"D-Link","2c:d0:5a":"D-Link","34:08:04":"D-Link",
    "34:31:c4":"D-Link","5c:d9:98":"D-Link","64:70:02":"D-Link","78:54:2e":"D-Link",
    "84:c9:b2":"D-Link","90:94:e4":"D-Link","9c:72:b9":"D-Link","a0:ab:1b":"D-Link",
    "b4:c7:99":"D-Link","bc:f6:85":"D-Link","c0:a0:bb":"D-Link","c8:be:19":"D-Link",
    "cc:b2:55":"D-Link","d8:eb:97":"D-Link","e4:6f:13":"D-Link","f0:7d:68":"D-Link",
    "f8:1a:67":"D-Link","fc:75:16":"D-Link",
    # ── Synology ─────────────────────────────────────────────────────────────
    "00:11:32":"Synology","2c:fd:a1":"Synology",  # 00:11:32 is also Synology
    "bc:ee:7b":"Synology",
    # ── QNAP ─────────────────────────────────────────────────────────────────
    "00:08:9b":"QNAP","00:08:9b":"QNAP","24:5e:be":"QNAP","68:63:7c":"QNAP",
    "d8:29:f8":"QNAP","00:90:a9":"QNAP",
    # ── Fortinet ─────────────────────────────────────────────────────────────
    "00:09:0f":"Fortinet","00:0b:86":"Fortinet","00:78:88":"Fortinet",
    "70:4c:a5":"Fortinet","90:6c:ac":"Fortinet",
    # ── Palo Alto Networks ────────────────────────────────────────────────────
    "00:1b:17":"Palo Alto","3c:4a:92":"HP",  # HP overrides Palo Alto for this prefix
    # ── Juniper ───────────────────────────────────────────────────────────────
    "00:12:1e":"Juniper","00:17:cb":"Juniper","00:19:e2":"Juniper","00:21:59":"Juniper",
    "00:23:9c":"Juniper","00:24:dc":"Juniper","00:26:88":"Juniper",
    "28:8a:1c":"Juniper","2c:6b:f5":"Juniper","3c:61:04":"Juniper",
    "40:b4:f0":"Juniper","4c:96:14":"Juniper","54:e0:32":"Juniper",
    "64:87:88":"Juniper","6c:b2:ae":"Juniper","84:18:88":"Juniper",
    "88:e0:f3":"Juniper","98:65:15":"Juniper","a4:50:46":"Juniper",
    "cc:e1:7f":"Juniper","f0:1c:2d":"Juniper","f4:a7:39":"Juniper",
    "fc:2f:40":"Juniper",
    # ── Aruba / HP Networking ─────────────────────────────────────────────────
    "00:0b:86":"Aruba","00:1a:1e":"Aruba","00:24:6c":"Aruba","04:bd:88":"Aruba",
    "08:26:97":"Aruba","0c:f8:93":"Aruba","18:64:72":"Aruba","1c:28:af":"Aruba",
    "20:4c:03":"Aruba","20:a6:cd":"Aruba","24:de:c6":"Aruba","2c:a8:35":"Aruba",
    "34:fc:b9":"Aruba","40:e3:d6":"Aruba","4c:6d:7f":"Aruba","58:8b:f3":"Aruba",
    "6c:f3:7f":"Aruba","70:88:6b":"Aruba","74:f8:db":"Aruba","84:d4:7e":"Aruba",
    "94:b4:0f":"Aruba","9c:1c:12":"Aruba","a8:bd:27":"Aruba","ac:a3:1e":"Aruba",
    "b0:5a:da":"Aruba","b4:5d:50":"Aruba","c4:01:7c":"Aruba","d8:c7:c8":"Aruba",
    "e8:26:89":"Aruba","ec:b3:18":"Aruba","f0:5c:19":"Aruba",
    # ── HP Printing ──────────────────────────────────────────────────────────
    "00:17:c8":"HP","00:1b:78":"HP","00:1f:29":"HP","00:21:5a":"HP",
    "00:24:81":"HP","18:a9:05":"HP","1c:c1:de":"HP","28:92:4a":"HP",
    "30:8d:99":"HP","38:ea:a7":"HP","3c:d9:2b":"HP","40:b8:9a":"HP",
    "48:0f:cf":"HP","70:5a:0f":"HP","78:ac:c0":"HP","94:57:a5":"HP",
    "a4:5d:36":"HP","b4:99:ba":"HP","d8:9d:67":"HP","e8:04:0b":"HP",
    # ── Canon ─────────────────────────────────────────────────────────────────
    "00:00:85":"Canon","00:1e:8f":"Canon","00:80:92":"Canon","3c:43:8e":"Canon",
    "4c:49:e3":"Canon","74:d0:2b":"Canon","90:ca:fa":"Canon","ac:41:76":"Canon",
    "b4:75:0e":"Canon","c4:ac:59":"Canon","d4:20:b0":"Canon","f4:81:39":"Canon",
    # ── Epson ─────────────────────────────────────────────────────────────────
    "00:00:48":"Epson","00:26:ab":"Epson","08:00:46":"Epson","3c:3a:ef":"Epson",
    "4c:f6:08":"Epson","60:55:f9":"Epson","64:eb:8c":"Epson","ac:18:26":"Epson",
    # ── Brother ───────────────────────────────────────────────────────────────
    "00:00:74":"Brother","00:1b:a9":"Brother","00:80:77":"Brother","00:c0:97":"Brother",
    "0c:98:38":"Brother","30:05:5c":"Brother","34:56:fe":"Brother","3c:56:a6":"Brother",
    "40:49:0f":"Brother","5c:96:9d":"Brother","70:77:81":"Brother","b8:2a:72":"Brother",
    "c8:47:0d":"Brother","d4:11:a3":"Brother","d8:9b:3b":"Brother","e0:06:e6":"Brother",
    # ── Kyocera ───────────────────────────────────────────────────────────────
    "00:60:67":"Kyocera","00:c0:ee":"Kyocera","08:00:46":"Kyocera","0c:7e:d2":"Kyocera",
    "a4:1f:72":"Kyocera",
    # ── Ricoh ────────────────────────────────────────────────────────────────
    "00:00:74":"Ricoh","00:00:78":"Ricoh","00:60:b0":"Ricoh","08:00:48":"Ricoh",
    "00:26:73":"Ricoh","2c:5b:e1":"Ricoh","ac:de:48":"Ricoh",
    # ── Xerox ─────────────────────────────────────────────────────────────────
    "00:00:aa":"Xerox","00:00:6b":"Xerox","00:00:f4":"Xerox","34:9c:cd":"Xerox",
    "38:1a:52":"Xerox","3c:6f:6c":"Xerox","44:1e:a1":"Xerox","60:f4:45":"Xerox",
    # ── Lexmark ───────────────────────────────────────────────────────────────
    "00:04:00":"Lexmark","00:04:00":"Lexmark","00:0d:87":"Lexmark","34:60:f9":"Lexmark",
    # ── Samsung ──────────────────────────────────────────────────────────────
    "00:00:f0":"Samsung","00:02:78":"Samsung","00:12:47":"Samsung","00:15:b9":"Samsung",
    "00:16:32":"Samsung","00:17:c9":"Samsung","00:1d:25":"Samsung","00:1e:7d":"Samsung",
    "00:21:19":"Samsung","00:23:99":"Samsung","00:24:54":"Samsung","00:26:37":"Samsung",
    "04:18:d6":"Samsung","08:08:c2":"Samsung","08:d4:2b":"Samsung","10:30:47":"Samsung",
    "10:d5:42":"Samsung","14:49:e0":"Samsung","18:3a:2d":"Samsung","1c:af:f7":"Samsung",
    "20:13:e0":"Samsung","24:4b:81":"Samsung","28:27:bf":"Samsung","2c:ae:2b":"Samsung",
    "30:19:66":"Samsung","34:14:5f":"Samsung","38:01:97":"Samsung","3c:8b:fe":"Samsung",
    "40:0e:85":"Samsung","44:6d:57":"Samsung","48:44:f7":"Samsung","4c:3c:16":"Samsung",
    "50:32:75":"Samsung","54:88:0e":"Samsung","58:ef:68":"Samsung","5c:0a:5b":"Samsung",
    "60:a1:0a":"Samsung","64:77:91":"Samsung","68:27:37":"Samsung","6c:2f:2c":"Samsung",
    "70:f9:27":"Samsung","78:1f:db":"Samsung","7c:0b:c6":"Samsung","80:65:6d":"Samsung",
    "84:25:db":"Samsung","88:32:9b":"Samsung","8c:71:f8":"Samsung","90:00:4e":"Samsung",
    "94:76:b7":"Samsung","98:52:b1":"Samsung","9c:02:98":"Samsung","a0:07:98":"Samsung",
    "a4:eb:d3":"Samsung","a8:06:00":"Samsung","ac:36:13":"Samsung","b0:72:bf":"Samsung",
    "b4:3a:28":"Samsung","b8:6c:e8":"Samsung","bc:14:ef":"Samsung","c0:bd:d1":"Samsung",
    "c4:42:02":"Samsung","c8:ba:94":"Samsung","cc:07:ab":"Samsung","d0:17:6a":"Samsung",
    "d4:88:90":"Samsung","d8:57:ef":"Samsung","dc:71:96":"Samsung","e0:62:90":"Samsung",
    "e4:e0:c5":"Samsung","e8:03:9a":"Samsung","ec:1f:72":"Samsung","f0:25:b7":"Samsung",
    "f4:7b:5e":"Samsung","f8:04:2e":"Samsung","fc:a1:3e":"Samsung",
    # ── Realtek ───────────────────────────────────────────────────────────────
    "00:e0:4c":"Realtek","52:54:00":"Realtek","54:ab:3a":"Realtek","e0:d5:5e":"Realtek",
    # ── APC / Schneider ───────────────────────────────────────────────────────
    "00:c0:b7":"APC","00:60:26":"APC","c8:cb:9e":"APC",
    # ── Supermicro ────────────────────────────────────────────────────────────
    "00:25:90":"Supermicro","00:30:48":"Supermicro","18:66:da":"Supermicro",
    # ── IBM ───────────────────────────────────────────────────────────────────
    "00:04:ac":"IBM","00:06:29":"IBM","00:09:6b":"IBM","00:0d:60":"IBM",
    "00:11:25":"IBM","00:14:5e":"IBM","00:17:ef":"IBM","00:21:5e":"IBM",
    "00:26:55":"IBM",
    # ── Google ────────────────────────────────────────────────────────────────
    "00:1a:11":"Google","3c:5a:b4":"Google","3c:61:04":"Google",
    "54:60:09":"Google","94:eb:cd":"Google","a4:77:33":"Google",
    "f4:f5:d8":"Google","f4:f5:e8":"Google",
    # ── Amazon ────────────────────────────────────────────────────────────────
    "00:bb:3a":"Amazon","18:74:2e":"Amazon","40:b4:cd":"Amazon","44:65:0d":"Amazon",
    "4c:ef:c0":"Amazon","50:dc:e7":"Amazon","68:37:e9":"Amazon","74:c2:46":"Amazon",
    "84:d6:d0":"Amazon","a0:02:dc":"Amazon","ac:63:be":"Amazon","b4:7c:9c":"Amazon",
    "cc:f7:35":"Amazon","d0:04:01":"Amazon","d0:f8:8c":"Amazon","e4:80:45":"Amazon",
    "f0:27:2d":"Amazon","f8:04:2e":"Amazon","fc:65:de":"Amazon",
    # ── Buffalo ───────────────────────────────────────────────────────────────
    "00:07:40":"Buffalo","00:08:9b":"Buffalo","00:0d:0b":"Buffalo",
    "00:16:01":"Buffalo","00:1d:73":"Buffalo","00:24:a5":"Buffalo",
    "10:6f:3f":"Buffalo","18:c0:4d":"Buffalo","1c:87:2c":"Buffalo",
    "28:3b:82":"Buffalo","2c:fd:a1":"Buffalo","30:85:a9":"Buffalo",
    "40:f2:01":"Buffalo","48:5b:39":"Buffalo","5c:57:c8":"Buffalo",
    "7c:dd:90":"Buffalo","80:35:c1":"Buffalo","90:f6:52":"Buffalo",
    "a8:92:0e":"Buffalo","c4:e9:84":"Buffalo","d8:50:e6":"Buffalo",
    # ── Linksys / Belkin ──────────────────────────────────────────────────────
    "00:06:25":"Linksys","00:0c:41":"Linksys","00:0f:66":"Linksys",
    "00:12:17":"Linksys","00:13:10":"Linksys","00:14:bf":"Linksys",
    "00:16:b6":"Linksys","00:18:39":"Linksys","00:18:f8":"Linksys",
    "00:1a:70":"Linksys","00:1c:10":"Linksys","00:1d:7e":"Linksys",
    "00:1e:e5":"Linksys","00:20:a6":"Linksys","00:21:29":"Linksys",
    "00:22:6b":"Linksys","00:25:9c":"Linksys","c0:c1:c0":"Linksys",
    # ── Mikrotik ─────────────────────────────────────────────────────────────
    "00:0c:42":"Mikrotik","18:fd:74":"Mikrotik","2c:c8:1b":"Mikrotik",
    "48:8f:5a":"Mikrotik","4c:5e:0c":"Mikrotik","6c:3b:6b":"Mikrotik",
    "74:4d:28":"Mikrotik","b8:69:f4":"Mikrotik","cc:2d:e0":"Mikrotik",
    "d4:ca:6d":"Mikrotik","dc:2c:6e":"Mikrotik","e4:8d:8c":"Mikrotik",
    # ── Hikvision ────────────────────────────────────────────────────────────
    "44:19:b6":"Hikvision","4c:bd:8f":"Hikvision","54:c4:15":"Hikvision",
    "8c:e7:48":"Hikvision","94:40:c9":"Hikvision","bc:ad:28":"Hikvision",
    "c4:2f:90":"Hikvision","c8:02:8f":"Hikvision",
    # ── Dahua ────────────────────────────────────────────────────────────────
    "3c:ef:8c":"Dahua","4c:11:bf":"Dahua","90:02:a9":"Dahua",
    "a4:14:37":"Dahua","e0:50:8b":"Dahua",
    # ── Axis (caméras IP) ─────────────────────────────────────────────────────
    "00:40:8c":"Axis","ac:cc:8e":"Axis","b8:a4:4f":"Axis",
}

def _oui_vendor(mac):
    """Retourne le fabricant depuis l'adresse MAC (préfixe OUI 24 bits).
    
    Priorité :
    1. Fichier oui.txt téléchargé (base IEEE complète ~60 000 entrées)
    2. Table embarquée _OUI (~930 entrées)
    """
    if not mac: return ""
    # Normaliser : d4-81-d7-xx → d4:81:d7
    prefix = mac[:8].lower().replace('-', ':').replace(' ', '')
    if len(prefix) < 8: return ""
    
    # 1. Essayer la table IEEE complète si chargée
    global _OUI_FULL
    if _OUI_FULL is None:
        _oui_load_full()
    if _OUI_FULL:
        v = _OUI_FULL.get(prefix, "")
        if v: return v
    
    # 2. Fallback table embarquée
    return _OUI.get(prefix, "")

# Cache de la table IEEE complète (None = pas encore tentée, {} = tentée mais vide)
_OUI_FULL = None


def _oui_path():
    """Emplacement persistant de oui.txt — à côté des autres données
    persistantes (base, uploads), PAS à côté du code.

    Bug réel trouvé en creusant l'imprécision du scan signalée en usage
    réel : ce chemin était jusqu'ici calculé depuis __file__, qui pointe en
    exécutable packagé (PyInstaller --onefile, voir parcinfo.spec) vers un
    dossier d'extraction TEMPORAIRE recréé à chaque lancement — jamais le
    dossier réel de l'exe. Résultat : même en suivant la documentation à la
    lettre (télécharger oui.txt et le placer à côté de ParcInfo.exe), AUCUN
    exécutable packagé ne parvenait jamais à charger la base complète —
    silencieusement replié sur les ~930 préfixes embarqués, pour toujours.
    """
    return os.path.join(_data_base, 'oui.txt')


def _oui_load_full():
    """Charge la base OUI IEEE depuis oui.txt si disponible."""
    global _OUI_FULL
    _OUI_FULL = {}
    oui_path = _oui_path()
    if not os.path.exists(oui_path):
        return
    try:
        count = 0
        with open(oui_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line = line.strip()
                # Format IEEE : "00-14-22   (hex)\t\tDell Inc."
                # ou format compact : "00:14:22\tDell Inc."
                if '(hex)' in line:
                    parts = line.split('(hex)')
                    if len(parts) == 2:
                        prefix_raw = parts[0].strip().replace('-', ':').lower()
                        vendor = parts[1].strip()
                        if len(prefix_raw) == 8 and vendor:
                            _OUI_FULL[prefix_raw] = vendor
                            count += 1
                elif re.match(r'^[0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2}\t', line):
                    parts = line.split('\t', 1)
                    if len(parts) == 2:
                        prefix_raw = parts[0].strip().replace('-', ':').lower()
                        vendor = parts[1].strip()
                        if vendor:
                            _OUI_FULL[prefix_raw] = vendor
                            count += 1
        if count > 0:
            app.logger.info(f"OUI database loaded: {count} entries from oui.txt")
    except Exception as e:
        app.logger.warning(f"Could not load oui.txt: {e}")
        _OUI_FULL = {}


_OUI_MAJ_AGE_MAX_JOURS = 30  # l'IEEE enregistre de nouveaux préfixes en continu


def _oui_telecharger(force=False):
    """Télécharge (ou met à jour) la base OUI IEEE complète (~60 000
    préfixes, ~5 Mo) vers le dossier de données persistant.

    `force=False` (démarrage, cron quotidien) : ne retélécharge pas si le
    fichier existe déjà et date de moins de _OUI_MAJ_AGE_MAX_JOURS — inutile
    de solliciter le serveur IEEE à chaque démarrage, le scan reste
    largement fonctionnel entre deux rafraîchissements. `force=True`
    (bouton "Mettre à jour" de la page Scan) télécharge toujours.

    Best-effort à dessein : toute erreur réseau laisse la base existante
    (embarquée ou précédemment téléchargée) intacte plutôt que de casser la
    détection en cours — jamais bloquant pour le scan.
    """
    import urllib.request
    oui_path = _oui_path()
    if not force and os.path.exists(oui_path):
        age_jours = (time.time() - os.path.getmtime(oui_path)) / 86400
        if age_jours < _OUI_MAJ_AGE_MAX_JOURS:
            return {'ok': True, 'skipped': True, 'raison': 'déjà à jour',
                    'count': len(_OUI_FULL) if _OUI_FULL else 0}

    tmp_path = oui_path + '.tmp'
    try:
        requete = urllib.request.Request(
            'https://standards-oui.ieee.org/oui/oui.txt',
            headers={'User-Agent': 'ParcInfo-OUI-Updater'})
        with urllib.request.urlopen(requete, timeout=30) as reponse:
            contenu = reponse.read()
        with open(tmp_path, 'wb') as f:
            f.write(contenu)
        os.replace(tmp_path, oui_path)  # remplacement atomique : jamais de fichier à moitié écrit
        _oui_load_full()
        try:
            cfg_set('oui_derniere_maj', _utcnow().isoformat())
        except Exception:
            pass
        logger.info('Base OUI mise à jour : %d préfixes', len(_OUI_FULL or {}))
        return {'ok': True, 'count': len(_OUI_FULL or {})}
    except Exception as e:
        logger.warning('Échec du téléchargement de la base OUI (%s) — base existante conservée', e)
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        return {'ok': False, 'erreur': str(e)}


# ── PING & DÉCOUVERTE ────────────────────────────────────────────────────────

def _run_hidden(cmd, **kwargs):
    """Lance un subprocess sans fenêtre console visible sur Windows."""
    if IS_WINDOWS:
        kwargs.setdefault('creationflags', subprocess.CREATE_NO_WINDOW)
    return subprocess.run(cmd, **kwargs)


def _tcp_probe_rapide(ip_str, ports, timeout=0.3):
    """Sonde plusieurs ports TCP EN PARALLÈLE, retourne dès que l'un répond
    (ou après épuisement de tous les autres).

    Remplace un ancien fallback séquentiel (un port après l'autre, chacun
    jusqu'à son propre timeout) dont le pire cas — hôte injoignable, aucun
    port ouvert, le cas le plus courant pour la grande majorité des
    adresses inutilisées d'un scan réseau — coûtait plusieurs secondes PAR
    HÔTE rien que pour ce fallback (jusqu'à 12 × 0.4s = 4.8s dans `_ping()`,
    8 × 1.0s = 8s dans `_ping_once()`), avant même le reste du sondage.
    Ici, le pire cas est borné au timeout d'UN SEUL port quel que soit le
    nombre de ports essayés — c'était le principal goulot d'étranglement du
    ping et du scan réseau, signalé comme trop lent en usage réel.
    """
    def _essai(port):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(timeout)
            ok = s.connect_ex((ip_str, port)) == 0
            s.close()
            return ok
        except Exception:
            return False
    # Pas de `with` ici : Executor.__exit__ appelle shutdown(wait=True), qui
    # bloquerait jusqu'à ce que TOUS les threads terminent — y compris ceux
    # qu'on n'attend plus une fois qu'un port a déjà répondu, un port ouvert
    # répond typiquement en quelques millisecondes pendant qu'un autre,
    # filtré, met tout le timeout à échouer. shutdown(wait=False) laisse les
    # threads restants se terminer seuls en arrière-plan, sans bloquer le
    # retour de la fonction.
    ex = concurrent.futures.ThreadPoolExecutor(max_workers=len(ports))
    futures = [ex.submit(_essai, p) for p in ports]
    trouve = False
    try:
        for f in concurrent.futures.as_completed(futures, timeout=timeout + 1):
            if f.result():
                trouve = True
                break
    except concurrent.futures.TimeoutError:
        pass
    finally:
        ex.shutdown(wait=False)
    return trouve

def _ping(ip_str):
    """Teste si un hôte est joignable.

    Stratégie en 2 étapes :
    1. Commande ping système (ICMP fiable, évite les faux positifs du raw socket)
    2. Fallback TCP sur ports courants si ping non disponible (en parallèle,
       voir _tcp_probe_rapide — c'est ce qui rend le scan réseau rapide
       même sur les nombreuses adresses inutilisées d'une plage /24)

    Note : le raw ICMP socket N'EST PAS utilisé car en scan parallèle,
    un socket peut intercepter la réponse ICMP destinée à un autre thread,
    générant des faux positifs.
    """
    # 1. Commande ping système — timeout subprocess resserré à 1.5s (la
    # commande elle-même est déjà bornée en interne par -w 500/-W 1) : les
    # 3s précédentes n'étaient qu'une marge de sécurité rarement nécessaire,
    # payée en pire cas sur chaque hôte injoignable.
    try:
        if IS_WINDOWS:
            cmd = ['ping', '-n', '1', '-w', '500', ip_str]
        else:
            cmd = ['ping', '-c', '1', '-W', '1', ip_str]
        result = _run_hidden(cmd, capture_output=True, timeout=1.5)
        if result.returncode == 0:
            return True
    except FileNotFoundError:
        pass  # ping non disponible, on passe au fallback TCP
    except Exception:
        pass
    # 2. Fallback TCP — un hôte vivant a forcément au moins un de ces ports
    # ouverts (ou fermé-mais-répondant, ce qui suffit à prouver sa présence)
    return _tcp_probe_rapide(ip_str, [80, 443, 22, 445, 135, 139, 3389, 8080, 53, 8443, 5000, 9100])

def _hostname(ip_str):
    """Résolution DNS inverse."""
    try:
        return socket.gethostbyaddr(ip_str)[0]
    except Exception:
        return ""

def _netbios_name(ip_str):
    """Requête NetBIOS Name Service (UDP 137) — retourne le nom NetBIOS de la machine."""
    try:
        # Paquet NBSTAT query conforme RFC 1002
        # Transaction ID aléatoire, NBSTAT query pour '*'
        import os as _os
        txid   = _os.urandom(2)
        # Nom encodé NetBIOS pour '*' (wildcard NBSTAT)
        # '*' = 0x2A, encodé en nibbles : 0x2A → 'CK', reste = 'AA'*15
        nb_name = b'CKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'  # 32 octets (16 * 2 nibbles)
        query = (
            txid +
            b'\x00\x00' +          # Flags: standard query
            b'\x00\x01' +          # QDCOUNT: 1 question
            b'\x00\x00' +          # ANCOUNT: 0
            b'\x00\x00' +          # NSCOUNT: 0
            b'\x00\x00' +          # ARCOUNT: 0
            b'\x20' +               # Longueur nom: 32
            nb_name +
            b'\x00' +               # Terminateur
            b'\x00\x21' +          # QTYPE: NBSTAT (0x21)
            b'\x00\x01'            # QCLASS: IN
        )
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(1.0)
        s.sendto(query, (ip_str, 137))
        data, _ = s.recvfrom(1024)
        s.close()
        # Parser la réponse : offset 56 = nombre de noms
        if len(data) > 57:
            nb_names = data[56]
            offset = 57
            for _ in range(nb_names):
                if offset + 18 > len(data):
                    break
                raw_name = data[offset:offset+15]
                flags    = data[offset+15:offset+18]
                name     = raw_name.decode('ascii', 'ignore').strip()
                # Flag byte: bit 7 = group name, on veut les noms individuels (type 0x00 = workstation)
                name_type = data[offset+15] if offset+15 < len(data) else 0xFF
                if name_type in (0x00, 0x20) and name and name != '\x00' * 15:
                    return name
                offset += 18
    except Exception:
        pass
    return ""

def _ttl_os_guess(ip_str):
    """Deviner l'OS par le TTL de la réponse ping.
    
    Utilise uniquement la commande ping système pour lire le TTL.
    Le raw ICMP socket n'est pas utilisé (risque de faux positifs en parallèle).
    """
    try:
        if IS_WINDOWS:
            r = _run_hidden(['ping', '-n', '1', ip_str],
                            capture_output=True, text=True, timeout=3)
            m = re.search(r'TTL[=\s]+(\d+)', r.stdout, re.IGNORECASE)
        else:
            r = _run_hidden(['ping', '-c', '1', ip_str],
                            capture_output=True, text=True, timeout=3)
            m = re.search(r'ttl[=\s]*(\d+)', r.stdout, re.IGNORECASE)
        if m:
            ttl = int(m.group(1))
            if ttl <= 64:  return 'Linux/Unix'
            if ttl <= 128: return 'Windows'
            return 'Network'
    except Exception:
        pass
    return ""

def _mac_from_arp(ip_str):
    """Récupère l'adresse MAC depuis la table ARP.
    
    Stratégie : lit la TABLE COMPLÈTE puis cherche la ligne correspondant à ip_str.
    Ne jamais passer l'IP en argument à arp — sur Windows, arp -a <ip> ne retourne
    rien si l'entrée n'est pas encore dans le cache au moment exact de l'appel.
    """
    mac_regex = re.compile(r'([0-9a-fA-F]{2}[:\-]){5}[0-9a-fA-F]{2}')
    
    # 1. /proc/net/arp — Linux, le plus fiable et rapide
    try:
        with open('/proc/net/arp', 'r') as f:
            for line in f.readlines()[1:]:
                parts = line.split()
                if len(parts) >= 4 and parts[0] == ip_str:
                    mac = parts[3]
                    if mac and mac not in ('00:00:00:00:00:00', '<incomplete>'):
                        return mac.lower()
    except Exception:
        pass
    
    # 2. arp -a (table complète) — Windows et Linux
    # IMPORTANT: utiliser une regex pour matcher l'IP exacte (éviter 192.168.1.1 dans 192.168.1.10)
    ip_pattern = re.compile(r'(?<![0-9])' + re.escape(ip_str) + r'(?![0-9])')
    bad_macs = {'ff:ff:ff:ff:ff:ff', '00:00:00:00:00:00'}
    try:
        r = _run_hidden(['arp', '-a'], capture_output=True, text=True, timeout=5)
        for line in r.stdout.splitlines():
            if ip_pattern.search(line):
                m = mac_regex.search(line)
                if m:
                    mac = m.group(0).replace('-', ':').lower()
                    if mac not in bad_macs and not mac.startswith('01:') and not mac.startswith('ff:'):
                        return mac
    except Exception:
        pass
    
    # 3. ip neigh show (Linux moderne)
    try:
        r = _run_hidden(['ip', 'neigh', 'show'], capture_output=True, text=True, timeout=5)
        for line in r.stdout.splitlines():
            if ip_pattern.search(line) and 'FAILED' not in line and 'INCOMPLETE' not in line:
                m = mac_regex.search(line)
                if m:
                    return m.group(0).lower()
    except Exception:
        pass
    
    return ""

def _scan_ports(ip_str):
    """Scan TCP des ports configurés."""
    raw = cfg_get('scan_ports',
                  '21,22,23,25,53,80,110,135,139,143,389,443,445,631,1433,3306,3389,5900,8080,8443,9100')
    try:
        PORTS = [int(p.strip()) for p in raw.split(',') if p.strip().isdigit()]
    except Exception:
        PORTS = []
    if not PORTS:
        PORTS = [21,22,23,25,53,80,443,445,3389,9100]
    try:
        timeout = float(cfg_get('ping_timeout', '0.4'))
    except Exception:
        timeout = 0.4
    def check(p):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(timeout)
            ok = s.connect_ex((ip_str, p)) == 0
            s.close()
            return p if ok else None
        except Exception:
            return None
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(PORTS), 50)) as ex:
        return sorted([p for p in ex.map(check, PORTS) if p])

# ── BANNIÈRES DE SERVICE ────────────────────────────────────────────────────
# Le scan ne savait dire que "port ouvert ou non" — un port 22 répond souvent
# "SSH-2.0-OpenSSH_8.9 Ubuntu" ou "...Cisco", un port web a un en-tête Server
# et un titre de page. Affine la détection de type/fabricant sans dépendance
# supplémentaire. Uniquement sur les ports déjà trouvés ouverts (pas de
# tentative sur un port fermé).
_PORTS_BANNIERE_TEXTE = {21, 22, 23, 25, 110, 143}   # saluent en premier, sans requête
_PORTS_BANNIERE_HTTP  = {80, 8000, 8008, 8080, 8888}
_PORTS_BANNIERE_HTTPS = {443, 8443}
# Ports de service serveur — utilisés pour ne retenir "server"/"srv" (mot-clé
# hostname/bannière trop fréquent sur du matériel non-serveur : "Print
# Server", "Web Server Login"...) que s'il est corroboré par un vrai service
# de serveur en écoute, pas la simple présence du mot dans un titre de page.
_PORTS_SERVEUR = {3306, 1433, 5432, 25, 587, 143, 993, 1521, 27017}


def _grab_banniere_texte(ip_str, port, timeout=1.2):
    """Bannière de salutation (SSH/FTP/SMTP/POP3/IMAP...), envoyée par le
    serveur dès la connexion, sans qu'on ait besoin d'écrire quoi que ce soit."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(timeout)
        s.connect((ip_str, port))
        data = s.recv(256)
        s.close()
        texte = re.sub(r'[\r\n\x00-\x1f]+', ' ', data.decode('utf-8', errors='replace')).strip()
        return texte[:200]
    except Exception:
        return ''


def _grab_banniere_http(ip_str, port, https=False, timeout=2.0):
    """En-tête Server + titre de page — beaucoup d'imprimantes/NAS/routeurs
    s'identifient dans leur page de connexion. HTTPS avec vérification
    désactivée : ces panneaux d'administration LAN utilisent presque
    toujours un certificat auto-signé, on ne fait ici que lire un en-tête,
    pas transmettre de secret."""
    import urllib.request
    try:
        ctx = None
        if https:
            import ssl
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
        url = f"{'https' if https else 'http'}://{ip_str}:{port}/"
        requete = urllib.request.Request(url, headers={'User-Agent': 'ParcInfo-Scan'})
        with urllib.request.urlopen(requete, timeout=timeout, context=ctx) as reponse:
            serveur = reponse.headers.get('Server', '')
            corps = reponse.read(4096).decode('utf-8', errors='replace')
        m = re.search(r'<title[^>]*>(.*?)</title>', corps, re.IGNORECASE | re.DOTALL)
        titre = re.sub(r'\s+', ' ', m.group(1)).strip() if m else ''
        return ' — '.join(x for x in (serveur, titre) if x)[:200]
    except Exception:
        return ''


def _grab_banniere(ip_str, port):
    if port in _PORTS_BANNIERE_HTTPS:
        return _grab_banniere_http(ip_str, port, https=True)
    if port in _PORTS_BANNIERE_HTTP:
        return _grab_banniere_http(ip_str, port, https=False)
    if port in _PORTS_BANNIERE_TEXTE:
        return _grab_banniere_texte(ip_str, port)
    return ''


def _grab_bannieres(ip_str, ports_ouverts):
    """Bannière de service pour chaque port ouvert pertinent. Retourne
    {port: texte}, sans les ports muets."""
    pertinents = [p for p in ports_ouverts
                  if p in _PORTS_BANNIERE_TEXTE or p in _PORTS_BANNIERE_HTTP or p in _PORTS_BANNIERE_HTTPS]
    if not pertinents:
        return {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(pertinents)) as ex:
        resultats = dict(zip(pertinents, ex.map(lambda p: _grab_banniere(ip_str, p), pertinents)))
    return {p: b for p, b in resultats.items() if b}


def _deviner_type(hostname, ports, os_guess="", vendor="", extra_signal="",
                   upnp_device_type="", mdns_service="", snmp_actif=False):
    """Détermine le type d'équipement depuis le hostname, les ports ouverts, l'OS et le fabricant.

    Plusieurs méthodes de détection (UPnP, mDNS, SNMP, fabricant MAC,
    hostname, ports ouverts, TTL) alimentent cette fonction avec des
    niveaux de confiance très différents — l'ordre des vérifications
    ci-dessous EST la hiérarchie de priorité, du plus fiable au moins
    fiable, pensée pour qu'un signal fort ne soit jamais écrasé par un
    signal plus faible arrivé après lui dans le code :

      1. Signaux structurés (deviceType UPnP, service mDNS) — un protocole
         qui se déclare lui-même ne laisse en général aucun doute.
      2. Fabricant MAC (préfixe OUI officiel) pour les familles où il est
         quasi-certain (Apple, puces IoT Espressif) ou les mots-clés de
         hostname/texte identifiant explicites (imprimante, NAS, routeur,
         switch, borne WiFi, objet connecté).
      3. Fabricant réseau reconnu sans indice de sous-type (mieux vaut une
         catégorie générique correcte qu'une supposition erronée).
      4. Le seul fait qu'un agent SNMP ait répondu (community "public") :
         un signal fort à lui seul — un PC/Mac de bureau n'expose
         pratiquement jamais SNMP — vérifié AVANT les heuristiques de port
         qui suivent, pour ne jamais se faire écraser par elles.
      5. Ports/protocoles ambigus (RDP est fiable seul, SMB ne l'est pas :
         Samba et le partage de fichiers macOS y répondent aussi).
      6. TTL (os_guess) en tout dernier repli : le signal le plus faible,
         seulement 3 paliers, incapable à lui seul de distinguer macOS de
         Linux ou un équipement réseau embarqué d'un serveur.

    `extra_signal` regroupe tout texte identifiant supplémentaire (bannières
    de service, sysDescr SNMP, nom UPnP/mDNS) : traité comme le hostname/
    fabricant pour la recherche de mots-clés.
    """
    # Signaux structurés d'abord : un deviceType UPnP ou un type de service
    # mDNS ne laisse en général aucun doute, contrairement à une simple
    # sous-chaîne de hostname.
    dt = (upnp_device_type or '').lower()
    if 'internetgatewaydevice' in dt:
        return 'Routeur/Pare-feu'
    if 'printer' in dt:
        return 'Imprimante'
    if 'mediaserver' in dt or 'nas' in dt:
        return 'NAS'
    svc = (mdns_service or '').lower()
    if svc in ('ipp', 'printer', 'pdl-datastream'):
        return 'Imprimante'
    if svc == 'smb':
        return 'Serveur'
    if svc in ('hap', 'shelly', 'esphomelib', 'matter', 'matterc'):
        return 'Objet connecté'

    v = (vendor or '').lower()
    h = (hostname + " " + vendor + " " + extra_signal).lower()
    # Imprimante
    if any(x in h for x in ['printer','print','canon','epson','brother','ricoh',
                              'xerox','kyocera','konica','lexmark','hp printer','jetdirect']):
        return 'Imprimante'
    if 9100 in ports or 631 in ports:
        return 'Imprimante'

    # NAS : gamme dédiée, fabricant identifiable au hostname/vendor — avant
    # le repli générique "Serveur" plus bas.
    if any(x in h for x in ['synology','qnap','readynas','truenas','freenas',
                              'terramaster',' nas']):
        return 'NAS'

    # Objets connectés (ESP32/ESP8266, prises/capteurs/relais) : signalé en
    # usage réel — un ESP32 ou une prise connectée ressortaient comme "PC".
    # Une liste de mots-clés hostname ne peut pas suivre la myriade de
    # marques commerciales (Tuya en particulier, revendu en marque blanche
    # sous des dizaines de noms) — le fabricant MAC officiel de la PUCE
    # radio (Espressif, très largement majoritaire sur ce segment — Sonoff,
    # Shelly et Tuya l'utilisent quasiment tous) est un signal bien plus
    # robuste,
    # complété par les quelques marques qui déposent leur propre OUI et par
    # les noms de firmware/marques les plus posés en hostname.
    if any(x in v for x in ['espressif', 'tuya smart', 'itead', 'allterco robotics',
                              'shenzhen sonoff', 'particle industries']):
        return 'Objet connecté'
    if any(x in h for x in ['tasmota', 'esphome', 'shelly', 'sonoff', 'tuya',
                              'esp32', 'esp8266',
                              'smart-plug', 'smart plug', 'smartplug',
                              'prise-connectee', 'prise connectee']):
        return 'Objet connecté'

    # Mac : le TTL seul ne distingue pas macOS de Linux (les deux à 64 par
    # défaut) — signalé en usage réel, un MacBook ressortait comme "machine
    # Linux". _scan_host affine déjà os_guess en 'macOS' via le fabricant
    # MAC officiel Apple ; le hostname mDNS/NetBIOS typique
    # ("Prenom-MacBook-Pro.local") est un second signal indépendant, tout
    # aussi fiable et disponible même sans lecture ARP.
    if os_guess == 'macOS' or 'apple' in v or any(
            x in h for x in ['macbook','imac','mac-mini','macmini','mac-pro','macstudio']):
        if not any(x in h for x in ['iphone','ipad','apple tv','watch',
                                      'airport','time capsule','homepod']):
            return 'MacBook'

    # Raspberry Pi / cartes mono-carte Linux : fabricant MAC identifiable,
    # avant les heuristiques de port plus bas — un Pi fait tourner Samba
    # (445) ou expose SSH (22) très couramment (Pi-hole, Home Assistant,
    # NAS DIY...), ce qui le ferait sinon retomber sur "PC (Windows)" via
    # la seule présence du port SMB.
    if any(x in v for x in ['raspberry pi']):
        return 'PC/Serveur (Linux)'

    # Équipements réseau — mots-clés explicites d'abord (sans ambiguïté sur
    # le sous-type), y compris les noms d'OS/firmware que le sysDescr SNMP
    # renvoie couramment en clair pour la famille pare-feu/routeur (FortiOS,
    # SonicOS, DD-WRT, OpenWrt, OPNsense — ceux-là ne tournent quasiment
    # jamais sur un switch), puis fabricants réseau reconnus en repli. Les
    # OS partagés entre routeurs ET switches chez un même fabricant (JUNOS,
    # IOS-XE, Comware, RouterOS, VRP...) restent volontairement hors de
    # cette liste : mieux vaut la catégorie générique "Switch/AP" plus bas
    # qu'un sous-type deviné à tort.
    if any(x in h for x in ['router',' gw ','gateway','firewall','pfsense',
                              'opnsense','fortigate','fortios','palo alto',
                              'checkpoint','sonicwall','sonicos','watchguard',
                              'edgerouter','edgemax','dd-wrt','openwrt']):
        return 'Routeur/Pare-feu'
    if any(x in h for x in ['switch',' sw-','usw-','poe switch']):
        return 'Switch'
    if any(x in h for x in ['cisco','juniper','extreme','3com','h3c','brocade',
                              'procurve']) and not any(x in h for x in ['server','srv']):
        return 'Switch'
    if any(x in h for x in ['ap-','borne','access point','wifi ap',' uap',
                              'accesspoint']):
        return 'Borne Wi-Fi'
    # Fabricant réseau reconnu (via le préfixe MAC officiel), mais sans
    # indice sur le sous-type précis dans le hostname : mieux vaut une
    # catégorie réseau générique correcte qu'une supposition erronée — une
    # borne Ubiquiti n'est pas un routeur juste parce que la marque en fait
    # aussi (c'était le cas avant : tout Ubiquiti/UniFi devenait
    # "Routeur/Pare-feu", même une borne WiFi ou un switch de la même gamme).
    if any(x in v for x in ['ubiquiti','mikrotik','tp-link','netgear','d-link',
                              'dlink','zyxel','ruckus','aruba','meraki',
                              'netonix','ruijie','huawei technologies',
                              'hewlett packard enterprise','draytek','tenda',
                              'grandstream']):
        return 'Switch/AP'
    # Même repli générique, cette fois via un nom d'OS/firmware réseau vu en
    # clair (typiquement le sysDescr SNMP) plutôt que le fabricant MAC —
    # ceux-ci tournent chez un même constructeur aussi bien sur des
    # routeurs que des switches, volontairement pas classés plus haut.
    if any(x in h for x in ['junos','ios-xe','comware','routeros','arubaos',
                              'aos-cx',' vrp ','huawei vrp']):
        return 'Switch/AP'

    # Serveur — mots-clés non ambigus (Exchange/vCenter/ESXi/contrôleur de
    # domaine) suffisent seuls ; "server"/"srv" nu est trop fréquent dans un
    # simple titre de page d'admin embarquée ("Print Server", "Web Server
    # Login" sur des switches/imprimantes bon marché non reconnus plus haut)
    # pour être pris au mot sans un port de service serveur à l'appui.
    if any(x in h for x in ['exchange','vcenter','esxi',' dc']):
        return 'Serveur'
    if any(x in h for x in ['server','srv']) and any(p in ports for p in _PORTS_SERVEUR):
        return 'Serveur'

    # PC Windows — RDP (3389) n'a pas d'équivalent légitime hors Windows,
    # signal fort à lui seul, prioritaire même sur SNMP ci-dessous (un
    # équipement réseau qui répondrait aussi en RDP serait de toute façon
    # une exception si rare qu'elle ne mérite pas de règle spéciale).
    if 3389 in ports:
        return 'PC (Windows)'
    # Un agent SNMP qui répond (community "public") est déjà un signal fort
    # à lui seul : un PC/Mac de bureau n'expose pratiquement jamais SNMP,
    # contrairement aux ports SMB/22 ci-dessous qui restent ambigus. Vérifié
    # AVANT ces heuristiques de port pour ne jamais se faire écraser par
    # elles — un switch/onduleur/imprimante SNMP dont aucun mot-clé
    # hostname/sysDescr n'a matché plus haut reste malgré tout bien plus
    # probablement du matériel réseau qu'un PC.
    if snmp_actif:
        return 'Switch/AP'
    # SMB (135/445) n'est PAS un signal Windows fiable à lui seul : Samba
    # (Linux) et le partage de fichiers natif de macOS répondent aussi sur
    # ces ports — signalé en usage réel, un Mac avec le partage de fichiers
    # activé ressortait comme "PC (Windows)". Le TTL garde la priorité s'il
    # pointe ailleurs (Apple déjà traité plus haut) ; Windows n'est retenu
    # via ces ports que faute de meilleur indice.
    if (135 in ports or 445 in ports) and os_guess not in ('Linux/Unix', 'macOS'):
        return 'PC (Windows)'
    # PC Linux
    if 22 in ports and 80 not in ports and 443 not in ports:
        if os_guess == 'Linux/Unix' or not os_guess:
            return 'PC/Serveur (Linux)'
    # OS fingerprint
    if os_guess == 'Network':
        return 'Switch/AP'
    if os_guess == 'macOS':
        return 'MacBook'
    if os_guess == 'Linux/Unix':
        return 'PC/Serveur (Linux)'
    if os_guess == 'Windows':
        return 'PC (Windows)'
    return 'PC'

def _enrich_from_wmi(ip_str, hostname_hint=""):
    """
    Tente d'enrichir les infos du scan via WMI (Windows Management Instrumentation).
    Récupère marque, modèle, RAM, CPU, etc. directement depuis la machine.

    Retourne un dict avec les infos enrichies, ou {} si WMI non disponible ou échec.

    Note: Nécessite :
    - WMI disponible (Windows uniquement)
    - Accès RPC à la machine (port 135)
    - Credentials valides (optionnel - sinon utilise credentials de session)
    """
    enriched = {}

    if _sys.platform != 'win32':
        # WMI n'est disponible que sur Windows
        return enriched

    try:
        import wmi
    except ImportError:
        # WMI non installé
        return enriched

    try:
        # Essayer une connexion WMI distante
        # Syntaxe : wmi.WMI("//IP/root/cimv2")
        c = wmi.WMI(f"//{ip_str}/root/cimv2")

        # Marque et modèle
        try:
            system = c.Win32_ComputerSystem()[0]
            enriched['brand'] = system.Manufacturer.strip() if system.Manufacturer else ''
            enriched['model'] = system.Model.strip() if system.Model else ''
        except Exception:
            pass

        # RAM
        try:
            mem = c.Win32_PhysicalMemory()
            total_ram_bytes = sum(int(m.Capacity) for m in mem)
            enriched['ram_gb'] = round(total_ram_bytes / (1024 ** 3), 1)
        except Exception:
            pass

        # CPU
        try:
            cpu = c.Win32_Processor()[0]
            enriched['cpu'] = cpu.Name.strip() if cpu.Name else ''
            enriched['cpu_cores'] = int(cpu.NumberOfCores) if hasattr(cpu, 'NumberOfCores') else None
        except Exception:
            pass

        # Disque
        try:
            disk = c.Win32_LogicalDisk(Name='C:')[0]
            enriched['disk_total_gb'] = round(int(disk.Size) / (1024 ** 3), 1)
        except Exception:
            pass

        # Numéro de série
        try:
            bios = c.Win32_SystemEnclosure()[0]
            enriched['serial_number'] = bios.SerialNumber.strip() if bios.SerialNumber else ''
        except Exception:
            pass

        app.logger.debug(f"WMI enrichment for {ip_str}: {len(enriched)} fields")

    except Exception as e:
        app.logger.debug(f"WMI enrichment failed for {ip_str}: {e}")
        pass

    return enriched

# ── DÉCOUVERTE UPnP (ssdp:all) ───────────────────────────────────────────────
# collector_core._upnp_decouvrir_passerelle (côté collecteur) ne cherche que
# la box Internet (ST=InternetGatewayDevice) depuis LE poste qui collecte.
# Ici, côté scan réseau, on cherche TOUT appareil UPnP du segment (NAS, TV
# connectées, serveurs média...) — une seule requête multicast pour tout le
# scan plutôt qu'une par hôte, les réponses portent chacune l'IP de leur
# émetteur.
def _ssdp_decouvrir_tout(timeout=3):
    """Retourne {ip: url_description_xml} pour chaque appareil UPnP qui répond."""
    message = (
        'M-SEARCH * HTTP/1.1\r\n'
        'HOST: 239.255.255.250:1900\r\n'
        'MAN: "ssdp:discover"\r\n'
        'MX: 2\r\n'
        'ST: ssdp:all\r\n'
        '\r\n'
    ).encode('utf-8')
    trouves = {}
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.settimeout(timeout)
            s.sendto(message, ('239.255.255.250', 1900))
            fin = time.time() + timeout
            while time.time() < fin:
                try:
                    data, (ip_source, _port) = s.recvfrom(4096)
                except socket.timeout:
                    break
                except Exception:
                    continue
                if ip_source in trouves:
                    continue
                texte = data.decode('utf-8', errors='ignore')
                for ligne in texte.split('\r\n'):
                    if ligne.upper().startswith('LOCATION:'):
                        trouves[ip_source] = ligne.split(':', 1)[1].strip()
                        break
    except Exception:
        pass
    return trouves


def _upnp_description_appareil(location_url):
    """Fabricant/modèle/nom/type depuis la description XML d'un appareil UPnP."""
    import urllib.request
    import xml.etree.ElementTree as ET
    try:
        requete = urllib.request.Request(location_url, headers={'User-Agent': 'ParcInfo-Scan'})
        with urllib.request.urlopen(requete, timeout=3) as reponse:
            xml_brut = reponse.read()
        racine = ET.fromstring(xml_brut)
    except Exception:
        return {}
    ns = {'u': 'urn:schemas-upnp-org:device-1-0'}

    def texte(chemin):
        el = racine.find('.//u:' + chemin, ns)
        return el.text.strip() if el is not None and el.text else ''

    resultat = {
        'manufacturer': texte('manufacturer'),
        'model_name': texte('modelName'),
        'friendly_name': texte('friendlyName'),
        'device_type': texte('deviceType'),
    }
    return {k: v for k, v in resultat.items() if v}


def _decouverte_upnp_reseau(timeout=3):
    """SSDP ssdp:all + description XML pour chaque appareil trouvé.
    Retourne {ip: {manufacturer, model_name, friendly_name, device_type}} —
    best-effort, {} si rien ne répond (fréquent : beaucoup d'appareils
    grand public seulement, UPnP souvent désactivé sur le matériel pro)."""
    emplacements = _ssdp_decouvrir_tout(timeout=timeout)
    if not emplacements:
        return {}
    resultat = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(emplacements), 10)) as ex:
        futurs = {ex.submit(_upnp_description_appareil, url): ip for ip, url in emplacements.items()}
        for f in concurrent.futures.as_completed(futurs):
            ip_source = futurs[f]
            try:
                description = f.result()
            except Exception:
                description = {}
            if description:
                resultat[ip_source] = description
    return resultat


# ── DÉCOUVERTE mDNS (imprimantes, Apple, Chromecast, NAS...) ────────────────
# zeroconf (déjà une dépendance) ne servait jusqu'ici qu'à retrouver d'autres
# instances ParcInfo (_register_mdns / collector_core.discover_parcinfo_mdns)
# — jamais à identifier le reste du matériel réseau. Types de service les
# plus courants sur un LAN grand public/PME.
_MDNS_TYPES_APPAREILS = [
    '_ipp._tcp.local.',            # Imprimantes IPP (la grande majorité aujourd'hui)
    '_printer._tcp.local.',        # Partage d'imprimante plus ancien
    '_pdl-datastream._tcp.local.', # Impression brute port 9100
    '_airplay._tcp.local.',        # Apple TV / récepteurs AirPlay
    '_raop._tcp.local.',           # AirPlay audio (ancien)
    '_googlecast._tcp.local.',     # Chromecast
    '_smb._tcp.local.',            # Partages SMB / NAS
    '_afpovertcp._tcp.local.',     # Partage Apple ancien (Time Capsule, vieux NAS)
    '_device-info._tcp.local.',    # Identification d'appareil Apple (Mac/iPhone)
    # Objets connectés — quasi tous bâtis sur ESP8266/ESP32, en marque
    # blanche sous des dizaines de noms commerciaux (Tuya en particulier) :
    # une liste de mots-clés hostname ne peut pas suivre, mais la plupart
    # s'annoncent en clair sur l'un de ces trois services mDNS.
    '_hap._tcp.local.',            # HomeKit (très large : prises, capteurs, ampoules...)
    '_shelly._tcp.local.',         # Shelly (prises/relais)
    '_esphomelib._tcp.local.',     # Firmware ESPHome (remplace souvent Tasmota/Tuya d'origine)
    '_matter._tcp.local.',         # Matter — standard smart-home unifié (2023+), adoption rapide
    '_matterc._udp.local.',        # Matter en cours d'appairage (commissionable)
    '_workstation._tcp.local.',    # Présence générique d'un poste (Linux/macOS/Windows Bonjour) —
                                    # capte un nom d'hôte là où NetBIOS/DNS inverse ne répondent pas
]


def _decouverte_mdns_reseau(timeout=3):
    """Parcourt les types de service mDNS les plus courants. Retourne
    {ip: {'nom':, 'service':}} — une seule session de navigation pour tout
    le scan, pas une par hôte.

    Limite connue (même que collector_core.discover_parcinfo_mdns) : un
    ParcInfo en conteneur Docker réseau « bridge » ne relaie généralement
    pas le trafic multicast — cette découverte reste alors silencieuse, le
    scan IP classique restant le repli.
    """
    if not MDNS_AVAILABLE:
        return {}
    try:
        from zeroconf import Zeroconf, ServiceBrowser
    except ImportError:
        return {}

    trouves = {}

    class _EcouteurAppareils:
        def add_service(self, zc, type_service, nom):
            if nom.startswith('ParcInfo'):
                return  # une autre instance ParcInfo, pas un appareil du parc
            try:
                info = zc.get_service_info(type_service, nom, timeout=1500)
            except Exception:
                info = None
            if not info:
                return
            try:
                adresses = info.parsed_addresses()
            except Exception:
                adresses = []
            for ip_str in adresses:
                if ip_str not in trouves:
                    trouves[ip_str] = {
                        'nom': nom.split('.')[0],
                        'service': type_service.rstrip('.').lstrip('_').split('._')[0],
                    }

        def update_service(self, *a, **k):
            pass

        def remove_service(self, *a, **k):
            pass

    zc = Zeroconf()
    try:
        ecouteur = _EcouteurAppareils()
        for type_service in _MDNS_TYPES_APPAREILS:
            try:
                ServiceBrowser(zc, type_service, ecouteur)
            except Exception:
                pass
        time.sleep(timeout)
    except Exception:
        pass
    finally:
        try:
            zc.close()
        except Exception:
            pass
    return trouves


# ── SNMP (sysDescr/sysName) ──────────────────────────────────────────────────
# Absent jusqu'ici (audit architecture : "toujours aucun SNMP") — pourtant le
# moyen le plus fiable d'identifier précisément un switch/imprimante/routeur
# managé : sysDescr renvoie en général le texte exact du constructeur
# ("Cisco IOS Software, C2960...", "HP ETHERNET MULTI-ENVIRONMENT..."). Client
# SNMPv1 GET minimal, encodage BER à la main (même logique que la requête DNS
# construite à la main pour la vérification DNS) : aucune dépendance
# supplémentaire pour une fonctionnalité de scan optionnelle. Communauté
# "public" en lecture seule uniquement, jamais d'écriture SNMP.
_OID_SYS_DESCR = '1.3.6.1.2.1.1.1.0'
_OID_SYS_NAME  = '1.3.6.1.2.1.1.5.0'


def _ber_longueur(n):
    if n < 0x80:
        return bytes([n])
    octets = []
    while n:
        octets.insert(0, n & 0xff)
        n >>= 8
    return bytes([0x80 | len(octets)]) + bytes(octets)


def _ber_entier(n):
    corps = n.to_bytes((n.bit_length() // 8) + 1, 'big', signed=True) if n else b'\x00'
    return b'\x02' + _ber_longueur(len(corps)) + corps


def _ber_chaine(s):
    b = s.encode('utf-8') if isinstance(s, str) else s
    return b'\x04' + _ber_longueur(len(b)) + b


def _ber_oid(oid_str):
    parties = [int(x) for x in oid_str.split('.')]
    octets = bytearray([parties[0] * 40 + parties[1]])
    for p in parties[2:]:
        if p == 0:
            octets.append(0)
            continue
        chunk = []
        while p:
            chunk.insert(0, p & 0x7f)
            p >>= 7
        for i in range(len(chunk) - 1):
            chunk[i] |= 0x80
        octets.extend(chunk)
    return b'\x06' + _ber_longueur(len(octets)) + bytes(octets)


def _ber_sequence(tag, contenu):
    return bytes([tag]) + _ber_longueur(len(contenu)) + contenu


def _ber_lire_tlv(data, pos):
    """Lit un TLV BER à partir de `pos`. Retourne (tag, valeur_brute, position_suivante)."""
    tag = data[pos]; pos += 1
    longueur = data[pos]; pos += 1
    if longueur & 0x80:
        n = longueur & 0x7f
        longueur = int.from_bytes(data[pos:pos + n], 'big')
        pos += n
    valeur = data[pos:pos + longueur]
    return tag, valeur, pos + longueur


def _snmp_get(ip_str, oids, communaute='public', timeout=0.8, port=161):
    """GET SNMPv1 minimal. Retourne {oid: texte} pour les OID qui ont répondu
    avec une chaîne (sysDescr/sysName sont toujours des OCTET STRING) — {}
    au moindre souci (agent absent, communauté refusée, timeout, réponse
    inattendue) : ce n'est qu'un signal supplémentaire, jamais bloquant.
    `port` n'existe que pour les tests (agent factice sur un port non
    privilégié) ; toujours 161 en usage réel."""
    try:
        varbinds = b''.join(_ber_sequence(0x30, _ber_oid(oid) + b'\x05\x00') for oid in oids)
        pdu_corps = _ber_entier(1) + _ber_entier(0) + _ber_entier(0) + _ber_sequence(0x30, varbinds)
        pdu = _ber_sequence(0xa0, pdu_corps)
        message = _ber_sequence(
            0x30, _ber_entier(0) + _ber_chaine(communaute) + pdu)

        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.settimeout(timeout)
            s.sendto(message, (ip_str, port))
            data, _ = s.recvfrom(2048)

        # SEQUENCE { version, communaute, GetResponse-PDU(0xA2) { ..., varbindlist } }
        _, corps, _ = _ber_lire_tlv(data, 0)
        pos = 0
        _, _version, pos = _ber_lire_tlv(corps, pos)
        _, _comm, pos = _ber_lire_tlv(corps, pos)
        tag_pdu, pdu_corps, _ = _ber_lire_tlv(corps, pos)
        if tag_pdu != 0xa2:
            return {}
        p = 0
        _, _reqid, p = _ber_lire_tlv(pdu_corps, p)
        _, err_status, p = _ber_lire_tlv(pdu_corps, p)
        _, _err_idx, p = _ber_lire_tlv(pdu_corps, p)
        if err_status and int.from_bytes(err_status, 'big', signed=True) != 0:
            return {}
        _, varbindlist, p = _ber_lire_tlv(pdu_corps, p)

        # Un agent SNMP répond dans le MÊME ordre que les OID demandés (RFC
        # 1157) — associer varbind #i à oids[i] directement, plutôt que de
        # tenter de ré-identifier l'OID retourné : un varbind en erreur (type
        # différent d'OCTET STRING) ne doit pas décaler les suivants.
        resultats = {}
        vp = 0
        i = 0
        while vp < len(varbindlist) and i < len(oids):
            tag_vb, vb_corps, vp = _ber_lire_tlv(varbindlist, vp)
            if tag_vb == 0x30:
                vbp = 0
                _tag_oid, _oid_brut, vbp = _ber_lire_tlv(vb_corps, vbp)
                tag_val, val_brut, vbp = _ber_lire_tlv(vb_corps, vbp)
                if tag_val == 0x04:  # OCTET STRING
                    texte = val_brut.decode('utf-8', errors='replace').strip()
                    if texte:
                        resultats[oids[i]] = texte
            i += 1
        return resultats
    except Exception:
        return {}


# ── SNMP walk (GETNEXT) — palier 3 du diagnostic réseau ──────────────────────
# _snmp_get ne fait qu'un GET sur des scalaires. Lire les compteurs par port
# d'un switch (ifTable / ifXTable / dot3StatsTable) demande de PARCOURIR une
# table : boucle GETNEXT jusqu'à sortir du sous-arbre demandé. Toujours en
# lecture seule, v1/v2c, best-effort ({} au moindre souci) — même philosophie
# que _snmp_get. Utilisé par network_diag._interroger_equipement().

def _ber_decoder_oid(brut):
    """Octets d'un OID BER → chaîne pointée ('1.3.6.1.2.1.2.2.1.2.1')."""
    if not brut:
        return ''
    premier = brut[0]
    parties = [str(premier // 40), str(premier % 40)]
    valeur = 0
    for octet in brut[1:]:
        valeur = (valeur << 7) | (octet & 0x7f)
        if not (octet & 0x80):
            parties.append(str(valeur))
            valeur = 0
    return '.'.join(parties)


# Sentinelles SNMPv2 (exception values) : l'OID existe mais pas de valeur.
_SNMP_SENTINELLES = {0x80: 'noSuchObject', 0x81: 'noSuchInstance', 0x82: 'endOfMibView'}


def _ber_decoder_valeur(tag, brut):
    """Décode une valeur de varbind SNMP → int | str | None (sentinelle)."""
    if tag in _SNMP_SENTINELLES:
        return None
    if tag == 0x02:  # INTEGER
        return int.from_bytes(brut, 'big', signed=True) if brut else 0
    if tag in (0x41, 0x42, 0x43, 0x46):  # Counter32, Gauge32/Unsigned32, TimeTicks, Counter64
        return int.from_bytes(brut, 'big', signed=False) if brut else 0
    if tag == 0x40:  # IpAddress
        return '.'.join(str(b) for b in brut)
    if tag == 0x06:  # OID
        return _ber_decoder_oid(brut)
    if tag == 0x05:  # NULL
        return None
    # 0x04 OCTET STRING et tout le reste : texte best-effort
    return brut.decode('utf-8', errors='replace').strip()


def _snmp_walk(ip_str, oid_base, communautes=('public',), timeout=1.2,
               max_vars=800, port=161):
    """Parcourt le sous-arbre `oid_base` par GETNEXT. Retourne
    {suffixe_oid: valeur} (suffixe = ce qui suit oid_base, ex. l'ifIndex).
    Essaie chaque communauté dans l'ordre ; s'arrête à endOfMibView, à la
    sortie du sous-arbre, sur erreur, ou à max_vars. {} si rien ne répond."""
    if isinstance(communautes, str):
        communautes = [communautes]
    prefixe = oid_base if oid_base.endswith('.') else oid_base + '.'
    for communaute in communautes:
        resultats = {}
        oid_courant = oid_base
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                s.settimeout(timeout)
                for _ in range(max_vars):
                    vb = _ber_sequence(0x30, _ber_oid(oid_courant) + b'\x05\x00')
                    pdu_corps = (_ber_entier(_snmp_walk_reqid()) + _ber_entier(0)
                                 + _ber_entier(0) + _ber_sequence(0x30, vb))
                    pdu = _ber_sequence(0xa1, pdu_corps)  # GetNextRequest-PDU
                    # version 1 (v2c) : GETBULK serait mieux mais GETNEXT marche partout
                    message = _ber_sequence(
                        0x30, _ber_entier(1) + _ber_chaine(communaute) + pdu)
                    s.sendto(message, (ip_str, port))
                    data, _ = s.recvfrom(4096)

                    _, corps, _ = _ber_lire_tlv(data, 0)
                    pos = 0
                    _, _v, pos = _ber_lire_tlv(corps, pos)
                    _, _c, pos = _ber_lire_tlv(corps, pos)
                    tag_pdu, pdu_corps_r, _ = _ber_lire_tlv(corps, pos)
                    if tag_pdu != 0xa2:
                        break
                    p = 0
                    _, _reqid, p = _ber_lire_tlv(pdu_corps_r, p)
                    _, err, p = _ber_lire_tlv(pdu_corps_r, p)
                    _, _ei, p = _ber_lire_tlv(pdu_corps_r, p)
                    if err and int.from_bytes(err, 'big', signed=True) != 0:
                        break
                    _, vblist, p = _ber_lire_tlv(pdu_corps_r, p)

                    vp = 0
                    _tag_vb, vb_corps, vp = _ber_lire_tlv(vblist, vp)
                    if _tag_vb != 0x30:
                        break
                    bp = 0
                    _to, oid_brut, bp = _ber_lire_tlv(vb_corps, bp)
                    tag_val, val_brut, bp = _ber_lire_tlv(vb_corps, bp)
                    oid_ret = _ber_decoder_oid(oid_brut)
                    if not oid_ret.startswith(prefixe) or tag_val == 0x82:
                        break  # sorti du sous-arbre / endOfMibView
                    resultats[oid_ret[len(prefixe):]] = _ber_decoder_valeur(tag_val, val_brut)
                    oid_courant = oid_ret
            if resultats:
                return resultats
        except Exception:
            continue
    return {}


_snmp_walk_reqid_compteur = [0]


def _snmp_walk_reqid():
    _snmp_walk_reqid_compteur[0] = (_snmp_walk_reqid_compteur[0] + 1) & 0x7fffffff
    return _snmp_walk_reqid_compteur[0] or 1


def _scan_host(ip_str, enrich_wmi=False, upnp_par_ip=None, mdns_par_ip=None):
    """Scanne un hôte : ping, hostname, NetBIOS, OS, ports, MAC, fabricant.

    Optionnellement enrichit avec WMI si enrich_wmi=True et que c'est une machine Windows
    accessible en RPC. `upnp_par_ip`/`mdns_par_ip` : résultats de la découverte
    UPnP/mDNS faite UNE fois pour tout le scan (_run_scan), pas par hôte —
    on y cherche simplement l'entrée pour CETTE ip.
    """
    ping_ok = _ping(ip_str)
    # Même si le ping échoue (pare-feu, très nombreux objets connectés qui
    # bloquent ICMP par défaut), l'OS a dû résoudre l'adresse MAC via ARP
    # pour tenter d'envoyer ce paquet sur le réseau local — un hôte qui
    # répond à cette résolution existe bel et bien, même silencieux côté
    # IP. Vérifier la table ARP même après un ping en échec révèle ces
    # appareils, jusqu'ici invisibles du scan (demandé explicitement : "une
    # image la plus complète possible du réseau, sans identification
    # manuelle par IP/MAC"). Un hôte ni pingable ni résolu en ARP est
    # considéré absent, comme avant.
    time.sleep(0.4)
    mac = _mac_from_arp(ip_str)
    if not ping_ok and not mac:
        time.sleep(0.4)
        mac = _mac_from_arp(ip_str)
        if not mac:
            return None
    # Après ping, laisser l'OS finir de peupler la table ARP si le premier
    # relevé n'a rien donné (hôte qui répond à l'IP mais MAC pas encore en
    # cache) — 0.5s est suffisant même sur les réseaux chargés.
    elif not mac:
        time.sleep(0.5)
        mac = _mac_from_arp(ip_str)
    # Lancer hostname + NetBIOS + OS + ports en parallèle
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as ex:
        f_hostname = ex.submit(_hostname,     ip_str)
        f_netbios  = ex.submit(_netbios_name, ip_str)
        f_os       = ex.submit(_ttl_os_guess, ip_str)
        f_ports    = ex.submit(_scan_ports,   ip_str)
        f_snmp     = ex.submit(_snmp_get,     ip_str, [_OID_SYS_DESCR, _OID_SYS_NAME])
        try: hostname  = f_hostname.result(timeout=5)
        except Exception: hostname = ""
        try: netbios   = f_netbios.result(timeout=5)
        except Exception: netbios = ""
        try: os_guess  = f_os.result(timeout=5)
        except Exception: os_guess = ""
        try: ports     = f_ports.result(timeout=15)
        except Exception: ports = []
        try: snmp_info = f_snmp.result(timeout=2)
        except Exception: snmp_info = {}
    # Toujours pas de MAC malgré les relevés précoces ? Les sondes
    # ci-dessus (hostname/NetBIOS/ports) ont généré du trafic supplémentaire
    # qui a pu déclencher la résolution ARP entre-temps — un dernier essai,
    # jamais un écrasement d'un MAC déjà trouvé.
    if not mac:
        mac = _mac_from_arp(ip_str)
        if not mac:
            time.sleep(0.5)
            mac = _mac_from_arp(ip_str)
    vendor    = _oui_vendor(mac)
    # Le TTL seul ne distingue pas macOS de Linux (les deux répondent à 64
    # par défaut) — signalé en usage réel : un MacBook ressortait comme
    # « Linux/Unix ». Le préfixe MAC officiel Apple, lui, ne laisse aucun
    # doute quand il est disponible ; affiné ici pour profiter à la fois de
    # la colonne OS affichée et de _deviner_type (qui reconnaît 'macOS').
    if os_guess == 'Linux/Unix' and 'apple' in (vendor or '').lower():
        os_guess = 'macOS'
    bannieres = _grab_bannieres(ip_str, ports)
    upnp_info = (upnp_par_ip or {}).get(ip_str) or {}
    mdns_info = (mdns_par_ip or {}).get(ip_str) or {}

    # sysName (SNMP) est en général le nom configuré à la main sur un
    # équipement managé — aussi fiable qu'un hostname DNS/NetBIOS, en repli
    # juste après eux (avant IGD UPnP/mDNS, moins spécifiques à CET appareil).
    snmp_sysname = (snmp_info or {}).get(_OID_SYS_NAME, '')
    snmp_sysdescr = (snmp_info or {}).get(_OID_SYS_DESCR, '')
    display_name = (netbios or hostname or snmp_sysname
                    or upnp_info.get('friendly_name') or mdns_info.get('nom') or ip_str)
    signal_supplementaire = ' '.join(bannieres.values())
    host_type = _deviner_type(
        display_name, ports, os_guess, vendor,
        extra_signal=' '.join([signal_supplementaire, snmp_sysdescr,
                               upnp_info.get('friendly_name', ''),
                               upnp_info.get('manufacturer', ''), mdns_info.get('nom', '')]),
        upnp_device_type=upnp_info.get('device_type', ''),
        mdns_service=mdns_info.get('service', ''),
        snmp_actif=bool(snmp_sysdescr or snmp_sysname),
    )

    result = {
        "ip":           ip_str,
        "hostname":     hostname,
        "netbios":      netbios,
        "display_name": display_name,
        "mac":          mac,
        "vendor":       vendor,
        "ports":        ports,
        "os_guess":     os_guess,
        "type":         host_type,
        # Présent mais ne répond pas à l'IP (pare-feu/ICMP bloqué) : détecté
        # uniquement via sa résolution ARP — l'appareil existe bel et bien,
        # simplement silencieux sur les sondes actives (ping/ports).
        "silencieux":   not ping_ok,
        "en_ligne":     True,
    }
    if bannieres:
        result["bannieres"] = bannieres
    if upnp_info:
        result["upnp"] = upnp_info
    if mdns_info:
        result["mdns"] = mdns_info
    if snmp_sysdescr or snmp_sysname:
        result["snmp"] = {'sysDescr': snmp_sysdescr, 'sysName': snmp_sysname}

    # Marque/modèle « les plus précis disponibles » : UPnP et mDNS
    # identifient déjà l'appareil lui-même (contrairement à vendor, qui ne
    # vient que du préfixe MAC — le fabricant de la puce réseau, pas
    # toujours celui de l'appareil). Le WMI ci-dessous, plus fiable encore
    # quand disponible, aura le dernier mot en écrasant ces valeurs.
    if upnp_info.get('manufacturer'):
        result['marque_detectee'] = upnp_info['manufacturer']
    if upnp_info.get('model_name'):
        result['modele_detectee'] = upnp_info['model_name']

    # Enrichissement WMI optionnel (pour Windows)
    if enrich_wmi and os_guess == "Windows":
        wmi_data = _enrich_from_wmi(ip_str, hostname)
        # Fusionner les données WMI (sans surcharger les données de scan)
        if wmi_data:
            result.update({
                'brand': wmi_data.get('brand', result.get('vendor', '')),
                'model': wmi_data.get('model', ''),
                'serial_number': wmi_data.get('serial_number', ''),
                'ram_gb': wmi_data.get('ram_gb', ''),
                'cpu': wmi_data.get('cpu', ''),
                'cpu_cores': wmi_data.get('cpu_cores', ''),
                'disk_total_gb': wmi_data.get('disk_total_gb', ''),
            })
            if wmi_data.get('brand'):
                result['marque_detectee'] = wmi_data['brand']
            if wmi_data.get('model'):
                result['modele_detectee'] = wmi_data['model']

    return result

def _run_scan(plages, nb_threads, enrich_wmi=False):
    global scan_status
    with scan_lock:
        scan_status = {"running": True, "progress": 0, "message": "Résolution des plages...", "results": [], "errors": [], "plages": plages}
    try:
        hosts = []
        for plage in plages:
            try:
                hosts += [str(ip) for ip in ipaddress.ip_network(plage.strip(), strict=False).hosts()]
            except Exception as e:
                with scan_lock:
                    scan_status["errors"].append(f"Plage invalide '{plage}': {e}")
        # Dédoublonner
        hosts = list(dict.fromkeys(hosts))

        # Découverte UPnP (ssdp:all) + mDNS (imprimantes, Apple, Chromecast,
        # NAS...) : UNE fois pour tout le scan, pas par hôte — et terminée
        # AVANT de démarrer le balayage IP, pour que chaque résultat soit
        # enrichi dès sa première apparition. Un petit balayage (peu d'IP)
        # peut finir en moins de temps que les ~3s que prend chacune de ces
        # découvertes ; les lancer après aurait laissé une course où les
        # premiers hôtes scannés n'auraient jamais leur enrichissement.
        with scan_lock:
            scan_status["message"] = "Découverte UPnP/mDNS..."
        decouvertes_upnp, decouvertes_mdns = {}, {}

        def _decouvrir_upnp():
            nonlocal decouvertes_upnp
            try:
                decouvertes_upnp = _decouverte_upnp_reseau()
            except Exception:
                pass

        def _decouvrir_mdns():
            nonlocal decouvertes_mdns
            try:
                decouvertes_mdns = _decouverte_mdns_reseau()
            except Exception:
                pass

        fils_decouverte = [threading.Thread(target=_decouvrir_upnp, daemon=True),
                           threading.Thread(target=_decouvrir_mdns, daemon=True)]
        for f in fils_decouverte:
            f.start()
        for f in fils_decouverte:
            f.join(timeout=6)

        total = len(hosts); found = []; scanned = [0]
        def on_done(future, ip):
            scanned[0] += 1
            try: result = future.result()
            except: result = None
            with scan_lock:
                scan_status["progress"] = int(scanned[0] / total * 100)
                scan_status["message"] = f"Progression : {scanned[0]}/{total} — {len(found)} trouvé(s)..."
                if result:
                    found.append(result)
                    scan_status["results"] = list(found)
        with concurrent.futures.ThreadPoolExecutor(max_workers=nb_threads) as executor:
            futures = {executor.submit(_scan_host, ip, enrich_wmi=enrich_wmi,
                                       upnp_par_ip=decouvertes_upnp, mdns_par_ip=decouvertes_mdns): ip
                      for ip in hosts}
            for f in concurrent.futures.as_completed(futures):
                on_done(f, futures[f])
        with scan_lock:
            scan_status.update({
                "progress": 100,
                "message": f"Terminé — {len(found)} appareil(s) détecté(s) sur {total} adresses",
                "running": False,
                "total_scanned": total,
            })
    except Exception as e:
        with scan_lock:
            scan_status.update({"message": f"Erreur : {e}", "running": False})

@app.route('/api/scan/client-suggere')
@login_required
def api_scan_client_suggere():
    """Liste des clients accessibles + suggestion pour la modale de sélection
    de client affichée avant chaque scan (voir scan_reseau.html).

    La suggestion n'est renvoyée que si un seul client, sans ambiguïté, a une
    plage IP locale (parc_general.plage_ip_locale) qui recoupe le réseau
    physique sur lequel CE poste ParcInfo se trouve actuellement — même
    logique que la surveillance ping (_reseaux_locaux_actuels /
    _appareil_sur_reseau_courant). Zéro ou plusieurs correspondances :
    aucune suggestion, la modale ne pré-coche rien.
    """
    clients = get_clients()
    reseaux_locaux = _reseaux_locaux_actuels()
    conn = get_db()
    matches = []
    for cl in clients:
        row = conn.execute(
            'SELECT plage_ip_locale FROM parc_general WHERE client_id=?', (cl['id'],)).fetchone()
        plage = ((row[0] if row else '') or '').strip()
        if plage and _appareil_sur_reseau_courant('', plage, reseaux_locaux):
            matches.append(cl['id'])
    conn.close()
    return jsonify({
        'clients': [{'id': cl['id'], 'nom': cl['nom']} for cl in clients],
        'client_actif': get_client_id(),
        'suggestion': matches[0] if len(matches) == 1 else None,
    })

@app.route('/api/scan/lancer', methods=['POST'])
@login_required
def lancer_scan():
    with scan_lock:
        if scan_status["running"]: return jsonify({"error":"Scan déjà en cours"}), 400
    data = request.json or {}
    # Support multiple ranges: "192.168.1.0/24,10.0.0.0/24" ou liste
    plage_raw = data.get('plage_ip', '192.168.1.0/24')
    if isinstance(plage_raw, list):
        plages = [p.strip() for p in plage_raw if p.strip()]
    else:
        plages = [p.strip() for p in plage_raw.split(',') if p.strip()]
    if not plages:
        plages = ['192.168.1.0/24']
    nb_threads = min(int(data.get('threads', 30)), 200)
    enrich_wmi = data.get('enrich_wmi', False)  # Optionnel - enrichir via WMI
    threading.Thread(target=_run_scan, args=(plages, nb_threads, enrich_wmi), daemon=True).start()
    return jsonify({"status": "started", "plages": plages, "enrich_wmi": enrich_wmi})

@app.route('/api/scan/status')
@login_required
def status_scan():
    with scan_lock: return jsonify(dict(scan_status))


# ─── DIAGNOSTIC RÉSEAU (module network_diag) ─────────────────────────────────

@app.route('/diag-reseau')
@login_required
def page_diag_reseau():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    if not get_client_access(cid):
        flash('Accès refusé', 'danger')
        return redirect(url_for('index'))
    conn = get_db()
    parc = row_to_dict(conn.execute(
        'SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    dernier_run = row_to_dict(conn.execute(
        'SELECT * FROM diag_reseau_runs WHERE client_id=? ORDER BY id DESC LIMIT 1', (cid,)).fetchone() or {})
    conn.close()
    return render_template('diag_reseau.html',
                           parc=parc, client=client, dernier_run=dernier_run,
                           etat_capture=network_diag.etat_capture(),
                           etat_moniteur=network_diag.etat_moniteur(),
                           peut_ecrire=can_write(cid),
                           clients=get_clients(), client_actif_id=cid)


@app.route('/api/diag-reseau/snapshot', methods=['POST'])
@login_required
def api_diag_snapshot():
    cid = get_client_id()
    if not can_write(cid):
        return jsonify({'error': 'Forbidden'}), 403
    data = request.json or {}
    plage = (data.get('plage') or '').strip()
    avec_capture = data.get('avec_capture')  # None = suit le réglage global
    if not network_diag.lancer_snapshot(cid, plage, avec_capture):
        return jsonify({'error': 'Un diagnostic est déjà en cours'}), 400
    return jsonify({'status': 'started'})


@app.route('/api/diag-reseau/snapshot/status')
@login_required
def api_diag_snapshot_status():
    st = network_diag.statut_snapshot()
    # N'exposer le statut que s'il concerne le client actif (ou aucun)
    if st.get('client_id') not in (None, get_client_id()):
        st = {'running': st.get('running'), 'progress': 0, 'message': 'Diagnostic en cours pour un autre client',
              'findings': [], 'avertissements': []}
    for f in st.get('findings', []):
        f['categorie_libelle'] = network_diag.libelle_categorie(f.get('categorie', ''))
    return jsonify(st)


@app.route('/api/diag-reseau/etat-capture')
@login_required
def api_diag_etat_capture():
    return jsonify(network_diag.etat_capture())


@app.route('/api/diag-reseau/snmp')
@login_required
def api_diag_snmp():
    cid = get_client_id()
    if not get_client_access(cid):
        return jsonify({'error': 'Forbidden'}), 403
    etat = network_diag.etat_snmp(cid)
    # Rattacher les évènements SNMP actifs à leur port pour surligner le tableau
    conn = get_db()
    evts = conn.execute(
        "SELECT categorie, details_json FROM diag_reseau_evenements "
        "WHERE client_id=? AND resolu=0 AND categorie IN "
        "('duplex_mismatch','port_crc','port_erreurs','port_sature','port_flapping','vitesse_reduite')",
        (cid,)).fetchall()
    conn.close()
    par_port = {}
    for cat, dj in evts:
        try:
            d = json.loads(dj or '{}')
        except Exception:
            continue
        cle = f"{d.get('equipement')}:{d.get('port_index')}"
        par_port.setdefault(cle, []).append({'categorie': cat,
                                             'libelle': network_diag.libelle_categorie(cat)})
    for eq in etat.get('equipements', []):
        for p in eq.get('ports', []):
            p['findings'] = par_port.get(f"{eq['ip']}:{p['port_index']}", [])
    return jsonify(etat)


@app.route('/api/diag-reseau/evenements')
@login_required
def api_diag_evenements():
    cid = get_client_id()
    if not get_client_access(cid):
        return jsonify({'error': 'Forbidden'}), 403
    page = max(1, int(request.args.get('page', 1) or 1))
    gravite = (request.args.get('gravite') or '').strip()
    categorie = (request.args.get('categorie') or '').strip()
    resolu = (request.args.get('resolu') or '').strip()  # '', '0', '1'
    where = ['client_id=?']
    params = [cid]
    if gravite:
        where.append('gravite=?'); params.append(gravite)
    if categorie:
        where.append('categorie=?'); params.append(categorie)
    if resolu in ('0', '1'):
        where.append('resolu=?'); params.append(int(resolu))
    where_sql = ' AND '.join('e.' + w for w in where)
    q = ('SELECT e.*, a.nom_machine AS appareil_nom '
         'FROM diag_reseau_evenements e '
         'LEFT JOIN appareils a ON a.id = e.appareil_id '
         'WHERE ' + where_sql +
         ' ORDER BY e.resolu ASC, CASE e.gravite WHEN "critique" THEN 0 WHEN "avertissement" THEN 1 ELSE 2 END,'
         ' e.derniere_occurrence DESC')
    rows, pagination = paginate(q, tuple(params), page)
    evenements = []
    for r in rows:
        d = row_to_dict(r)
        try:
            d['details'] = json.loads(d.get('details_json') or '{}')
        except Exception:
            d['details'] = {}
        d['categorie_libelle'] = network_diag.libelle_categorie(d.get('categorie', ''))
        evenements.append(d)
    return jsonify({'evenements': evenements, 'pagination': pagination})


@app.route('/api/diag-reseau/evenement/<int:id>/resoudre', methods=['POST'])
@login_required
def api_diag_resoudre(id):
    cid = get_client_id()
    if not can_write(cid):
        return jsonify({'error': 'Forbidden'}), 403
    user = get_auth_user()
    conn = get_db()
    row = conn.execute('SELECT titre FROM diag_reseau_evenements WHERE id=? AND client_id=?',
                       (id, cid)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not Found'}), 404
    reouvrir = bool((request.json or {}).get('reouvrir'))
    conn.execute(
        'UPDATE diag_reseau_evenements SET resolu=?, date_resolu=? WHERE id=? AND client_id=?',
        (0 if reouvrir else 1, None if reouvrir else _utcnow().isoformat(), id, cid))
    log_history(conn, cid, 'diag_reseau', id, row[0],
                'DIAG_REOUVRIR' if reouvrir else 'DIAG_RESOUDRE')
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/diag-reseau/surveillance', methods=['POST'])
@login_required
def api_diag_surveillance():
    cid = get_client_id()
    if not can_write(cid):
        return jsonify({'error': 'Forbidden'}), 403
    user = get_auth_user()
    actif = '1' if (request.json or {}).get('actif') else '0'
    cfg_set('diag_surveillance_active', actif)
    conn = get_db()
    log_history(conn, cid, 'diag_reseau', 0, 'Surveillance réseau',
                'DIAG_SURVEILLANCE', 'activée' if actif == '1' else 'désactivée')
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'actif': actif == '1', 'etat': network_diag.etat_moniteur()})

def _fmt_go(valeur):
    """'16' / 16.0 -> '16 Go' ; '512.5' -> '512.5 Go'. Vide si valeur absente."""
    if valeur in (None, ''):
        return ''
    try:
        nombre = float(valeur)
    except (TypeError, ValueError):
        return ''
    return f"{int(nombre)} Go" if nombre == int(nombre) else f"{nombre} Go"


# Types pour lesquels une position dans la baie de brassage a du sens —
# audit architecture : parc_general et la baie décrivaient le même matériel
# sans jamais se recouper ; le scan sait maintenant identifier ces types
# avec marque/modèle, autant s'en servir pour suggérer une entrée plutôt
# que de laisser la baie 100% manuelle. Jamais imposé : juste suggéré, la
# position physique dans le rack reste à saisir (elle ne peut pas être
# déduite d'un scan réseau).
_TYPES_SUGGESTION_BAIE = {'Switch', 'Routeur/Pare-feu', 'Switch/AP', 'NAS'}


@app.route('/api/scan/importer', methods=['POST'])
@login_required
def importer_scan():
    cid = get_client_id()
    items = request.json.get('appareils', [])
    conn = get_db(); now = _utcnow().isoformat()
    importes = 0; mis_a_jour = 0
    suggestions_baie = []
    for item in items:
        ip        = item.get('ip', '')
        ports_str = ','.join(str(p) for p in item.get('ports', []))
        nom       = item.get('netbios') or item.get('display_name') or item.get('hostname') or ip
        dns       = item.get('hostname', '')
        mac       = item.get('mac', '')
        vendor    = item.get('vendor', '')
        # Marque : la plus précise disponible (WMI/UPnP identifient l'appareil
        # lui-même — marque_detectee, posée par _scan_host) sinon repli sur le
        # fabricant de la puce réseau déduit du préfixe MAC.
        marque    = (item.get('marque_detectee') or item.get('brand') or '').strip() \
                    or (vendor.split('/')[0].strip() if vendor else '')
        modele       = (item.get('modele_detectee') or item.get('model') or '').strip()
        numero_serie = (item.get('serial_number') or '').strip()
        cpu          = (item.get('cpu') or '').strip()
        ram          = _fmt_go(item.get('ram_gb'))
        stockage     = _fmt_go(item.get('disk_total_gb'))
        existing  = conn.execute('SELECT id FROM appareils WHERE client_id=? AND adresse_ip=?', (cid, ip)).fetchone()
        if existing:
            app_id = existing[0]
            # Une valeur déjà présente (saisie à la main, ou posée par une
            # collecte précédente) n'est jamais écrasée par une simple
            # détection de scan — même principe que adresse_mac ci-dessous,
            # déjà en place.
            conn.execute(
                '''UPDATE appareils SET en_ligne=1, dernier_ping=?, ports_ouverts=?,
                   adresse_mac=COALESCE(NULLIF(adresse_mac,""),?),
                   marque=COALESCE(NULLIF(marque,""),?),
                   modele=COALESCE(NULLIF(modele,""),?),
                   numero_serie=COALESCE(NULLIF(numero_serie,""),?),
                   cpu=COALESCE(NULLIF(cpu,""),?),
                   ram=COALESCE(NULLIF(ram,""),?),
                   stockage=COALESCE(NULLIF(stockage,""),?),
                   date_maj=? WHERE client_id=? AND adresse_ip=?''',
                (now, ports_str, mac, marque, modele, numero_serie, cpu, ram, stockage,
                 now, cid, ip))
            mis_a_jour += 1
            log_history(conn, cid, 'appareil', app_id, nom, 'Auto-remplissage (scan réseau)',
                        {'source': 'scan-reseau', 'ip': ip})
        else:
            conn.execute(
                '''INSERT INTO appareils (client_id,adresse_ip,nom_machine,nom_dns,adresse_mac,marque,modele,
                   numero_serie,cpu,ram,stockage,type_appareil,
                   ports_ouverts,en_ligne,dernier_ping,decouvert_scan,statut,date_creation,date_maj)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,1,'actif',?,?)''',
                (cid, ip, nom, dns, mac, marque, modele, numero_serie, cpu, ram, stockage,
                 item.get('type', 'PC'), ports_str, now, now, now))
            app_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            importes += 1
            log_history(conn, cid, 'appareil', app_id, nom, 'Création (scan réseau)',
                        {'source': 'scan-reseau', 'ip': ip, 'mac': mac})

        type_detecte = item.get('type', 'PC')
        if type_detecte in _TYPES_SUGGESTION_BAIE:
            deja_dans_baie = conn.execute(
                'SELECT 1 FROM baie_slots WHERE appareil_id=?', (app_id,)).fetchone()
            if not deja_dans_baie:
                suggestions_baie.append({'id': app_id, 'nom': nom, 'type': type_detecte,
                                         'marque': marque, 'modele': modele})
    conn.commit(); conn.close()
    return jsonify({"importes": importes, "total": len(items), "suggestions_baie": suggestions_baie})


def _sync_collector_peripherals(conn, cid, appareil_id, monitors, printers, usb_devices=None):
    """Crée dans l'inventaire les périphériques remontés par le collecteur.

    Couvre les écrans, les imprimantes et les périphériques USB. Idempotent :
    chaque collecte relance la même liste, il ne doit donc rien se créer en
    double. La clé de dédoublonnage est le numéro de série quand il existe (les
    écrans EDID en fournissent un), sinon le couple marque/modèle pour ce
    client. Les périphériques USB disposent en plus d'une identité VID:PID
    stockée dans `source_usb_id`, plus stable qu'un libellé que Windows peut
    formuler différemment d'une version à l'autre.

    L'utilisateur affecté à la fiche appareil est reporté sur les fiches
    périphériques rattachées : un périphérique branché sur une machine est de
    fait utilisé par la personne à qui cette machine est affectée.

    Retourne le nombre de périphériques réellement créés.
    """
    created = 0
    now = _utcnow().isoformat()

    _row_user = conn.execute(
        'SELECT utilisateur FROM appareils WHERE id=?', (appareil_id,)).fetchone()
    _utilisateur_texte = _row_user[0] if _row_user else ''
    utilisateur_id = _resolve_utilisateur_id(conn, cid, _utilisateur_texte)

    def upsert(categorie, marque, modele, numero_serie, description, usb_id=''):
        nonlocal created
        marque, modele = (marque or '').strip(), (modele or '').strip()
        numero_serie = (numero_serie or '').strip()
        if not modele and not marque:
            return

        existing = None
        if usb_id:
            # Identité VID:PID (+ série) : on retrouve le même matériel d'une
            # collecte à l'autre. Sans numéro de série l'identité est limitée à
            # la machine, sinon deux souris identiques sur deux postes
            # fusionneraient en une seule fiche.
            if numero_serie:
                existing = conn.execute(
                    'SELECT id FROM peripheriques WHERE client_id=? AND source_usb_id=?',
                    (cid, usb_id)).fetchone()
            else:
                existing = conn.execute(
                    'SELECT p.id FROM peripheriques p'
                    ' JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id'
                    ' WHERE p.client_id=? AND p.source_usb_id=? AND pa.appareil_id=?',
                    (cid, usb_id, appareil_id)).fetchone()
        if not existing:
            if numero_serie:
                existing = conn.execute(
                    'SELECT id FROM peripheriques WHERE client_id=? AND categorie=? AND numero_serie=?',
                    (cid, categorie, numero_serie)).fetchone()
            else:
                existing = conn.execute(
                    'SELECT id FROM peripheriques WHERE client_id=? AND categorie=? AND marque=? AND modele=?',
                    (cid, categorie, marque, modele)).fetchone()

        if existing:
            pid = existing[0]
            if usb_id:
                conn.execute(
                    'UPDATE peripheriques SET source_usb_id=?, date_maj=? WHERE id=?',
                    (usb_id, now, pid))
        else:
            # Le lien à l'appareil vit uniquement dans la table pivot
            # peripheriques_appareils (ci-dessous), jamais dans la colonne
            # historique peripheriques.appareil_id — plus lue nulle part,
            # sa tenir à jour ici ne faisait qu'entretenir un doublon.
            conn.execute(
                '''INSERT INTO peripheriques
                   (client_id, utilisateur_id, categorie, marque, modele,
                    numero_serie, description, statut, source_usb_id, date_creation, date_maj)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (cid, utilisateur_id, categorie, marque,
                 modele, numero_serie, description, 'actif', usb_id, now, now))
            pid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            created += 1
            log_history(conn, cid, 'peripherique', pid,
                        f"{marque} {modele}".strip() or categorie,
                        'Création (collecteur système)',
                        {'source': 'system-info-collector', 'appareil_id': appareil_id})

        # Rattachement à la machine, que le périphérique soit neuf ou déjà connu
        conn.execute(
            'INSERT OR IGNORE INTO peripheriques_appareils (peripherique_id, appareil_id) VALUES (?,?)',
            (pid, appareil_id))

    try:
        for mon in monitors[:20]:
            if not isinstance(mon, dict):
                continue
            # Un EDID incomplet (dummy HDMI, KVM, splitter) ne donne qu'un
            # fabricant illisible : sans modèle, la fiche n'a aucune valeur
            if not (mon.get('model') or '').strip():
                continue
            year = mon.get('year')
            upsert('Ecran', mon.get('manufacturer'), mon.get('model'),
                   mon.get('serial_number'),
                   f"Détecté automatiquement{f' — année {year}' if year else ''}")

        for pr in printers[:20]:
            if not isinstance(pr, dict):
                continue
            # Print to PDF, XPS, fax, OneNote… ne sont pas du matériel
            if pr.get('virtual'):
                continue
            # Win32_Printer ne sépare pas marque et modèle : le nom porte tout
            details = []
            if pr.get('driver'):
                details.append(f"Pilote : {pr['driver']}")
            if pr.get('port'):
                details.append(f"Port : {pr['port']}")
            if pr.get('network'):
                details.append('Imprimante réseau')
            upsert('Imprimante', '', pr.get('name'), '',
                   ' — '.join(['Détectée automatiquement'] + details))

        for dev in (usb_devices or [])[:200]:
            if not isinstance(dev, dict):
                continue
            nom = (dev.get('inventory_name') or dev.get('name') or '').strip()
            if not nom:
                continue
            identite = _usb_identity(dev)
            upsert(dev.get('categorie') or 'Autre', dev.get('manufacturer') or '', nom,
                   dev.get('serial') or '',
                   'Détecté automatiquement par le collecteur (USB %s)' % identite,
                   usb_id=identite)

        # Report de l'utilisateur sur tous les périphériques rattachés, y
        # compris ceux créés lors d'une collecte précédente.
        _propager_utilisateur_aux_peripheriques(
            conn, appareil_id, cid, '', _utilisateur_texte)

        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("Création des périphériques depuis le collecteur")
        return 0

    return created


# Nombre de relevés conservés par appareil. Au-delà, les plus anciens partent :
# la tendance se lit sur les dernières semaines, pas sur l'historique complet.
COLLECTES_CONSERVEES = 60


def _enregistrer_collecte(conn, client_id, appareil_id, data):
    """Ajoute un relevé à l'historique de l'appareil.

    Ne conserve que ce qui se compare d'une collecte à l'autre. La liste des
    logiciels est réduite à « nom|version » : de quoi calculer les ajouts et
    retraits sans stocker l'inventaire complet à chaque passage.
    """
    try:
        logiciels = data.get('installed_software') or []
        empreintes = []
        for logiciel in logiciels:
            if isinstance(logiciel, dict):
                empreintes.append('%s|%s' % (logiciel.get('name', ''),
                                             logiciel.get('version', '')))
            else:
                empreintes.append('%s|' % logiciel)

        maintenant = _utcnow()
        horodatage = maintenant.isoformat(timespec='seconds')
        # La clé descend à la microseconde alors que l'horodatage affiché reste
        # à la seconde : deux collectes rapprochées auraient porté la même clé,
        # et la seconde aurait remplacé la première sans que rien ne le dise.
        cle = '%s|%s' % (appareil_id, maintenant.isoformat(timespec='microseconds'))

        conn.execute(
            '''INSERT OR REPLACE INTO collectes
               (cle, appareil_id, client_id, horodatage, disque_total_go,
                disque_utilise_go, disque_libre_go, ram_go, nb_logiciels,
                logiciels, os_version, cpu, numero_serie, nb_maj_attente,
                nb_peripheriques_erreur, date_maj)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (cle, appareil_id, client_id, horodatage,
             data.get('disk_total_gb'), data.get('disk_used_gb'), data.get('disk_free_gb'),
             data.get('ram_gb'), len(logiciels),
             json.dumps(empreintes, ensure_ascii=False),
             data.get('os_version') or '', data.get('cpu') or '',
             data.get('serial_number') or '',
             len(data.get('pending_updates') or []) or None,
             data.get('problem_devices_count'), horodatage))

        conn.execute(
            '''DELETE FROM collectes WHERE appareil_id=? AND cle NOT IN
               (SELECT cle FROM collectes WHERE appareil_id=?
                ORDER BY horodatage DESC LIMIT ?)''',
            (appareil_id, appareil_id, COLLECTES_CONSERVEES))
        max_j = int(cfg_get('collectes_max_jours') or 0)
        if max_j > 0:
            limite = (maintenant - timedelta(days=max_j)).isoformat(timespec='seconds')
            conn.execute('DELETE FROM collectes WHERE appareil_id=? AND horodatage < ?',
                        (appareil_id, limite))
        return True
    except Exception:
        # L'historique est un confort : son échec ne doit pas faire perdre la
        # collecte elle-même, qui vient d'aboutir.
        logger.exception('Relevé non enregistré pour l\'appareil %s', appareil_id)
        return False


def _enregistrer_cles_bitlocker(conn, client_id, appareil_id, cles):
    """Stocke les clés de récupération, chiffrées comme les mots de passe.

    La clé de la table mêle appareil et volume : une nouvelle collecte remplace
    la clé du même volume au lieu d'empiler des doublons, et deux instances
    n'entrent pas en collision comme le ferait un identifiant auto-incrémenté.
    """
    if not cles:
        return 0
    crypto = _get_crypto_shared()
    horodatage = _utcnow().isoformat(timespec='seconds')
    enregistrees = 0
    for entree in cles:
        if not isinstance(entree, dict):
            continue
        valeur = (entree.get('cle') or '').strip()
        if not valeur:
            continue
        volume = (entree.get('volume') or '').strip()
        identifiant = (entree.get('identifiant') or '').strip()
        conn.execute(
            '''INSERT OR REPLACE INTO cles_recuperation
               (cle, appareil_id, client_id, volume, identifiant, protection,
                chiffrement, valeur, date_maj) VALUES (?,?,?,?,?,?,?,?,?)''',
            ('%s|%s|%s' % (appareil_id, volume, identifiant),
             appareil_id, client_id, volume, identifiant,
             (entree.get('protection') or '').strip(),
             (entree.get('chiffrement') or '').strip(),
             crypto.encrypt(valeur), horodatage))
        enregistrees += 1
    return enregistrees


def _tendance_disque(releves):
    """Projette la date de saturation à partir des relevés d'espace disque.

    Régression linéaire simple sur (jours, espace libre). Ne conclut qu'à partir
    de trois relevés couvrant au moins une semaine : sur deux points rapprochés,
    la moindre variation donnerait une date absurde.
    """
    points = []
    for r in releves:
        libre = r.get('disque_libre_go')
        if libre is None:
            continue
        try:
            quand = datetime.fromisoformat(r['horodatage'])
        except (ValueError, TypeError, KeyError):
            continue
        points.append((quand, float(libre)))

    if len(points) < 3:
        return None
    points.sort(key=lambda p: p[0])
    etendue_jours = (points[-1][0] - points[0][0]).total_seconds() / 86400
    if etendue_jours < 7:
        return None

    origine = points[0][0]
    xs = [(q - origine).total_seconds() / 86400 for q, _ in points]
    ys = [v for _, v in points]
    n = len(points)
    moy_x, moy_y = sum(xs) / n, sum(ys) / n
    denominateur = sum((x - moy_x) ** 2 for x in xs)
    if denominateur == 0:
        return None
    pente = sum((x - moy_x) * (y - moy_y) for x, y in zip(xs, ys)) / denominateur

    # Pente positive : l'espace libre augmente, il n'y a rien à annoncer.
    if pente >= -0.01:
        return {'go_par_jour': round(pente, 2), 'saturation': None}

    libre_actuel = ys[-1]
    jours_restants = libre_actuel / abs(pente)
    if jours_restants > 3650:
        return {'go_par_jour': round(pente, 2), 'saturation': None}
    return {
        'go_par_jour': round(pente, 2),
        'jours_restants': int(jours_restants),
        'saturation': (points[-1][0] + timedelta(days=jours_restants)).date().isoformat(),
    }


def _comparer_collectes(recent, precedent):
    """Différences entre deux relevés : logiciels et matériel."""
    def _charger(releve):
        try:
            return set(json.loads(releve.get('logiciels') or '[]'))
        except (ValueError, TypeError):
            return set()

    logiciels_recents, logiciels_avant = _charger(recent), _charger(precedent)

    def _nom(empreinte):
        nom, _, version = empreinte.partition('|')
        return '%s %s' % (nom, version) if version else nom

    changements = []
    for champ, libelle in (('os_version', "Version du système"),
                           ('cpu', 'Processeur'),
                           ('numero_serie', 'Numéro de série'),
                           ('ram_go', 'Mémoire (Go)')):
        avant, apres = precedent.get(champ), recent.get(champ)
        if avant and apres and str(avant) != str(apres):
            changements.append({'champ': libelle, 'avant': avant, 'apres': apres})

    return {
        'ajoutes': sorted(_nom(e) for e in (logiciels_recents - logiciels_avant)),
        'retires': sorted(_nom(e) for e in (logiciels_avant - logiciels_recents)),
        'materiel': changements,
    }


def historique_appareil(conn, client_id, appareil_id):
    """Relevés d'un appareil, avec tendance et comparaison des deux derniers."""
    releves = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM collectes WHERE appareil_id=? AND client_id=? '
        'ORDER BY horodatage DESC LIMIT ?',
        (appareil_id, client_id, COLLECTES_CONSERVEES)).fetchall()]
    if not releves:
        return None

    resultat = {
        'releves': releves,
        'nombre': len(releves),
        'depuis': releves[-1]['horodatage'],
        'tendance': _tendance_disque(releves),
        'comparaison': None,
    }
    if len(releves) >= 2:
        resultat['comparaison'] = _comparer_collectes(releves[0], releves[1])
        resultat['compare_a'] = releves[1]['horodatage']
    return resultat


def _normaliser_libelle(valeur):
    """Minuscules sans accents ni ponctuation, pour comparer des libellés."""
    import unicodedata
    texte = unicodedata.normalize('NFKD', str(valeur or '')).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', ' ', texte.lower()).strip()


def _rapprocher_liste(valeur, entrees):
    """Entrée de la liste qui correspond le mieux à la valeur détectée.

    Les listes sont curées (« Windows Defender / Microsoft Defender »), la
    valeur détectée ne l'est pas (« Windows Defender ») : on compare sur les
    mots, et l'entrée qui en partage le plus l'emporte.
    """
    cible = set(_normaliser_libelle(valeur).split())
    if not cible:
        return ''
    meilleur, score_max = '', 0
    for entree in entrees or []:
        mots = set(_normaliser_libelle(entree).split())
        score = len(cible & mots)
        # Un seul mot commun ne suffit que s'il couvre l'entrée entière,
        # sinon « Security » rapprocherait n'importe quoi de n'importe quoi.
        if score > score_max and (score > 1 or mots <= cible or cible <= mots):
            meilleur, score_max = entree, score
    return meilleur


def champs_deduits_du_collecteur(conn, client_id, rapport, existant=None):
    """Champs de la fiche appareil que la collecte permet de renseigner.

    Ne renvoie que ce qui manque : une valeur saisie par un technicien prime
    toujours sur une valeur déduite, comme pour le type d'appareil.
    """
    existant = existant or {}
    deduits = {}

    def poser(colonne, valeur):
        if valeur and not (existant.get(colonne) or '').strip():
            deduits[colonne] = valeur

    # Utilisateur : la session ouverte au moment de la collecte. Le domaine est
    # retiré, la fiche attend un nom de personne, pas un identifiant complet.
    session_ouverte = (rapport.get('logged_on_user') or '').strip()
    if session_ouverte:
        poser('utilisateur', session_ouverte.split('\\')[-1].split('@')[0])

    # Antivirus : le collecteur écrivait dans la colonne `antivirus`, que le
    # formulaire n'affiche pas — d'où des champs Marque et Nom restés vides
    # alors que la fiche système annonçait « Windows Defender ».
    produits = rapport.get('antivirus_products') or []
    detecte = (produits[0].get('name') if produits else '') or rapport.get('antivirus') or ''
    detecte = detecte.split(',')[0].strip()
    if detecte:
        poser('av_marque', _rapprocher_liste(detecte, get_liste('marques_antivirus')) or detecte)
        poser('av_nom', _rapprocher_liste(detecte, get_liste('noms_antivirus')) or detecte)

    # EDR et RMM : détectés parmi les services (au mieux — voir
    # collector_core._AGENTS_EDR/_AGENTS_RMM), avec leur marque et leur nom
    # déjà mis en forme, sans passage par _rapprocher_liste.
    edr = (rapport.get('edr_agents') or [None])[0]
    if edr:
        poser('edr_marque', edr.get('marque'))
        poser('edr_nom', edr.get('nom'))
    rmm = (rapport.get('remote_support_agents') or [None])[0]
    if rmm:
        poser('rmm_marque', rmm.get('marque'))
        poser('rmm_nom', rmm.get('nom'))

    # AnyDesk : l'identifiant se lit directement sur le poste
    # (anydesk.exe --get-id), sans dépendre d'une saisie a posteriori.
    if rapport.get('anydesk_id'):
        poser('anydesk_id', rapport['anydesk_id'])

    # Logiciels métier : ceux de la liste du client effectivement installés.
    try:
        references = _get_logiciels_metier_list(conn, client_id) or []
    except Exception:
        references = []
    if references:
        installes = {_normaliser_libelle(l.get('name') if isinstance(l, dict) else l)
                     for l in (rapport.get('installed_software') or [])}
        trouves = [ref for ref in references
                   if any(_normaliser_libelle(ref) and _normaliser_libelle(ref) in nom
                          for nom in installes)]
        if trouves:
            poser('logiciels', json.dumps(trouves, ensure_ascii=False))

    return deduits


#: Marqueur de rattrapage, pour ne le faire qu'une fois par base. Le suffixe
#: de version augmente quand `champs_deduits_du_collecteur` apprend à en
#: déduire davantage : sans cela, les collectes déjà en base — reçues avant
#: l'ajout — ne profiteraient jamais du nouveau rattrapage.
_CLE_RATTRAPAGE_FICHES = '_fiches_completees_v2'
_rattrapage_fait = False


def completer_fiches_existantes():
    """Renseigne a posteriori les fiches déjà collectées.

    Les appareils collectés avant cette version ont leur rapport en base mais
    des champs restés vides : sans ce rattrapage, il faudrait relancer le
    collecteur sur chaque poste pour une donnée déjà présente. Ne touche que
    les cases vides, et ne s'exécute qu'une fois.
    """
    global _rattrapage_fait
    if _rattrapage_fait:
        return 0
    _rattrapage_fait = True

    conn = get_db()
    try:
        if (cfg_get(_CLE_RATTRAPAGE_FICHES, '') or '').strip():
            return 0

        completes = 0
        lignes = conn.execute(
            'SELECT * FROM appareils '
            "WHERE rapport_systeme_json IS NOT NULL AND rapport_systeme_json != ''"
        ).fetchall()
        for ligne in lignes:
            appareil = row_to_dict(ligne)
            try:
                rapport = json.loads(appareil.get('rapport_systeme_json') or '{}')
            except (ValueError, TypeError):
                continue
            if not isinstance(rapport, dict):
                continue
            deduits = champs_deduits_du_collecteur(
                conn, appareil.get('client_id'), rapport, appareil)
            if not deduits:
                continue
            conn.execute('UPDATE appareils SET %s WHERE id=?'
                         % ', '.join('%s=?' % c for c in deduits),
                         list(deduits.values()) + [appareil['id']])
            completes += 1

        conn.commit()
        cfg_set(_CLE_RATTRAPAGE_FICHES, _utcnow().isoformat(timespec='seconds'))
        if completes:
            logger.info('Fiches appareils complétées depuis les collectes déjà reçues : %d',
                        completes)
        return completes
    except Exception:
        logger.exception('Rattrapage des fiches appareils (sans conséquence)')
        return 0
    finally:
        conn.close()


#: Marqueur de rattrapage, pour ne le faire qu'une fois par base — voir
#: _TRACKED_JOURNAL['baie_slot_ports'] ci-dessus : le trigger de
#: journalisation ne couvre que les écritures futures, il ne journalise pas
#: rétroactivement les ports déjà créés avant ce correctif. Sans ce
#: rattrapage, un port/câblage créé avant la mise à jour resterait invisible
#: sur les autres instances jusqu'à sa prochaine modification manuelle.
_CLE_RATTRAPAGE_BAIE_PORTS = '_baie_slot_ports_journalises_v1'
_rattrapage_baie_ports_fait = False


def rattraper_sync_baie_slot_ports():
    """Journalise (une seule fois) les baie_slot_ports déjà en base comme
    autant d'INSERT — permet à la sync bidirectionnelle existante de les
    propager vers Turso puis vers les autres instances au prochain cycle,
    sans mécanisme de sync séparé."""
    global _rattrapage_baie_ports_fait
    if _rattrapage_baie_ports_fait:
        return 0
    _rattrapage_baie_ports_fait = True

    conn = get_db()
    try:
        if (cfg_get(_CLE_RATTRAPAGE_BAIE_PORTS, '') or '').strip():
            return 0

        ids = [r[0] for r in conn.execute('SELECT id FROM baie_slot_ports').fetchall()]
        now = _utcnow().isoformat()
        for rid in ids:
            conn.execute(
                "DELETE FROM _sync_journal WHERE tbl='baie_slot_ports' AND record_id=? AND action='INSERT'",
                (rid,))
            conn.execute(
                "INSERT INTO _sync_journal (tbl, record_id, action, timestamp) VALUES ('baie_slot_ports', ?, 'INSERT', ?)",
                (rid, now))
        conn.commit()
        cfg_set(_CLE_RATTRAPAGE_BAIE_PORTS, _utcnow().isoformat(timespec='seconds'))
        if ids:
            logger.info('Rattrapage sync baie_slot_ports : %d port(s) journalisé(s)', len(ids))
        return len(ids)
    except Exception:
        logger.exception('Rattrapage sync baie_slot_ports (sans conséquence)')
        return 0
    finally:
        conn.close()


#: Même mécanique que _CLE_RATTRAPAGE_BAIE_PORTS ci-dessus, pour les prises
#: murales issues de la migration dans init_db() — ajouter la table au
#: trigger de journalisation ne couvre que les écritures FUTURES, pas les
#: lignes que la migration vient d'insérer directement en base.
_CLE_RATTRAPAGE_BAIE_PRISES_MURALES = '_baie_prises_murales_journalisees_v1'
_rattrapage_baie_prises_murales_fait = False


def rattraper_sync_baie_prises_murales():
    """Journalise (une seule fois) les baie_prises_murales déjà en base
    comme autant d'INSERT — même rôle que rattraper_sync_baie_slot_ports()
    pour la table sœur."""
    global _rattrapage_baie_prises_murales_fait
    if _rattrapage_baie_prises_murales_fait:
        return 0
    _rattrapage_baie_prises_murales_fait = True

    conn = get_db()
    try:
        if (cfg_get(_CLE_RATTRAPAGE_BAIE_PRISES_MURALES, '') or '').strip():
            return 0

        ids = [r[0] for r in conn.execute('SELECT id FROM baie_prises_murales').fetchall()]
        now = _utcnow().isoformat()
        for rid in ids:
            conn.execute(
                "DELETE FROM _sync_journal WHERE tbl='baie_prises_murales' AND record_id=? AND action='INSERT'",
                (rid,))
            conn.execute(
                "INSERT INTO _sync_journal (tbl, record_id, action, timestamp) VALUES ('baie_prises_murales', ?, 'INSERT', ?)",
                (rid, now))
        conn.commit()
        cfg_set(_CLE_RATTRAPAGE_BAIE_PRISES_MURALES, _utcnow().isoformat(timespec='seconds'))
        if ids:
            logger.info('Rattrapage sync baie_prises_murales : %d prise(s) journalisée(s)', len(ids))
        return len(ids)
    except Exception:
        logger.exception('Rattrapage sync baie_prises_murales (sans conséquence)')
        return 0
    finally:
        conn.close()


def jeton_collecteur_valide(cid=None):
    """Vrai si la requête du collecteur est autorisée.

    Tant qu'aucun jeton global n'est configuré, tout passe : c'était le
    comportement depuis toujours, et le durcir sans prévenir couperait les
    collecteurs déjà déployés. Dès qu'un jeton global est renseigné dans la
    configuration, il devient obligatoire — sans lui, n'importe qui atteignant
    le serveur peut créer ou modifier des appareils et déposer des fichiers.
    Il agit comme un jeton maître : valide, il donne accès à tous les clients.

    Si aucun jeton global n'est configuré, un client peut malgré tout se
    protéger individuellement en renseignant son propre jeton (fiche client) :
    tant qu'il reste vide, ce client suit le comportement permissif historique
    (aucune régression pour les déploiements existants qui ignorent ce champ) ;
    une fois renseigné, seul ce jeton (passé en paramètre `cid`) autorise les
    requêtes ciblant CE client — ça referme la faille « un jeton valide pour
    un client donne accès à tous les autres », sans rien casser pour ceux qui
    n'ont configuré qu'un jeton global unique.
    """
    attendu = (cfg_get('collecteur_token', '') or '').strip()
    recu = (request.headers.get('X-Collector-Token')
            or request.headers.get('Authorization', '')).strip()
    if recu.lower().startswith('bearer '):
        recu = recu[7:].strip()

    if attendu:
        # Comparaison sur les octets : compare_digest refuse les chaînes
        # contenant des caractères non-ASCII, et un jeton accentué faisait
        # répondre 500 au lieu de 401 — la fonction censée protéger l'API
        # la cassait.
        if not recu or not secrets.compare_digest(recu.encode('utf-8'),
                                                  attendu.encode('utf-8')):
            logger.warning('Collecteur refusé : jeton absent ou invalide (ip=%s, chemin=%s)',
                           request.remote_addr, request.path)
            return False
        return True

    if cid:
        conn = get_db()
        row = conn.execute('SELECT collecteur_token FROM clients WHERE id=?', (cid,)).fetchone()
        conn.close()
        jeton_client = ((row[0] if row else '') or '').strip()
        if jeton_client:
            if not recu or not secrets.compare_digest(recu.encode('utf-8'),
                                                       jeton_client.encode('utf-8')):
                logger.warning('Collecteur refusé : jeton dédié requis pour client_id=%s (ip=%s)',
                               cid, request.remote_addr)
                return False

    return True


@app.route('/api/device-info', methods=['POST'])
def api_device_info():
    """
    Endpoint pour recevoir les infos système du collecteur autonome.

    POST /api/device-info
    {
        "mac_address": "00:1A:2B:3C:4D:5E",
        "hostname": "DESKTOP-ABC123",
        "ip_addresses": ["192.168.1.100"],
        "os_name": "Windows",
        "os_version": "11",
        "brand": "Dell",
        "model": "Latitude 5420",
        "serial_number": "ABC123XYZ",
        "ram_gb": 16,
        "cpu": "Intel Core i7-1185G7",
        "cpu_cores": 4,
        "disk_total_gb": 512,
        "antivirus": "Windows Defender",
        "installed_software": ["Python", "VS Code", ...],
        "timestamp": "2026-08-03T12:34:56.789012"
    }

    Matching :
    1. MAC address si fourni
    2. IP (première de ip_addresses)
    3. Crée si aucun match
    """
    if not jeton_collecteur_valide():
        return jsonify({"status": "error", "message": "Jeton collecteur requis"}), 401
    try:
        data = request.json or {}

        mac_address = data.get('mac_address', '').upper()
        ip_addresses = data.get('ip_addresses', [])
        ip_address = ip_addresses[0] if ip_addresses else ''
        hostname = data.get('hostname', '')

        # Le collecteur DOIT spécifier le client cible
        # Stratégies :
        # 1. Paramètre client_id dans le JSON
        # 2. Token d'authentification (valide un user qui a accès)
        # 3. Client par défaut (fallback)

        cid = None

        # 1. Essayer client_id du JSON
        cid_param = data.get('client_id')
        if cid_param:
            try:
                cid = int(cid_param)
                # Vérifier que ce client existe
                conn_test = get_db()
                exists = conn_test.execute('SELECT id FROM clients WHERE id=?', (cid,)).fetchone()
                conn_test.close()
                if not exists:
                    return jsonify({"status": "error", "message": f"Client ID {cid} not found"}), 404
            except (ValueError, TypeError):
                return jsonify({"status": "error", "message": "Invalid client_id parameter"}), 400

        # 1b. Essayer client_name (fallback)
        if not cid:
            client_name = data.get('client_name')
            if client_name:
                conn_name = get_db()
                client_row = conn_name.execute('SELECT id FROM clients WHERE nom=?', (client_name,)).fetchone()
                conn_name.close()
                if client_row:
                    cid = client_row[0]
                else:
                    return jsonify({"status": "error", "message": f"Client '{client_name}' not found"}), 404

        # 2. Si pas de client_id, vérifier token d'auth
        if not cid:
            token = data.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
            if token:
                conn_token = get_db()
                # Chercher un user avec ce token (simple vérification)
                # Dans une vraie implémentation, générer des tokens API
                # Pour l'instant, utiliser une clé secrète simple
                api_secret = cfg_get('api_collector_secret', '')
                if api_secret and token == api_secret:
                    # Token valide - utiliser le premier client de l'admin
                    user = conn_token.execute('SELECT id FROM auth_users WHERE role="admin" LIMIT 1').fetchone()
                    if user:
                        client = conn_token.execute('SELECT id FROM clients WHERE auth_user_id=? LIMIT 1',
                                                   (user[0],)).fetchone()
                        if client:
                            cid = client[0]
                conn_token.close()

        # 3. Fallback : client "Découverte réseau" ou premier client
        if not cid:
            conn_fall = get_db()
            # Chercher client "Découverte réseau"
            c = conn_fall.execute('SELECT id FROM clients WHERE nom LIKE ?', ('%Découverte%',)).fetchone()
            if c:
                cid = c[0]
            else:
                # Créer le client s'il n'existe pas
                conn_fall.execute(
                    'INSERT INTO clients (nom, contact, email, date_creation, date_maj) VALUES (?,?,?,?,?)',
                    ('Découverte réseau', 'auto', 'auto@parcinfo.local', _utcnow().isoformat(), _utcnow().isoformat())
                )
                conn_fall.commit()
                cid = conn_fall.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn_fall.close()

            # Avertir que le client "Découverte réseau" est utilisé
            app.logger.warning(f"Device info received without explicit client_id - using default 'Découverte réseau' (ID: {cid}). Specify 'client_id' in request to target a specific client.")

        if not cid:
            return jsonify({"status": "error", "message": "No valid client found - specify client_id or configure collector token"}), 400

        if not jeton_collecteur_valide(cid):
            return jsonify({"status": "error", "message": "Jeton collecteur requis pour ce client"}), 401

        # Marque et modèle
        brand = data.get('brand', '')
        model = data.get('model', '')
        serial = data.get('serial_number', '')

        # Infos système (séparées et complètes)
        os_name = data.get('os_name', '').strip()
        os_version = data.get('os_version', '').strip()
        ram_gb = data.get('ram_gb', '')
        cpu = data.get('cpu', '')
        cpu_cores = data.get('cpu_cores', '')
        disk_gb = data.get('disk_total_gb', '')
        antivirus = data.get('antivirus', '')
        gpu = data.get('gpu', '')
        dns_name = (data.get('dns_name') or '').strip()
        device_type = (data.get('device_type') or '').strip()
        # IP publique + opérateur : comme adresse_ip, une valeur qui peut
        # légitimement changer d'une collecte à l'autre (poste itinérant,
        # IP dynamique) — toujours resynchronisée, jamais figée sur la
        # première collecte (voir plus bas, même traitement qu'adresse_ip).
        public_ip = (data.get('public_ip') or '').strip()
        public_ip_isp = (data.get('public_ip_isp') or '').strip()

        # Ports TCP en écoute → colonne ports_ouverts (format identique au scan réseau)
        open_ports = data.get('open_ports') or []
        ports_str = ', '.join(str(p) for p in open_ports[:200]) if open_ports else ''

        # Logiciels (liste complète - garde-fou à 2000 entrées contre un payload aberrant)
        software_list = data.get('installed_software', [])
        software_json = json.dumps(software_list[:2000], ensure_ascii=False) if software_list else ''

        # Périphériques USB (même garde-fou contre un payload aberrant)
        usb_devices = data.get('usb_devices') or []
        if not isinstance(usb_devices, list):
            usb_devices = []
        usb_devices = usb_devices[:200]

        # Licences avec clé complète récupérée par le collecteur
        collected_licenses = data.get('licenses') or []
        if not isinstance(collected_licenses, list):
            collected_licenses = []

        # Snapshot complet du collecteur (mémoire par slot, licences, SMART, écrans…)
        # Plafonné à 1 Mo : au-delà, c'est un payload aberrant, pas un inventaire
        system_report = data.get('system_report') or {}
        # Les clés de récupération BitLocker déverrouillent les disques : elles
        # sont extraites du rapport avant tout stockage. Le rapport, lui, est
        # conservé tel quel en base et repris dans le PDF joint à l'appareil —
        # un secret n'a rien à y faire.
        cles_bitlocker = []
        if isinstance(system_report, dict) and system_report.get('bitlocker_keys'):
            cles_bitlocker = system_report.pop('bitlocker_keys') or []

        system_report_json = ''
        if system_report:
            try:
                candidate = json.dumps(system_report, ensure_ascii=False)
                if len(candidate) <= 1_000_000:
                    system_report_json = candidate
                else:
                    app.logger.warning(
                        f"system_report ignoré ({len(candidate)} octets > 1 Mo) pour {hostname or mac_address}")
            except (TypeError, ValueError):
                app.logger.warning("system_report non sérialisable - ignoré")

        now = _utcnow().isoformat()
        conn = get_db()

        # Stratégie de matching :
        # 1. MAC address si connu
        # 2. IP address si fourni
        # 3. Hostname sinon (risqué)
        existing = None

        if mac_address:
            existing = conn.execute(
                'SELECT id FROM appareils WHERE client_id=? AND adresse_mac=?',
                (cid, mac_address)
            ).fetchone()

        if not existing and ip_address:
            existing = conn.execute(
                'SELECT id FROM appareils WHERE client_id=? AND adresse_ip=?',
                (cid, ip_address)
            ).fetchone()

        if not existing and hostname:
            existing = conn.execute(
                'SELECT id FROM appareils WHERE client_id=? AND nom_machine=?',
                (cid, hostname)
            ).fetchone()

        # Construire le nom machine
        device_name = hostname or ip_address or f"Device-{mac_address[:8]}"

        if existing:
            # Mise à jour
            app_id = existing[0]
            old_row = conn.execute('SELECT * FROM appareils WHERE id=?', (app_id,)).fetchone()
            old_data = row_to_dict(old_row) if old_row else {}

            # Mettre à jour les champs techniques fournis par le collecteur (toujours
            # resynchronisés - ces valeurs reflètent l'état réel de la machine et ne
            # doivent pas rester bloquées sur la première valeur enregistrée)
            updates = []
            params = []

            # Marque/modèle/n° de série ne peuvent pas varier sur une machine
            # réelle (contrairement à l'IP, l'OS, la RAM, etc.) : une collecte
            # ne doit jamais écraser une correction manuelle par une valeur
            # mal détectée. Rempli uniquement si le champ est encore vide.
            marque_actuelle = (old_data.get('marque') or '').strip()
            if brand and not marque_actuelle:
                updates.append('marque=?')
                params.append(brand)

            modele_actuel = (old_data.get('modele') or '').strip()
            if model and not modele_actuel:
                updates.append('modele=?')
                params.append(model)

            serial_actuel = (old_data.get('numero_serie') or '').strip()
            if serial and not serial_actuel:
                updates.append('numero_serie=?')
                params.append(serial)

            if ip_address:
                updates.append('adresse_ip=?')
                params.append(ip_address)

            if mac_address:
                updates.append('adresse_mac=?')
                params.append(mac_address)

            if public_ip:
                updates.append('adresse_ip_publique=?')
                params.append(public_ip)

            if public_ip_isp:
                updates.append('operateur_ip_publique=?')
                params.append(public_ip_isp)

            # Un appareil créé avant que le hostname soit connu (résolution
            # échouée, ou carte réseau pas encore identifiée) reste nommé
            # d'après son IP ou « Device-XXXXXXXX » pour toujours : rien dans
            # la boucle normale ne corrige `nom_machine` après la création.
            # Un vrai hostname obtenu depuis remplace ce repli — jamais un
            # nom que quelqu'un a choisi à la main.
            nom_actuel = (old_data.get('nom_machine') or '').strip()
            est_repli = bool(re.match(r'^\d{1,3}(\.\d{1,3}){3}$', nom_actuel)) \
                or nom_actuel.startswith('Device-')
            if hostname and est_repli and hostname != nom_actuel:
                updates.append('nom_machine=?')
                params.append(hostname)

            if os_name:
                updates.append('os=?')
                params.append(os_name)

            if os_version:
                updates.append('version_os=?')
                params.append(os_version)

            if ram_gb:
                updates.append('ram=?')
                params.append(str(ram_gb))

            if cpu:
                updates.append('cpu=?')
                params.append(cpu)

            if disk_gb:
                updates.append('stockage=?')
                params.append(str(disk_gb))

            if antivirus:
                updates.append('antivirus=?')
                params.append(antivirus)

            if gpu:
                updates.append('carte_graphique=?')
                params.append(gpu)

            if dns_name:
                updates.append('nom_dns=?')
                params.append(dns_name)

            if ports_str:
                updates.append('ports_ouverts=?')
                params.append(ports_str)

            # Le type est déduit du châssis SMBIOS. On ne l'écrase que s'il est
            # vide ou resté sur le 'PC' générique posé par les anciennes versions :
            # un type corrigé à la main par un technicien doit primer.
            if device_type and (old_data.get('type_appareil') or 'PC') == 'PC':
                updates.append('type_appareil=?')
                params.append(device_type)

            if software_json:
                updates.append('logiciels_installes_json=?')
                params.append(software_json)

            if system_report_json:
                updates.append('rapport_systeme_json=?')
                params.append(system_report_json)

            # Champs de la fiche que la collecte permet de renseigner : ils
            # n'étaient jamais alimentés, alors que la donnée était sous les yeux
            # dans la fiche système. Seules les cases encore vides sont remplies.
            for colonne, valeur in champs_deduits_du_collecteur(
                    conn, cid, system_report or data, old_data).items():
                updates.append('%s=?' % colonne)
                params.append(valeur)

            # Toujours mettre à jour la date de dernière synchronisation
            updates.append('derniere_synchro=?')
            params.append(now)

            if updates:
                query = f"UPDATE appareils SET {', '.join(updates)} WHERE id=? AND client_id=?"
                params.extend([app_id, cid])
                conn.execute(query, params)

                log_history(
                    conn, cid, 'appareil', app_id, device_name, 'Auto-remplissage (collecteur)',
                    {'source': 'system-info-collector', 'fields_updated': len(updates)}
                )

            conn.commit()
            message = f"Appareil mis à jour (ID: {app_id})"
            action = 'updated'
        else:
            # Création
            conn.execute(
                '''INSERT INTO appareils
                   (client_id, nom_machine, nom_dns, adresse_ip, adresse_mac, marque, modele, numero_serie,
                    os, version_os, ram, cpu, stockage, antivirus, carte_graphique, ports_ouverts,
                    logiciels_installes_json, rapport_systeme_json,
                    adresse_ip_publique, operateur_ip_publique,
                    type_appareil, statut, decouvert_scan, en_ligne, derniere_synchro, date_creation, date_maj)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (cid, device_name, dns_name, ip_address, mac_address, brand, model, serial,
                 os_name, os_version, str(ram_gb) if ram_gb else '', cpu, str(disk_gb) if disk_gb else '',
                 antivirus, gpu, ports_str, software_json, system_report_json,
                 public_ip, public_ip_isp,
                 device_type or 'PC', 'actif', 0, 1, now, now, now)
            )
            app_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

            # Même remplissage qu'en mise à jour : sur un appareil neuf, toutes
            # les cases concernées sont vides, donc toutes sont renseignées.
            deduits = champs_deduits_du_collecteur(conn, cid, system_report or data)
            if deduits:
                conn.execute(
                    'UPDATE appareils SET %s WHERE id=?'
                    % ', '.join('%s=?' % c for c in deduits),
                    list(deduits.values()) + [app_id])

            log_history(
                conn, cid, 'appareil', app_id, device_name, 'Création (collecteur système)',
                {'source': 'system-info-collector', 'mac': mac_address, 'ip': ip_address}
            )

            conn.commit()
            message = f"Nouvel appareil créé (ID: {app_id})"
            action = 'created'

        # Relevé horodaté : c'est lui qui permettra de comparer cette collecte
        # aux suivantes, une fois celle-ci écrasée dans la fiche appareil.
        if _enregistrer_collecte(conn, cid, app_id, data):
            conn.commit()

        # Clés de récupération BitLocker, chiffrées et hors du rapport.
        cles_enregistrees = 0
        if cles_bitlocker:
            try:
                cles_enregistrees = _enregistrer_cles_bitlocker(
                    conn, cid, app_id, cles_bitlocker)
                conn.commit()
                if cles_enregistrees:
                    log_history(conn, cid, 'appareil', app_id, device_name,
                                'Clés de récupération BitLocker relevées',
                                {'volumes': cles_enregistrees})
                    conn.commit()
            except Exception:
                logger.exception('Clés BitLocker non enregistrées (appareil %s)', app_id)

        # Licences dont la clé complète a été récupérée
        lic_ajoutees = 0
        if collected_licenses:
            lic_ajoutees = _sync_licences_from_collector(
                conn, cid, app_id, collected_licenses)
            if lic_ajoutees:
                log_history(
                    conn, cid, 'appareil', app_id, device_name,
                    'Licences relevées (collecteur)', {'ajoutees': lic_ajoutees})
                conn.commit()

        # Écrans, imprimantes et périphériques USB détectés → inventaire
        peripherals_created = _sync_collector_peripherals(
            conn, cid, app_id, data.get('monitors') or [], data.get('printers') or [],
            usb_devices)
        conn.commit()

        conn.close()

        app.logger.info(f"Device info received: {device_name} ({action}) - MAC: {mac_address}, IP: {ip_address}")

        return jsonify({
            "status": "success",
            "action": action,
            "device_id": app_id,
            "client_id": cid,
            "message": message,
            "mac_address": mac_address,
            "ip_address": ip_address,
            "hostname": hostname,
            "peripherals_created": peripherals_created,
            "licences_ajoutees": lic_ajoutees
        }), 200

    except Exception as e:
        app.logger.exception("Error in /api/device-info")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/device-info/wifi-credentials', methods=['POST'])
def api_device_info_wifi_credentials():
    """
    Reçoit les réseaux Wi-Fi enregistrés sur un poste, remontés par un appel
    séparé de /api/device-info (voir collector_core.py:send_wifi_credentials_to_parcinfo) :
    contrairement au reste de la collecte, ces données ne transitent jamais par
    le snapshot système (rapport_systeme_json, PDF, visibles en clair) — elles
    vont directement dans la table identifiants, chiffrées.

    POST /api/device-info/wifi-credentials
    {
        "device_id": 42,
        "client_id": 3,
        "profiles": [
            {"ssid": "Bureau-Principal", "authentification": "WPA2",
             "chiffrement": "AES", "password": "..."},
            {"ssid": "Invites", "authentification": "Ouvert"}
        ]
    }
    Le mot de passe est optionnel par entrée : absent, l'identifiant existant
    n'est pas touché ; nouveau, il est chiffré avant stockage.
    """
    if not jeton_collecteur_valide():
        return jsonify({"status": "error", "message": "Jeton collecteur requis"}), 401
    try:
        data = request.json or {}
        try:
            cid = int(data.get('client_id'))
        except (TypeError, ValueError):
            return jsonify({"status": "error", "message": "client_id invalide"}), 400

        if not jeton_collecteur_valide(cid):
            return jsonify({"status": "error", "message": "Jeton collecteur requis pour ce client"}), 401

        profiles = data.get('profiles') or []
        if not isinstance(profiles, list):
            return jsonify({"status": "error", "message": "profiles doit être une liste"}), 400

        conn = get_db()
        if not conn.execute('SELECT id FROM clients WHERE id=?', (cid,)).fetchone():
            conn.close()
            return jsonify({"status": "error", "message": f"Client ID {cid} introuvable"}), 404

        crees, maj = _sync_wifi_credentials_from_collector(conn, cid, profiles)
        if crees or maj:
            device_id = data.get('device_id')
            nom_appareil = 'Collecteur'
            if device_id:
                row = conn.execute(
                    'SELECT nom_machine FROM appareils WHERE id=? AND client_id=?',
                    (device_id, cid)).fetchone()
                if row and row[0]:
                    nom_appareil = row[0]
            log_history(conn, cid, 'appareil', device_id or 0, nom_appareil,
                        'Réseaux Wi-Fi synchronisés (collecteur)',
                        {'crees': crees, 'mis_a_jour': maj})
            conn.commit()
        conn.close()

        return jsonify({"status": "success", "created": crees, "updated": maj}), 200
    except Exception as e:
        app.logger.exception("Error in /api/device-info/wifi-credentials")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/instance-info', methods=['GET'])
def api_instance_info():
    """
    Endpoint PUBLIC (aucune authentification) pour identifier cette instance.

    Utilisé par le collecteur pour étiqueter les instances trouvées lors
    d'un balayage réseau (port 3456/5010) par le nom de la machine qui les
    héberge, plutôt que par leur seule adresse IP — le TXT record mDNS
    (_register_mdns ci-dessous) expose déjà les mêmes informations sans
    authentification à quiconque écoute sur le réseau, mais mDNS ne
    traverse pas le réseau "bridge" de Docker : une instance conteneurisée
    (typiquement exposée sur le port 5010, voir docker-compose.yml) n'est
    alors visible qu'via un balayage de ports direct, qui a besoin de cette
    route pour obtenir un nom lisible plutôt que l'IP nue.
    """
    try:
        hostname = socket.gethostname()
    except Exception:
        hostname = ''
    return jsonify({
        'hostname': hostname,
        'version': APP_VERSION,
        'docker': bool(os.environ.get('RUNNING_IN_DOCKER')),
    }), 200


@app.route('/api/clients-public', methods=['GET'])
def api_clients_public():
    """
    Endpoint PUBLIC pour récupérer la liste des clients (sans authentification).

    Utilisé par le collecteur système GUI pour afficher les clients disponibles.

    Retourne :
    [
      {"id": 1, "nom": "Mon Entreprise"},
      {"id": 2, "nom": "Client A"},
      ...
    ]
    """
    # Sans jeton configuré, la liste reste ouverte comme auparavant. Dès qu'un
    # jeton existe, elle est protégée : les noms de vos clients n'ont pas à être
    # lisibles par quiconque atteint le serveur.
    if not jeton_collecteur_valide():
        return jsonify({"status": "error", "message": "Jeton collecteur requis"}), 401
    try:
        conn = get_db()
        clients = [
            {"id": row[0], "nom": row[1]}
            for row in conn.execute('SELECT id, nom FROM clients ORDER BY nom').fetchall()
        ]
        conn.close()

        if not clients:
            # S'il n'y a aucun client, créer un défaut
            conn = get_db()
            conn.execute(
                "INSERT INTO clients (nom, date_creation, date_maj) VALUES (?, ?, ?)",
                ("Client par défaut", _utcnow().isoformat(), _utcnow().isoformat())
            )
            conn.commit()
            new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.close()
            clients = [{"id": new_id, "nom": "Client par défaut"}]

        # Suggestion : si l'une des adresses MAC fournies correspond à un
        # appareil déjà enregistré, le collecteur peut présélectionner son
        # client au lieu de laisser l'utilisateur deviner. Plusieurs adresses
        # possibles (paramètre répété) : une machine avec plusieurs cartes
        # (VPN, Hyper-V/WSL, VirtualBox…) n'a pas de « bonne » adresse
        # évidente côté collecteur — voir get_all_mac_addresses() dans
        # collector_core.py — donc on compare à toutes. On ne divulgue que
        # l'identifiant et le nom du client — déjà publics sur cet endpoint —
        # et rien n'est révélé quand la machine est inconnue.
        macs = [m.strip().upper() for m in request.args.getlist('mac') if m.strip()]
        suggestion = None
        if macs:
            conn = get_db()
            placeholders = ','.join('?' * len(macs))
            row = conn.execute(
                'SELECT c.id, c.nom FROM appareils a JOIN clients c ON c.id = a.client_id'
                f' WHERE UPPER(a.adresse_mac) IN ({placeholders}) ORDER BY a.date_maj DESC LIMIT 1',
                macs).fetchone()
            conn.close()
            if row:
                suggestion = {'id': row[0], 'nom': row[1]}

        if suggestion is not None:
            # Réponse enrichie : les collecteurs antérieurs continuent de lire
            # une liste, les nouveaux lisent l'objet.
            return jsonify({'clients': clients, 'suggested_client': suggestion}), 200

        return jsonify(clients), 200

    except Exception as e:
        app.logger.exception("Error in /api/clients-public")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/device-info/upload-report', methods=['POST'])
def api_device_info_upload_report():
    """
    Endpoint pour uploader le rapport HTML complet d'un appareil.

    Paramètres (multipart form-data):
    - device_id: ID de l'appareil (obligatoire)
    - client_id: ID du client (obligatoire)
    - report: fichier HTML (obligatoire)
    """
    if not jeton_collecteur_valide():
        return jsonify({"status": "error", "message": "Jeton collecteur requis"}), 401
    try:
        device_id = request.form.get('device_id')
        client_id = request.form.get('client_id')
        report_file = request.files.get('report')

        if not device_id or not client_id or not report_file:
            return jsonify({"status": "error", "message": "Missing parameters: device_id, client_id, report"}), 400

        try:
            device_id = int(device_id)
            client_id = int(client_id)
        except (ValueError, TypeError):
            return jsonify({"status": "error", "message": "Invalid device_id or client_id"}), 400

        if not jeton_collecteur_valide(client_id):
            return jsonify({"status": "error", "message": "Jeton collecteur requis pour ce client"}), 401

        # Ce point d'entrée est ouvert aux collecteurs : il ne doit accepter que
        # ce que ceux-ci produisent, un rapport PDF ou son repli HTML.
        ok_rapport, motif_rapport = verifier_fichier(report_file, {'pdf', 'html', 'htm'})
        if not ok_rapport:
            return jsonify({"status": "error", "message": motif_rapport}), 400

        conn = get_db()

        # Vérifier que l'appareil existe et appartient au client
        appareil = conn.execute(
            'SELECT id, nom_machine FROM appareils WHERE id=? AND client_id=?',
            (device_id, client_id)
        ).fetchone()

        if not appareil:
            conn.close()
            return jsonify({"status": "error", "message": "Device not found"}), 404

        # Lire le contenu du fichier
        try:
            file_content = report_file.read()
        except Exception as e:
            conn.close()
            return jsonify({"status": "error", "message": f"Error reading file: {str(e)}"}), 400

        # Écrire le fichier sur disque, comme TOUTES les autres routes d'upload.
        # Indispensable pour la synchronisation multi-machines : contenu_blob est
        # exclu de la réplication des lignes (_BLOB_SYNC_EXCLUDE), et le transfert
        # des fichiers vers Turso (_push_documents_to_turso) lit le contenu DEPUIS
        # LE DISQUE. Un rapport stocké uniquement en blob local n'était donc jamais
        # transférable : les autres machines voyaient la ligne dans la liste mais
        # obtenaient "Fichier introuvable" à l'ouverture.
        # Le préfixe "app<id>_" est aussi requis par _cleanup_orphaned_files() pour
        # reconnaître le fichier comme géré par l'application.
        safe = secure_filename(report_file.filename or 'rapport.pdf') or 'rapport.pdf'
        unique = f"app{device_id}_{int(time.time())}_{safe}"
        try:
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            with open(os.path.join(UPLOAD_FOLDER, unique), 'wb') as _f:
                _f.write(file_content)
        except Exception as e:
            conn.close()
            app.logger.exception("Error saving report file to disk")
            return jsonify({"status": "error", "message": f"Error saving file: {str(e)}"}), 500

        # Insérer le document dans la table documents_appareils
        # (contenu_blob conservé en plus du fichier : l'aperçu reste instantané sur
        # cette machine, et le fichier disque alimente la synchronisation)
        now = _utcnow().isoformat()
        try:
            conn.execute('''
                INSERT INTO documents_appareils
                (appareil_id, client_id, nom, description, type_doc, nom_fichier, taille, contenu_blob, date_upload)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                device_id,
                client_id,
                f"Rapport Système - {now}",
                "Rapport PDF/HTML collecté par système-info-collector",
                "rapport_system",
                unique,
                len(file_content),
                file_content,
                now
            ))
        except Exception as e:
            conn.close()
            app.logger.exception("Error inserting document")
            return jsonify({"status": "error", "message": f"Error inserting document: {str(e)}"}), 500

        conn.commit()
        doc_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.close()

        app.logger.info(f"Report uploaded for device {device_id} (client {client_id}) - doc_id: {doc_id}")

        return jsonify({
            "status": "success",
            "message": "Report uploaded successfully",
            "device_id": device_id,
            "document_id": doc_id,
            "hostname": appareil[1]
        }), 200

    except Exception as e:
        app.logger.exception("Error in /api/device-info/upload-report")
        return jsonify({"status": "error", "message": str(e)}), 500


def _send_collector_bundle(entry_script, archive_name):
    """Renvoie une archive ZIP contenant le script d'entrée et collector_core.py.

    Les collecteurs ne sont plus des fichiers autonomes : toute la logique de
    collecte vit dans collector_core.py, partagé par les deux. Servir le seul
    script d'entrée livrerait un collecteur qui échoue à l'import.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files = [entry_script, 'collector_core.py']

    missing = [f for f in files if not os.path.exists(os.path.join(base_dir, f))]
    if missing:
        return jsonify({"error": f"Collecteur incomplet, fichier(s) manquant(s): {', '.join(missing)}"}), 404

    try:
        import zipfile
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
            for name in files:
                archive.write(os.path.join(base_dir, name), arcname=name)
            archive.writestr('LISEZMOI.txt', (
                "Collecteur systeme ParcInfo\r\n"
                "==========================\r\n\r\n"
                "Decompressez les deux fichiers dans le MEME dossier, puis lancez :\r\n"
                f"    python {entry_script} --server <URL> --client-id <ID>\r\n\r\n"
                "collector_core.py contient la logique de collecte : il doit rester\r\n"
                "a cote du script principal.\r\n\r\n"
                "Lancez de preference en administrateur : sans elevation, le SMART\r\n"
                "detaille, le TPM, BitLocker et la cle OEM ne sont pas lisibles.\r\n"
            ))
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=archive_name,
                         mimetype='application/zip')
    except Exception as e:
        app.logger.exception(f"Error building collector bundle for {entry_script}")
        return jsonify({"error": str(e)}), 500


@app.route('/download/system-info-collector', methods=['GET'])
def download_collector():
    """
    Endpoint pour télécharger le collecteur CLI (ligne de commande).

    Utilisation :
    - /download/system-info-collector → archive ZIP (script CLI + collector_core.py)
    """
    return _send_collector_bundle('system-info-collector.py', 'system-info-collector.zip')


@app.route('/download/system-info-collector-gui', methods=['GET'])
def download_collector_gui():
    """
    Endpoint pour télécharger le collecteur GUI (interface graphique).

    Utilisation :
    - /download/system-info-collector-gui → archive ZIP (script GUI + collector_core.py)
    """
    return _send_collector_bundle('system-info-collector-gui.py', 'system-info-collector-gui.zip')


# --- PERIPHERIQUES -----------------------------------------------------------

# Colonnes triables pour l'inventaire périphériques
_PERIPH_SORT_COLS = {
    'cat':      'p.categorie, p.marque, p.modele',
    'marque':   'p.marque, p.modele',
    'app':      'nom_machine_lié, p.marque',
    'user':     'p.utilisateur_nom, p.marque',
    'loc':      'p.localisation, p.marque',
    'garantie': "CASE WHEN p.date_fin_garantie='' OR p.date_fin_garantie IS NULL THEN '9999-99-99' ELSE p.date_fin_garantie END",
    'statut':   'p.statut, p.marque',
}

@app.route('/peripheriques')
@login_required
def liste_peripheriques():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    page        = request.args.get('page', 1, type=int)
    filtre_cats = request.args.getlist('cat')
    sort_col    = request.args.get('sort', 'cat')
    sort_dir    = request.args.get('dir', 'asc')
    filtre_stat = request.args.get('statut', '')
    filtre_app  = request.args.get('appareil', '')
    q = ("SELECT p.*,"
         " u.prenom || ' ' || u.nom as utilisateur_nom,"
         " s.nom as service_nom, s.couleur as service_couleur,"
         " (SELECT COUNT(*) FROM contrats_peripheriques cp JOIN contrats ct"
         "  ON cp.contrat_id=ct.id WHERE cp.peripherique_id=p.id AND ct.client_id=p.client_id) as nb_contrats,"
         " (SELECT COUNT(*) FROM documents_peripheriques dp WHERE dp.peripherique_id=p.id) as nb_docs"
         " FROM peripheriques p"
         " LEFT JOIN utilisateurs u ON p.utilisateur_id = u.id"
         " LEFT JOIN services s ON u.service_id = s.id"
         " WHERE p.client_id=?")
    params = [cid]
    if filtre_cats:
        ph = ','.join('?' * len(filtre_cats))
        q += f' AND p.categorie IN ({ph})'
        params.extend(filtre_cats)
    if filtre_stat: q += ' AND p.statut=?';      params.append(filtre_stat)
    if filtre_app:
        q += ' AND p.id IN (SELECT peripherique_id FROM peripheriques_appareils WHERE appareil_id=?)'
        params.append(int(filtre_app))
    order_expr_p = _PERIPH_SORT_COLS.get(sort_col, 'p.categorie, p.marque, p.modele')
    dir_p = 'DESC' if sort_dir == 'desc' else 'ASC'
    q += f' ORDER BY {order_expr_p} {dir_p}'
    rows, pagination = paginate(q, tuple(params), page)
    periph = [fmt_garantie_periph(row_to_dict(r)) for r in rows]
    # Enrichir avec les appareils liés (via pivot)
    if periph:
        pid_list = ','.join(str(p['id']) for p in periph)
        conn2 = get_db()
        app_rows = conn2.execute(
            f"SELECT pa.peripherique_id, a.id, a.nom_machine, a.adresse_ip, a.type_appareil"
            f" FROM peripheriques_appareils pa"
            f" JOIN appareils a ON pa.appareil_id = a.id"
            f" WHERE pa.peripherique_id IN ({pid_list})").fetchall()
        conn2.close()
        app_map = {}
        for r in app_rows:
            app_map.setdefault(r[0], []).append({'id': r[1], 'nom_machine': r[2], 'adresse_ip': r[3], 'type_appareil': r[4]})
        for p in periph:
            p['appareils_lies'] = app_map.get(p['id'], [])
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,adresse_ip,type_appareil FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    cats_utilisees = [r[0] for r in conn.execute(
        'SELECT DISTINCT categorie FROM peripheriques WHERE client_id=? ORDER BY categorie', (cid,)).fetchall()]
    stats = {
        'total': conn.execute('SELECT COUNT(*) FROM peripheriques WHERE client_id=?', (cid,)).fetchone()[0],
        'actif': conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='actif'", (cid,)).fetchone()[0],
        'stock': conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='stock'", (cid,)).fetchone()[0],
        'hors_service': conn.execute("SELECT COUNT(*) FROM peripheriques WHERE client_id=? AND statut='hors_service'", (cid,)).fetchone()[0],
    }
    conn.close()
    return render_template('peripheriques.html', peripheriques=periph, appareils=appareils,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_peripheriques'), cats_utilisees=cats_utilisees,
                           filtre_cats=filtre_cats, filtre_stat=filtre_stat, filtre_app=filtre_app,
                           sort_col=sort_col, sort_dir=sort_dir,
                           stats=stats, pagination=pagination)

@app.route('/peripherique/nouveau', methods=['GET','POST'])
@login_required
def nouveau_peripherique():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_peripheriques'))
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,adresse_ip,type_appareil FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    utilisateurs = [row_to_dict(r) for r in conn.execute(
        "SELECT id,prenom,nom FROM utilisateurs WHERE client_id=? AND statut='actif' ORDER BY nom", (cid,)).fetchall()]
    conn.close()
    if request.method == 'POST':
        now = datetime.now().isoformat()
        conn = get_db()
        vals = _extract_periph(cid, request.form)
        conn.execute(("INSERT INTO peripheriques"
            " (client_id,utilisateur_id,categorie,marque,modele,numero_serie,description,"
            "localisation,statut,date_achat,duree_garantie,date_fin_garantie,fournisseur,prix_achat,"
            "numero_commande,notes,date_creation,date_maj)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"), vals + (now, now))
        new_pid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        # Liens appareils N:N
        app_ids = request.form.getlist('appareil_ids')
        for aid in app_ids:
            try:
                conn.execute("INSERT OR IGNORE INTO peripheriques_appareils (peripherique_id, appareil_id) VALUES (?,?)",
                             (new_pid, int(aid)))
            except Exception:
                pass
        nom_p = (request.form.get('marque','') + ' ' + request.form.get('modele','')).strip() or 'Nouveau périphérique'
        log_history(conn, cid, 'peripherique', new_pid, nom_p, 'Création')
        conn.commit(); conn.close()
        flash('Peripherique ajoute', 'success')
        return redirect(url_for('liste_peripheriques'))
    pre_app = request.args.get('appareil_id', '')
    return render_template('form_peripherique.html', peripherique=None, action='Ajouter',
                           appareils=appareils, utilisateurs=utilisateurs,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_peripheriques'), pre_appareil_id=pre_app,
                           linked_app_ids=[int(pre_app)] if pre_app else [])

@app.route('/peripherique/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_peripherique(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_peripheriques'))
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,adresse_ip,type_appareil FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    utilisateurs = [row_to_dict(r) for r in conn.execute(
        "SELECT id,prenom,nom FROM utilisateurs WHERE client_id=? AND statut='actif' ORDER BY nom", (cid,)).fetchall()]
    if request.method == 'POST':
        now = datetime.now().isoformat()
        _old = row_to_dict(conn.execute('SELECT * FROM peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        vals = _extract_periph(cid, request.form)
        conn.execute(("UPDATE peripheriques SET"
            " client_id=?,utilisateur_id=?,categorie=?,marque=?,modele=?,numero_serie=?,"
            "description=?,localisation=?,statut=?,date_achat=?,duree_garantie=?,date_fin_garantie=?,"
            "fournisseur=?,prix_achat=?,numero_commande=?,notes=?,date_maj=? WHERE id=? AND client_id=?"),
            vals + (now, id, cid))
        # Mettre à jour les liens appareils N:N
        app_ids = request.form.getlist('appareil_ids')
        conn.execute("DELETE FROM peripheriques_appareils WHERE peripherique_id=?", (id,))
        for aid in app_ids:
            try:
                conn.execute("INSERT OR IGNORE INTO peripheriques_appareils (peripherique_id, appareil_id) VALUES (?,?)",
                             (id, int(aid)))
            except Exception:
                pass
        nom = (request.form.get('marque','') + ' ' + request.form.get('modele','')).strip() or f'Périphérique #{id}'
        _cols_p = _ENTITE_COLS['peripherique']
        _details_p = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_p},
                                  {k: str(request.form.get(k,'') or '') for k in _cols_p})
        log_history(conn, cid, 'peripherique', id, nom, 'Modification', _details_p)
        conn.commit(); conn.close()
        flash('Peripherique mis a jour', 'success')
        return redirect(url_for('liste_peripheriques'))
    p = fmt_garantie_periph(row_to_dict(
        conn.execute('SELECT * FROM peripheriques WHERE id=? AND client_id=?', (id, cid)).fetchone() or {}))
    docs_per = [row_to_dict(r) for r in conn.execute(
        'SELECT id, peripherique_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_peripheriques WHERE peripherique_id=? ORDER BY date_upload DESC', (id,)).fetchall()]
    for d in docs_per:
        d['taille_fmt'] = human_size(d.get('taille', 0))
    # Appareils déjà liés à ce périphérique
    linked_app_ids = [r[0] for r in conn.execute(
        "SELECT appareil_id FROM peripheriques_appareils WHERE peripherique_id=?", (id,)).fetchall()]

    # Fetch related interventions
    interventions = [fmt_intervention(row_to_dict(r)) for r in conn.execute(
        'SELECT i.* FROM interventions i JOIN interventions_peripheriques ip ON i.id=ip.intervention_id '
        'WHERE ip.peripherique_id=? AND i.statut != ? ORDER BY i.date_intervention DESC LIMIT 10',
        (id, 'archivee')).fetchall()]

    # UNION avec baie_prises_murales — voir le même commentaire côté fiche
    # appareil (fonction editer_appareil).
    ports_baie = [row_to_dict(r) for r in conn.execute(
        '''SELECT bp.numero, bs.id AS slot_id, bs.baie_nom, bs.position,
                  bs.nom_custom, bs.type_equipement, 'port' AS origine
           FROM baie_slot_ports bp JOIN baie_slots bs ON bp.slot_id=bs.id
           WHERE bp.peripherique_id=? AND bs.client_id=?
           UNION ALL
           SELECT pm.numero, bs.id AS slot_id, bs.baie_nom, bs.position,
                  bs.nom_custom, bs.type_equipement, 'prise_murale' AS origine
           FROM baie_prises_murales pm JOIN baie_slots bs ON pm.slot_id=bs.id
           WHERE pm.peripherique_id=? AND bs.client_id=?
           ORDER BY baie_nom, position, numero''', (id, cid, id, cid)).fetchall()]

    conn.close()
    return render_template('form_peripherique.html', peripherique=p, documents=docs_per, action='Modifier',
                           appareils=appareils, utilisateurs=utilisateurs,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_peripheriques'), pre_appareil_id='',
                           linked_app_ids=linked_app_ids, interventions=interventions,
                           ports_baie=ports_baie)

@app.route('/peripherique/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_peripherique(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_peripheriques'))
    cid = get_client_id()
    conn = get_db()
    conn.execute('PRAGMA foreign_keys = ON')
    p = row_to_dict(conn.execute('SELECT marque,modele FROM peripheriques WHERE id=?',(id,)).fetchone() or {})
    nom_p = (p.get('marque','') + ' ' + p.get('modele','')).strip() or '?'
    log_history(conn, cid, 'peripherique', id, nom_p, 'Suppression')
    # Tables sans FK déclarée (ajoutées via ALTER TABLE) — non couvertes par le
    # PRAGMA ci-dessus, nettoyage manuel nécessaire pour éviter les orphelins.
    conn.execute('UPDATE identifiants SET peripherique_id=NULL WHERE peripherique_id=? AND client_id=?', (id, cid))
    conn.execute('UPDATE baie_slots SET peripherique_id=NULL WHERE peripherique_id=? AND client_id=?', (id, cid))
    conn.execute('''UPDATE baie_slot_ports SET peripherique_id=NULL WHERE peripherique_id=? AND slot_id IN
        (SELECT id FROM baie_slots WHERE client_id=?)''', (id, cid))
    conn.execute('''UPDATE baie_prises_murales SET peripherique_id=NULL WHERE peripherique_id=? AND slot_id IN
        (SELECT id FROM baie_slots WHERE client_id=?)''', (id, cid))
    conn.execute('DELETE FROM peripheriques WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Peripherique supprime', 'info')
    return redirect(url_for('liste_peripheriques'))

@app.route('/api/peripheriques/appareil/<int:app_id>')
@login_required
def api_periph_appareil(app_id):
    cid = get_client_id()
    conn = get_db()
    rows = [row_to_dict(r) for r in conn.execute(
        ("SELECT p.*, u.prenom || ' ' || u.nom as utilisateur_nom"
         " FROM peripheriques p"
         " JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id"
         " LEFT JOIN utilisateurs u ON p.utilisateur_id = u.id"
         " WHERE pa.appareil_id=? AND p.client_id=? ORDER BY p.categorie"),
        (app_id, cid)).fetchall()]
    conn.close()
    return jsonify(rows)

def _extract_periph(cid, f):
    user_id = int(f.get('utilisateur_id') or 0) or None
    prix = None
    try:
        prix = float(f['prix_achat']) if f.get('prix_achat') else None
    except:
        pass
    duree = 0
    try:
        duree = int(f['duree_garantie']) if f.get('duree_garantie') else 0
    except:
        pass
    return (cid, user_id,
            f.get('categorie',''), f.get('marque',''), f.get('modele',''),
            f.get('numero_serie',''), f.get('description',''), f.get('localisation',''),
            f.get('statut','actif'), f.get('date_achat',''), duree,
            f.get('date_fin_garantie',''), f.get('fournisseur',''), prix,
            f.get('numero_commande',''), f.get('notes',''))



# --- CONTRATS & ABONNEMENTS --------------------------------------------------

PERIODICITES = ['mensuel', 'trimestriel', 'semestriel', 'annuel', 'pluriannuel', 'unique']

@app.route('/contrats')
@login_required
def liste_contrats():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    page          = request.args.get('page', 1, type=int)
    filtre_type   = request.args.get('type', '')
    filtre_stat   = request.args.get('statut', '')
    filtre_app_id = request.args.get('appareil', '')
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    q = 'SELECT * FROM contrats WHERE client_id=?'
    params = [cid]
    if filtre_type:   q += ' AND type_contrat=?'; params.append(filtre_type)
    if filtre_stat:   q += ' AND statut=?'; params.append(filtre_stat)
    if filtre_app_id:
        q = ('SELECT DISTINCT c.* FROM contrats c '
             'JOIN contrats_appareils ca ON c.id=ca.contrat_id '
             'WHERE c.client_id=? AND ca.appareil_id=?')
        params = [cid, int(filtre_app_id)]
        if filtre_type: q += ' AND c.type_contrat=?'; params.append(filtre_type)
        if filtre_stat: q += ' AND c.statut=?'; params.append(filtre_stat)
    q += ' ORDER BY date_fin, titre'
    rows, pagination = paginate(q, tuple(params), page)
    contrats = [fmt_contrat(row_to_dict(r)) for r in rows]
    # Nom de l'appareil filtré pour l'afficher dans la page
    filtre_app_nom = ''
    if filtre_app_id:
        a = conn.execute('SELECT nom_machine FROM appareils WHERE id=?', (int(filtre_app_id),)).fetchone()
        if a: filtre_app_nom = a[0]
    # Compter les elements lies pour chaque contrat
    for ct in contrats:
        ct['nb_appareils'] = conn.execute(
            'SELECT COUNT(*) FROM contrats_appareils WHERE contrat_id=?', (ct['id'],)).fetchone()[0]
        ct['nb_peripheriques'] = conn.execute(
            'SELECT COUNT(*) FROM contrats_peripheriques WHERE contrat_id=?', (ct['id'],)).fetchone()[0]
        ct['nb_docs'] = conn.execute(
            'SELECT COUNT(*) FROM documents_contrats WHERE contrat_id=?', (ct['id'],)).fetchone()[0]
    types_utilises = [r[0] for r in conn.execute(
        'SELECT DISTINCT type_contrat FROM contrats WHERE client_id=? ORDER BY type_contrat', (cid,)).fetchall()]
    stats = {
        'total':   conn.execute('SELECT COUNT(*) FROM contrats WHERE client_id=?', (cid,)).fetchone()[0],
        'actif':   conn.execute("SELECT COUNT(*) FROM contrats WHERE client_id=? AND statut='actif'", (cid,)).fetchone()[0],
        'expire':  conn.execute("SELECT COUNT(*) FROM contrats WHERE client_id=? AND statut='expire'", (cid,)).fetchone()[0],
        'resilie': conn.execute("SELECT COUNT(*) FROM contrats WHERE client_id=? AND statut='resilie'", (cid,)).fetchone()[0],
    }
    # Alertes: contrats expirant bientot
    alertes = [ct for ct in contrats if ct['expire_bientot'] or ct['expire_depasse']]
    conn.close()
    return render_template('contrats.html', contrats=contrats, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           types_contrats=get_liste_cached('types_contrats'), types_utilises=types_utilises,
                           periodicites=PERIODICITES, stats=stats, alertes=alertes,
                           filtre_type=filtre_type, filtre_stat=filtre_stat,
                           filtre_app_id=filtre_app_id, filtre_app_nom=filtre_app_nom,
                           pagination=pagination)

@app.route('/contrat/nouveau', methods=['GET','POST'])
@login_required
def nouveau_contrat():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_contrats'))
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,type_appareil,adresse_ip FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id,categorie,marque,modele,description FROM peripheriques WHERE client_id=? ORDER BY categorie,marque', (cid,)).fetchall()]
    conn.close()
    if request.method == 'POST':
        errs = validate_form([
            ('titre',       'str',   True),
            ('date_debut',  'date',  False),
            ('date_fin',    'date',  False),
            ('email_fournisseur', 'email', False),
        ], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        f = request.form; now = datetime.now().isoformat()
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""INSERT INTO contrats (client_id,titre,type_contrat,fournisseur,contact_fournisseur,
            email_fournisseur,telephone_fournisseur,numero_contrat,date_debut,date_fin,
            reconduction_auto,preavis_jours,montant_ht,periodicite,description,notes,
            statut,date_creation,date_maj) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            _extract_contrat(cid, f) + (now, now))
        cid_contrat = cur.lastrowid
        # Liaisons appareils
        for app_id in request.form.getlist('appareils_lies'):
            try:
                conn.execute('INSERT INTO contrats_appareils (contrat_id,appareil_id) VALUES (?,?)', (cid_contrat, int(app_id)))
            except: pass
        # Liaisons peripheriques
        for per_id in request.form.getlist('peripheriques_lies'):
            try:
                conn.execute('INSERT INTO contrats_peripheriques (contrat_id,peripherique_id) VALUES (?,?)', (cid_contrat, int(per_id)))
            except: pass
        new_cid2 = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_history(conn, cid, 'contrat', new_cid2, request.form.get('titre','') or 'Nouveau contrat', 'Création')
        conn.commit(); conn.close()
        flash('Contrat créé', 'success')
        return redirect(url_for('detail_contrat', id=cid_contrat))
    return render_template('form_contrat.html', contrat=None, action='Nouveau',
                           appareils=appareils, peripheriques=peripheriques,
                           appareils_lies=[], peripheriques_lies=[],
                           client=client, clients=get_clients(), client_actif_id=cid,
                           types_contrats=get_liste_cached('types_contrats'), periodicites=PERIODICITES)

@app.route('/contrat/<int:id>', methods=['GET'])
@login_required
def detail_contrat(id):
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    ct = fmt_contrat(row_to_dict(
        conn.execute('SELECT * FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {}))
    appareils_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT a.* FROM appareils a JOIN contrats_appareils ca ON a.id=ca.appareil_id WHERE ca.contrat_id=?', (id,)).fetchall()]
    periph_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT p.* FROM peripheriques p JOIN contrats_peripheriques cp ON p.id=cp.peripherique_id WHERE cp.contrat_id=?', (id,)).fetchall()]
    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, contrat_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload, sync_status FROM documents_contrats WHERE contrat_id=? ORDER BY date_upload DESC', (id,)).fetchall()]
    for d in docs: d['taille_fmt'] = human_size(d.get('taille', 0))

    # Fetch related interventions
    interventions = [fmt_intervention(row_to_dict(r)) for r in conn.execute(
        'SELECT * FROM interventions WHERE contrat_id=? AND statut != ? ORDER BY date_intervention DESC LIMIT 10',
        (id, 'archivee')).fetchall()]

    conn.close()
    return render_template('detail_contrat.html', contrat=ct, appareils_lies=appareils_lies,
                           periph_lies=periph_lies, docs=docs, interventions=interventions,
                           client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/contrat/<int:id>/editer', methods=['GET','POST'])
@login_required
def editer_contrat(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_contrats'))
    cid = get_client_id()
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id,nom_machine,type_appareil,adresse_ip FROM appareils WHERE client_id=? ORDER BY nom_machine', (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id,categorie,marque,modele,description FROM peripheriques WHERE client_id=? ORDER BY categorie,marque', (cid,)).fetchall()]
    if request.method == 'POST':
        errs = validate_form([
            ('titre',       'str',   True),
            ('date_debut',  'date',  False),
            ('date_fin',    'date',  False),
            ('email_fournisseur', 'email', False),
        ], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        f = request.form; now = datetime.now().isoformat()
        _old = row_to_dict(conn.execute('SELECT * FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
        conn.execute("""UPDATE contrats SET titre=?,type_contrat=?,fournisseur=?,contact_fournisseur=?,
            email_fournisseur=?,telephone_fournisseur=?,numero_contrat=?,date_debut=?,date_fin=?,
            reconduction_auto=?,preavis_jours=?,montant_ht=?,periodicite=?,description=?,notes=?,
            statut=?,date_maj=? WHERE id=? AND client_id=?""",
            _extract_contrat(cid, f)[1:] + (now, id, cid))
        # Reset liaisons
        conn.execute('DELETE FROM contrats_appareils WHERE contrat_id=?', (id,))
        conn.execute('DELETE FROM contrats_peripheriques WHERE contrat_id=?', (id,))
        for app_id in request.form.getlist('appareils_lies'):
            try: conn.execute('INSERT INTO contrats_appareils (contrat_id,appareil_id) VALUES (?,?)', (id, int(app_id)))
            except: pass
        for per_id in request.form.getlist('peripheriques_lies'):
            try: conn.execute('INSERT INTO contrats_peripheriques (contrat_id,peripherique_id) VALUES (?,?)', (id, int(per_id)))
            except: pass
        _cols_c = _ENTITE_COLS['contrat']
        _details_c = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_c},
                                  {k: str(f.get(k,'') or '') for k in _cols_c})
        log_history(conn, cid, 'contrat', id, f.get('titre','') or f'Contrat #{id}', 'Modification', _details_c)
        conn.commit(); conn.close()
        flash('Contrat mis à jour', 'success')
        return redirect(url_for('detail_contrat', id=id))
    ct = fmt_contrat(row_to_dict(
        conn.execute('SELECT * FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {}))
    appareils_lies = [r[0] for r in conn.execute(
        'SELECT appareil_id FROM contrats_appareils WHERE contrat_id=?', (id,)).fetchall()]
    periph_lies = [r[0] for r in conn.execute(
        'SELECT peripherique_id FROM contrats_peripheriques WHERE contrat_id=?', (id,)).fetchall()]
    conn.close()
    return render_template('form_contrat.html', contrat=ct, action='Modifier',
                           appareils=appareils, peripheriques=peripheriques,
                           appareils_lies=appareils_lies, peripheriques_lies=periph_lies,
                           client=client, clients=get_clients(), client_actif_id=cid,
                           types_contrats=get_liste_cached('types_contrats'), periodicites=PERIODICITES)

@app.route('/contrat/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_contrat(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_contrats'))
    cid = get_client_id()
    conn = get_db()
    conn.execute('PRAGMA foreign_keys = ON')
    c = row_to_dict(conn.execute('SELECT titre FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    log_history(conn, cid, 'contrat', id, c.get('titre','?'), 'Suppression')
    # av_contrat_id/edr_contrat_id/rmm_contrat_id ont été ajoutées via ALTER TABLE
    # (SQLite ne permet pas d'y déclarer de FK) — nettoyage manuel nécessaire.
    conn.execute(
        'UPDATE appareils SET av_contrat_id=NULL WHERE av_contrat_id=?', (id,))
    conn.execute(
        'UPDATE appareils SET edr_contrat_id=NULL WHERE edr_contrat_id=?', (id,))
    conn.execute(
        'UPDATE appareils SET rmm_contrat_id=NULL WHERE rmm_contrat_id=?', (id,))
    conn.execute('DELETE FROM contrats WHERE id=? AND client_id=?', (id, cid))
    conn.commit(); conn.close()
    flash('Contrat supprimé', 'info')
    return redirect(url_for('liste_contrats'))

@app.route('/contrat/<int:id>/document/upload', methods=['POST'])
@login_required
def upload_doc_contrat(id):
    cid = get_client_id()
    if 'fichier' not in request.files:
        return redirect(url_for('detail_contrat', id=id))
    f = request.files['fichier']
    ok, motif = verifier_fichier(f)
    if not ok:
        flash(motif, 'danger')
        return redirect(url_for('detail_contrat', id=id))
    safe = secure_filename(f.filename)
    if not safe:
        flash('Nom de fichier invalide', 'danger')
        return redirect(url_for('detail_contrat', id=id))
    unique = f"ctr{id}_{int(time.time())}_{safe}"
    save_path = os.path.join(UPLOAD_FOLDER, unique)
    logger.info(f"Upload document contrat #{id}: saving to {save_path}")
    try:
        f.save(save_path)
        taille = os.path.getsize(save_path)
        logger.info(f"Upload document contrat #{id}: saved {taille} bytes")
    except Exception as e:
        logger.exception(f"Upload document contrat #{id}: save FAILED → {save_path}")
        flash(f'Erreur lors de la sauvegarde du fichier : {e}', 'danger')
        return redirect(url_for('detail_contrat', id=id))

    nom = request.form.get('nom','') or f.filename
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute('INSERT INTO documents_contrats (contrat_id,client_id,nom,description,type_doc,nom_fichier,taille,date_upload,contenu_blob,sync_status,date_sync) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                 (id, cid, nom, request.form.get('description',''), request.form.get('type_doc',''), unique, taille, now, None, 'local', ''))

    # Log document upload
    ctr_title = conn.execute('SELECT titre FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone()
    ctr_name = ctr_title[0] if ctr_title else f'Contrat #{id}'
    log_history(conn, cid, 'contrat', id, ctr_name, 'Ajout de document',
                _diff_json({}, {'nom': nom, 'fichier': unique, 'type_doc': request.form.get('type_doc','')}))

    conn.commit(); conn.close()
    flash(f'Document ajouté', 'success')
    return redirect(url_for('detail_contrat', id=id))

@app.route('/contrat/document/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_doc_contrat(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_contrats'))
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute('SELECT id, contrat_id, client_id, nom, nom_fichier, contenu_blob FROM documents_contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    ctr_id = doc.get('contrat_id', 0)
    if doc:
        conn.execute('DELETE FROM documents_contrats WHERE id=?', (id,))

        # Log document deletion
        ctr_title = conn.execute('SELECT titre FROM contrats WHERE id=? AND client_id=?', (ctr_id, cid)).fetchone()
        ctr_name = ctr_title[0] if ctr_title else f'Contrat #{ctr_id}'
        log_history(conn, cid, 'contrat', ctr_id, ctr_name, 'Suppression de document',
                    _diff_json({'nom': doc.get('nom', ''), 'fichier': doc.get('nom_fichier', '')}, {}))

        conn.commit()
        try: os.remove(os.path.join(UPLOAD_FOLDER, doc['nom_fichier']))
        except: pass
    conn.close()
    return redirect(url_for('detail_contrat', id=ctr_id))

@app.route('/contrat/document/<int:id>/apercu')
@login_required
def apercu_doc_contrat(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute('SELECT id, contrat_id, client_id, nom, nom_fichier, contenu_blob FROM documents_contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not doc: return 'Not found', 404

    # Préférer servir depuis BLOB si disponible (synced)
    if doc.get('contenu_blob'):
        return send_file(
            io.BytesIO(doc['contenu_blob']),
            as_attachment=False,
            download_name=doc.get('nom_fichier', 'document')
        )

    # Fallback: servir depuis fichier local
    try:
        return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=False)
    except Exception:
        return f"Fichier introuvable : {doc.get('nom_fichier', '?')}", 404

@app.route('/contrat/document/<int:id>/telecharger')
@login_required
def telecharger_doc_contrat(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute('SELECT id, contrat_id, client_id, nom, nom_fichier, contenu_blob FROM documents_contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not doc: return 'Not found', 404

    # Préférer servir depuis BLOB si disponible (synced)
    if doc.get('contenu_blob'):
        return send_file(
            io.BytesIO(doc['contenu_blob']),
            as_attachment=True,
            download_name=doc['nom']
        )

    # Fallback: servir depuis fichier local
    return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=True, download_name=doc['nom'])

def _contrats_appareil(conn, app_id, cid):
    """Tous les contrats liés à un appareil, quel que soit le mécanisme de
    liaison : la table pivot contrats_appareils (rattachement général) ET
    les colonnes dédiées av_contrat_id/edr_contrat_id/rmm_contrat_id
    (contrat spécifique à l'antivirus/EDR/RMM) — ces dernières n'étaient
    jusqu'ici jamais reprises ici, un contrat antivirus pouvait donc être
    invisible dans « Contrats liés » tant qu'il n'était pas *aussi* ajouté
    manuellement au pivot. Chaque contrat ressort avec la liste des rôles
    par lesquels il est rattaché (peut en cumuler plusieurs)."""
    entrees = {}
    for row in conn.execute(
            'SELECT c.* FROM contrats c JOIN contrats_appareils ca ON c.id=ca.contrat_id '
            'WHERE ca.appareil_id=? AND c.client_id=?', (app_id, cid)).fetchall():
        d = row_to_dict(row)
        entrees[d['id']] = {'contrat': d, 'roles': ['Général']}

    appareil = row_to_dict(conn.execute(
        'SELECT av_contrat_id, edr_contrat_id, rmm_contrat_id FROM appareils WHERE id=? AND client_id=?',
        (app_id, cid)).fetchone() or {})
    for colonne, role in (('av_contrat_id', 'Antivirus'), ('edr_contrat_id', 'EDR'), ('rmm_contrat_id', 'RMM')):
        contrat_id = appareil.get(colonne)
        if not contrat_id:
            continue
        if contrat_id in entrees:
            entrees[contrat_id]['roles'].append(role)
        else:
            row = conn.execute('SELECT * FROM contrats WHERE id=? AND client_id=?', (contrat_id, cid)).fetchone()
            if row:
                entrees[contrat_id] = {'contrat': row_to_dict(row), 'roles': [role]}

    resultat = []
    for entree in entrees.values():
        d = fmt_contrat(entree['contrat'])
        d['roles'] = entree['roles']
        resultat.append(d)
    resultat.sort(key=lambda d: d.get('titre') or '')
    return resultat


@app.route('/api/contrats/appareil/<int:app_id>')
@login_required
def api_contrats_appareil(app_id):
    cid = get_client_id()
    conn = get_db()
    resultat = _contrats_appareil(conn, app_id, cid)
    conn.close()
    return jsonify(resultat)


def _identifiants_appareil(conn, app_id, cid):
    """Identifiants explicitement rattachés à cet appareil (identifiants.appareil_id),
    avec détection d'un éventuel conflit face aux champs rapides de la fiche
    (user_login/user_password, admin_login/admin_password) : deux silos de
    credentials pour la même machine, sans rapprochement jusqu'ici (audit du
    2026-08-24). Le mot de passe n'est jamais renvoyé ici — seul un
    booléen/label de conflit l'est ; la valeur en clair reste réservée au
    déchiffrement à la demande (/api/identifiant/<id>/mdp), comme partout
    ailleurs dans l'app."""
    appareil = row_to_dict(conn.execute(
        'SELECT user_login, user_password, admin_login, admin_password '
        'FROM appareils WHERE id=? AND client_id=?', (app_id, cid)).fetchone() or {})
    crypto = _get_crypto_shared()
    resultat = []
    for row in conn.execute(
            'SELECT * FROM identifiants WHERE appareil_id=? AND client_id=? ORDER BY categorie, nom',
            (app_id, cid)).fetchall():
        d = row_to_dict(row)
        mdp_dechiffre = crypto.decrypt(d['mot_de_passe']) if d.get('mot_de_passe') else ''
        conflit = None
        for prefixe, label in (('user', 'Login utilisateur'), ('admin', 'Login administrateur')):
            login_rapide = (appareil.get(f'{prefixe}_login') or '').strip()
            pwd_rapide = appareil.get(f'{prefixe}_password') or ''
            if login_rapide and d.get('login') and login_rapide.lower() == d['login'].strip().lower():
                if pwd_rapide and mdp_dechiffre and pwd_rapide != mdp_dechiffre:
                    conflit = label
                break
        resultat.append({
            'id': d['id'], 'nom': d['nom'], 'categorie': d['categorie'],
            'login': d['login'], 'conflit': conflit,
        })
    return resultat


@app.route('/api/identifiants/appareil/<int:app_id>')
@login_required
def api_identifiants_appareil(app_id):
    cid = get_client_id()
    conn = get_db()
    resultat = _identifiants_appareil(conn, app_id, cid)
    conn.close()
    return jsonify(resultat)

@app.route('/api/contrats/peripherique/<int:per_id>')
@login_required
def api_contrats_peripherique(per_id):
    cid = get_client_id()
    conn = get_db()
    rows = [row_to_dict(r) for r in conn.execute(
        'SELECT c.* FROM contrats c JOIN contrats_peripheriques cp ON c.id=cp.contrat_id WHERE cp.peripherique_id=? AND c.client_id=?',
        (per_id, cid)).fetchall()]
    conn.close()
    return jsonify([fmt_contrat(r) for r in rows])

def _extract_contrat(cid, f):
    montant = None
    try: montant = float(f['montant_ht']) if f.get('montant_ht') else None
    except: pass
    preavis = 30
    try: preavis = int(f.get('preavis_jours') or 30)
    except: pass
    return (cid, f.get('titre',''), f.get('type_contrat',''), f.get('fournisseur',''),
            f.get('contact_fournisseur',''), f.get('email_fournisseur',''), f.get('telephone_fournisseur',''),
            f.get('numero_contrat',''), f.get('date_debut',''), f.get('date_fin',''),
            1 if f.get('reconduction_auto') else 0, preavis, montant,
            f.get('periodicite','annuel'), f.get('description',''), f.get('notes',''),
            f.get('statut','actif'))


def _generate_maintenance_series(conn, maint_id, cid, date_planifiee, recurrence, date_fin_recurrence, created_by, f, lookahead_date=None):
    """Génère les occurrences récurrentes d'une maintenance jusqu'à lookahead_date (ou date_fin_recurrence si None)"""
    if not recurrence or recurrence == '':
        return

    from datetime import datetime, timedelta

    try:
        start_date = datetime.strptime(date_planifiee, '%Y-%m-%d')
        end_date = datetime.strptime(date_fin_recurrence, '%Y-%m-%d') if date_fin_recurrence else None
    except:
        logger.warning(f"Erreur parsing dates pour récurrence: {date_planifiee}, {date_fin_recurrence}")
        return

    if not end_date or end_date <= start_date:
        return

    # Utiliser lookahead_date si fourni (sinon utiliser date_fin_recurrence)
    if lookahead_date:
        try:
            cutoff_date = datetime.strptime(lookahead_date, '%Y-%m-%d') if isinstance(lookahead_date, str) else lookahead_date
        except:
            cutoff_date = end_date
    else:
        cutoff_date = end_date

    # Calculer les dates futures selon le type de récurrence
    occurrences = []

    if recurrence == 'hebdomadaire':
        current_date = start_date + timedelta(days=7)
        while current_date <= cutoff_date:
            occurrences.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=7)
    elif recurrence == 'mensuelle':
        current_date = start_date
        while True:
            # Ajouter 1 mois
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
            if current_date > cutoff_date:
                break
            occurrences.append(current_date.strftime('%Y-%m-%d'))
    elif recurrence == 'annuelle':
        current_date = start_date
        while True:
            current_date = current_date.replace(year=current_date.year + 1)
            if current_date > cutoff_date:
                break
            occurrences.append(current_date.strftime('%Y-%m-%d'))
    else:
        return

    # Extraire appareil_id et peripherique_id
    appareil_id = None
    try: appareil_id = int(f.get('appareil_id')) if f.get('appareil_id') else None
    except: pass
    peripherique_id = None
    try: peripherique_id = int(f.get('peripherique_id')) if f.get('peripherique_id') else None
    except: pass

    # Insérer les occurrences
    for occ_date in occurrences:
        conn.execute(
            '''INSERT INTO maintenances
            (client_id, appareil_id, peripherique_id, type_maintenance, description,
             date_planifiee, date_realisee, heure_debut, heure_fin, responsable, notes,
             statut, recurrence, date_fin_recurrence, parent_id, created_by, updated_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (cid, appareil_id, peripherique_id, f.get('type_maintenance',''), f.get('description',''),
             occ_date, '', f.get('heure_debut',''), f.get('heure_fin',''), f.get('responsable',''),
             f.get('notes',''), 'programmee', recurrence, date_fin_recurrence, maint_id, created_by, created_by)
        )

    conn.commit()


def _regenerate_all_maintenance_occurrences():
    """Cron job: génère les occurrences futures pour toutes les maintenances récurrentes"""
    try:
        from datetime import datetime, timedelta
        conn = get_db()

        # Récupérer toutes les maintenances avec récurrence actives
        maintenances = conn.execute(
            '''SELECT id, client_id, date_planifiee, recurrence, date_fin_recurrence,
                      appareil_id, peripherique_id, type_maintenance, description,
                      heure_debut, heure_fin, responsable, notes
               FROM maintenances
               WHERE recurrence IS NOT NULL AND recurrence != ''
                     AND date_fin_recurrence IS NOT NULL
                     AND parent_id IS NULL
               ORDER BY id'''
        ).fetchall()

        for maint in maintenances:
            m = row_to_dict(maint)
            # Vérifier la dernière occurrence générée
            last_occ = conn.execute(
                'SELECT MAX(date_planifiee) FROM maintenances WHERE parent_id=?',
                (m['id'],)
            ).fetchone()[0]

            if not last_occ:
                last_occ = m['date_planifiee']

            # Lookahead: générer jusqu'à aujourd'hui + 28 jours
            lookahead_date = (datetime.now() + timedelta(days=28)).strftime('%Y-%m-%d')
            last_occ_dt = datetime.strptime(last_occ, '%Y-%m-%d')

            # Si la dernière occurrence est < lookahead, générer les nouvelles
            if last_occ_dt < datetime.strptime(lookahead_date, '%Y-%m-%d'):
                # Construire un fake form dict pour _generate_maintenance_series
                fake_form = {
                    'type_maintenance': m['type_maintenance'],
                    'description': m['description'],
                    'appareil_id': m['appareil_id'],
                    'peripherique_id': m['peripherique_id'],
                    'heure_debut': m['heure_debut'],
                    'heure_fin': m['heure_fin'],
                    'responsable': m['responsable'],
                    'notes': m['notes'],
                }
                _generate_maintenance_series(conn, m['id'], m['client_id'],
                                            last_occ, m['recurrence'],
                                            m['date_fin_recurrence'],
                                            1,  # created_by=1 (cron job)
                                            fake_form, lookahead_date=lookahead_date)

        conn.close()
        logger.info(f"Cron job: {len(maintenances)} maintenances récurrentes vérifiées")
    except Exception as e:
        logger.exception(f"Erreur dans cron job de régénération: {e}")


def _send_email(to_email, subject, body):
    """Envoie un email via SMTP. Retourne True si succès."""
    try:
        smtp_server = cfg_get('smtp_server', '')
        smtp_port = int(cfg_get('smtp_port', '587'))
        smtp_login = cfg_get('smtp_login', '')
        smtp_password = cfg_get('smtp_password', '')
        from_email = cfg_get('from_email', smtp_login)

        if not all([smtp_server, smtp_login, smtp_password]):
            logger.warning('SMTP non configuré - notification ignorée')
            return False

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = to_email
        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_login, smtp_password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()

        return True
    except Exception as e:
        logger.error(f'Erreur envoi email: {e}')
        return False


def _notify_upcoming_maintenances():
    """Cron job: envoie notifications des maintenances à venir (3 jours)"""
    try:
        conn = get_db()

        # Récupérer maintenances des 3 prochains jours
        today = date.today().isoformat()
        in_3_days = (date.today() + timedelta(days=3)).isoformat()

        # NOT EXISTS sans condition de date : une maintenance déjà notifiée une
        # fois ne doit plus jamais l'être une seconde fois pour la même
        # occurrence, même si elle reste dans la fenêtre de préavis de 3 jours
        # plusieurs jours de suite (notification_date >= today se réinitialisait
        # chaque jour puisque « today » change à chaque exécution — la même
        # maintenance recevait donc un email par jour pendant 3 jours au lieu
        # d'un seul).
        maintenances = conn.execute('''
            SELECT m.id, m.date_planifiee, m.type_maintenance, m.description,
                   m.responsable, m.statut, a.nom_machine, p.categorie, p.marque
            FROM maintenances m
            LEFT JOIN appareils a ON m.appareil_id = a.id
            LEFT JOIN peripheriques p ON m.peripherique_id = p.id
            WHERE m.statut = 'programmee'
              AND m.date_planifiee BETWEEN ? AND ?
              AND NOT EXISTS (
                SELECT 1 FROM maintenance_notifications
                WHERE maintenance_id=m.id
              )
            ORDER BY m.date_planifiee
        ''', (today, in_3_days)).fetchall()

        for maint in maintenances:
            m = row_to_dict(maint)
            recipient = m.get('responsable', '')

            if recipient and '@' in recipient:
                subject = f"⚙️ Maintenance à venir: {m['type_maintenance']} - {m['date_planifiee']}"

                equipment = m.get('nom_machine') or f"{m.get('categorie', '')} {m.get('marque', '')}"
                body = f"""<html><body style="font-family: Arial;">
                <h2>Notification de maintenance</h2>
                <p><strong>Date:</strong> {m['date_planifiee']}</p>
                <p><strong>Type:</strong> {m['type_maintenance']}</p>
                <p><strong>Équipement:</strong> {equipment or '—'}</p>
                <p><strong>Description:</strong> {m.get('description', '—')}</p>
                <p><em>Veuillez confirmer l'exécution dans ParcInfo</em></p>
                </body></html>"""

                if _send_email(recipient, subject, body):
                    # Enregistrer la notification envoyée
                    conn.execute(
                        'INSERT INTO maintenance_notifications (maintenance_id, notification_date) VALUES (?, ?)',
                        (m['id'], today)
                    )
                    conn.commit()
                    logger.info(f"Notification envoyée pour maintenance {m['id']} à {recipient}")

        conn.close()
    except Exception as e:
        logger.exception(f'Erreur notification maintenances: {e}')


def _extract_maintenance(cid, f, user_id):
    appareil_id = None
    try: appareil_id = int(f.get('appareil_id')) if f.get('appareil_id') else None
    except: pass
    peripherique_id = None
    try: peripherique_id = int(f.get('peripherique_id')) if f.get('peripherique_id') else None
    except: pass
    contrat_id = None
    try: contrat_id = int(f.get('contrat_id')) if f.get('contrat_id') else None
    except: pass
    return (cid, appareil_id, peripherique_id, contrat_id, f.get('type_maintenance',''), f.get('description',''),
            f.get('date_planifiee',''), f.get('date_realisee',''), f.get('heure_debut',''),
            f.get('heure_fin',''), f.get('responsable',''), f.get('notes',''),
            f.get('statut','programmee'), f.get('recurrence'), f.get('date_fin_recurrence'),
            None, user_id, user_id)


def _format_maintenance_for_list(rows):
    """Formate maintenances pour affichage liste"""
    result = []
    for r in rows:
        m = row_to_dict(r)
        m['statut_label'] = {
            'programmee': 'Programmée',
            'realisee': 'Réalisée',
            'reportee': 'Reportée',
            'annulee': 'Annulée'
        }.get(m.get('statut', ''), m.get('statut', ''))
        m['type_label'] = (m.get('type_maintenance') or '').title()
        _format_date_field(m, 'date_planifiee')
        if m.get('date_realisee'):
            _format_date_field(m, 'date_realisee')
        result.append(m)
    return result


# --- INTERVENTIONS ----------------------------------------------------------

@app.route('/interventions')
@login_required
def liste_interventions():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '')
    filtre_type = request.args.get('type', '')
    filtre_statut = request.args.get('statut', '')
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})

    # Build query
    query = 'SELECT * FROM interventions WHERE client_id=? AND statut != ?'
    params = [cid, 'archivee']

    if q:
        query += ' AND titre LIKE ?'
        params.append(f'%{q}%')
    if filtre_type:
        query += ' AND type_intervention=?'
        params.append(filtre_type)
    if filtre_statut:
        query += ' AND statut=?'
        params.append(filtre_statut)

    query += ' ORDER BY date_intervention DESC'

    rows, pagination = paginate(query, tuple(params), page)
    interventions = [fmt_intervention(row_to_dict(r)) for r in rows]

    # Stats
    stats = {
        'total': conn.execute(
            'SELECT COUNT(*) FROM interventions WHERE client_id=? AND statut != ?', (cid, 'archivee')).fetchone()[0],
        'planifiee': conn.execute(
            "SELECT COUNT(*) FROM interventions WHERE client_id=? AND statut='planifiee'", (cid,)).fetchone()[0],
        'en_cours': conn.execute(
            "SELECT COUNT(*) FROM interventions WHERE client_id=? AND statut='en_cours'", (cid,)).fetchone()[0],
        'completee': conn.execute(
            "SELECT COUNT(*) FROM interventions WHERE client_id=? AND statut='completee'", (cid,)).fetchone()[0],
    }

    filtre_stat = ''
    if filtre_statut:
        filtre_stat = filtre_statut

    types_interventions = get_liste('types_interventions')
    conn.close()

    return render_template('interventions.html', interventions=interventions, client=client,
                          clients=get_clients(), client_actif_id=cid,
                          pagination=pagination, stats=stats,
                          filtre_type=filtre_type, filtre_statut=filtre_statut, filtre_stat=filtre_stat,
                          types_interventions=types_interventions)

@app.route('/intervention/nouveau', methods=['GET', 'POST'])
@login_required
def nouveau_intervention():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_interventions'))

    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))

    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine, type_appareil, adresse_ip FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie, marque',
        (cid,)).fetchall()]
    contrats = [row_to_dict(r) for r in conn.execute(
        'SELECT id, titre FROM contrats WHERE client_id=? ORDER BY titre', (cid,)).fetchall()]
    types_interventions = get_liste('types_interventions')

    if request.method == 'POST':
        errs = validate_form([
            ('titre', 'str', True),
            ('type_intervention', 'str', True),
            ('date_intervention', 'date', True),
            ('description', 'str', True),
        ], request.form)

        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)

        f = request.form
        user = get_auth_user()
        now = datetime.now().isoformat()

        conn = get_db()
        cur = conn.cursor()

        contrat_id = None
        try:
            contrat_id = int(f.get('contrat_id')) if f.get('contrat_id') else None
        except:
            pass

        cout_ht = None
        try:
            cout_ht = float(f.get('cout_ht')) if f.get('cout_ht') else None
        except:
            pass

        duree_minutes = 0
        try:
            duree_minutes = int(f.get('duree_minutes')) if f.get('duree_minutes') else 0
        except:
            pass

        cur.execute("""INSERT INTO interventions
            (client_id, titre, type_intervention, description, notes,
             date_intervention, heure_debut, heure_fin, duree_minutes,
             technicien_nom, technicien_email, statut, contrat_id, cout_ht, devise,
             date_creation, date_maj, auth_user_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (cid, f.get('titre', ''), f.get('type_intervention', ''),
             f.get('description', ''), f.get('notes', ''),
             f.get('date_intervention', ''), f.get('heure_debut', ''), f.get('heure_fin', ''),
             duree_minutes, f.get('technicien_nom', ''), f.get('technicien_email', ''),
             f.get('statut', 'completee'), contrat_id, cout_ht, 'EUR',
             now, now, user['id']))

        intv_id = cur.lastrowid

        # Link appareils
        for app_id in request.form.getlist('appareils_lies'):
            try:
                conn.execute('INSERT INTO interventions_appareils (intervention_id, appareil_id) VALUES (?,?)',
                           (intv_id, int(app_id)))
            except:
                pass

        # Link peripheriques
        for per_id in request.form.getlist('peripheriques_lies'):
            try:
                conn.execute('INSERT INTO interventions_peripheriques (intervention_id, peripherique_id) VALUES (?,?)',
                           (intv_id, int(per_id)))
            except:
                pass

        log_history(conn, cid, 'intervention', intv_id, f.get('titre', '') or f'Intervention #{intv_id}', 'Création')
        conn.commit()
        conn.close()
        flash('Intervention créée', 'success')
        return redirect(url_for('detail_intervention', id=intv_id))

    conn.close()
    return render_template('form_intervention.html', intervention=None,
                          appareils=appareils, appareils_lies=[],
                          peripheriques=peripheriques, peripheriques_lies=[],
                          contrats=contrats, types_interventions=types_interventions,
                          client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/intervention/<int:id>')
@login_required
def detail_intervention(id):
    cid = get_client_id()
    conn = get_db()

    intv = fmt_intervention(row_to_dict(
        conn.execute('SELECT * FROM interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {}))

    if not intv:
        conn.close()
        return 'Intervention non trouvée', 404

    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})

    appareils_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT a.* FROM appareils a JOIN interventions_appareils ia ON a.id=ia.appareil_id WHERE ia.intervention_id=?',
        (id,)).fetchall()]

    peripheriques_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT p.* FROM peripheriques p JOIN interventions_peripheriques ip ON p.id=ip.peripherique_id WHERE ip.intervention_id=?',
        (id,)).fetchall()]

    docs = [row_to_dict(r) for r in conn.execute(
        'SELECT id, intervention_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload FROM documents_interventions WHERE intervention_id=? ORDER BY date_upload DESC', (id,)).fetchall()]
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))

    # Get contrat title if exists
    if intv.get('contrat_id'):
        contrat = conn.execute('SELECT titre FROM contrats WHERE id=?', (intv['contrat_id'],)).fetchone()
        if contrat:
            intv['contrat_titre'] = contrat[0]

    conn.close()
    return render_template('detail_intervention.html', intervention=intv,
                          appareils_lies=appareils_lies, peripheriques_lies=peripheriques_lies,
                          docs=docs, client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/intervention/<int:id>/editer', methods=['GET', 'POST'])
@login_required
def editer_intervention(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_interventions'))

    cid = get_client_id()
    conn = get_db()

    intv = fmt_intervention(row_to_dict(
        conn.execute('SELECT * FROM interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {}))

    if not intv:
        conn.close()
        return 'Intervention non trouvée', 404

    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine, type_appareil, adresse_ip FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie, marque',
        (cid,)).fetchall()]
    contrats = [row_to_dict(r) for r in conn.execute(
        'SELECT id, titre FROM contrats WHERE client_id=? ORDER BY titre', (cid,)).fetchall()]
    types_interventions = get_liste('types_interventions')

    if request.method == 'POST':
        errs = validate_form([
            ('titre', 'str', True),
            ('type_intervention', 'str', True),
            ('date_intervention', 'date', True),
            ('description', 'str', True),
        ], request.form)

        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)

        f = request.form
        user = get_auth_user()
        now = datetime.now().isoformat()

        # Fetch old values for comparison
        _old = row_to_dict(conn.execute('SELECT * FROM interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})

        contrat_id = None
        try:
            contrat_id = int(f.get('contrat_id')) if f.get('contrat_id') else None
        except:
            pass

        cout_ht = None
        try:
            cout_ht = float(f.get('cout_ht')) if f.get('cout_ht') else None
        except:
            pass

        duree_minutes = 0
        try:
            duree_minutes = int(f.get('duree_minutes')) if f.get('duree_minutes') else 0
        except:
            pass

        conn.execute("""UPDATE interventions SET
            titre=?, type_intervention=?, description=?, notes=?,
            date_intervention=?, heure_debut=?, heure_fin=?, duree_minutes=?,
            technicien_nom=?, technicien_email=?, statut=?, contrat_id=?, cout_ht=?, devise=?,
            date_maj=? WHERE id=? AND client_id=?""",
            (f.get('titre', ''), f.get('type_intervention', ''),
             f.get('description', ''), f.get('notes', ''),
             f.get('date_intervention', ''), f.get('heure_debut', ''), f.get('heure_fin', ''),
             duree_minutes, f.get('technicien_nom', ''), f.get('technicien_email', ''),
             f.get('statut', 'completee'), contrat_id, cout_ht, 'EUR',
             now, id, cid))

        # Reset liaisons
        conn.execute('DELETE FROM interventions_appareils WHERE intervention_id=?', (id,))
        conn.execute('DELETE FROM interventions_peripheriques WHERE intervention_id=?', (id,))

        for app_id in request.form.getlist('appareils_lies'):
            try:
                conn.execute('INSERT INTO interventions_appareils (intervention_id, appareil_id) VALUES (?,?)',
                           (id, int(app_id)))
            except:
                pass

        for per_id in request.form.getlist('peripheriques_lies'):
            try:
                conn.execute('INSERT INTO interventions_peripheriques (intervention_id, peripherique_id) VALUES (?,?)',
                           (id, int(per_id)))
            except:
                pass

        # Record change details
        _cols_i = _ENTITE_COLS['intervention']
        _details_i = _diff_json({k: str(_old.get(k,'') or '') for k in _cols_i},
                                 {k: str(f.get(k,'') or '') for k in _cols_i})
        log_history(conn, cid, 'intervention', id, f.get('titre', '') or f'Intervention #{id}', 'Modification', _details_i)
        conn.commit()
        conn.close()
        flash('Intervention mise à jour', 'success')
        return redirect(url_for('detail_intervention', id=id))

    appareils_lies = [r[0] for r in conn.execute(
        'SELECT appareil_id FROM interventions_appareils WHERE intervention_id=?', (id,)).fetchall()]
    peripheriques_lies = [r[0] for r in conn.execute(
        'SELECT peripherique_id FROM interventions_peripheriques WHERE intervention_id=?', (id,)).fetchall()]

    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    conn.close()

    return render_template('form_intervention.html', intervention=intv,
                          appareils=appareils, appareils_lies=appareils_lies,
                          peripheriques=peripheriques, peripheriques_lies=peripheriques_lies,
                          contrats=contrats, types_interventions=types_interventions,
                          client=client, clients=get_clients(), client_actif_id=cid)

@app.route('/intervention/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_intervention(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_interventions'))

    cid = get_client_id()
    conn = get_db()

    # Soft delete: set status to archivee
    conn.execute('UPDATE interventions SET statut=? WHERE id=? AND client_id=?',
                ('archivee', id, cid))
    log_history(conn, cid, 'intervention', id, f'Intervention #{id}', 'Archivage')
    conn.commit()
    conn.close()

    flash('Intervention archivée', 'info')
    return redirect(url_for('liste_interventions'))

@app.route('/intervention/<int:id>/document/upload', methods=['POST'])
@login_required
def upload_doc_intervention(id):
    cid = get_client_id()
    if 'fichier' not in request.files:
        return redirect(url_for('detail_intervention', id=id))

    f = request.files['fichier']
    ok, motif = verifier_fichier(f)
    if not ok:
        flash(motif, 'danger')
        return redirect(url_for('detail_intervention', id=id))

    unique = f"intv{id}_{int(time.time())}_{secure_filename(f.filename)}"
    save_path = os.path.join(UPLOAD_FOLDER, unique)
    f.save(save_path)

    nom = request.form.get('nom', '') or f.filename
    now = datetime.now().isoformat()

    conn = get_db()
    conn.execute(
        'INSERT INTO documents_interventions (intervention_id, client_id, nom, description, type_doc, nom_fichier, taille, date_upload) VALUES (?,?,?,?,?,?,?,?)',
        (id, cid, nom, request.form.get('description', ''), request.form.get('type_doc', ''),
         unique, os.path.getsize(save_path), now))

    # Log document upload
    user = get_auth_user()
    intv_title = conn.execute('SELECT titre FROM interventions WHERE id=? AND client_id=?', (id, cid)).fetchone()
    intv_name = intv_title[0] if intv_title else f'Intervention #{id}'
    log_history(conn, cid, 'intervention', id, intv_name, 'Ajout de document',
                _diff_json({}, {'nom': nom, 'fichier': unique, 'type_doc': request.form.get('type_doc', '')}))
    conn.commit()
    conn.close()

    flash('Document ajouté', 'success')
    return redirect(url_for('detail_intervention', id=id))

@app.route('/intervention/document/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_doc_intervention(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_interventions'))

    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, intervention_id, client_id, nom, nom_fichier FROM documents_interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    intv_id = doc.get('intervention_id', 0)

    if doc:
        conn.execute('DELETE FROM documents_interventions WHERE id=?', (id,))

        # Log document deletion
        intv_title = conn.execute('SELECT titre FROM interventions WHERE id=? AND client_id=?', (intv_id, cid)).fetchone()
        intv_name = intv_title[0] if intv_title else f'Intervention #{intv_id}'
        log_history(conn, cid, 'intervention', intv_id, intv_name, 'Suppression de document',
                    _diff_json({'nom': doc.get('nom', ''), 'fichier': doc.get('nom_fichier', '')}, {}))

        conn.commit()
        try:
            os.remove(os.path.join(UPLOAD_FOLDER, doc['nom_fichier']))
        except:
            pass

    conn.close()
    return redirect(url_for('detail_intervention', id=intv_id))

@app.route('/intervention/document/<int:id>/apercu')
@login_required
def apercu_doc_intervention(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, intervention_id, client_id, nom, nom_fichier, contenu_blob FROM documents_interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()

    if not doc:
        return 'Not found', 404

    if doc.get('contenu_blob'):
        return send_file(io.BytesIO(doc['contenu_blob']), as_attachment=False, download_name=doc.get('nom_fichier', 'document'))

    fichier_path = os.path.join(UPLOAD_FOLDER, doc['nom_fichier'])
    if not os.path.exists(fichier_path):
        return f"Fichier introuvable : {doc.get('nom_fichier', '?')}", 404
    return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=False)

@app.route('/intervention/document/<int:id>/telecharger')
@login_required
def telecharger_doc_intervention(id):
    cid = get_client_id()
    conn = get_db()
    doc = row_to_dict(conn.execute(
        'SELECT id, intervention_id, client_id, nom, nom_fichier, contenu_blob FROM documents_interventions WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()

    if not doc:
        return 'Not found', 404

    if doc.get('contenu_blob'):
        return send_file(io.BytesIO(doc['contenu_blob']), as_attachment=True, download_name=doc['nom'])

    fichier_path = os.path.join(UPLOAD_FOLDER, doc['nom_fichier'])
    if not os.path.exists(fichier_path):
        return f"Fichier introuvable : {doc.get('nom_fichier', '?')}", 404
    return send_from_directory(UPLOAD_FOLDER, doc['nom_fichier'], as_attachment=True, download_name=doc['nom'])


@app.route('/identifiant/<int:id>/popup')
@login_required
def popup_identifiant(id):
    cid = get_client_id()
    conn = get_db()
    ident = row_to_dict(conn.execute(
        'SELECT * FROM identifiants WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    conn.close()
    if not ident:
        return 'Identifiant introuvable', 404
    # ✅ Déchiffrer le mot de passe
    if ident.get('mot_de_passe'):
        crypto = _get_crypto_shared()
        ident['mot_de_passe'] = crypto.decrypt(ident['mot_de_passe']) or ident['mot_de_passe']
    # Format dates
    today = date.today()
    if ident.get('date_expiration'):
        try:
            df = date.fromisoformat(ident['date_expiration'])
            ident['date_expiration_fmt'] = df.strftime('%d/%m/%Y')
            delta = (df - today).days
            ident['expire_depasse']  = delta < 0
            ident['expire_bientot']  = 0 <= delta <= 30
        except (ValueError, TypeError):
            ident['date_expiration_fmt'] = ident['date_expiration']
    return render_template('popup_identifiant.html', ident=ident)


# --- MAINTENANCE ----------------------------------------------------------

@app.route('/maintenances')
@login_required
def liste_maintenances():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))

    page = request.args.get('page', 1, type=int)
    filtre_type = request.args.get('type_maintenance', '')
    filtre_statut = request.args.get('statut', '')
    filtre_appareil = request.args.get('appareil_id', '')
    filtre_date_debut = request.args.get('date_debut', '')
    filtre_date_fin = request.args.get('date_fin', '')
    filtre_responsable = request.args.get('responsable', '')

    # Tri
    sort_by = request.args.get('sort_by', 'date_planifiee')
    sort_order = request.args.get('sort_order', 'desc').lower()

    # Valider la colonne de tri (whitelist)
    allowed_sort_cols = ['date_planifiee', 'type_maintenance', 'responsable', 'statut', 'date_realisee']
    if sort_by not in allowed_sort_cols:
        sort_by = 'date_planifiee'
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'

    clients_sel = request.args.getlist('clients_selection')
    if not clients_sel:
        clients_sel = None

    conn = get_db()

    # Build query with JOINs to get appareil/peripherique/contrat names
    query = '''SELECT m.*,
               a.nom_machine as appareil_nom,
               p.categorie as peripherique_categorie, p.marque as peripherique_marque, p.modele as peripherique_modele,
               c.type_contrat as contrat_type, c.fournisseur as contrat_fournisseur
        FROM maintenances m
        LEFT JOIN appareils a ON m.appareil_id = a.id
        LEFT JOIN peripheriques p ON m.peripherique_id = p.id
        LEFT JOIN contrats c ON m.contrat_id = c.id
        WHERE m.client_id IN ({})'''.format(
        ','.join(['?'] * len([c['id'] for c in get_clients_for_filter(clients_sel)])))

    params = [c['id'] for c in get_clients_for_filter(clients_sel)]

    if filtre_type:
        query += ' AND type_maintenance=?'
        params.append(filtre_type)
    if filtre_statut:
        query += ' AND statut=?'
        params.append(filtre_statut)
    if filtre_appareil:
        query += ' AND appareil_id=?'
        params.append(int(filtre_appareil))
    if filtre_date_debut:
        query += ' AND date_planifiee>=?'
        params.append(filtre_date_debut)
    if filtre_date_fin:
        query += ' AND date_planifiee<=?'
        params.append(filtre_date_fin)
    if filtre_responsable:
        query += ' AND responsable LIKE ?'
        params.append(f'%{filtre_responsable}%')

    query += f' ORDER BY {sort_by} {sort_order.upper()}'

    rows, pagination = paginate(query, tuple(params), page)
    maintenances = _format_maintenance_for_list(rows)

    # Stats
    types_maintenance = get_liste('types_maintenance')
    statuts_maintenance = get_liste('statuts_maintenance')

    # Get appareils for filter
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]

    clients = get_clients()
    conn.close()

    return render_template('liste_maintenances.html',
                          maintenances=maintenances, client_actif_id=cid,
                          pagination=pagination, clients=clients,
                          types_maintenance=types_maintenance,
                          statuts_maintenance=statuts_maintenance,
                          appareils=appareils,
                          filtre_type=filtre_type, filtre_statut=filtre_statut,
                          filtre_appareil=filtre_appareil, filtre_date_debut=filtre_date_debut,
                          filtre_date_fin=filtre_date_fin, filtre_responsable=filtre_responsable,
                          sort_by=sort_by, sort_order=sort_order)


@app.route('/maintenance/nouveau', methods=['GET', 'POST'])
@login_required
def nouveau_maintenance():
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_maintenances'))

    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))

    user = get_auth_user()
    conn = get_db()

    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie',
        (cid,)).fetchall()]
    contrats = [row_to_dict(r) for r in conn.execute(
        'SELECT id, type_contrat, fournisseur, date_debut, date_fin FROM contrats WHERE client_id=? ORDER BY date_debut DESC',
        (cid,)).fetchall()]
    types_maintenance = get_liste('types_maintenance')

    if request.method == 'POST':
        errs = validate_form([
            ('type_maintenance', 'str', True),
            ('date_planifiee', 'date', True),
        ], request.form)

        if errs:
            for e in errs: flash(e, 'danger')
            conn.close()
            return redirect(request.url)

        f = request.form
        now = datetime.now().isoformat()

        params = _extract_maintenance(cid, f, user['id'])

        cur = conn.cursor()
        try:
            cur.execute(
                '''INSERT INTO maintenances
                (client_id, appareil_id, peripherique_id, contrat_id, type_maintenance, description,
                 date_planifiee, date_realisee, heure_debut, heure_fin, responsable, notes,
                 statut, recurrence, date_fin_recurrence, parent_id, created_by, updated_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                params)
            maint_id = cur.lastrowid
            conn.commit()

            # Générer les occurrences si récurrence définie
            if f.get('recurrence'):
                # Lookahead: générer 4 semaines à l'avance
                lookahead_date = (datetime.strptime(f.get('date_planifiee'), '%Y-%m-%d') + timedelta(days=28)).strftime('%Y-%m-%d')
                _generate_maintenance_series(conn, maint_id, cid, f.get('date_planifiee'),
                                            f.get('recurrence'), f.get('date_fin_recurrence'),
                                            user['id'], f, lookahead_date=lookahead_date)

            log_history(conn, cid, 'maintenance', maint_id, f.get('description','Maintenance'),
                       'Création', {'type': f.get('type_maintenance'), 'date': f.get('date_planifiee')})
            conn.commit()
            if f.get('recurrence'):
                flash('Maintenance créée avec occurrences des 4 prochaines semaines', 'success')
            else:
                flash('Maintenance programmée créée', 'success')
        except Exception as e:
            conn.rollback()
            logger.exception('Erreur création maintenance')
            flash(f'Erreur : {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('liste_maintenances'))

    techniciens = [row_to_dict(r) for r in conn.execute(
        "SELECT id, nom, prenom FROM auth_users WHERE role != 'admin' AND actif=1 ORDER BY nom, prenom"
        ).fetchall()]
    conn.close()
    return render_template('form_maintenance.html', appareils=appareils, peripheriques=peripheriques,
                          contrats=contrats, types_maintenance=types_maintenance, techniciens=techniciens, action='Créer')


@app.route('/maintenance/<int:id>/editer', methods=['GET', 'POST'])
@login_required
def editer_maintenance(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_maintenances'))

    cid = get_client_id()
    user = get_auth_user()
    conn = get_db()

    maint = row_to_dict(conn.execute(
        'SELECT * FROM maintenances WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})

    if not maint:
        conn.close()
        flash('Maintenance introuvable', 'danger')
        return redirect(url_for('liste_maintenances'))

    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie, marque, modele FROM peripheriques WHERE client_id=? ORDER BY categorie',
        (cid,)).fetchall()]
    contrats = [row_to_dict(r) for r in conn.execute(
        'SELECT id, type_contrat, fournisseur, date_debut, date_fin FROM contrats WHERE client_id=? ORDER BY date_debut DESC',
        (cid,)).fetchall()]
    types_maintenance = get_liste('types_maintenance')

    if request.method == 'POST':
        errs = validate_form([
            ('type_maintenance', 'str', True),
            ('date_planifiee', 'date', True),
        ], request.form)

        if errs:
            for e in errs: flash(e, 'danger')
            conn.close()
            return redirect(request.url)

        f = request.form
        now = datetime.now().isoformat()

        try:
            appareil_id = int(f.get('appareil_id')) if f.get('appareil_id') else None
        except:
            appareil_id = None
        try:
            peripherique_id = int(f.get('peripherique_id')) if f.get('peripherique_id') else None
        except:
            peripherique_id = None
        try:
            contrat_id = int(f.get('contrat_id')) if f.get('contrat_id') else None
        except:
            contrat_id = None

        cur = conn.cursor()
        try:
            cur.execute(
                '''UPDATE maintenances
                SET type_maintenance=?, description=?, date_planifiee=?, date_realisee=?,
                    heure_debut=?, heure_fin=?, responsable=?, notes=?,
                    statut=?, recurrence=?, date_fin_recurrence=?,
                    appareil_id=?, peripherique_id=?, contrat_id=?, updated_by=?, date_maj=?
                WHERE id=? AND client_id=?''',
                (f.get('type_maintenance'), f.get('description'), f.get('date_planifiee'),
                 f.get('date_realisee'), f.get('heure_debut'), f.get('heure_fin'),
                 f.get('responsable'), f.get('notes'), f.get('statut'),
                 f.get('recurrence'), f.get('date_fin_recurrence'),
                 appareil_id, peripherique_id, contrat_id, user['id'], now,
                 id, cid))
            conn.commit()
            log_history(conn, cid, 'maintenance', id, f.get('description','Maintenance'),
                       'Modification', {'type': f.get('type_maintenance')})
            conn.commit()
            flash('Maintenance mise à jour', 'success')
        except Exception as e:
            conn.rollback()
            logger.exception('Erreur édition maintenance')
            flash(f'Erreur : {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('liste_maintenances'))

    techniciens = [row_to_dict(r) for r in conn.execute(
        "SELECT id, nom, prenom FROM auth_users WHERE role != 'admin' AND actif=1 ORDER BY nom, prenom"
        ).fetchall()]
    conn.close()
    return render_template('form_maintenance.html', maint=maint, appareils=appareils,
                          peripheriques=peripheriques, contrats=contrats,
                          types_maintenance=types_maintenance,
                          techniciens=techniciens, action='Éditer')


@app.route('/maintenance/<int:id>/confirmer', methods=['POST'])
@login_required
def confirmer_maintenance(id):
    if not can_write():
        return jsonify({'error': 'Accès en lecture seule — modification non autorisée'}), 403

    cid = get_client_id()
    user = get_auth_user()
    conn = get_db()

    maint = row_to_dict(conn.execute(
        'SELECT * FROM maintenances WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})

    if not maint:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    now = datetime.now().isoformat()
    today = date.today().isoformat()

    try:
        cur = conn.cursor()
        cur.execute(
            'UPDATE maintenances SET statut=?, date_realisee=?, updated_by=?, date_maj=? WHERE id=? AND client_id=?',
            ('realisee', today, user['id'], now, id, cid))
        conn.commit()
        log_history(conn, cid, 'maintenance', id, maint.get('description','Maintenance'),
                   'Confirmation', {'ancien_statut': maint.get('statut'), 'nouveau_statut': 'realisee'})
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        logger.exception('Erreur confirmation maintenance')
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/maintenance/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_maintenance(id):
    if not can_write():
        flash('Accès en lecture seule — modification non autorisée', 'danger')
        return redirect(url_for('liste_maintenances'))

    cid = get_client_id()
    user = get_auth_user()
    conn = get_db()

    maint = row_to_dict(conn.execute(
        'SELECT * FROM maintenances WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})

    if not maint:
        conn.close()
        flash('Maintenance introuvable', 'danger')
        return redirect(url_for('liste_maintenances'))

    now = datetime.now().isoformat()

    try:
        cur = conn.cursor()
        cur.execute(
            'UPDATE maintenances SET statut=?, updated_by=?, date_maj=? WHERE id=? AND client_id=?',
            ('annulee', user['id'], now, id, cid))
        conn.commit()
        log_history(conn, cid, 'maintenance', id, maint.get('description','Maintenance'),
                   'Suppression', {})
        conn.commit()
        flash('Maintenance annulée', 'success')
    except Exception as e:
        conn.rollback()
        logger.exception('Erreur suppression maintenance')
        flash(f'Erreur : {str(e)}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('liste_maintenances'))


@app.route('/rapport/maintenances')
@login_required
def rapport_maintenances():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))

    types_maintenance = get_liste('types_maintenance')
    clients = get_clients()

    # Filtres
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    type_maint = request.args.get('type_maintenance', '')
    statut = request.args.get('statut', '')
    clients_filter = request.args.get('clients_filter', 'current')

    # Déterminer clients sélectionnés
    if clients_filter == 'all':
        client_ids = [c['id'] for c in clients]
    else:
        client_ids = [cid]

    # Requête SQL filtrée
    conn = get_db()
    query = '''SELECT m.*,
                      a.nom_machine AS appareil_nom,
                      p.marque AS peripherique_marque, p.modele AS peripherique_modele, p.categorie AS peripherique_categorie,
                      c.type_contrat AS contrat_type, c.fournisseur AS contrat_fournisseur
               FROM maintenances m
               LEFT JOIN appareils a ON m.appareil_id = a.id
               LEFT JOIN peripheriques p ON m.peripherique_id = p.id
               LEFT JOIN contrats c ON m.contrat_id = c.id
               WHERE m.client_id IN ({})'''.format(','.join(['?'] * len(client_ids)))
    params = client_ids[:]

    if date_debut:
        query += ' AND m.date_planifiee >= ?'
        params.append(date_debut)
    if date_fin:
        query += ' AND m.date_planifiee <= ?'
        params.append(date_fin)
    if type_maint:
        query += ' AND m.type_maintenance = ?'
        params.append(type_maint)
    if statut:
        query += ' AND m.statut = ?'
        params.append(statut)

    query += ' ORDER BY m.date_planifiee DESC'

    rows = conn.execute(query, params).fetchall()
    maintenances = [row_to_dict(r) for r in rows]
    maintenances = _format_maintenance_for_list(maintenances)

    # Statistiques
    total = len(maintenances)
    realisees = sum(1 for m in maintenances if m['statut'] == 'realisee')
    attente = sum(1 for m in maintenances if m['statut'] == 'programmee')
    reportees = sum(1 for m in maintenances if m['statut'] == 'reportee')

    conn.close()

    return render_template('rapport_maintenance.html',
                          clients=clients,
                          types_maintenance=types_maintenance,
                          maintenances=maintenances,
                          stats={'total': total, 'realisees': realisees, 'attente': attente, 'reportees': reportees})


@app.route('/rapport/maintenances/pdf')
@login_required
def rapport_maintenances_pdf():
    """Génère un PDF du rapport de maintenances avec mise en page professionnelle"""
    if not REPORTLAB_AVAILABLE:
        flash('La bibliothèque reportlab n\'est pas installée. Installez-la avec: pip install reportlab', 'danger')
        return redirect(request.referrer or url_for('rapport_maintenances'))

    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))

    # Récupérer les mêmes filtres que le rapport HTML
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    type_maint = request.args.get('type_maintenance', '')
    statut = request.args.get('statut', '')
    clients_filter = request.args.get('clients_filter', 'current')

    clients = get_clients()
    if clients_filter == 'all':
        client_ids = [c['id'] for c in clients]
    else:
        client_ids = [cid]

    # Construire la requête SQL
    conn = get_db()
    query = '''SELECT m.*,
                      a.nom_machine AS appareil_nom,
                      p.marque AS peripherique_marque, p.modele AS peripherique_modele, p.categorie AS peripherique_categorie,
                      c.type_contrat AS contrat_type, c.fournisseur AS contrat_fournisseur
               FROM maintenances m
               LEFT JOIN appareils a ON m.appareil_id = a.id
               LEFT JOIN peripheriques p ON m.peripherique_id = p.id
               LEFT JOIN contrats c ON m.contrat_id = c.id
               WHERE m.client_id IN ({})'''.format(','.join(['?'] * len(client_ids)))
    params = client_ids[:]

    if date_debut:
        query += ' AND m.date_planifiee >= ?'
        params.append(date_debut)
    if date_fin:
        query += ' AND m.date_planifiee <= ?'
        params.append(date_fin)
    if type_maint:
        query += ' AND m.type_maintenance = ?'
        params.append(type_maint)
    if statut:
        query += ' AND m.statut = ?'
        params.append(statut)

    query += ' ORDER BY m.date_planifiee DESC'

    rows = conn.execute(query, params).fetchall()
    maintenances = [row_to_dict(r) for r in rows]
    maintenances = _format_maintenance_for_list(maintenances)

    # Calculer statistiques
    total = len(maintenances)
    realisees = sum(1 for m in maintenances if m['statut'] == 'realisee')
    attente = sum(1 for m in maintenances if m['statut'] == 'programmee')
    reportees = sum(1 for m in maintenances if m['statut'] == 'reportee')

    conn.close()

    try:
        # Créer le PDF avec reportlab
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0a0d12'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#00c9ff'),
            spaceAfter=3,
            fontName='Helvetica-Bold'
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#0a0d12'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )

        # Titre
        story.append(Paragraph('📊 Rapport Maintenances', title_style))
        story.append(Paragraph('Synthèse des opérations de maintenance', subtitle_style))
        story.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 15))

        # Statistiques
        stats_data = [
            ['Total Opérations', f'{total}'],
            ['Réalisées', f'{realisees}'],
            ['En Attente', f'{attente}'],
            ['Reportées', f'{reportees}']
        ]
        stats_table = Table(stats_data, colWidths=[3*inch, 1*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#e8f5ff')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd')),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))

        # Tableau détail
        story.append(Paragraph('Détail des Opérations', heading_style))

        if maintenances:
            # Préparer les données du tableau avec des Paragraphs pour meilleur enroulement
            table_data = [
                ['Date Planifiée', 'Type', 'Description', 'Responsable', 'Statut', 'Réalisée le']
            ]

            # Style pour le contenu des cellules
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8,
                leading=10,
                alignment=0  # LEFT
            )

            for m in maintenances[:50]:  # Limiter à 50 lignes pour la lisibilité
                # Tronquer la description à 80 caractères max pour éviter débordement
                description = m.get('description', '')[:80] or '—'

                table_data.append([
                    Paragraph(m.get('date_planifiee_fmt', ''), cell_style),
                    Paragraph(m.get('type_label', ''), cell_style),
                    Paragraph(description, cell_style),
                    Paragraph(m.get('responsable', '') or '—', cell_style),
                    Paragraph(m.get('statut_label', ''), cell_style),
                    Paragraph(m.get('date_realisee_fmt', '') or '—', cell_style)
                ])

            # Créer le tableau avec meilleures largeurs et hauteurs
            table = Table(table_data, colWidths=[1.0*inch, 0.8*inch, 2.2*inch, 0.9*inch, 0.8*inch, 1.0*inch])
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a0d12')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),

                # Contenu
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 1), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

                # Grille et bordures
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd')),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#0a0d12')),

                # Hauteur minimale des lignes
                ('ROWHEIGHT', (0, 0), (-1, -1), None),  # Auto-hauteur
                ('ROWHEIGHT', (0, 1), (-1, -1), 35),  # Minimum 35 points pour le contenu
            ]))
            story.append(table)
        else:
            story.append(Paragraph('Aucune maintenance trouvée pour les critères spécifiés.', styles['Italic']))

        # Pied de page
        story.append(Spacer(1, 20))
        story.append(Paragraph('ParcInfo — Rapport de Maintenance — Document généré automatiquement',
                              styles['Normal']))

        # Générer le PDF
        doc.build(story)
        pdf_buffer.seek(0)

        # Créer la réponse
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="rapport-maintenance-{datetime.now().strftime("%Y%m%d-%H%M%S")}.pdf"'
        return response

    except Exception as e:
        logger.exception('Erreur génération PDF rapport')
        flash(f'Erreur lors de la génération du PDF : {str(e)}', 'danger')
        return redirect(request.referrer or url_for('rapport_maintenances',
                                date_debut=date_debut,
                                date_fin=date_fin,
                                type_maintenance=type_maint,
                                statut=statut,
                                clients_filter=clients_filter))


@app.route('/maintenance/historique')
@login_required
def historique_maintenance():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))

    conn = get_db()

    # Statistiques par type de maintenance
    types_stats = conn.execute('''
        SELECT type_maintenance, COUNT(*) as count,
               SUM(CASE WHEN statut='realisee' THEN 1 ELSE 0 END) as realisees
        FROM maintenances WHERE client_id=? GROUP BY type_maintenance
    ''', (cid,)).fetchall()

    # Statistiques par mois (12 derniers mois)
    months_stats = conn.execute('''
        SELECT strftime('%Y-%m', date_planifiee) as mois, COUNT(*) as count
        FROM maintenances WHERE client_id=? AND date_planifiee >= date('now', '-12 months')
        GROUP BY strftime('%Y-%m', date_planifiee) ORDER BY mois
    ''', (cid,)).fetchall()

    # Statistiques par statut
    status_stats = conn.execute('''
        SELECT statut, COUNT(*) as count FROM maintenances WHERE client_id=?
        GROUP BY statut
    ''', (cid,)).fetchall()

    # Total maintenances
    total = conn.execute('SELECT COUNT(*) FROM maintenances WHERE client_id=?', (cid,)).fetchone()[0]
    realisees = conn.execute('SELECT COUNT(*) FROM maintenances WHERE client_id=? AND statut="realisee"', (cid,)).fetchone()[0]
    taux_realisation = int((realisees / total * 100) if total > 0 else 0)

    conn.close()

    # Formater données pour Chart.js
    types_labels = [row[0] for row in types_stats]
    types_data = [row[1] for row in types_stats]

    months_labels = [row[0] for row in months_stats]
    months_data = [row[1] for row in months_stats]

    status_map = {'programmee': 'Programmée', 'realisee': 'Réalisée', 'reportee': 'Reportée', 'annulee': 'Annulée'}
    status_labels = [status_map.get(row[0], row[0]) for row in status_stats]
    status_data = [row[1] for row in status_stats]

    import json
    return render_template('historique_maintenance.html',
        types_labels=json.dumps(types_labels),
        types_data=json.dumps(types_data),
        months_labels=json.dumps(months_labels),
        months_data=json.dumps(months_data),
        status_labels=json.dumps(status_labels),
        status_data=json.dumps(status_data),
        total=total,
        realisees=realisees,
        taux_realisation=taux_realisation)


@app.route('/maintenances/export.csv')
@login_required
def export_maintenances_csv():
    cid = get_client_id()
    filtre_date_debut = request.args.get('date_debut', '')
    filtre_date_fin = request.args.get('date_fin', '')
    filtre_type = request.args.get('type_maintenance', '')
    filtre_statut = request.args.get('statut', '')
    clients_filter = request.args.get('clients_filter', 'current')

    # Déterminer clients sélectionnés
    if clients_filter == 'all':
        client_ids = [c['id'] for c in get_clients()]
    else:
        client_ids = [cid]

    conn = get_db()

    query = '''SELECT m.*, a.nom_machine AS appareil_nom, p.marque AS peripherique_marque, p.modele AS peripherique_modele
               FROM maintenances m
               LEFT JOIN appareils a ON m.appareil_id = a.id
               LEFT JOIN peripheriques p ON m.peripherique_id = p.id
               WHERE m.client_id IN ({})'''.format(','.join(['?'] * len(client_ids)))
    params = client_ids[:]

    if filtre_date_debut:
        query += ' AND m.date_planifiee >= ?'
        params.append(filtre_date_debut)
    if filtre_date_fin:
        query += ' AND m.date_planifiee <= ?'
        params.append(filtre_date_fin)
    if filtre_type:
        query += ' AND m.type_maintenance = ?'
        params.append(filtre_type)
    if filtre_statut:
        query += ' AND m.statut = ?'
        params.append(filtre_statut)

    query += ' ORDER BY m.date_planifiee DESC'

    rows = conn.execute(query, params).fetchall()
    maintenances = [row_to_dict(r) for r in rows]
    conn.close()

    # Build CSV
    import io
    output = io.StringIO()
    output.write('\ufeff')  # BOM UTF-8
    output.write('Date planifiée;Appareil;Périphérique;Type;Description;Responsable;Statut;Date réalisée;Notes\n')

    for m in maintenances:
        app_name = ''
        if m.get('appareil_id'):
            c2 = get_db()
            a = c2.execute('SELECT nom_machine FROM appareils WHERE id=?', (m['appareil_id'],)).fetchone()
            if a: app_name = a[0]
            c2.close()

        periph_name = ''
        if m.get('peripherique_id'):
            c2 = get_db()
            p = c2.execute('SELECT categorie, marque FROM peripheriques WHERE id=?', (m['peripherique_id'],)).fetchone()
            if p: periph_name = f"{p[0]} {p[1]}"
            c2.close()

        output.write(f"{m.get('date_planifiee','')};{app_name};{periph_name};{m.get('type_maintenance','')};")
        output.write(f"{m.get('description','')};{m.get('responsable','')};{m.get('statut','')};")
        output.write(f"{m.get('date_realisee','')};{m.get('notes','')}\n")

    response = app.response_class(output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = f'attachment; filename=maintenances_{cid}_{date.today().isoformat()}.csv'
    return response


@app.route('/maintenances/export.xlsx')
@login_required
def export_maintenances_xlsx():
    cid = get_client_id()
    filtre_date_debut = request.args.get('date_debut', '')
    filtre_date_fin = request.args.get('date_fin', '')
    filtre_type = request.args.get('type_maintenance', '')
    filtre_statut = request.args.get('statut', '')
    clients_filter = request.args.get('clients_filter', 'current')

    if clients_filter == 'all':
        client_ids = [c['id'] for c in get_clients()]
    else:
        client_ids = [cid]

    conn = get_db()
    query = '''SELECT m.*, a.nom_machine AS appareil_nom, p.marque AS peripherique_marque, p.modele AS peripherique_modele
               FROM maintenances m
               LEFT JOIN appareils a ON m.appareil_id = a.id
               LEFT JOIN peripheriques p ON m.peripherique_id = p.id
               WHERE m.client_id IN ({})'''.format(','.join(['?'] * len(client_ids)))
    params = client_ids[:]
    if filtre_date_debut:
        query += ' AND m.date_planifiee >= ?'
        params.append(filtre_date_debut)
    if filtre_date_fin:
        query += ' AND m.date_planifiee <= ?'
        params.append(filtre_date_fin)
    if filtre_type:
        query += ' AND m.type_maintenance = ?'
        params.append(filtre_type)
    if filtre_statut:
        query += ' AND m.statut = ?'
        params.append(filtre_statut)
    query += ' ORDER BY m.date_planifiee DESC'

    maintenances = [row_to_dict(r) for r in conn.execute(query, params).fetchall()]
    conn.close()

    lignes = []
    for m in maintenances:
        app_name, periph_name = '', ''
        if m.get('appareil_id'):
            c2 = get_db()
            a = c2.execute('SELECT nom_machine FROM appareils WHERE id=?', (m['appareil_id'],)).fetchone()
            if a: app_name = a[0]
            c2.close()
        if m.get('peripherique_id'):
            c2 = get_db()
            p = c2.execute('SELECT categorie, marque FROM peripheriques WHERE id=?', (m['peripherique_id'],)).fetchone()
            if p: periph_name = f"{p[0]} {p[1]}"
            c2.close()
        lignes.append([
            m.get('date_planifiee', ''), app_name, periph_name, m.get('type_maintenance', ''),
            m.get('description', ''), m.get('responsable', ''), m.get('statut', ''),
            m.get('date_realisee', ''), m.get('notes', ''),
        ])

    headers = ['Date planifiée', 'Appareil', 'Périphérique', 'Type', 'Description',
               'Responsable', 'Statut', 'Date réalisée', 'Notes']
    filename = f'maintenances_{cid}_{date.today().isoformat()}.xlsx'
    return _xlsx_response(headers, lignes, filename, 'Maintenances')


# --- LISTES PERSONNALISABLES -------------------------------------------------

def _liste_est_initialisee(conn, nom: str) -> bool:
    """Indique si une liste a déjà été persistée en DB.
    Lit d'abord le flag dans le cache config (lecture seule, pas d'écriture sur conn),
    puis fait un COUNT en fallback pour les DBs existantes sans flag.
    IMPORTANT : n'appelle jamais cfg_set ici — conn peut avoir un verrou d'écriture actif."""
    if cfg_get(f'_list_init_{nom}', '0') == '1':
        return True
    # Fallback legacy : lignes existantes = liste déjà initialisée
    # Le flag sera posé par l'appelant après conn.commit() pour éviter le deadlock SQLite
    nb = conn.execute('SELECT COUNT(*) FROM config_listes WHERE nom_liste=?', (nom,)).fetchone()[0]
    return nb > 0

def _initialiser_liste(conn, nom: str, exclure: str = None):
    """Écrit tous les defaults d'une liste sur `conn` (sauf `exclure`).
    N'appelle PAS cfg_set — l'appelant doit le faire APRÈS conn.commit()
    pour éviter le deadlock SQLite (deux connexions en écriture simultanée)."""
    for i, v in enumerate(LISTE_DEFAULTS.get(nom, [])):
        if v == exclure:
            continue
        try:
            conn.execute('INSERT OR IGNORE INTO config_listes (nom_liste,valeur,ordre) VALUES (?,?,?)', (nom, v, i))
        except Exception:
            pass

@app.route('/api/listes/<nom>', methods=['GET'])
@login_required
def api_liste_get(nom):
    if nom not in LISTE_DEFAULTS:
        return jsonify({'error': 'Liste inconnue'}), 404
    return jsonify({'nom': nom, 'valeurs': get_liste(nom), 'defaults': LISTE_DEFAULTS[nom]})

@app.route('/api/listes/<nom>/ajouter', methods=['POST'])
@login_required
def api_liste_ajouter(nom):
    if nom not in LISTE_DEFAULTS:
        return jsonify({'error': 'Liste inconnue'}), 404
    valeur = (request.json or {}).get('valeur', '').strip()
    if not valeur:
        return jsonify({'error': 'Valeur vide'}), 400

    # ── Opération principale sur la DB active ────────────────────────────────
    conn = get_db()
    deja_init = _liste_est_initialisee(conn, nom)
    if not deja_init:
        # Première utilisation : persister les defaults pour préserver l'ordre
        _initialiser_liste(conn, nom)
    ordre = conn.execute('SELECT COALESCE(MAX(ordre),0)+1 FROM config_listes WHERE nom_liste=?', (nom,)).fetchone()[0]
    try:
        conn.execute('INSERT INTO config_listes (nom_liste,valeur,ordre) VALUES (?,?,?)', (nom, valeur, ordre))
        conn.commit()
    except Exception:
        conn.close()
        return jsonify({'error': 'Valeur déjà existante'}), 409
    conn.close()

    # ── Propagation croisée (évite les écarts entre local et Turso) ─────────
    try:
        _t_url   = cfg_get('turso_url',   '').strip()
        _t_token = cfg_get('turso_token', '').strip()
        _db_type = cfg_get('db_type', 'local')
        if _t_url and _t_token:
            if _db_type in ('local', 'sync'):
                from database import TursoConnection as _TC
                _TC(_t_url, _t_token).execute(
                    'INSERT OR IGNORE INTO config_listes (nom_liste,valeur,ordre) VALUES (?,?,?)',
                    (nom, valeur, ordre))
            elif _db_type == 'turso':
                from database import get_local_db as _get_local
                _lconn = _get_local()
                _lconn.execute('INSERT OR IGNORE INTO config_listes (nom_liste,valeur,ordre) VALUES (?,?,?)',
                               (nom, valeur, ordre))
                _lconn.commit(); _lconn.close()
    except Exception:
        logger.warning('Propagation croisée ajout config_listes échouée', exc_info=True)

    # ── Flag posé APRÈS commit : conn fermée, aucun verrou actif → pas de deadlock ──
    if not deja_init:
        cfg_set(f'_list_init_{nom}', '1')
    return jsonify({'ok': True, 'valeurs': get_liste(nom)})

@app.route('/api/listes/<nom>/supprimer', methods=['POST'])
@login_required
def api_liste_supprimer(nom):
    if nom not in LISTE_DEFAULTS:
        return jsonify({'error': 'Liste inconnue'}), 404
    valeur = (request.json or {}).get('valeur', '').strip()
    logger.info('[liste_supprimer] liste=%s valeur=%r db_type=%s', nom, valeur, cfg_get('db_type'))

    # ── 1. Opération principale sur la DB active (locale ou Turso selon config) ──
    conn = get_db()
    deja_init = _liste_est_initialisee(conn, nom)
    logger.info('[liste_supprimer] deja_init=%s', deja_init)
    if not deja_init:
        # Première suppression : persister tous les defaults SAUF la valeur supprimée
        _initialiser_liste(conn, nom, exclure=valeur)
        logger.info('[liste_supprimer] liste initialisée sans %r', valeur)
    else:
        # Liste déjà persistée : suppression directe
        conn.execute('DELETE FROM config_listes WHERE nom_liste=? AND valeur=?', (nom, valeur))
        logger.info('[liste_supprimer] DELETE exécuté sur DB principale')
    conn.commit()
    conn.close()

    # ── 2. Propagation croisée pour éviter que la sync bidirectionnelle réinjecte la valeur ──
    #    Si mode local/sync → aussi supprimer sur Turso
    #    Si mode turso → aussi supprimer sur SQLite local
    try:
        _t_url   = cfg_get('turso_url',   '').strip()
        _t_token = cfg_get('turso_token', '').strip()
        _db_type = cfg_get('db_type', 'local')
        if _t_url and _t_token:
            if _db_type in ('local', 'sync'):
                # DB principale = local → propager sur Turso
                from database import TursoConnection as _TC
                _turso = _TC(_t_url, _t_token)
                _turso.execute('DELETE FROM config_listes WHERE nom_liste=? AND valeur=?', (nom, valeur))
                logger.info('[liste_supprimer] DELETE propagé vers Turso pour %r', valeur)
            elif _db_type == 'turso':
                # DB principale = Turso → propager sur SQLite local
                from database import get_local_db as _get_local
                _lconn = _get_local()
                _lconn.execute('DELETE FROM config_listes WHERE nom_liste=? AND valeur=?', (nom, valeur))
                _lconn.commit(); _lconn.close()
                logger.info('[liste_supprimer] DELETE propagé vers SQLite local pour %r', valeur)
    except Exception:
        logger.warning('Propagation croisée config_listes échouée', exc_info=True)

    # ── 3. Flag posé APRÈS commit+close → aucun conflit de verrou SQLite ─────
    if not deja_init:
        cfg_set(f'_list_init_{nom}', '1')
    valeurs_apres = get_liste(nom)
    logger.info('[liste_supprimer] liste après suppression (%d éléments)', len(valeurs_apres))
    return jsonify({'ok': True, 'valeurs': valeurs_apres})

@app.route('/api/listes/<nom>/reset', methods=['POST'])
@login_required
def api_liste_reset(nom):
    if nom not in LISTE_DEFAULTS:
        return jsonify({'error': 'Liste inconnue'}), 404
    # Supprimer des deux côtés pour éviter la réinjection par la sync
    conn = get_db()
    conn.execute('DELETE FROM config_listes WHERE nom_liste=?', (nom,))
    conn.commit(); conn.close()
    try:
        _t_url   = cfg_get('turso_url',   '').strip()
        _t_token = cfg_get('turso_token', '').strip()
        _db_type = cfg_get('db_type', 'local')
        if _t_url and _t_token:
            if _db_type in ('local', 'sync'):
                from database import TursoConnection as _TC
                _TC(_t_url, _t_token).execute('DELETE FROM config_listes WHERE nom_liste=?', (nom,))
            elif _db_type == 'turso':
                from database import get_local_db as _get_local
                _lconn = _get_local()
                _lconn.execute('DELETE FROM config_listes WHERE nom_liste=?', (nom,))
                _lconn.commit(); _lconn.close()
    except Exception:
        logger.warning('Propagation croisée reset config_listes échouée', exc_info=True)
    # Remettre le flag à 0 APRÈS commit : la liste est de nouveau "non initialisée"
    cfg_set(f'_list_init_{nom}', '0')
    cfg_invalidate()
    return jsonify({'ok': True, 'valeurs': LISTE_DEFAULTS[nom]})


@app.route('/api/services', methods=['GET'])
@login_required
def api_services_get():
    cid = get_client_id()
    if not cid: return jsonify({'error': 'no client'}), 400
    conn = get_db()
    rows = conn.execute('SELECT id,nom,couleur,responsable FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()
    conn.close()
    return jsonify({'services': [{'id':r[0],'nom':r[1],'couleur':r[2],'responsable':r[3]} for r in rows]})

@app.route('/api/services/ajouter', methods=['POST'])
@login_required
def api_services_ajouter():
    cid = get_client_id()
    if not cid: return jsonify({'error': 'no client'}), 400
    nom = (request.json or {}).get('nom', '').strip()
    if not nom: return jsonify({'error': 'Nom vide'}), 400
    now = _utcnow().isoformat()
    conn = get_db()
    try:
        conn.execute('INSERT INTO services (client_id,nom,couleur,ordre,date_creation,date_maj) VALUES (?,?,?,?,?,?)',
            (cid, nom, '#6a8aaa', 0, now, now))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500
    rows = conn.execute('SELECT id,nom,couleur,responsable FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()
    conn.close()
    return jsonify({'ok': True, 'services': [{'id':r[0],'nom':r[1],'couleur':r[2],'responsable':r[3]} for r in rows]})

@app.route('/api/services/supprimer', methods=['POST'])
@login_required
def api_services_supprimer():
    cid = get_client_id()
    if not cid: return jsonify({'error': 'no client'}), 400
    sid = (request.json or {}).get('id')
    if not sid: return jsonify({'error': 'id manquant'}), 400
    conn = get_db()
    conn.execute('DELETE FROM services WHERE id=? AND client_id=?', (sid, cid))
    conn.commit()
    rows = conn.execute('SELECT id,nom,couleur,responsable FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()
    conn.close()
    return jsonify({'ok': True, 'services': [{'id':r[0],'nom':r[1],'couleur':r[2],'responsable':r[3]} for r in rows]})

# --- WATCHDOG PING -----------------------------------------------------------
# Thread de surveillance : ping tous les appareils avec une IP toutes les N sec.
#
# Signalé en usage réel : ParcInfo tourne souvent sur un poste qui se déplace
# physiquement d'un client à l'autre (portable technicien), pas sur un serveur
# fixe. Ce thread pingeait jusqu'ici TOUS les appareils de TOUS les clients à
# chaque cycle, sans savoir sur quel réseau il se trouvait réellement — dès
# qu'on quitte le réseau d'un client, ses appareils passaient à « hors ligne »
# non pas parce qu'ils le sont, mais parce que le ping échoue depuis un autre
# réseau. Chaque cycle détermine maintenant les réseaux locaux actuels de CE
# poste et ne ping (donc ne modifie en base) que les appareils plausiblement
# sur l'un d'eux — les autres gardent leur dernier statut connu, jamais
# écrasé par un faux « hors ligne ».

PING_INTERVAL = 60   # secondes entre deux cycles complets
# Timeout par port lors du fallback TCP (_tcp_probe_rapide) — les ports sont
# désormais sondés EN PARALLÈLE plutôt qu'un par un, donc cette valeur borne
# le pire cas total (hôte injoignable) plutôt que de s'accumuler par port ;
# resserré de 1.0s à 0.4s au passage (correctif de lenteur signalé en usage
# réel), toujours assez patient pour un réseau local normal.
PING_TIMEOUT  = 0.4  # timeout par tentative ping
PING_WORKERS  = 30   # threads simultanes

_ping_cache = {}           # { appareil_id: {en_ligne, ts, ip} }
_ping_cache_lock = threading.Lock()
_watchdog_state  = {'running': False, 'last_cycle': None, 'cycle_count': 0}


def _reseaux_locaux_actuels():
    """Réseaux (/24, best-effort) sur lesquels CE poste se trouve actuellement.

    Un poste technicien a souvent plusieurs interfaces actives à la fois
    (Ethernet + Wi-Fi) : socket.gethostbyname_ex() (même technique que
    collector_core.get_ip_addresses(), côté collecteur) remonte les adresses
    IP de toutes les interfaces connues du système, pas une seule — sans
    dépendance supplémentaire (pas de psutil/netifaces).

    En Docker (RUNNING_IN_DOCKER=1, déploiement officiellement supporté —
    voir Dockerfile / README), cette détection n'a aucun sens : l'IP vue
    par le conteneur est celle du réseau bridge interne (ex. 172.17.0.2),
    jamais celle du LAN du client (ex. 192.168.1.0/24), même quand le
    conteneur PEUT parfaitement joindre ce LAN via le NAT sortant du hôte.
    Sans ce court-circuit, _appareil_sur_reseau_courant() rejette alors
    silencieusement TOUS les appareils à chaque cycle : le watchdog tourne
    (le compteur de cycles avance) mais ne ping plus jamais rien, et le
    statut en ligne/hors ligne affiché reste figé sur sa dernière valeur
    connue — exactement le symptôme d'un appareil réellement en ligne
    affiché à tort comme hors ligne. Retourner un set vide déclenche le
    même repli que « résolution DNS indisponible » : on ping tout, comme
    avant l'introduction de ce filtre (comportement historique)."""
    if os.environ.get('RUNNING_IN_DOCKER'):
        return set()
    reseaux = set()
    try:
        ips = socket.gethostbyname_ex(socket.gethostname())[2]
    except Exception:
        ips = []
    for ip in ips:
        if ip.startswith('127.'):
            continue
        try:
            reseaux.add(ipaddress.ip_network(f'{ip}/24', strict=False))
        except Exception:
            pass
    return reseaux


def _appareil_sur_reseau_courant(ip_appareil, plage_client, reseaux_locaux):
    """Vrai si CE poste est plausiblement sur le même réseau que l'appareil visé.

    Compare les réseaux locaux actuels à la plage IP configurée du client
    (parc_general.plage_ip_locale, la source la plus fiable quand elle est
    renseignée) ET au repli /24 déduit de l'IP de l'appareil lui-même (utile
    quand ce champ est resté sur sa valeur par défaut, ou vide). `overlaps()`
    reste vrai même si l'un des réseaux comparés est un sous-ensemble de
    l'autre (ex : client configuré en /22, poste sur un /24 dedans).

    `reseaux_locaux` vide (résolution DNS locale indisponible) : on ne peut
    rien affirmer sur notre propre position — mieux vaut pinger comme avant
    (comportement historique) que perdre toute surveillance.
    """
    if not reseaux_locaux:
        return True
    cibles = []
    if plage_client:
        try:
            cibles.append(ipaddress.ip_network(plage_client, strict=False))
        except Exception:
            pass
    try:
        cibles.append(ipaddress.ip_network(f'{ip_appareil}/24', strict=False))
    except Exception:
        pass
    return any(reseau_local.overlaps(cible) for reseau_local in reseaux_locaux for cible in cibles)

def _ping_once(ip_str):
    # Timeout subprocess resserré à 1.5s (voir _ping()/_tcp_probe_rapide un
    # peu plus haut dans le fichier — même correctif de performance, cette
    # fonction est le pendant "un seul hôte à la demande" utilisée par le
    # bouton Ping de la baie de brassage et par le watchdog en tâche de
    # fond) ; fallback TCP en parallèle plutôt que port par port, seul
    # vrai goulot d'étranglement du ping quand l'ICMP est filtré (pare-feu
    # Windows par défaut).
    try:
        cmd = ['ping','-n','1','-w','500',ip_str] if IS_WINDOWS else ['ping','-c','1','-W','1',ip_str]
        if _run_hidden(cmd, capture_output=True, timeout=1.5).returncode == 0:
            return True
    except Exception:
        pass
    return _tcp_probe_rapide(ip_str, [80, 443, 22, 445, 3389, 8080, 53, 135, 139], timeout=PING_TIMEOUT)

def _ping_worker(item):
    """Ping puis, si ça répond, vérifie via ARP que c'est bien NOTRE
    appareil (par MAC) qui a répondu sur cette IP — pas un autre.

    _appareil_sur_reseau_courant() ne fait que restreindre AUX appareils
    plausibles (même plage IP) : ça ne suffit pas si deux clients partagent
    la même plage (192.168.1.0/24 par défaut, très courant), ou si une IP a
    été redistribuée par DHCP à une autre machine entre deux visites — dans
    ces cas, l'IP répond mais pour un appareil différent de celui enregistré.
    La MAC, elle, identifie la machine sans ambiguïté (une même IP ne
    prouve rien, une même MAC si). `en_ligne=None` = résultat inconclusif
    (une autre MAC a répondu) : ni « en ligne » ni « hors ligne », le
    dernier statut connu reste inchangé.
    """
    aid, ip, mac_attendu = item
    try:
        en_ligne = _ping_once(ip)
    except Exception:
        en_ligne = False
    mac_vue = ''
    if en_ligne:
        time.sleep(0.5)  # laisser l'OS peupler sa table ARP après le ping
        mac_vue = _mac_from_arp(ip)
        if mac_attendu and mac_vue and mac_vue.lower() != mac_attendu.lower():
            en_ligne = None
    return aid, ip, en_ligne, mac_vue, datetime.now().isoformat()

def _watchdog_cycle():
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT a.id, a.adresse_ip, a.adresse_mac, p.plage_ip_locale "
            "FROM appareils a LEFT JOIN parc_general p ON p.client_id = a.client_id "
            "WHERE a.adresse_ip != '' AND a.adresse_ip IS NOT NULL"
        ).fetchall()
        conn.close()
    except Exception:
        return
    if not rows:
        return
    reseaux_locaux = _reseaux_locaux_actuels()
    items = [(r[0], r[1], r[2]) for r in rows
            if _appareil_sur_reseau_courant(r[1], r[3], reseaux_locaux)]
    # Aucun appareil connu sur le réseau actuel (poste ailleurs que chez un
    # client suivi) : rien à pinger, et surtout rien à écraser en base — mais
    # le cycle a bien eu lieu, l'horodatage ci-dessous avance quand même
    # (sans quoi la pastille « dernier cycle » de la topbar semblerait figée
    # dès qu'on quitte un réseau client, alors que la surveillance tourne
    # normalement).
    if items:
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=PING_WORKERS) as ex:
            for res in ex.map(_ping_worker, items):
                results.append(res)
        try:
            conn = get_db()
            for aid, ip, en_ligne, mac_vue, ts in results:
                if en_ligne is None:
                    # MAC différente de celle attendue : une autre machine
                    # répond sur cette IP (chevauchement de plages entre
                    # clients, ou réattribution DHCP). Inconclusif — on ne
                    # touche pas au dernier statut connu.
                    app.logger.debug(
                        "Watchdog : IP %s répond mais MAC inattendue (appareil #%s) — ignoré", ip, aid)
                    continue
                conn.execute(
                    "UPDATE appareils SET en_ligne=?, dernier_ping=?, "
                    "adresse_mac=COALESCE(NULLIF(adresse_mac,''),?) WHERE id=?",
                    (1 if en_ligne else 0, ts, mac_vue, aid))
                with _ping_cache_lock:
                    _ping_cache[aid] = {'en_ligne': en_ligne, 'ts': ts, 'ip': ip}
            conn.commit(); conn.close()
        except Exception:
            pass
    with _ping_cache_lock:
        # UTC + 'Z' explicite : sans indicateur de fuseau, le JS qui affiche cette
        # heure (new Date(...).toLocaleTimeString(...) dans base.html) l'interprète
        # comme une heure locale déjà correcte et ne fait aucune conversion — d'où un
        # décalage identique au fuseau du serveur vs celui du navigateur (même bug
        # que _sync_state['last_sync'], voir database.py:sync_once).
        _watchdog_state['last_cycle'] = datetime.now(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z')
        _watchdog_state['cycle_count'] += 1

def _watchdog_loop():
    _watchdog_state['running'] = True
    time.sleep(5)
    while True:
        try:
            _watchdog_cycle()
        except Exception:
            pass
        time.sleep(PING_INTERVAL)

_wd_thread = threading.Thread(target=_watchdog_loop, daemon=True, name='PingWatchdog')
_wd_thread.start()

@app.route('/api/ping/statuts')
@login_required
def api_ping_statuts():
    cid = get_client_id()
    conn = get_db()
    rows = conn.execute(
        "SELECT id, en_ligne, dernier_ping, adresse_ip FROM appareils WHERE client_id=?", (cid,)
    ).fetchall()
    conn.close()
    result = {}
    for r in rows:
        aid = r[0]
        with _ping_cache_lock:
            cached = _ping_cache.get(aid)
        if cached:
            result[str(aid)] = {'en_ligne': cached['en_ligne'], 'ts': cached['ts'], 'ip': cached['ip']}
        else:
            result[str(aid)] = {'en_ligne': bool(r[1]), 'ts': r[2] or '', 'ip': r[3] or ''}
    return jsonify({
        'statuts': result,
        'last_cycle': _watchdog_state['last_cycle'],
        'cycle_count': _watchdog_state['cycle_count'],
        'interval': PING_INTERVAL
    })

@app.route('/api/ping/summary')
@login_required
def api_ping_summary():
    '''Résumé léger pour la topbar : nb en ligne, nb total, dernière mise à jour'''
    cid = get_client_id()
    if not cid: return jsonify({'en_ligne': 0, 'total': 0})
    conn = get_db()
    total   = conn.execute("SELECT COUNT(*) FROM appareils WHERE client_id=? AND statut='actif' AND adresse_ip!=''", (cid,)).fetchone()[0]
    en_ligne = conn.execute("SELECT COUNT(*) FROM appareils WHERE client_id=? AND en_ligne=1", (cid,)).fetchone()[0]
    conn.close()
    return jsonify({
        'en_ligne':  en_ligne,
        'total':     total,
        'last_cycle': _watchdog_state.get('last_cycle'),
    })

@app.route('/api/ping/force', methods=['POST'])
@login_required
def api_ping_force():
    threading.Thread(target=_watchdog_cycle, daemon=True).start()
    return jsonify({'started': True})

@app.route('/api/ping/appareil/<int:id>')
@login_required
def api_ping_appareil(id):
    cid = get_client_id()
    conn = get_db()
    row = conn.execute('SELECT id, adresse_ip FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone()
    conn.close()
    if not row or not row[1]:
        return jsonify({'error': 'Appareil sans IP'}), 400
    aid, ip = row[0], row[1]
    en_ligne = _ping_once(ip)
    ts = datetime.now().isoformat()
    conn = get_db()
    conn.execute('UPDATE appareils SET en_ligne=?, dernier_ping=? WHERE id=?', (1 if en_ligne else 0, ts, aid))
    conn.commit(); conn.close()
    with _ping_cache_lock:
        _ping_cache[aid] = {'en_ligne': en_ligne, 'ts': ts, 'ip': ip}
    return jsonify({'id': aid, 'ip': ip, 'en_ligne': en_ligne, 'ts': ts})

# ─── HELPERS ─────────────────────────────────────────────────────────────────

# Types d'appareils qui doivent apparaître automatiquement dans les périphériques
def _sync_licences_from_collector(conn, client_id, appareil_id, licences):
    """Enregistre les clés de licence remontées par le collecteur.

    Seules les licences dont la clé complète a été récupérée arrivent ici : une
    licence numérique Windows ou un Office Click-to-Run n'expose aucune clé, il
    n'y aurait rien à stocker.

    Les lignes existantes ne sont jamais modifiées ni supprimées — une licence
    saisie à la main reste intacte, et la même clé n'est pas ajoutée deux fois.
    Le formulaire de la fiche appareil réenregistre l'ensemble des licences à
    chaque sauvegarde (`_save_licences`), ce qui suffit à les faire persister
    puisqu'elles y sont rendues comme les autres.

    Retourne le nombre de licences ajoutées.
    """
    if not licences:
        return 0

    existantes = set()
    for row in conn.execute(
            'SELECT cle_licence, editeur, produit FROM licences_appareils WHERE appareil_id=?',
            (appareil_id,)).fetchall():
        cle = (row[0] or '').strip().upper()
        if cle:
            existantes.add(cle)

    now = _utcnow().isoformat()
    ajoutees = 0
    for lic in licences[:50]:
        if not isinstance(lic, dict):
            continue
        # Nommage de collector_core : `name` / `full_key`.
        cle = (lic.get('full_key') or '').strip().upper()
        produit = (lic.get('name') or '').strip()
        if not cle or not produit or cle in existantes:
            continue
        existantes.add(cle)
        conn.execute(
            'INSERT INTO licences_appareils'
            ' (appareil_id, client_id, editeur, produit, cle_licence, contrat_id, date_creation)'
            ' VALUES (?,?,?,?,?,NULL,?)',
            (appareil_id, client_id, (lic.get('editeur') or 'Microsoft').strip(),
             produit, cle, now))
        ajoutees += 1
    return ajoutees


# netsh (authentification WLAN) → libellé du select wifi_securite de
# form_identifiant.html. Ce que collector_core.py:_win_wifi_profiles() ne
# reconnaît pas déjà est renvoyé tel quel côté collecteur, pas retenté ici.
_WIFI_SECURITE_VALIDES = {'WPA2', 'WPA3', 'WPA', 'WEP', 'Ouvert'}


def _sync_wifi_credentials_from_collector(conn, client_id, profiles):
    """Enregistre les réseaux Wi-Fi remontés par le collecteur comme identifiants.

    Un réseau déjà connu pour ce client (même SSID) est mis à jour plutôt que
    dupliqué : nom/login/description saisis à la main restent intacts. Le mot
    de passe existant n'est écrasé que si CE relevé en apporte un nouveau —
    une collecte sans la case « mots de passe » cochée ne le vide jamais.

    Retourne (créés, mis_à_jour).
    """
    if not profiles:
        return 0, 0

    existants = {}
    for row in conn.execute(
            "SELECT id, wifi_ssid FROM identifiants WHERE client_id=? AND categorie='Wi-Fi'",
            (client_id,)).fetchall():
        ssid = (row[1] or '').strip()
        if ssid:
            existants[ssid] = row[0]

    crypto = _get_crypto_shared()
    now = _utcnow().isoformat()
    crees = maj = 0

    for profil in profiles[:100]:
        if not isinstance(profil, dict):
            continue
        ssid = (profil.get('ssid') or '').strip()
        if not ssid:
            continue
        securite = profil.get('authentification') or 'WPA2'
        if securite not in _WIFI_SECURITE_VALIDES:
            securite = 'WPA2'
        mot_de_passe_brut = profil.get('password') or ''

        if ssid in existants:
            ident_id = existants[ssid]
            if mot_de_passe_brut:
                conn.execute(
                    'UPDATE identifiants SET wifi_securite=?, mot_de_passe=?, date_maj=? '
                    'WHERE id=? AND client_id=?',
                    (securite, crypto.encrypt(mot_de_passe_brut), now, ident_id, client_id))
            else:
                conn.execute(
                    'UPDATE identifiants SET wifi_securite=?, date_maj=? WHERE id=? AND client_id=?',
                    (securite, now, ident_id, client_id))
            maj += 1
        else:
            mdp_chiffre = crypto.encrypt(mot_de_passe_brut) if mot_de_passe_brut else ''
            conn.execute(
                '''INSERT INTO identifiants
                   (client_id,categorie,nom,login,mot_de_passe,description,wifi_ssid,wifi_securite,
                    date_creation,date_maj)
                   VALUES (?,'Wi-Fi',?,?,?,?,?,?,?,?)''',
                (client_id, ssid, ssid, mdp_chiffre,
                 'Détecté automatiquement par le collecteur', ssid, securite, now, now))
            existants[ssid] = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            crees += 1

    return crees, maj


def _normalize_name(text):
    """Normalise un nom pour comparaison : casse, accents et espaces multiples."""
    import unicodedata
    text = unicodedata.normalize('NFD', str(text or ''))
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    return ' '.join(text.lower().split())


def _utilisateurs_pour_formulaire(conn, client_id):
    """Noms affichables (datalist) + variantes normalisées (détection de non-
    correspondance côté JS) des utilisateurs actifs du client.

    Mêmes variantes que _resolve_utilisateur_id ci-dessous (Prénom Nom, Nom
    Prénom, Nom seul) : le formulaire peut ainsi signaler côté client, sans
    aller-retour serveur, qu'une saisie ne correspondra à aucune fiche —
    piste « champ Utilisateur en texte libre » de l'audit architecture
    (une coquille ou un surnom cassait le rattachement automatique aux
    périphériques sans le moindre signalement).
    """
    rows = conn.execute(
        "SELECT prenom, nom FROM utilisateurs WHERE client_id=? AND statut='actif' ORDER BY nom",
        (client_id,)).fetchall()
    noms = []
    variantes = set()
    for prenom, nom in rows:
        prenom, nom = prenom or '', nom or ''
        affichage = ('%s %s' % (prenom, nom)).strip()
        if affichage:
            noms.append(affichage)
        for candidat in ('%s %s' % (prenom, nom), '%s %s' % (nom, prenom), nom):
            v = _normalize_name(candidat)
            if v:
                variantes.add(v)
    return noms, sorted(variantes)


def _services_pour_formulaire(conn, client_id):
    """Noms des services du client, pour le datalist du champ « Service /
    Département » de la fiche appareil — même mécanique que le datalist
    utilisateurs juste à côté : `appareils.service` reste un champ texte
    libre (pas de service_id), mais suggérer les services déjà créés évite
    qu'un même service se retrouve orthographié de plusieurs façons."""
    rows = conn.execute(
        "SELECT nom FROM services WHERE client_id=? ORDER BY ordre, nom", (client_id,)).fetchall()
    return [r[0] for r in rows if r[0]]


def _resolve_utilisateur_id(conn, client_id, texte):
    """Retrouve l'utilisateur final correspondant au texte libre de la fiche appareil.

    `appareils.utilisateur` est un champ texte libre alors que
    `peripheriques.utilisateur_id` est une clé étrangère vers `utilisateurs` :
    le rapprochement se fait sur le nom, dans les deux ordres possibles
    (« Jean Dupont » comme « Dupont Jean »).

    Aucun utilisateur n'est créé si rien ne correspond : laisser le champ vide
    vaut mieux qu'inventer une fiche utilisateur à partir d'une saisie libre.
    """
    cible = _normalize_name(texte)
    if not cible:
        return None
    rows = conn.execute(
        'SELECT id, nom, prenom FROM utilisateurs WHERE client_id=?', (client_id,)).fetchall()
    for r in rows:
        uid, nom, prenom = r[0], r[1] or '', r[2] or ''
        candidats = {
            _normalize_name('%s %s' % (prenom, nom)),
            _normalize_name('%s %s' % (nom, prenom)),
            _normalize_name(nom),
        }
        candidats.discard('')
        if cible in candidats:
            return uid
    return None


def _resolve_service_id(conn, client_id, texte):
    """Retrouve le service correspondant au texte libre de la fiche appareil
    (appareils.service) — même principe que _resolve_utilisateur_id
    ci-dessus, en plus simple (pas de variantes prénom/nom à essayer)."""
    cible = _normalize_name(texte)
    if not cible:
        return None
    rows = conn.execute('SELECT id, nom FROM services WHERE client_id=?', (client_id,)).fetchall()
    for sid, nom in rows:
        if _normalize_name(nom) == cible:
            return sid
    return None


def _propager_utilisateur_aux_peripheriques(conn, appareil_id, client_id,
                                            ancien_texte, nouveau_texte):
    """Reporte l'utilisateur de la fiche appareil sur ses périphériques.

    Un périphérique rattaché à une machine est, dans les faits, utilisé par la
    personne à qui la machine est affectée. La propagation ne touche que les
    périphériques sans utilisateur ou portant encore l'ancien titulaire : une
    affectation saisie à la main sur un périphérique précis n'est jamais écrasée.
    """
    nouvel_id = _resolve_utilisateur_id(conn, client_id, nouveau_texte)
    ancien_id = _resolve_utilisateur_id(conn, client_id, ancien_texte)
    if nouvel_id is None and ancien_id is None:
        return 0

    # Périphériques liés, via la colonne directe comme via la table pivot.
    rows = conn.execute(
        'SELECT DISTINCT p.id, p.utilisateur_id FROM peripheriques p'
        ' LEFT JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id'
        ' WHERE p.client_id=? AND (p.appareil_id=? OR pa.appareil_id=?)',
        (client_id, appareil_id, appareil_id)).fetchall()

    now = _utcnow().isoformat()
    touches = 0
    for pid, courant in rows:
        if courant is None or (ancien_id is not None and courant == ancien_id):
            if courant == nouvel_id:
                continue
            conn.execute('UPDATE peripheriques SET utilisateur_id=?, date_maj=? WHERE id=?',
                         (nouvel_id, now, pid))
            touches += 1
    return touches


def _usb_identity(device):
    """Clé d'identité stable d'un périphérique USB.

    Sert à retrouver le même matériel d'une collecte à l'autre plutôt que d'en
    recréer une fiche à chaque exécution du collecteur.
    """
    vid = (device.get('vid') or '').upper()
    pid = (device.get('pid') or '').upper()
    serial = (device.get('serial') or '').strip()
    if vid and serial:
        return '%s:%s:%s' % (vid, pid, serial)
    if vid:
        return '%s:%s' % (vid, pid)
    return 'name:' + _normalize_name(device.get('inventory_name') or device.get('name'))


_APPAREIL_PERIPH_MAP = {
    'Imprimante':              'Imprimante',
    'Imprimante multifonction':'Imprimante multifonction',
    'NAS':                     'Disque dur externe',
}

def _sync_appareil_to_periph(conn, appareil_id, client_id):
    """
    Si l'appareil est de type Imprimante / NAS, crée ou met à jour
    l'entrée correspondante dans la table périphériques.
    Le lien est maintenu via la table pivot peripheriques_appareils.
    """
    a = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=?', (appareil_id,)).fetchone() or {})
    if not a:
        return
    categorie = _APPAREIL_PERIPH_MAP.get(a.get('type_appareil', ''))
    if not categorie:
        return
    now = _utcnow().isoformat()
    # Chercher via table pivot
    existing = conn.execute(
        'SELECT p.id FROM peripheriques p'
        ' JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id'
        ' WHERE pa.appareil_id=? AND p.client_id=? AND p.categorie=?',
        (appareil_id, client_id, categorie)).fetchone()
    if existing:
        conn.execute('''UPDATE peripheriques SET
            marque=?, modele=?, numero_serie=?, localisation=?, statut=?,
            date_achat=?, duree_garantie=?, date_fin_garantie=?, fournisseur=?, date_maj=?
            WHERE id=?''',
            (a.get('marque',''), a.get('modele',''), a.get('numero_serie',''),
             a.get('localisation',''), a.get('statut','actif'),
             a.get('date_achat',''), a.get('duree_garantie',0),
             a.get('date_fin_garantie',''), a.get('fournisseur',''), now, existing[0]))
    else:
        conn.execute(
            '''INSERT INTO peripheriques
               (client_id, appareil_id, categorie, marque, modele, numero_serie, localisation,
                statut, date_achat, duree_garantie, date_fin_garantie, fournisseur,
                utilisateur_id, description, prix_achat, numero_commande, notes,
                date_creation, date_maj)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,NULL,'',NULL,'','',?,?)''',
            (client_id, appareil_id, categorie,
             a.get('marque',''), a.get('modele',''), a.get('numero_serie',''),
             a.get('localisation',''), a.get('statut','actif'),
             a.get('date_achat',''), a.get('duree_garantie',0),
             a.get('date_fin_garantie',''), a.get('fournisseur',''), now, now))
        new_pid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute("INSERT OR IGNORE INTO peripheriques_appareils (peripherique_id, appareil_id) VALUES (?,?)",
                     (new_pid, appareil_id))

SW_COURANTS_GROUPS = [
    ('Bureautique',       ['Microsoft 365', 'Microsoft Office', 'LibreOffice', 'Google Workspace']),
    ('Navigateurs',       ['Google Chrome', 'Mozilla Firefox', 'Microsoft Edge', 'Safari']),
    ('Communication',     ['Microsoft Teams', 'Zoom', 'Slack', 'Skype', 'Discord', 'WhatsApp']),
    ('Email',             ['Microsoft Outlook', 'Mozilla Thunderbird']),
    ('Cloud / Stockage',  ['OneDrive', 'SharePoint', 'Dropbox', 'Google Drive', 'iCloud Drive']),
    ('Utilitaires',       ['7-Zip', 'WinRAR', 'VLC', 'Notepad++', 'Adobe Reader', 'Adobe Acrobat Pro', 'PDF24']),
    ('Accès distant',     ['AnyDesk', 'TeamViewer', 'RealVNC', 'Remote Desktop', 'mRemoteNG']),
    ('Développement',     ['Visual Studio Code', 'Git', 'Python', 'Node.js', 'Docker', 'Postman']),
]
# Ensemble plat pour détecter les logiciels personnalisés
SW_COURANTS_ALL = set(sw for _, items in SW_COURANTS_GROUPS for sw in items)


def _extract_form(f):
    prix = None
    try: prix = float(f['prix_achat']) if f.get('prix_achat') else None
    except: pass
    duree = 0
    try: duree = int(f['duree_garantie']) if f.get('duree_garantie') else 0
    except: pass
    av_contrat_id = None
    try: av_contrat_id = int(f['av_contrat_id']) if f.get('av_contrat_id') else None
    except: pass
    edr_contrat_id = None
    try: edr_contrat_id = int(f['edr_contrat_id']) if f.get('edr_contrat_id') else None
    except: pass
    rmm_contrat_id = None
    try: rmm_contrat_id = int(f['rmm_contrat_id']) if f.get('rmm_contrat_id') else None
    except: pass
    return (f.get('nom_machine',''), f.get('type_appareil',''), f.get('marque',''), f.get('modele',''),
            f.get('numero_serie',''), f.get('adresse_ip',''), f.get('adresse_mac',''), f.get('nom_dns',''),
            f.get('utilisateur',''), f.get('service',''), f.get('localisation',''),
            f.get('date_achat',''), duree, f.get('date_fin_garantie',''), f.get('fournisseur',''),
            prix, f.get('numero_commande',''), f.get('os',''), f.get('version_os',''),
            f.get('ram',''), f.get('cpu',''), f.get('stockage',''), f.get('carte_graphique',''),
            f.get('statut','actif'), f.get('notes',''),
            f.get('user_login',''), f.get('user_password',''),
            f.get('admin_login',''), f.get('admin_password',''),
            f.get('anydesk_id',''), f.get('anydesk_password',''),
            f.get('av_marque',''), f.get('av_nom',''),
            f.get('av_date_debut',''), f.get('av_date_fin',''), av_contrat_id,
            f.get('edr_marque',''), f.get('edr_nom',''),
            f.get('edr_date_fin',''), edr_contrat_id,
            f.get('rmm_marque',''), f.get('rmm_nom',''),
            f.get('rmm_agent_id',''), f.get('rmm_date_fin',''), rmm_contrat_id,
            json.dumps(f.getlist('logiciels'), ensure_ascii=False) if f.getlist('logiciels') else '[]',
            1 if f.get('garantie_alerte_ignoree') else 0)


# ─── HISTORIQUE ───────────────────────────────────────────────────────────────

# ── JOURNAL : DIFF & ANNULATION ──────────────────────────────────────────────

import json as _hist_json

_HIST_SENSITIVE = {'user_password', 'admin_password', 'anydesk_password', 'mot_de_passe'}

_HIST_LABELS = {
    'nom_machine':'Nom machine','type_appareil':'Type','marque':'Marque','modele':'Modèle',
    'numero_serie':'N° série','adresse_ip':'Adresse IP','adresse_mac':'Adresse MAC',
    'nom_dns':'Nom DNS','utilisateur':'Utilisateur','service':'Service',
    'localisation':'Localisation','date_achat':'Date achat',
    'duree_garantie':'Garantie (mois)','date_fin_garantie':'Fin garantie',
    'fournisseur':'Fournisseur','prix_achat':'Prix HT','numero_commande':'N° cmd',
    'os':'OS','version_os':'Version OS','ram':'RAM','cpu':'CPU','stockage':'Stockage',
    'carte_graphique':'GPU','statut':'Statut','notes':'Notes',
    'user_login':'Login user','user_password':'MDP user',
    'admin_login':'Login admin','admin_password':'MDP admin',
    'anydesk_id':'AnyDesk ID','anydesk_password':'AnyDesk MDP',
    'categorie':'Catégorie','description':'Description',
    'appareil_id':'Appareil attaché','utilisateur_id':'Utilisateur attaché',
    'peripherique_id':'Périphérique attaché',
    'nom':'Nom','login':'Login','mot_de_passe':'Mot de passe','url':'URL',
    'date_expiration':'Expiration','wifi_ssid':'SSID Wi-Fi','wifi_securite':'Sécurité Wi-Fi',
    'titre':'Titre','type_contrat':'Type contrat','contact_fournisseur':'Contact',
    'email_fournisseur':'Email fournisseur','telephone_fournisseur':'Tél. fournisseur',
    'numero_contrat':'N° contrat','date_debut':'Date début','date_fin':'Date fin',
    'reconduction_auto':'Reconduction auto','preavis_jours':'Préavis (j)',
    'montant_ht':'Montant HT','periodicite':'Périodicité',
    'prenom':'Prénom','poste':'Poste','email':'Email','telephone':'Téléphone',
    'login_windows':'Login Windows','login_mail':'Login mail','service_id':'Service',
}

# Colonnes métier par entité — utilisées pour le diff et la restauration
_ENTITE_COLS = {
    'appareil': ['nom_machine','type_appareil','marque','modele','numero_serie',
        'adresse_ip','adresse_mac','nom_dns','utilisateur','service','localisation',
        'date_achat','duree_garantie','date_fin_garantie','fournisseur','prix_achat',
        'numero_commande','os','version_os','ram','cpu','stockage','carte_graphique',
        'statut','notes','user_login','user_password','admin_login','admin_password',
        'anydesk_id','anydesk_password',
        'av_marque','av_nom','av_date_debut','av_date_fin','av_contrat_id',
        'edr_marque','edr_nom','edr_date_fin','edr_contrat_id',
        'rmm_marque','rmm_nom','rmm_agent_id','rmm_date_fin','rmm_contrat_id'],
    # service_id/utilisateur_id (sur appareils) volontairement absents ici
    # pour la même raison qu'appareil_id ci-dessous : ce sont des colonnes
    # dérivées, recalculées automatiquement depuis service/utilisateur (déjà
    # suivis) à chaque sauvegarde — les garder ferait doublonner le même
    # changement dans l'historique.
    # appareil_id (colonne historique) volontairement absente : le lien
    # appareil<->périphérique vit désormais uniquement dans la table pivot
    # peripheriques_appareils (N:N), jamais écrite par ce formulaire — la
    # garder ici ne faisait que produire de faux écarts dans l'historique
    # (champ form « appareil_ids » au pluriel, jamais « appareil_id »).
    'peripherique': ['categorie','marque','modele','numero_serie','description','localisation',
        'statut','date_achat','duree_garantie','date_fin_garantie','fournisseur','prix_achat',
        'numero_commande','notes','utilisateur_id'],
    'identifiant': ['categorie','nom','login','mot_de_passe','url','description','notes',
        'date_expiration','wifi_ssid','wifi_securite','appareil_id','peripherique_id','utilisateur_id'],
    'contrat': ['titre','type_contrat','fournisseur','contact_fournisseur','email_fournisseur',
        'telephone_fournisseur','numero_contrat','date_debut','date_fin','reconduction_auto',
        'preavis_jours','montant_ht','periodicite','description','notes','statut'],
    'utilisateur': ['prenom','nom','poste','email','telephone','login_windows','login_mail',
        'statut','notes','service_id'],
    'intervention': ['titre','type_intervention','description','notes','date_intervention',
        'heure_debut','heure_fin','duree_minutes','technicien_nom','technicien_email',
        'statut','contrat_id','cout_ht','devise'],
}

_ENTITE_TABLE = {
    'appareil':'appareils','peripherique':'peripheriques',
    'identifiant':'identifiants','contrat':'contrats','utilisateur':'utilisateurs',
    'intervention':'interventions',
}


def _diff_json(avant: dict, apres: dict) -> str:
    """Compare deux dicts métier et retourne JSON {avant:{…},apres:{…}} des seuls champs modifiés.
    Les champs sensibles sont remplacés par ••••."""
    da, dp = {}, {}
    for k in set(avant) | set(apres):
        v1 = str(avant.get(k, '') or '').strip()
        v2 = str(apres.get(k, '') or '').strip()
        if v1 != v2:
            if k in _HIST_SENSITIVE:
                da[k] = '••••' if v1 else ''
                dp[k] = '••••' if v2 else ''
            else:
                da[k] = v1
                dp[k] = v2
    return _hist_json.dumps({'avant': da, 'apres': dp}, ensure_ascii=False) if da else ''


@app.route('/historique/<int:hist_id>/annuler', methods=['POST'])
@login_required
def annuler_historique(hist_id):
    """Restaure une entité à son état avant une modification enregistrée dans l'historique."""
    if not can_write():
        flash('Accès en lecture seule — annulation non autorisée', 'danger')
        return redirect(url_for('page_historique'))
    cid = get_client_id()
    conn = get_db()
    h = row_to_dict(conn.execute(
        'SELECT * FROM historique WHERE id=? AND client_id=?', (hist_id, cid)).fetchone() or {})
    if not h:
        flash('Entrée introuvable', 'danger')
        conn.close(); return redirect(url_for('page_historique'))
    if h.get('action') != 'Modification':
        flash('Seules les modifications peuvent être annulées', 'warning')
        conn.close(); return redirect(url_for('page_historique'))
    try:
        details = _hist_json.loads(h.get('details') or '{}')
        avant = details.get('avant', {})
    except Exception:
        avant = {}
    if not avant:
        flash('Données avant-modification non disponibles pour cette entrée', 'warning')
        conn.close(); return redirect(url_for('page_historique'))
    table = _ENTITE_TABLE.get(h.get('entite', ''))
    allowed_cols = set(_ENTITE_COLS.get(h.get('entite', ''), []))
    if not table:
        flash("Type d'entité non supporté", 'danger')
        conn.close(); return redirect(url_for('page_historique'))
    exists = conn.execute(
        f'SELECT id FROM {table} WHERE id=? AND client_id=?', (h['entite_id'], cid)).fetchone()
    if not exists:
        flash("L'élément n'existe plus, impossible d'annuler", 'danger')
        conn.close(); return redirect(url_for('page_historique'))
    cols = [k for k in avant if k in allowed_cols]
    if not cols:
        flash('Aucune donnée à restaurer', 'warning')
        conn.close(); return redirect(url_for('page_historique'))
    now = _utcnow().isoformat()
    set_clause = ', '.join(f'"{c}"=?' for c in cols) + ', date_maj=?'
    vals = [avant[c] for c in cols] + [now, h['entite_id'], cid]
    conn.execute(f'UPDATE {table} SET {set_clause} WHERE id=? AND client_id=?', vals)
    ts = (h.get('date_action', '')[:16] or '').replace('T', ' ')
    log_history(conn, cid, h['entite'], h['entite_id'], h['entite_nom'], 'Annulation',
                _hist_json.dumps({'message': f'Restauration état avant modification du {ts}'},
                                 ensure_ascii=False))
    conn.commit(); conn.close()
    flash(f'Modification annulée — « {h["entite_nom"]} » restauré à l\'état précédent', 'success')
    return redirect(url_for('page_historique'))


@app.route('/historique/<int:hist_id>/supprimer', methods=['POST'])
@login_required
def supprimer_entree_historique(hist_id):
    """Supprime une entrée individuelle du journal d'historique."""
    cid = get_client_id()
    if not cid:
        return jsonify({'ok': False, 'message': 'Aucun client actif'}), 400
    if not can_write(cid):
        return jsonify({'ok': False, 'message': 'Accès en lecture seule — suppression non autorisée'}), 403
    conn = get_db()
    row = conn.execute('SELECT id FROM historique WHERE id=? AND (client_id=? OR client_id=0)', (hist_id, cid)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'message': 'Entrée introuvable'}), 404
    conn.execute('DELETE FROM historique WHERE id=?', (hist_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/historique/vider-erreurs', methods=['POST'])
@login_required
def vider_erreurs_historique():
    """Supprime toutes les entrées d'erreur système du journal."""
    cid = get_client_id()
    if not cid:
        return jsonify({'ok': False, 'message': 'Aucun client actif'}), 400
    if not can_write(cid):
        return jsonify({'ok': False, 'message': 'Accès en lecture seule — suppression non autorisée'}), 403
    conn = get_db()
    # Compte des lignes à supprimer, pour le retourner dans la réponse
    ids = [r[0] for r in conn.execute(
        "SELECT id FROM historique WHERE (client_id=? OR client_id=0) AND action='Erreur'", (cid,)).fetchall()]
    if ids:
        conn.execute(
            "DELETE FROM historique WHERE (client_id=? OR client_id=0) AND action='Erreur'", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'nb': len(ids)})


@app.route('/historique')
@login_required
def page_historique():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    filtre = request.args.get('entite', '')
    limit  = int(request.args.get('limit', 200))
    if filtre:
        rows = conn.execute(
            "SELECT * FROM historique WHERE (client_id=? OR client_id=0) AND entite=? ORDER BY date_action DESC LIMIT ?",
            (cid, filtre, limit)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM historique WHERE (client_id=? OR client_id=0) ORDER BY date_action DESC LIMIT ?",
            (cid, limit)).fetchall()
    hist = [row_to_dict(r) for r in rows]
    conn.close()
    return render_template('historique.html', hist=hist, client=client,
                           filtre=filtre, can_write_flag=can_write(),
                           clients=get_clients(), client_actif_id=cid)

@app.route('/api/historique/entite/<entite>/<int:entite_id>')
@login_required
def api_historique_entite(entite, entite_id):
    cid = get_client_id()
    conn = get_db()
    rows = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM historique WHERE client_id=? AND entite=? AND entite_id=? ORDER BY date_action DESC LIMIT 20",
        (cid, entite, entite_id)).fetchall()]
    conn.close()
    return jsonify(rows)


import csv, io as _io

# ─── COLONNES EXPORT/IMPORT ──────────────────────────────────────────────────

# Appareils : toutes les colonnes métier (pas id, client_id, ping interne)
COLS_APPAREILS = [
    'nom_machine','type_appareil','marque','modele','numero_serie',
    'adresse_ip','adresse_mac','nom_dns','utilisateur','service','localisation',
    'date_achat','duree_garantie','date_fin_garantie','fournisseur','prix_achat',
    'numero_commande','os','version_os','ram','cpu','stockage','statut',
    'ports_ouverts','notes','user_login','user_password',
    'admin_login','admin_password','anydesk_id','anydesk_password',
    'date_creation','date_maj',
]

# Périphériques
COLS_PERIPHERIQUES = [
    'categorie','marque','modele','numero_serie','description','localisation',
    'statut','date_achat','duree_garantie','date_fin_garantie','fournisseur',
    'prix_achat','numero_commande','notes','date_creation','date_maj',
]

def _xlsx_response(headers, rows, filename, sheet_name='Export'):
    """Construit un classeur Excel (.xlsx) à partir d'en-têtes + lignes et le
    retourne en pièce jointe. Même forme de données que les exports CSV
    existants (liste de colonnes, lignes itérables) pour rester cohérent."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or 'Export'  # 31 car. max imposé par Excel
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    largeurs = [len(str(h)) for h in headers]
    for row in rows:
        values = [v if v is not None else '' for v in row]
        ws.append(values)
        for i, v in enumerate(values):
            largeurs[i] = max(largeurs[i], min(len(str(v)), 60))
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeur + 2

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return app.response_class(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'})


# ─── EXPORT CSV APPAREILS ────────────────────────────────────────────────────

@app.route('/appareils/export.csv')
@login_required
def export_appareils_csv():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    rows = conn.execute(
        f"SELECT {','.join(COLS_APPAREILS)} FROM appareils WHERE client_id=? ORDER BY nom_machine",
        (cid,)).fetchall()
    conn.close()
    out = _io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(COLS_APPAREILS)
    for r in rows:
        w.writerow([str(v) if v is not None else '' for v in r])
    bom = '\ufeff'
    resp = app.response_class(bom + out.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=appareils_export.csv'})
    return resp

@app.route('/appareils/export.xlsx')
@login_required
def export_appareils_xlsx():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    rows = conn.execute(
        f"SELECT {','.join(COLS_APPAREILS)} FROM appareils WHERE client_id=? ORDER BY nom_machine",
        (cid,)).fetchall()
    conn.close()
    return _xlsx_response(COLS_APPAREILS, rows, 'appareils_export.xlsx', 'Appareils')

# ─── EXPORT CSV PÉRIPHÉRIQUES ────────────────────────────────────────────────

@app.route('/peripheriques/export.csv')
@login_required
def export_peripheriques_csv():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    rows = conn.execute(
        f"SELECT {','.join(COLS_PERIPHERIQUES)} FROM peripheriques WHERE client_id=? ORDER BY categorie,marque,modele",
        (cid,)).fetchall()
    conn.close()
    out = _io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(COLS_PERIPHERIQUES)
    for r in rows:
        w.writerow([str(v) if v is not None else '' for v in r])
    bom = '\ufeff'
    resp = app.response_class(bom + out.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=peripheriques_export.csv'})
    return resp

@app.route('/peripheriques/export.xlsx')
@login_required
def export_peripheriques_xlsx():
    cid = get_client_id()
    if not cid: return redirect(url_for('nouveau_client'))
    conn = get_db()
    rows = conn.execute(
        f"SELECT {','.join(COLS_PERIPHERIQUES)} FROM peripheriques WHERE client_id=? ORDER BY categorie,marque,modele",
        (cid,)).fetchall()
    conn.close()
    return _xlsx_response(COLS_PERIPHERIQUES, rows, 'peripheriques_export.xlsx', 'Périphériques')

# ─── IMPORT CSV APPAREILS ────────────────────────────────────────────────────

@app.route('/appareils/import', methods=['POST'])
@login_required
def import_appareils_csv():
    cid = get_client_id()
    if not cid: return redirect(url_for('liste_appareils'))
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('liste_appareils'))
    f = request.files['fichier']
    if not f.filename.lower().endswith('.csv'):
        flash('Le fichier doit être au format CSV', 'danger')
        return redirect(url_for('liste_appareils'))
    try:
        content_bytes = f.read()
        # Handle BOM
        text = content_bytes.decode('utf-8-sig')
        reader = csv.DictReader(_io.StringIO(text), delimiter=';')
        
        # Validate header
        if not reader.fieldnames:
            flash('Fichier CSV vide ou invalide', 'danger')
            return redirect(url_for('liste_appareils'))
        
        missing = [c for c in ['nom_machine'] if c not in reader.fieldnames]
        if missing:
            flash(f'Colonnes manquantes : {", ".join(missing)}. Utilisez le CSV exporté comme modèle.', 'danger')
            return redirect(url_for('liste_appareils'))
        
        now = _utcnow().isoformat()
        conn = get_db()
        inserted = updated = errors = 0
        
        for row in reader:
            try:
                nom = row.get('nom_machine','').strip()
                if not nom: continue
                
                # Vérifier si un appareil avec ce nom existe déjà
                existing = conn.execute(
                    'SELECT id FROM appareils WHERE client_id=? AND nom_machine=?',
                    (cid, nom)).fetchone()
                
                prix = None
                try: prix = float(row.get('prix_achat','')) if row.get('prix_achat','').strip() else None
                except: pass
                duree = 0
                try: duree = int(row.get('duree_garantie','') or 0)
                except: pass
                
                vals = [
                    row.get('nom_machine','').strip(),
                    row.get('type_appareil','').strip(),
                    row.get('marque','').strip(),
                    row.get('modele','').strip(),
                    row.get('numero_serie','').strip(),
                    row.get('adresse_ip','').strip(),
                    row.get('adresse_mac','').strip(),
                    row.get('nom_dns','').strip(),
                    row.get('utilisateur','').strip(),
                    row.get('service','').strip(),
                    row.get('localisation','').strip(),
                    row.get('date_achat','').strip(),
                    duree,
                    row.get('date_fin_garantie','').strip(),
                    row.get('fournisseur','').strip(),
                    prix,
                    row.get('numero_commande','').strip(),
                    row.get('os','').strip(),
                    row.get('version_os','').strip(),
                    row.get('ram','').strip(),
                    row.get('cpu','').strip(),
                    row.get('stockage','').strip(),
                    row.get('statut','actif').strip() or 'actif',
                    row.get('ports_ouverts','').strip(),
                    row.get('notes','').strip(),
                    row.get('user_login','').strip(),
                    row.get('user_password','').strip(),
                    row.get('admin_login','').strip(),
                    row.get('admin_password','').strip(),
                    row.get('anydesk_id','').strip(),
                    row.get('anydesk_password','').strip(),
                ]
                
                if existing:
                    conn.execute(
                        f"""UPDATE appareils SET
                            nom_machine=?,type_appareil=?,marque=?,modele=?,numero_serie=?,
                            adresse_ip=?,adresse_mac=?,nom_dns=?,utilisateur=?,service=?,localisation=?,
                            date_achat=?,duree_garantie=?,date_fin_garantie=?,fournisseur=?,prix_achat=?,
                            numero_commande=?,os=?,version_os=?,ram=?,cpu=?,stockage=?,statut=?,
                            ports_ouverts=?,notes=?,user_login=?,user_password=?,
                            admin_login=?,admin_password=?,anydesk_id=?,anydesk_password=?,
                            date_maj=? WHERE client_id=? AND id=?""",
                        vals + [now, cid, existing[0]])
                    updated += 1
                    log_history(conn, cid, 'appareil', existing[0], nom, 'Mise à jour (import CSV)',
                                {'source': 'import-csv'})
                else:
                    conn.execute(
                        f"""INSERT INTO appareils
                            (nom_machine,type_appareil,marque,modele,numero_serie,
                            adresse_ip,adresse_mac,nom_dns,utilisateur,service,localisation,
                            date_achat,duree_garantie,date_fin_garantie,fournisseur,prix_achat,
                            numero_commande,os,version_os,ram,cpu,stockage,statut,
                            ports_ouverts,notes,user_login,user_password,
                            admin_login,admin_password,anydesk_id,anydesk_password,
                            client_id,date_creation,date_maj)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        vals + [cid, now, now])
                    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    inserted += 1
                    log_history(conn, cid, 'appareil', new_id, nom, 'Création (import CSV)',
                                {'source': 'import-csv'})
            except Exception as e:
                errors += 1
        
        conn.commit(); conn.close()
        msg = f'Import terminé : {inserted} ajouté(s), {updated} mis à jour'
        if errors: msg += f', {errors} erreur(s)'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Erreur lors de l\'import : {str(e)}', 'danger')
    return redirect(url_for('liste_appareils'))

# ─── IMPORT CSV PÉRIPHÉRIQUES ────────────────────────────────────────────────

@app.route('/peripheriques/import', methods=['POST'])
@login_required
def import_peripheriques_csv():
    cid = get_client_id()
    if not cid: return redirect(url_for('liste_peripheriques'))
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('liste_peripheriques'))
    f = request.files['fichier']
    if not f.filename.lower().endswith('.csv'):
        flash('Le fichier doit être au format CSV', 'danger')
        return redirect(url_for('liste_peripheriques'))
    try:
        text = f.read().decode('utf-8-sig')
        reader = csv.DictReader(_io.StringIO(text), delimiter=';')
        if not reader.fieldnames or 'categorie' not in reader.fieldnames:
            flash('Colonne "categorie" manquante. Utilisez le CSV exporté comme modèle.', 'danger')
            return redirect(url_for('liste_peripheriques'))
        
        now = _utcnow().isoformat()
        conn = get_db()
        inserted = updated = errors = 0
        
        for row in reader:
            try:
                cat = row.get('categorie','').strip()
                marque = row.get('marque','').strip()
                modele = row.get('modele','').strip()
                if not cat: continue
                
                prix = None
                try: prix = float(row.get('prix_achat','')) if row.get('prix_achat','').strip() else None
                except: pass
                duree = 0
                try: duree = int(row.get('duree_garantie','') or 0)
                except: pass
                
                # Identifier par categorie+marque+modele+serie
                serie = row.get('numero_serie','').strip()
                existing = None
                if serie:
                    existing = conn.execute(
                        'SELECT id FROM peripheriques WHERE client_id=? AND numero_serie=? AND numero_serie!=""',
                        (cid, serie)).fetchone()
                if not existing and marque and modele:
                    existing = conn.execute(
                        'SELECT id FROM peripheriques WHERE client_id=? AND categorie=? AND marque=? AND modele=?',
                        (cid, cat, marque, modele)).fetchone()
                
                vals = [
                    cat, marque, modele, serie,
                    row.get('description','').strip(),
                    row.get('localisation','').strip(),
                    row.get('statut','actif').strip() or 'actif',
                    row.get('date_achat','').strip(),
                    duree,
                    row.get('date_fin_garantie','').strip(),
                    row.get('fournisseur','').strip(),
                    prix,
                    row.get('numero_commande','').strip(),
                    row.get('notes','').strip(),
                ]
                
                if existing:
                    conn.execute(
                        """UPDATE peripheriques SET
                            categorie=?,marque=?,modele=?,numero_serie=?,description=?,
                            localisation=?,statut=?,date_achat=?,duree_garantie=?,
                            date_fin_garantie=?,fournisseur=?,prix_achat=?,numero_commande=?,
                            notes=?,date_maj=? WHERE client_id=? AND id=?""",
                        vals + [now, cid, existing[0]])
                    updated += 1
                    log_history(conn, cid, 'peripherique', existing[0], f"{marque} {modele}".strip() or cat,
                                'Mise à jour (import CSV)', {'source': 'import-csv'})
                else:
                    conn.execute(
                        """INSERT INTO peripheriques
                            (categorie,marque,modele,numero_serie,description,localisation,
                            statut,date_achat,duree_garantie,date_fin_garantie,fournisseur,
                            prix_achat,numero_commande,notes,client_id,date_creation,date_maj)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        vals + [cid, now, now])
                    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    inserted += 1
                    log_history(conn, cid, 'peripherique', new_id, f"{marque} {modele}".strip() or cat,
                                'Création (import CSV)', {'source': 'import-csv'})
            except Exception as e:
                errors += 1
        
        conn.commit(); conn.close()
        msg = f'Import terminé : {inserted} ajouté(s), {updated} mis à jour'
        if errors: msg += f', {errors} erreur(s)'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Erreur lors de l\'import : {str(e)}', 'danger')
    return redirect(url_for('liste_peripheriques'))


import json as _json, zipfile as _zipfile, tempfile as _tempfile, shutil as _shutil, io as _io2


# ─── KNOWLEDGE BASE ──────────────────────────────────────────────────────────

@app.route('/kb')
@login_required
def page_kb():
    conn = get_db()
    cats = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM kb_categories ORDER BY ordre, nom').fetchall()]
    articles = [row_to_dict(r) for r in conn.execute(
        'SELECT a.*, c.nom as cat_nom, c.icone as cat_icone FROM kb_articles a '
        'JOIN kb_categories c ON a.categorie_id=c.id ORDER BY c.ordre, a.titre').fetchall()]
    conn.close()
    return render_template('kb.html', cats=cats, articles=articles,
                           clients=get_clients(), client_actif_id=get_client_id())

@app.route('/api/kb/search')
@login_required
def api_kb_search():
    q = request.args.get('q', '').lower().strip()
    conn = get_db()
    results = [row_to_dict(r) for r in conn.execute(
        "SELECT a.id, a.titre, a.tags, a.categorie_id, c.nom as cat_nom, c.icone as cat_icone "
        "FROM kb_articles a JOIN kb_categories c ON a.categorie_id=c.id "
        "WHERE lower(a.titre) LIKE ? OR lower(a.contenu) LIKE ? OR lower(a.tags) LIKE ? "
        "ORDER BY c.ordre, a.titre",
        (f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()]
    conn.close()
    return jsonify(results)

@app.route('/api/kb/article/<int:id>')
@login_required
def api_kb_article(id):
    conn = get_db()
    a = row_to_dict(conn.execute(
        'SELECT a.*, c.nom as cat_nom FROM kb_articles a '
        'JOIN kb_categories c ON a.categorie_id=c.id WHERE a.id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(a)

@app.route('/api/kb/article', methods=['POST'])
@login_required
def api_kb_create_article():
    f = request.json or {}
    now = _utcnow().isoformat()
    conn = get_db()
    conn.execute('INSERT INTO kb_articles (categorie_id,titre,contenu,tags,date_creation,date_maj) VALUES (?,?,?,?,?,?)',
        (f.get('categorie_id'), f.get('titre',''), f.get('contenu',''), f.get('tags',''), now, now))
    conn.commit()
    id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    a = row_to_dict(conn.execute('SELECT * FROM kb_articles WHERE id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(a)

@app.route('/api/kb/article/<int:id>', methods=['PUT'])
@login_required
def api_kb_update_article(id):
    f = request.json or {}
    now = _utcnow().isoformat()
    conn = get_db()
    conn.execute('UPDATE kb_articles SET categorie_id=?,titre=?,contenu=?,tags=?,date_maj=? WHERE id=?',
        (f.get('categorie_id'), f.get('titre',''), f.get('contenu',''), f.get('tags',''), now, id))
    conn.commit()
    a = row_to_dict(conn.execute('SELECT * FROM kb_articles WHERE id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(a)

@app.route('/api/kb/article/<int:id>', methods=['DELETE'])
@login_required
def api_kb_delete_article(id):
    conn = get_db()
    conn.execute('DELETE FROM kb_articles WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/kb/categories', methods=['GET'])
@login_required
def api_kb_categories():
    conn = get_db()
    cats = [row_to_dict(r) for r in conn.execute('SELECT * FROM kb_categories ORDER BY ordre,nom').fetchall()]
    conn.close()
    return jsonify(cats)

@app.route('/api/kb/category', methods=['POST'])
@login_required
def api_kb_create_category():
    f = request.json or {}
    conn = get_db()
    conn.execute('INSERT INTO kb_categories (nom,icone,ordre) VALUES (?,?,?)',
        (f.get('nom','Nouvelle categorie'), f.get('icone','📋'), f.get('ordre',99)))
    conn.commit()
    id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    cat = row_to_dict(conn.execute('SELECT * FROM kb_categories WHERE id=?', (id,)).fetchone() or {})
    conn.close()
    return jsonify(cat)

@app.route('/api/kb/category/<int:id>', methods=['DELETE'])
@login_required
def api_kb_delete_category(id):
    conn = get_db()
    conn.execute('DELETE FROM kb_articles WHERE categorie_id=?', (id,))
    conn.execute('DELETE FROM kb_categories WHERE id=?', (id,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ─── EXPORT / IMPORT GLOBAL ──────────────────────────────────────────────────
#
# Stratégie : export JSON complet de toutes les tables, sans les ids (on
# réassigne à l'import). Les fichiers uploadés (documents, photos de baie)
# sont inclus dans un ZIP. Toutes les tables sont exportées dynamiquement
# — si de nouvelles tables sont ajoutées à init_db(), elles seront
# automatiquement incluses à l'export.
#
# Tables exclues de l'export/import :
# ─── EXPORT / IMPORT ────────────────────────────────────────────────────────
#
# Portées disponibles :
#   scope=user   → tous les clients dont l'user est propriétaire + tables globales
#   scope=client → uniquement le client actif
#
# Les nouvelles tables avec colonne client_id sont incluses automatiquement.

TABLES_PAR_CLIENT   = ['clients','parc_general','appareils','peripheriques',
                        'identifiants','services','types_droits','droits_utilisateurs',
                        'baie_slots','baie_photos','utilisateurs','historique']
TABLES_FK_APPAREILS = ['documents_appareils','contrats_appareils']
TABLES_FK_PERIPH    = ['documents_peripheriques','contrats_peripheriques']
TABLES_FK_CONTRAT   = ['documents_contrats']
TABLES_GLOBALES     = ['config','config_listes','outils','kb_categories','kb_articles']
TABLES_AUTH         = ['auth_users','client_partages']

def _table_columns(conn, table):
    return [r[1] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()]

def _all_user_tables(conn):
    return [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
    ).fetchall()]

def _build_export(conn, client_ids, uid, scope_label):
    """Construit le dict d'export filtré pour les client_ids donnés."""
    if not client_ids:
        client_ids = [-1]
    all_tables = set(_all_user_tables(conn))
    data = {'_version':3, '_exported_at':_utcnow().isoformat(),
            '_app':'ParcInfo', '_scope':scope_label, 'tables':{}}

    ph = ','.join(['?' for _ in client_ids])

    # Tables filtrées — 'clients' par id, les autres par client_id
    for t in TABLES_PAR_CLIENT:
        if t not in all_tables: continue
        cols = _table_columns(conn, t)
        if t == 'clients':
            rows = conn.execute(f'SELECT * FROM clients WHERE id IN ({ph})', client_ids).fetchall()
        elif 'client_id' in cols:
            rows = conn.execute(f'SELECT * FROM {t} WHERE client_id IN ({ph})', client_ids).fetchall()
        else:
            rows = conn.execute(f'SELECT * FROM {t}').fetchall()
        data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}

    # IDs dans le périmètre pour les FK
    app_ids  = [r[0] for r in conn.execute(f'SELECT id FROM appareils WHERE client_id IN ({ph})', client_ids).fetchall()]
    peri_ids = [r[0] for r in conn.execute(f'SELECT id FROM peripheriques WHERE client_id IN ({ph})', client_ids).fetchall()]
    ctr_ids  = [r[0] for r in conn.execute(f'SELECT id FROM contrats WHERE client_id IN ({ph})', client_ids).fetchall()]

    for t in TABLES_FK_APPAREILS:
        if t not in all_tables: continue
        cols = _table_columns(conn, t)
        fk   = 'appareil_id' if 'appareil_id' in cols else None
        rows = conn.execute(f'SELECT * FROM {t} WHERE {fk} IN ({",".join(["?"]*len(app_ids))})', app_ids).fetchall()                if fk and app_ids else []
        data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}

    for t in TABLES_FK_PERIPH:
        if t not in all_tables: continue
        cols = _table_columns(conn, t)
        if t == 'documents_peripheriques' and peri_ids:
            rows = conn.execute(f'SELECT * FROM {t} WHERE peripherique_id IN ({",".join(["?"]*len(peri_ids))})', peri_ids).fetchall()
        elif t == 'contrats_peripheriques' and ctr_ids:
            rows = conn.execute(f'SELECT * FROM {t} WHERE contrat_id IN ({",".join(["?"]*len(ctr_ids))})', ctr_ids).fetchall()
        else: rows = []
        data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}

    for t in TABLES_FK_CONTRAT:
        if t not in all_tables or not ctr_ids: continue
        cols = _table_columns(conn, t)
        rows = conn.execute(f'SELECT * FROM {t} WHERE contrat_id IN ({",".join(["?"]*len(ctr_ids))})', ctr_ids).fetchall()
        data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}

    if 'contrats' in all_tables:
        cols = _table_columns(conn, 'contrats')
        rows = conn.execute(f'SELECT * FROM contrats WHERE client_id IN ({ph})', client_ids).fetchall()
        data['tables']['contrats'] = {'columns':cols, 'rows':[list(r) for r in rows]}

    for t in TABLES_GLOBALES:
        if t not in all_tables: continue
        cols = _table_columns(conn, t)
        rows = conn.execute(f'SELECT * FROM {t}').fetchall()
        data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}

    # Nouvelles tables inconnues avec client_id — incluses automatiquement
    known = set(TABLES_PAR_CLIENT+TABLES_FK_APPAREILS+TABLES_FK_PERIPH+
                TABLES_FK_CONTRAT+TABLES_GLOBALES+TABLES_AUTH+['contrats'])
    for t in all_tables:
        if t in known: continue
        cols = _table_columns(conn, t)
        if 'client_id' in cols:
            rows = conn.execute(f'SELECT * FROM {t} WHERE client_id IN ({ph})', client_ids).fetchall()
            data['tables'][t] = {'columns':cols, 'rows':[list(r) for r in rows]}
    return data

def _get_doc_filenames(conn, client_ids):
    ph = ','.join(['?' for _ in client_ids])
    fnames = set()
    for t, col in [('documents_appareils','nom_fichier'),('documents_peripheriques','nom_fichier'),
                   ('documents_contrats','nom_fichier'),('baie_photos','nom_fichier')]:
        try:
            for r in conn.execute(f'SELECT {col} FROM {t} WHERE client_id IN ({ph})', client_ids).fetchall():
                if r[0]: fnames.add(r[0])
        except: pass
    return fnames

def _make_json_response(data, fname):
    out = _json.dumps(data, ensure_ascii=False, indent=2, default=str)
    return app.response_class(out, mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename={fname}.json'})

def _make_zip_response(data, fname, file_names):
    json_str = _json.dumps(data, ensure_ascii=False, indent=2, default=str)
    buf = _io2.BytesIO()
    with _zipfile.ZipFile(buf, 'w', _zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{fname}/data.json', json_str)
        for fn in file_names:
            fp = os.path.join(UPLOAD_FOLDER, fn)
            if os.path.isfile(fp):
                zf.write(fp, f'{fname}/uploads/{fn}')
    buf.seek(0)
    return app.response_class(buf.read(), mimetype='application/zip',
        headers={'Content-Disposition': f'attachment; filename={fname}.zip'})

def _export_client_ids_for_user(conn, uid):
    role = (conn.execute('SELECT role FROM auth_users WHERE id=?', (uid,)).fetchone() or ['user'])[0]
    if role == 'admin':
        return [r[0] for r in conn.execute('SELECT id FROM clients ORDER BY id').fetchall()]
    return [r[0] for r in conn.execute('SELECT id FROM clients WHERE auth_user_id=?', (uid,)).fetchall()]

@app.route('/export/global.json')
@login_required
def export_global_json():
    uid   = session.get('auth_user_id')
    scope = request.args.get('scope', 'user')
    cid   = get_client_id()
    conn  = get_db()
    from datetime import date as _date
    today = _date.today().isoformat()
    if scope == 'client' and cid:
        cl    = row_to_dict(conn.execute('SELECT nom FROM clients WHERE id=?', (cid,)).fetchone() or {})
        slug  = cl.get('nom','client').replace(' ','_')[:30]
        data  = _build_export(conn, [cid], uid, f'client:{cid}')
        fname = f'parcinfo_{slug}_{today}'
    else:
        cids  = _export_client_ids_for_user(conn, uid)
        data  = _build_export(conn, cids, uid, f'user:{uid}')
        fname = f'parcinfo_user_{today}'
    conn.close()
    return _make_json_response(data, fname)

@app.route('/export/global.zip')
@login_required
def export_global_zip():
    uid   = session.get('auth_user_id')
    scope = request.args.get('scope', 'user')
    cid   = get_client_id()
    conn  = get_db()
    from datetime import date as _date
    today = _date.today().isoformat()
    if scope == 'client' and cid:
        cl     = row_to_dict(conn.execute('SELECT nom FROM clients WHERE id=?', (cid,)).fetchone() or {})
        slug   = cl.get('nom','client').replace(' ','_')[:30]
        data   = _build_export(conn, [cid], uid, f'client:{cid}')
        fnames = _get_doc_filenames(conn, [cid])
        fname  = f'parcinfo_{slug}_{today}'
    else:
        cids   = _export_client_ids_for_user(conn, uid)
        data   = _build_export(conn, cids, uid, f'user:{uid}')
        fnames = _get_doc_filenames(conn, cids)
        fname  = f'parcinfo_user_{today}'
    conn.close()
    return _make_zip_response(data, fname, fnames)

@app.route('/import/global', methods=['POST'])
@login_required
def import_global():
    mode  = request.form.get('mode', 'merge')
    scope = request.form.get('scope', 'user')
    uid   = session.get('auth_user_id')
    cid   = get_client_id()
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('parc_general'))
    f_up  = request.files['fichier']
    fname = f_up.filename.lower()
    try:
        json_str = None
        if fname.endswith('.zip'):
            buf = _io2.BytesIO(f_up.read())
            with _zipfile.ZipFile(buf) as zf:
                jfiles = [n for n in zf.namelist() if n.endswith('/data.json') or n == 'data.json']
                if not jfiles:
                    flash('data.json introuvable dans le ZIP', 'danger')
                    return redirect(url_for('parc_general'))
                json_str = zf.read(jfiles[0]).decode('utf-8')
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                for uf in zf.namelist():
                    if '/uploads/' in uf and not uf.endswith('/'):
                        fn = uf.split('/uploads/')[-1]
                        if fn:
                            with zf.open(uf) as src, open(os.path.join(UPLOAD_FOLDER, fn), 'wb') as dst:
                                dst.write(src.read())
        elif fname.endswith('.json'):
            json_str = f_up.read().decode('utf-8')
        else:
            flash('Format non supporté — utilisez .json ou .zip', 'danger')
            return redirect(url_for('parc_general'))

        data        = _json.loads(json_str)
        tables_data = data.get('tables', {})
        if not tables_data:
            flash('Fichier JSON invalide ou vide', 'danger')
            return redirect(url_for('parc_general'))

        conn = get_db()
        conn.execute('PRAGMA foreign_keys = OFF')

        if mode == 'reset':
            if scope == 'client' and cid:
                for t in reversed(TABLES_PAR_CLIENT):
                    if 'client_id' in _table_columns(conn, t):
                        try: conn.execute(f'DELETE FROM {t} WHERE client_id=?', (cid,))
                        except: pass
            else:
                my_cids = _export_client_ids_for_user(conn, uid)
                if my_cids:
                    ph = ','.join(['?' for _ in my_cids])
                    for t in reversed(TABLES_PAR_CLIENT):
                        if 'client_id' in _table_columns(conn, t):
                            try: conn.execute(f'DELETE FROM {t} WHERE client_id IN ({ph})', my_cids)
                            except: pass
                for t in TABLES_GLOBALES:
                    try: conn.execute(f'DELETE FROM {t}')
                    except: pass

        stats = {}
        for table, tdata in tables_data.items():
            cols = tdata.get('columns', [])
            rows = tdata.get('rows', [])
            if not cols or not rows: stats[table] = 0; continue
            try: conn.execute(f'SELECT 1 FROM {table} LIMIT 1')
            except: stats[table] = 0; continue
            inserted = 0
            for row in rows:
                if len(row) != len(cols): continue
                row_dict = dict(zip(cols, row))
                ph_v     = ','.join(['?' for _ in cols])
                col_str  = ','.join([f'"{c}"' for c in cols])
                if mode == 'merge':
                    pk     = row_dict.get('id') or row_dict.get('cle')
                    id_col = 'id' if 'id' in row_dict else ('cle' if 'cle' in row_dict else None)
                    if pk is not None and id_col:
                        if conn.execute(f'SELECT 1 FROM {table} WHERE {id_col}=?', (pk,)).fetchone():
                            continue
                try:
                    conn.execute(f'INSERT OR IGNORE INTO {table} ({col_str}) VALUES ({ph_v})', row)
                    inserted += 1
                except Exception: pass
            stats[table] = inserted

        conn.execute('PRAGMA foreign_keys = ON')
        conn.commit(); conn.close()
        total    = sum(stats.values())
        non_zero = {t:n for t,n in stats.items() if n>0}
        portee   = 'client courant' if scope == 'client' else 'tous vos clients'
        msg = f'Import {mode} ({portee}) — {total} entrée(s) importée(s)'
        if non_zero:
            detail = ', '.join(f'{t}:{n}' for t,n in list(non_zero.items())[:6])
            msg += f' ({detail}{"..." if len(non_zero)>6 else ""})'
        flash(msg, 'success')
    except Exception as e:
        flash(f"Erreur lors de l'import : {str(e)}", 'danger')
    return redirect(url_for('parc_general'))

# ─── AUTHENTIFICATION ─────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def page_login():
    if session.get('auth_user_id'):
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        ip = request.remote_addr or '0.0.0.0'
        if not _check_rate_limit(ip):
            logger.warning('Login rate limit dépassé pour %s', ip)
            error = 'Trop de tentatives. Réessayez dans 5 minutes.'
            return render_template('login.html', error=error, next=request.args.get('next',''))
        login = request.form.get('login','').strip()
        pwd   = request.form.get('password','')
        conn  = get_db()
        u = row_to_dict(conn.execute(
            'SELECT * FROM auth_users WHERE login=? AND actif=1', (login,)).fetchone() or {})
        conn.close()
        ok, needs_rehash = _check_pwd(pwd, u.get('password_hash','')) if u else (False, False)
        if ok:
            _reset_attempts(ip)
            if needs_rehash:
                conn2 = get_db()
                conn2.execute('UPDATE auth_users SET password_hash=? WHERE id=?',
                              (_hash_pwd(pwd), u['id']))
                conn2.commit(); conn2.close()
            session['auth_user_id'] = u['id']
            session['auth_user_nom'] = (u.get('prenom','') + ' ' + u.get('nom','')).strip() or u['login']
            session['auth_user_role'] = u.get('role','user')
            session['login_time'] = _utcnow().isoformat()
            from urllib.parse import urlparse
            raw_next = request.form.get('next') or request.args.get('next') or '/'
            parsed = urlparse(raw_next)
            next_url = raw_next if (not parsed.netloc and raw_next.startswith('/')) else '/'
            if u.get('must_change_password'):
                # Le changement de mot de passe forcé reste sur la page bureau
                # (aucune version mobile de ce formulaire) — mais la destination
                # d'origine est conservée pour y renvoyer une fois le mot de
                # passe changé, sinon un login depuis /m atterrissait sur le
                # bureau sans retour possible vers la version mobile.
                return redirect(url_for('page_profil', next=next_url))
            return redirect(next_url)
        _record_failed_attempt(ip)
        error = 'Identifiants incorrects'
    if error is None:
        # login.html est une page autonome (pas d'extends base.html) : elle
        # n'affiche jamais get_flashed_messages(). Des messages utiles
        # (timeout de session dans login_required, redirections diverses)
        # étaient donc flashés puis silencieusement perdus.
        _flashes = get_flashed_messages()
        if _flashes:
            error = _flashes[0]
    return render_template('login.html', error=error,
                           next=request.args.get('next',''))

@app.route('/logout')
def page_logout():
    session.clear()
    return redirect(url_for('page_login'))

def _mode_execution():
    """'windows' | 'macos' | 'docker' | 'sources' — détermine ce que /apropos
    peut proposer (Quitter/Redémarrer n'ont de sens que sur un exécutable)."""
    if getattr(_sys, 'frozen', False):
        return 'macos' if platform.system() == 'Darwin' else 'windows'
    return 'docker' if os.environ.get('RUNNING_IN_DOCKER') else 'sources'


@app.route('/apropos')
@login_required
def apropos():
    user = get_auth_user()
    mode = _mode_execution()
    est_frozen = mode in ('windows', 'macos')

    # != 'local', pas seulement == 'turso' : le mode le plus courant en usage
    # réel (Docker/PC/Mac qui se synchronisent entre eux) est 'sync' — base
    # locale au quotidien, synchronisée en tâche de fond avec Turso — pas
    # 'turso' (chaque requête interroge Turso directement, sans base locale).
    # Cette page affichait « Turso n'est pas configuré » sur toute instance en
    # mode 'sync', qui pourtant synchronise activement : la carte de statut
    # censée servir à diagnostiquer un souci de sync était donc invisible
    # exactement là où elle aurait le plus servi.
    turso_actif = cfg_get('db_type', 'local') != 'local'
    sync_state = None
    if turso_actif:
        from database import get_sync_state
        sync_state = get_sync_state()

    uptime_s = int(time.time() - _APP_DEMARRAGE)
    jours, reste = divmod(uptime_s, 86400)
    heures, reste = divmod(reste, 3600)
    minutes = reste // 60
    if jours:
        uptime_fmt = '%d j %d h' % (jours, heures)
    elif heures:
        uptime_fmt = '%d h %d min' % (heures, minutes)
    else:
        uptime_fmt = '%d min' % minutes

    return render_template(
        'apropos.html',
        mode=mode,
        peut_arreter=est_frozen and bool(user),
        turso_actif=turso_actif, sync_state=sync_state,
        host=request.host, url_racine=request.url_root,
        uptime_fmt=uptime_fmt,
    )


@app.route('/apropos/quitter', methods=['POST'])
@login_required
def apropos_quitter():
    # Ouvert à tout compte connecté, pas seulement admin : sur un poste local
    # (exécutable portable), c'est souvent le seul moyen de fermer proprement
    # l'application — pas d'icône de barre système sur macOS (désactivée,
    # voir launcher.run_systray), pas toujours d'icône dans le Dock non plus
    # (signalé en usage réel, macOS Intel). L'auteur est tracé.
    user = get_auth_user()
    if not user:
        return jsonify({'ok': False, 'error': 'Non authentifié'}), 403
    if _mode_execution() not in ('windows', 'macos'):
        return jsonify({'ok': False, 'error': "Indisponible hors exécutable (Docker/sources)"}), 400
    logger.info("Arrêt de ParcInfo demandé par %s", user['login'])
    from launcher import quitter_application
    quitter_application(logger)
    return jsonify({'ok': True, 'message': 'Arrêt en cours…'})


@app.route('/apropos/redemarrer', methods=['POST'])
@login_required
def apropos_redemarrer():
    # Voir apropos_quitter ci-dessus : ouvert à tout compte connecté.
    user = get_auth_user()
    if not user:
        return jsonify({'ok': False, 'error': 'Non authentifié'}), 403
    if _mode_execution() not in ('windows', 'macos'):
        return jsonify({'ok': False, 'error': "Indisponible hors exécutable (Docker/sources)"}), 400
    logger.info("Redémarrage de ParcInfo demandé par %s", user['login'])
    from launcher import redemarrer_application
    redemarrer_application(logger)
    return jsonify({'ok': True, 'message': 'Redémarrage en cours…'})


@app.route('/profil', methods=['GET','POST'])
@login_required
def page_profil():
    u = get_auth_user()
    if not u:
        return redirect(url_for('page_login'))
    if request.method == 'POST':
        errs = validate_form([('email', 'email', False)], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(url_for('page_profil'))
        conn = get_db()
        now  = _utcnow().isoformat()
        nom  = request.form.get('nom','').strip()
        prenom = request.form.get('prenom','').strip()
        email  = request.form.get('email','').strip()
        pwd    = request.form.get('password','')
        pwd2   = request.form.get('password2','')
        # Logo upload
        logo_fichier = u.get('logo_fichier','')
        if 'logo' in request.files and request.files['logo'].filename:
            logo = request.files['logo']
            # L'extension venait telle quelle du nom fourni : n'importe quel
            # type de fichier atterrissait dans le dossier des pièces jointes.
            ok_logo, motif_logo = verifier_fichier(logo, ALLOWED_IMAGE_EXTENSIONS)
            if not ok_logo:
                flash(motif_logo, 'danger')
                return redirect(request.referrer or url_for('admin_users'))
            ext  = extension_de(logo.filename)
            fname = f"logo_user{u['id']}_{int(time.time())}.{ext}"
            logo.save(os.path.join(UPLOAD_FOLDER, fname))
            logo_fichier = fname
        from urllib.parse import urlparse as _urlparse
        raw_next = request.form.get('next') or ''
        _parsed_next = _urlparse(raw_next)
        next_url = raw_next if (raw_next and not _parsed_next.netloc and raw_next.startswith('/')) else ''
        if pwd:
            if pwd != pwd2:
                flash('Les mots de passe ne correspondent pas', 'danger')
                return redirect(url_for('page_profil', next=next_url) if next_url else url_for('page_profil'))
            conn.execute('UPDATE auth_users SET nom=?,prenom=?,email=?,password_hash=?,logo_fichier=?,must_change_password=0,date_maj=? WHERE id=?',
                (nom, prenom, email, _hash_pwd(pwd), logo_fichier, now, u['id']))
        else:
            if u.get('must_change_password'):
                flash('Vous devez définir un nouveau mot de passe.', 'danger')
                conn.close()
                return redirect(url_for('page_profil', next=next_url) if next_url else url_for('page_profil'))
            conn.execute('UPDATE auth_users SET nom=?,prenom=?,email=?,logo_fichier=?,date_maj=? WHERE id=?',
                (nom, prenom, email, logo_fichier, now, u['id']))
        conn.commit(); conn.close()
        session['auth_user_nom'] = (prenom + ' ' + nom).strip() or u['login']
        flash('Profil mis à jour', 'success')
        return redirect(next_url) if next_url else redirect(url_for('page_profil'))
    return render_template('profil.html', u=u,
                           clients=get_clients(), client_actif_id=get_client_id())

# ─── ADMIN UTILISATEURS ───────────────────────────────────────────────────────

@app.route('/admin/utilisateurs')
@login_required
def admin_utilisateurs():
    u = get_auth_user()
    if not u or u.get('role') != 'admin':
        flash('Acces reserve a l\'administrateur', 'danger')
        return redirect(url_for('index'))
    conn = get_db()
    users = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM auth_users ORDER BY role DESC, nom').fetchall()]
    conn.close()
    return render_template('admin_users.html', users=users,
                           clients=get_clients(), client_actif_id=get_client_id())

@app.route('/admin/utilisateur/nouveau', methods=['GET','POST'])
@login_required
def admin_nouvel_utilisateur():
    u = get_auth_user()
    if not u or u.get('role') != 'admin':
        return redirect(url_for('index'))
    if request.method == 'POST':
        login  = request.form.get('login','').strip()
        pwd    = request.form.get('password','')
        nom    = request.form.get('nom','').strip()
        prenom = request.form.get('prenom','').strip()
        email  = request.form.get('email','').strip()
        role   = request.form.get('role','user')
        if not login or not pwd:
            flash('Login et mot de passe requis', 'danger')
            return redirect(request.url)
        errs = validate_form([('email', 'email', False)], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        conn = get_db()
        exists = conn.execute('SELECT id FROM auth_users WHERE login=?', (login,)).fetchone()
        if exists:
            conn.close()
            flash('Ce login est déjà utilisé', 'danger')
            return redirect(request.url)
        now = _utcnow().isoformat()
        conn.execute('INSERT INTO auth_users (login,password_hash,nom,prenom,email,role,actif,date_creation,date_maj) VALUES (?,?,?,?,?,?,1,?,?)',
            (login, _hash_pwd(pwd), nom, prenom, email, role, now, now))
        conn.commit(); conn.close()
        flash(f'Utilisateur {login} créé', 'success')
        return redirect(url_for('admin_utilisateurs'))
    return render_template('admin_user_form.html', edit_user=None,
                           clients=get_clients(), client_actif_id=get_client_id())

@app.route('/admin/utilisateur/<int:uid>/editer', methods=['GET','POST'])
@login_required
def admin_editer_utilisateur(uid):
    current = get_auth_user()
    if not current or current.get('role') != 'admin':
        return redirect(url_for('index'))
    conn = get_db()
    edit_user = row_to_dict(conn.execute('SELECT * FROM auth_users WHERE id=?', (uid,)).fetchone() or {})
    if not edit_user:
        conn.close(); flash('Utilisateur introuvable', 'danger')
        return redirect(url_for('admin_utilisateurs'))
    if request.method == 'POST':
        errs = validate_form([('email', 'email', False)], request.form)
        if errs:
            for e in errs: flash(e, 'danger')
            return redirect(request.url)
        now  = _utcnow().isoformat()
        nom    = request.form.get('nom','').strip()
        prenom = request.form.get('prenom','').strip()
        email  = request.form.get('email','').strip()
        role   = request.form.get('role','user')
        actif  = 1 if request.form.get('actif') else 0
        pwd    = request.form.get('password','')
        if pwd:
            conn.execute('UPDATE auth_users SET nom=?,prenom=?,email=?,role=?,actif=?,password_hash=?,date_maj=? WHERE id=?',
                (nom, prenom, email, role, actif, _hash_pwd(pwd), now, uid))
        else:
            conn.execute('UPDATE auth_users SET nom=?,prenom=?,email=?,role=?,actif=?,date_maj=? WHERE id=?',
                (nom, prenom, email, role, actif, now, uid))
        conn.commit(); conn.close()
        flash('Utilisateur mis à jour', 'success')
        return redirect(url_for('admin_utilisateurs'))
    conn.close()
    return render_template('admin_user_form.html', edit_user=edit_user,
                           clients=get_clients(), client_actif_id=get_client_id())

@app.route('/admin/utilisateur/<int:uid>/supprimer', methods=['POST'])
@login_required
def admin_supprimer_utilisateur(uid):
    current = get_auth_user()
    if not current or current.get('role') != 'admin':
        return redirect(url_for('index'))
    if uid == current['id']:
        flash('Impossible de supprimer son propre compte', 'danger')
        return redirect(url_for('admin_utilisateurs'))
    conn = get_db()
    # Réattribuer les clients à l'admin
    admin = conn.execute("SELECT id FROM auth_users WHERE role='admin' AND id!=?", (uid,)).fetchone()
    if admin:
        conn.execute('UPDATE clients SET auth_user_id=? WHERE auth_user_id=?', (admin[0], uid))
    conn.execute('DELETE FROM client_partages WHERE auth_user_id=?', (uid,))
    conn.execute('DELETE FROM auth_users WHERE id=?', (uid,))
    conn.commit(); conn.close()
    flash('Utilisateur supprimé', 'info')
    return redirect(url_for('admin_utilisateurs'))


@app.route('/admin/email-config', methods=['GET','POST'])
@login_required
def admin_email_config():
    user = get_auth_user()
    if not user or user.get('role') != 'admin':
        return redirect(url_for('index'))

    if request.method == 'POST':
        cfg_set('smtp_server', request.form.get('smtp_server', ''))
        cfg_set('smtp_port', request.form.get('smtp_port', '587'))
        cfg_set('smtp_login', request.form.get('smtp_login', ''))
        cfg_set('smtp_password', request.form.get('smtp_password', ''))
        cfg_set('from_email', request.form.get('from_email', ''))
        flash('Paramètres email sauvegardés', 'success')
        return redirect(url_for('admin_email_config'))

    return render_template('admin_email_config.html',
        smtp_server=cfg_get('smtp_server', ''),
        smtp_port=cfg_get('smtp_port', '587'),
        smtp_login=cfg_get('smtp_login', ''),
        from_email=cfg_get('from_email', ''))


# ─── PARTAGE DE CLIENTS ───────────────────────────────────────────────────────

@app.route('/client/<int:cid>/partager', methods=['GET','POST'])
@login_required
def partager_client(cid):
    u = get_auth_user()
    if not u:
        return redirect(url_for('page_login'))
    if get_client_access(cid) != 'proprietaire':
        flash('Seul le propriétaire peut partager ce client', 'danger')
        return redirect(url_for('liste_clients'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    partages = [row_to_dict(r) for r in conn.execute(
        'SELECT cp.*, au.login, au.nom, au.prenom FROM client_partages cp JOIN auth_users au ON cp.auth_user_id=au.id WHERE cp.client_id=?',
        (cid,)).fetchall()]
    all_users = [row_to_dict(r) for r in conn.execute(
        'SELECT id,login,nom,prenom FROM auth_users WHERE id!=? AND actif=1 ORDER BY nom',
        (u['id'],)).fetchall()]
    if request.method == 'POST':
        action = request.form.get('action')
        now = _utcnow().isoformat()
        if action == 'ajouter':
            target_uid = request.form.get('user_id')
            niveau     = request.form.get('niveau','lecture')
            if target_uid:
                conn.execute('INSERT OR REPLACE INTO client_partages (client_id,auth_user_id,niveau,date_partage) VALUES (?,?,?,?)',
                    (cid, int(target_uid), niveau, now))
                conn.commit()
                flash('Partage ajouté', 'success')
        elif action == 'supprimer':
            partage_id = request.form.get('partage_id')
            if partage_id:
                conn.execute('DELETE FROM client_partages WHERE id=? AND client_id=?', (int(partage_id), cid))
                conn.commit()
                flash('Partage supprimé', 'info')
        conn.close()
        return redirect(url_for('partager_client', cid=cid))
    conn.close()
    return render_template('partage_client.html', client=client, partages=partages,
                           all_users=all_users, clients=get_clients(), client_actif_id=get_client_id())

@app.route('/user/logo/<path:filename>')
@login_required
def user_logo(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# ── TABLEAU DE BORD ──────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    return redirect(url_for('index'))


# ── CACHE STATS (Admin) ──────────────────────────────────────────────────────

@app.route('/api/cache/stats')
@login_required
def cache_stats():
    """Retourne les statistiques du cache."""
    user = get_auth_user()
    if user['role'] != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    cache_mgr = get_cache_manager()
    stats = cache_mgr.stats()

    return jsonify({
        'entries': stats['entries'],
        'total_hits': stats['total_hits'],
        'avg_hits_per_entry': round(stats['avg_hits'], 2),
        'message': f"✅ Cache: {stats['entries']} entrées, {stats['total_hits']} hits"
    })


@app.route('/api/cache/invalidate', methods=['POST'])
@login_required
def cache_invalidate():
    """Invalide le cache (Admin uniquement)."""
    user = get_auth_user()
    if user['role'] != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    pattern = request.form.get('pattern', '')
    invalidate_cache_pattern(pattern)

    return jsonify({'ok': True, 'message': f'Cache invalidé: {pattern or "tout"}'})


# ── RECHERCHE FULL-TEXT ET AUTOCOMPLETE ──────────────────────────────────────

@app.route('/api/search')
@login_required
def api_search():
    """Recherche globale multi-entités."""
    query = request.args.get('q', '').strip()
    client_id = get_client_id()
    limit = min(int(request.args.get('limit', 20)), 100)

    if not query or len(query) < 2:
        return jsonify({
            'appareils': [],
            'contrats': [],
            'utilisateurs': [],
            'services': [],
            'peripheriques': [],
            'identifiants': [],
            'total': 0,
            'query': query
        })

    try:
        results = search_global(query, client_id, limit)
        return jsonify(results)
    except Exception as e:
        logger.exception(f"Search error for query='{query}'")
        return jsonify({'error': str(e)}), 500


@app.route('/api/autocomplete/<entity_type>')
@login_required
def api_autocomplete(entity_type):
    """Autocomplete pour un type d'entité spécifique."""
    query = request.args.get('q', '').strip()
    client_id = get_client_id()
    limit = min(int(request.args.get('limit', 10)), 50)

    if not query or len(query) < 1:
        return jsonify([])

    try:
        results = search_autocomplete(query, client_id, entity_type, limit)
        return jsonify(results)
    except Exception as e:
        logger.exception(f"Autocomplete error for entity_type='{entity_type}', query='{query}'")
        return jsonify({'error': str(e)}), 500


# ── GESTIONNAIRES D'ERREURS ──────────────────────────────────────────────────

@app.errorhandler(404)
def page_not_found(e):
    cid = get_client_id()
    return render_template('404.html',
        clients=get_clients(), client_actif_id=cid,
        client=row_to_dict(get_db().execute('SELECT * FROM clients WHERE id=?',(cid,)).fetchone() or {}) if cid else {}
    ), 404

@app.errorhandler(500)
def internal_error(e):
    import traceback as _tb
    try:
        cid = session.get('client_id')
        if cid:
            from database import get_local_db
            _conn = get_local_db()
            log_error(_conn, int(cid), request.url, e, _tb.format_exc())
            _conn.commit(); _conn.close()
    except Exception:
        pass
    cid = get_client_id()
    return render_template('500.html',
        clients=get_clients() if cid else [], client_actif_id=cid,
        client={}
    ), 500


@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    """Capture les exceptions non gérées, les logue et retourne une 500."""
    import traceback as _tb
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e   # Laisser Flask gérer les erreurs HTTP normales (404, 403…)
    try:
        cid = session.get('client_id')
        if cid:
            from database import get_local_db
            _conn = get_local_db()
            log_error(_conn, int(cid), request.url, e, _tb.format_exc())
            _conn.commit(); _conn.close()
    except Exception:
        pass
    logger.exception('Exception non gérée sur %s %s', request.method, request.url)
    cid = get_client_id()
    return render_template('500.html',
        clients=get_clients() if cid else [], client_actif_id=cid,
        client={}
    ), 500


# ─── GENERATEUR QR CODES POUR ETIQUETTES ────────────────────────────────────

@app.route('/qrcode-labels', methods=['GET'])
@login_required
def qrcode_labels():
    """Affiche formulaire pour générer étiquettes QR AVERY J8159."""
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))

    conn = get_db()
    appareils = [row_to_dict(r) for r in conn.execute(
        'SELECT id, nom_machine, adresse_ip, user_login FROM appareils WHERE client_id=? ORDER BY nom_machine',
        (cid,)).fetchall()]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        'SELECT id, categorie || \' - \' || marque || \' \' || modele as nom '
        'FROM peripheriques WHERE client_id=? ORDER BY categorie, marque, modele',
        (cid,)).fetchall()]
    conn.close()

    return render_template('qrcode_generator.html',
        appareils=appareils, peripheriques=peripheriques,
        clients=get_clients(), client_actif_id=cid)


@app.route('/qrcode-labels/fields', methods=['POST'])
@login_required
def qrcode_fields():
    """Retourne les champs disponibles pour un appareil/périphérique."""
    cid = get_client_id()
    asset_type = request.form.get('asset_type', 'appareil')
    asset_id = request.form.get('asset_id', '', type=int)

    if not cid or not asset_id:
        return jsonify({'error': 'Invalid request'}), 400

    conn = get_db()

    try:
        if asset_type == 'appareil':
            row = conn.execute(
                'SELECT * FROM appareils WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
            asset = row_to_dict(row) if row else {}
        else:  # peripherique
            row = conn.execute(
                'SELECT * FROM peripheriques WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
            asset = row_to_dict(row) if row else {}
    finally:
        conn.close()

    if not asset:
        return jsonify({'error': 'Asset not found'}), 404

    # Decrypt credentials if present
    crypto = _get_crypto_shared()
    if 'user_password' in asset and asset['user_password']:
        try:
            asset['user_password'] = crypto.decrypt(asset['user_password'])
        except:
            asset['user_password'] = ''
    if 'admin_password' in asset and asset['admin_password']:
        try:
            asset['admin_password'] = crypto.decrypt(asset['admin_password'])
        except:
            asset['admin_password'] = ''

    # Return available fields
    fields = {field: asset.get(field, '') for field in asset.keys()
              if asset.get(field) not in [None, '', 0, False]}

    return jsonify({'asset': asset, 'fields': fields})


@app.route('/qrcode-labels/preview', methods=['POST'])
@login_required
def qrcode_preview():
    """Génère aperçu du label en image PNG."""
    import os
    import io
    import tempfile
    from qrcode_helper import create_label_image

    cid = get_client_id()
    if not cid:
        return jsonify({'error': 'Not authenticated'}), 403

    import json
    asset_type = request.form.get('asset_type', 'appareil')
    asset_id = request.form.get('asset_id', '', type=int)
    selected_fields = request.form.getlist('selected_fields')
    custom_text = request.form.get('custom_text', '')[:200]
    logo_file = request.files.get('logo')

    # Parse template parameters
    template = {
        # QR Code settings
        'qrSize': request.form.get('qrSize', 'medium'),
        'qrPosition': request.form.get('qrPosition', 'center'),
        'qrColor': request.form.get('qrColor', '#000000'),
        'qrBgColor': request.form.get('qrBgColor', '#ffffff'),
        'qrBorder': int(request.form.get('qrBorder', 1)),
        # Logo settings
        'logoSize': float(request.form.get('logoSize', 8)),  # in mm
        'logoPosition': request.form.get('logoPosition', 'auto'),
        'logoOpacity': int(request.form.get('logoOpacity', 100)),
        'logoBorder': request.form.get('logoBorder', 'false').lower() == 'true',
        'logoBorderColor': request.form.get('logoBorderColor', '#000000'),
        'logoBorderWidth': int(request.form.get('logoBorderWidth', 1)),
        # Header text settings
        'customHeader': request.form.get('customHeader', ''),
        'headerSize': int(request.form.get('headerSize', 12)),
        'headerColor': request.form.get('headerColor', '#000000'),
        'headerFont': request.form.get('headerFont', 'arial'),
        'headerAlign': request.form.get('headerAlign', 'left'),
        # Asset fields text settings
        'assetSize': int(request.form.get('assetSize', 10)),
        'assetColor': request.form.get('assetColor', '#000000'),
        'assetFont': request.form.get('assetFont', 'arial'),
        'assetAlign': request.form.get('assetAlign', 'left'),
        # Footer text settings
        'customFooter': request.form.get('customFooter', ''),
        'footerSize': int(request.form.get('footerSize', 10)),
        'footerColor': request.form.get('footerColor', '#000000'),
        'footerFont': request.form.get('footerFont', 'arial'),
        'footerAlign': request.form.get('footerAlign', 'left'),
        # Background
        'bgColor': request.form.get('bgColor', '#ffffff')
    }

    # Get asset
    conn = get_db()
    try:
        if asset_type == 'appareil':
            row = conn.execute(
                'SELECT * FROM appareils WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
        else:
            row = conn.execute(
                'SELECT * FROM peripheriques WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
        asset = row_to_dict(row) if row else {}
    finally:
        conn.close()

    if not asset:
        return jsonify({'error': 'Asset not found'}), 404

    # Filter and decrypt
    crypto = _get_crypto_shared()
    filtered_asset = {'type': asset_type, 'id': asset_id}

    for field in selected_fields:
        if field in asset:
            value = asset[field]
            if field in ['user_password', 'admin_password'] and value:
                try:
                    value = crypto.decrypt(value)
                except:
                    value = ''
            filtered_asset[field] = value

    # Save logo if provided
    logo_path = None
    if logo_file and logo_file.filename:
        try:
            _, logo_path = tempfile.mkstemp(suffix='.png')
            logo_file.save(logo_path)
        except:
            logo_path = None

    # Generate label image
    try:
        logger.info(f"Template customFooter: '{template.get('customFooter', '')}'")
        logger.info(f"Custom text param: '{custom_text}'")
        logger.info(f"Calling create_label_image with: asset_data={filtered_asset}, template={template}")
        label_img = create_label_image(filtered_asset, logo_path, custom_text, template)
        logger.info(f"create_label_image returned: {type(label_img)}")
        if isinstance(label_img, str):
            logger.error(f"CRITICAL: create_label_image returned string: {label_img}")
    except Exception as e:
        logger.exception(f"Exception in create_label_image")
        return jsonify({'error': f"Label creation failed: {str(e)}"}), 500
    finally:
        if logo_path:
            try:
                os.remove(logo_path)
            except:
                pass

    # Verify we have an image
    if isinstance(label_img, str):
        logger.error(f"ERROR: label_img is string: {label_img}")
        return jsonify({'error': f"Label creation returned error: {label_img}"}), 500

    if not isinstance(label_img, Image.Image):
        logger.error(f"ERROR: label_img is {type(label_img)}, not PIL Image!")
        return jsonify({'error': f"Label image is invalid type: {type(label_img)}"}), 500

    # Convert to PNG bytes
    img_bytes = io.BytesIO()
    label_img.save(img_bytes, format='PNG')
    img_bytes.seek(0)

    return send_file(img_bytes, mimetype='image/png')


@app.route('/qrcode-labels/generate', methods=['POST'])
@login_required
def qrcode_generate():
    """Génère PDF pour impression sur AVERY J8159."""
    import os
    import json
    import tempfile
    from qrcode_helper import create_label_image, create_pdf_sheet

    cid = get_client_id()
    if not cid:
        return jsonify({'error': 'Not authenticated'}), 403

    if not can_write(cid):
        return jsonify({'error': 'Forbidden'}), 403

    asset_type = request.form.get('asset_type', 'appareil')
    asset_id = request.form.get('asset_id', '', type=int)
    selected_fields_json = request.form.get('fields', '[]')
    custom_text = request.form.get('custom_text', '')[:200]
    positions_json = request.form.get('positions', '{}')
    logo_file = request.files.get('logo')

    # Parse template parameters
    template = {
        # QR Code settings
        'qrSize': request.form.get('qrSize', 'medium'),
        'qrPosition': request.form.get('qrPosition', 'center'),
        'qrColor': request.form.get('qrColor', '#000000'),
        'qrBgColor': request.form.get('qrBgColor', '#ffffff'),
        'qrBorder': int(request.form.get('qrBorder', 1)),
        # Logo settings
        'logoSize': float(request.form.get('logoSize', 8)),  # in mm
        'logoPosition': request.form.get('logoPosition', 'auto'),
        'logoOpacity': int(request.form.get('logoOpacity', 100)),
        'logoBorder': request.form.get('logoBorder', 'false').lower() == 'true',
        'logoBorderColor': request.form.get('logoBorderColor', '#000000'),
        'logoBorderWidth': int(request.form.get('logoBorderWidth', 1)),
        # Header text settings
        'customHeader': request.form.get('customHeader', ''),
        'headerSize': int(request.form.get('headerSize', 12)),
        'headerColor': request.form.get('headerColor', '#000000'),
        'headerFont': request.form.get('headerFont', 'arial'),
        'headerAlign': request.form.get('headerAlign', 'left'),
        # Asset fields text settings
        'assetSize': int(request.form.get('assetSize', 10)),
        'assetColor': request.form.get('assetColor', '#000000'),
        'assetFont': request.form.get('assetFont', 'arial'),
        'assetAlign': request.form.get('assetAlign', 'left'),
        # Footer text settings
        'customFooter': request.form.get('customFooter', ''),
        'footerSize': int(request.form.get('footerSize', 10)),
        'footerColor': request.form.get('footerColor', '#000000'),
        'footerFont': request.form.get('footerFont', 'arial'),
        'footerAlign': request.form.get('footerAlign', 'left'),
        # Background
        'bgColor': request.form.get('bgColor', '#ffffff')
    }

    try:
        selected_fields = json.loads(selected_fields_json)
        positions_dict = json.loads(positions_json)
        positions_dict = {int(k): int(v) for k, v in positions_dict.items()}
    except:
        return jsonify({'error': 'Invalid parameters'}), 400

    # Validate positions
    if not positions_dict or any(p < 1 or p > 24 for p in positions_dict.keys()):
        return jsonify({'error': 'Invalid positions (1-24)'}), 400

    # Get asset
    conn = get_db()
    try:
        if asset_type == 'appareil':
            row = conn.execute(
                'SELECT * FROM appareils WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
        else:
            row = conn.execute(
                'SELECT * FROM peripheriques WHERE id=? AND client_id=?',
                (asset_id, cid)).fetchone()
        asset = row_to_dict(row) if row else {}
    finally:
        conn.close()

    if not asset:
        return jsonify({'error': 'Asset not found'}), 404

    # Filter and decrypt
    crypto = _get_crypto_shared()
    filtered_asset = {'type': asset_type, 'id': asset_id}

    for field in selected_fields:
        if field in asset:
            value = asset[field]
            if field in ['user_password', 'admin_password'] and value:
                try:
                    value = crypto.decrypt(value)
                except:
                    value = ''
            filtered_asset[field] = value

    # Save logo if provided
    logo_path = None
    if logo_file and logo_file.filename:
        try:
            _, logo_path = tempfile.mkstemp(suffix='.png')
            logo_file.save(logo_path)
        except:
            logo_path = None

    # Generate label image
    try:
        label_img = create_label_image(filtered_asset, logo_path, custom_text, template)
        pdf_bytes = create_pdf_sheet(label_img, positions_dict)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if logo_path:
            try:
                os.remove(logo_path)
            except:
                pass

    # Log action
    user = get_auth_user()
    log_conn = get_db()
    try:
        asset_name = asset.get('nom_machine') or asset.get('nom') or f'{asset_type}#{asset_id}'
        log_history(log_conn, cid, asset_type.upper(), asset_id, asset_name,
            'GENERATE_QR_LABELS',
            json.dumps({'fields': selected_fields, 'positions': len(positions_dict)}))
        log_conn.commit()
    finally:
        log_conn.close()

    # Return PDF
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'label_{asset_id}_{_utcnow().strftime("%Y%m%d")}.pdf')


# ─── MDNS REGISTRATION ────────────────────────────────────────────────────────
_mdns_instance = None
_mdns_service_name = None  # nom réellement enregistré — _unregister_mdns doit annoncer EXACTEMENT ce nom, pas en reconstruire un autre


def _nom_service_mdns(hostname):
    """Nom de service mDNS unique par instance (hostname inclus).

    Plusieurs instances ParcInfo (Docker + PC + Mac, par exemple) qui
    s'annonceraient toutes sous le même nom entreraient en conflit sur le
    réseau — une seule serait annoncée, les autres resteraient invisibles à
    la découverte, sans qu'aucune erreur ne le signale (zeroconf ne prévient
    pas d'un conflit après coup, il choisit juste un nom différent ou perd
    silencieusement la course). Le nom fixe `"ParcInfo._http._tcp.local."`
    d'origine ne permettait donc de découvrir qu'une seule instance à la fois
    sur un même réseau.
    """
    propre = re.sub(r'[^A-Za-z0-9 _-]', '', hostname or '').strip() or 'Poste'
    return f"ParcInfo sur {propre}._http._tcp.local."


def _register_mdns(port=3456):
    """Annonce cette instance sur le réseau local via mDNS.

    Sert à deux choses : l'ancien raccourci http://parcinfo.local (ne
    fonctionne de façon fiable qu'avec une seule instance sur le réseau —
    limite déjà présente avant ce changement, pas introduite ici) et la
    découverte automatique par le collecteur (get_all_mac_addresses() côté
    collecteur, ServiceBrowser côté client) — celle-ci, elle, fonctionne bien
    avec plusieurs instances grâce au nom de service unique par poste.
    """
    global _mdns_instance, _mdns_service_name
    if not MDNS_AVAILABLE:
        return

    try:
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        nom_service = _nom_service_mdns(hostname)

        info = ServiceInfo(
            "_http._tcp.local.",
            name=nom_service,
            addresses=[socket.inet_aton(local_ip)],
            port=port,
            properties={
                "path": "/",
                "version": APP_VERSION,
                "description": "IT Asset Management",
                "hostname": hostname,
                "docker": "1" if os.environ.get('RUNNING_IN_DOCKER') else "0",
            },
            server="parcinfo.local."
        )

        _mdns_instance = Zeroconf()
        _mdns_instance.register_service(info)
        _mdns_service_name = nom_service
        logger.info(f"✅ mDNS registered: {nom_service} → http://{local_ip}:{port}")
    except Exception as e:
        logger.warning(f"⚠️ mDNS registration failed: {e}")

def _unregister_mdns(port=3456):
    """Unregister mDNS service on shutdown"""
    global _mdns_instance
    if _mdns_instance:
        try:
            _mdns_instance.unregister_service(ServiceInfo(
                "_http._tcp.local.",
                name=_mdns_service_name or _nom_service_mdns(socket.gethostname()),
                addresses=[],
                port=port
            ))
            _mdns_instance.close()
        except Exception as e:
            logger.warning(f"⚠️ mDNS unregistration failed: {e}")


# ─── MOBILE (PWA lecture seule) ──────────────────────────────────────────────
#
# Interface optimisée smartphone, dédiée à la consultation : aucune route ici
# n'accepte POST/PUT/DELETE (une seule exception : le changement de client
# actif, un simple aiguillage de session sans écriture métier). Elle réutilise
# les mêmes helpers ACL (get_client_id/get_client_access/get_clients) et de
# formatage (fmt_appareils/fmt_contrat) que les pages bureau pour rester
# cohérente avec elles sans dupliquer leur logique.
#
# Les mots de passe suivent la même politique que le reste de l'app :
# masqués par défaut, révélés à la demande. Pour les identifiants chiffrés
# (table `identifiants`), on ne déchiffre jamais côté serveur avant l'appel
# explicite à /api/identifiant/<id>/mdp (plus strict que la liste bureau, qui
# déchiffre puis ne les affiche pas). Les mots de passe d'appareil et le Wi-Fi
# du parc général ne sont pas chiffrés en base (comportement déjà existant
# côté bureau) : ils sont simplement masqués côté client via JS.

_MOBILE_LIST_CAP = 500


@app.route('/m')
@app.route('/m/')
@login_required
def mobile_home():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    return redirect(url_for('mobile_dashboard'))


@app.route('/m/manifest.webmanifest')
def mobile_manifest():
    return send_from_directory(
        os.path.join(app.static_folder, 'mobile'), 'manifest.webmanifest',
        mimetype='application/manifest+json')


@app.route('/m/sw.js')
def mobile_service_worker():
    return send_from_directory(
        os.path.join(app.static_folder, 'mobile'), 'sw.js',
        mimetype='application/javascript')


@app.route('/m/client/<int:id>/selectionner')
@login_required
def mobile_selectionner_client(id):
    if not get_client_access(id):
        flash('Accès refusé à ce client', 'danger')
        return redirect(url_for('mobile_dashboard'))
    session['client_id'] = id
    return redirect(url_for('mobile_dashboard'))


@app.route('/m/dashboard')
@login_required
def mobile_dashboard():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    today = date.today()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    stats = _compute_client_dashboard_stats(conn, cid, today)
    alerts = _compute_alerts_for_client(conn, cid, today)
    valeur_parc = sum(a.get('prix_achat') or 0 for a in stats['appareils'])
    conn.close()
    return render_template('mobile/dashboard.html', client=client, clients=get_clients(),
                           client_actif_id=cid, stats=stats, alerts=alerts, valeur_parc=valeur_parc)


@app.route('/m/appareils')
@login_required
def mobile_appareils():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    rows = conn.execute(
        f'SELECT {_colonnes_appareils_liste()} FROM appareils a WHERE a.client_id=? '
        f'ORDER BY a.nom_machine ASC LIMIT ?', (cid, _MOBILE_LIST_CAP)).fetchall()
    appareils = fmt_appareils([row_to_dict(r) for r in rows])
    total = conn.execute('SELECT COUNT(*) FROM appareils WHERE client_id=?', (cid,)).fetchone()[0]
    conn.close()
    return render_template('mobile/appareils.html', appareils=appareils, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           types_appareils=get_liste_cached('types_appareils'),
                           total=total)


@app.route('/m/appareil/<int:id>')
@login_required
def mobile_appareil_detail(id):
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    a = row_to_dict(conn.execute('SELECT * FROM appareils WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if not a:
        conn.close()
        flash('Appareil introuvable', 'danger')
        return redirect(url_for('mobile_appareils'))
    a = fmt_appareils([a])[0]
    peripheriques = [row_to_dict(r) for r in conn.execute(
        '''SELECT p.* FROM peripheriques p
           JOIN peripheriques_appareils pa ON pa.peripherique_id = p.id
           WHERE pa.appareil_id=? ORDER BY p.categorie''', (id,)).fetchall()]
    contrats_lies = [fmt_contrat(row_to_dict(r)) for r in conn.execute(
        '''SELECT c.* FROM contrats c
           JOIN contrats_appareils ca ON ca.contrat_id=c.id
           WHERE ca.appareil_id=? ORDER BY c.titre''', (id,)).fetchall()]
    docs = [row_to_dict(r) for r in conn.execute(
        '''SELECT id, nom, description, type_doc, nom_fichier, taille, date_upload
           FROM documents_appareils WHERE appareil_id=? ORDER BY date_upload DESC''', (id,)).fetchall()]
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
    licences = [row_to_dict(r) for r in conn.execute(
        '''SELECT editeur, produit, cle_licence FROM licences_appareils
           WHERE appareil_id=? ORDER BY id''', (id,)).fetchall()]
    # Les clés BitLocker restent chiffrées ici : seules les métadonnées
    # (volume, protection) voyagent jusqu'à la page, la valeur ne part que
    # sur demande via /api/appareil/<id>/cle-bitlocker (déjà auditée).
    cles_bitlocker = [row_to_dict(r) for r in conn.execute(
        '''SELECT volume, identifiant, protection, chiffrement FROM cles_recuperation
           WHERE appareil_id=? AND client_id=? ORDER BY volume''', (id, cid)).fetchall()]
    conn.close()
    return render_template('mobile/appareil_detail.html', appareil=a, peripheriques=peripheriques,
                           contrats_lies=contrats_lies, docs=docs, licences=licences,
                           cles_bitlocker=cles_bitlocker,
                           client=client, clients=get_clients(), client_actif_id=cid)


@app.route('/m/baie')
@login_required
def mobile_baie():
    """Consultation lecture seule de la baie de brassage — demandé : c'est
    justement l'écran qu'un technicien veut voir depuis son téléphone,
    debout devant l'armoire, pas de quoi le modifier depuis là. Pas de
    grille glisser-déposer (impraticable sur petit écran) : une liste
    verticale par position U, dans l'ordre du rack."""
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    baies = [r[0] for r in conn.execute(
        "SELECT DISTINCT COALESCE(baie_nom,'Baie principale') FROM baie_slots WHERE client_id=? ORDER BY 1",
        (cid,)).fetchall()]
    if not baies:
        baies = ['Baie principale']
    baie_nom = request.args.get('baie') or baies[0]
    if baie_nom not in baies:
        baie_nom = baies[0]
    cond = "(s.baie_nom=? OR s.baie_nom IS NULL)" if baie_nom == 'Baie principale' else "s.baie_nom=?"
    slots = [row_to_dict(r) for r in conn.execute(
        f'''SELECT s.*, a.nom_machine, a.type_appareil, a.adresse_ip, a.en_ligne,
                  p.categorie AS p_categorie, p.marque AS p_marque, p.modele AS p_modele
           FROM baie_slots s LEFT JOIN appareils a ON s.appareil_id=a.id
                             LEFT JOIN peripheriques p ON s.peripherique_id=p.id
           WHERE s.client_id=? AND {cond} ORDER BY s.position''', (cid, baie_nom)).fetchall()]
    for s in slots:
        s['ports'] = _ports_avec_details(conn, s['id']) if s.get('nb_ports') else []
        # Un port "actif" l'est aussi via sa prise murale (bandeau RJ) — voir
        # _prises_murales_avec_details ; p.piece/p.nom_cible directement sur
        # le port restent utiles pour tout élément qui n'est pas un bandeau.
        s['ports_actifs'] = [p for p in s['ports'] if p.get('nom_cible') or p.get('piece')
                             or (p.get('prise_murale') and (p['prise_murale'].get('nom_cible') or p['prise_murale'].get('piece')))]
        s['nom_affiche'] = s.get('nom_custom') or s.get('nom_machine') or (
            ' '.join(filter(None, [s.get('p_categorie'), s.get('p_marque'), s.get('p_modele')]))) \
            or s.get('type_equipement') or f"U{s['position']}"
    conn.close()
    return render_template('mobile/baie.html', client=client, clients=get_clients(), client_actif_id=cid,
                           slots=slots, baies=baies, baie_nom=baie_nom)


@app.route('/m/contrats')
@login_required
def mobile_contrats():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    rows = conn.execute(
        'SELECT * FROM contrats WHERE client_id=? ORDER BY date_fin, titre LIMIT ?',
        (cid, _MOBILE_LIST_CAP)).fetchall()
    contrats = [fmt_contrat(row_to_dict(r)) for r in rows]
    total = conn.execute('SELECT COUNT(*) FROM contrats WHERE client_id=?', (cid,)).fetchone()[0]
    conn.close()
    return render_template('mobile/contrats.html', contrats=contrats, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           types_contrats=get_liste_cached('types_contrats'), total=total)


@app.route('/m/contrat/<int:id>')
@login_required
def mobile_contrat_detail(id):
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    ct = row_to_dict(conn.execute('SELECT * FROM contrats WHERE id=? AND client_id=?', (id, cid)).fetchone() or {})
    if not ct:
        conn.close()
        flash('Contrat introuvable', 'danger')
        return redirect(url_for('mobile_contrats'))
    ct = fmt_contrat(ct)
    appareils_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT a.* FROM appareils a JOIN contrats_appareils ca ON a.id=ca.appareil_id WHERE ca.contrat_id=?',
        (id,)).fetchall()]
    periph_lies = [row_to_dict(r) for r in conn.execute(
        'SELECT p.* FROM peripheriques p JOIN contrats_peripheriques cp ON p.id=cp.peripherique_id WHERE cp.contrat_id=?',
        (id,)).fetchall()]
    docs = [row_to_dict(r) for r in conn.execute(
        '''SELECT id, nom, description, type_doc, nom_fichier, taille, date_upload
           FROM documents_contrats WHERE contrat_id=? ORDER BY date_upload DESC''', (id,)).fetchall()]
    for d in docs:
        d['taille_fmt'] = human_size(d.get('taille', 0))
    interventions = [fmt_intervention(row_to_dict(r)) for r in conn.execute(
        'SELECT * FROM interventions WHERE contrat_id=? AND statut != ? ORDER BY date_intervention DESC LIMIT 10',
        (id, 'archivee')).fetchall()]
    conn.close()
    return render_template('mobile/contrat_detail.html', contrat=ct, appareils_lies=appareils_lies,
                           periph_lies=periph_lies, docs=docs, interventions=interventions,
                           client=client, clients=get_clients(), client_actif_id=cid)


@app.route('/m/utilisateurs')
@login_required
def mobile_utilisateurs():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    users = [row_to_dict(r) for r in conn.execute(
        '''SELECT u.*, s.nom as service_nom, s.couleur as service_couleur
           FROM utilisateurs u LEFT JOIN services s ON u.service_id=s.id
           WHERE u.client_id=? ORDER BY s.nom, u.nom, u.prenom LIMIT ?''',
        (cid, _MOBILE_LIST_CAP)).fetchall()]
    services = [row_to_dict(r) for r in conn.execute(
        'SELECT * FROM services WHERE client_id=? ORDER BY ordre,nom', (cid,)).fetchall()]
    total = conn.execute('SELECT COUNT(*) FROM utilisateurs WHERE client_id=?', (cid,)).fetchone()[0]
    conn.close()
    return render_template('mobile/utilisateurs.html', utilisateurs=users, services=services,
                           client=client, clients=get_clients(), client_actif_id=cid, total=total)


@app.route('/m/identifiants')
@login_required
def mobile_identifiants():
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    rows = conn.execute(
        'SELECT * FROM identifiants WHERE client_id=? ORDER BY categorie, nom LIMIT ?',
        (cid, _MOBILE_LIST_CAP)).fetchall()
    # Volontairement PAS de déchiffrement ici : mot_de_passe reste chiffré
    # dans le contexte du template, révélé uniquement via l'appel AJAX à
    # /api/identifiant/<id>/mdp (cf. note de section plus haut).
    ids_ = [row_to_dict(r) for r in rows]
    for i in ids_:
        if i.get('date_expiration'):
            try:
                d = date.fromisoformat(i['date_expiration'])
                i['expire_bientot'] = (d - date.today()).days <= 30
                i['expire_depasse'] = d < date.today()
                i['date_expiration_fmt'] = d.strftime('%d/%m/%Y')
            except Exception:
                i['expire_bientot'] = i['expire_depasse'] = False
                i['date_expiration_fmt'] = ''
        else:
            i['expire_bientot'] = i['expire_depasse'] = False
            i['date_expiration_fmt'] = ''
    parc = row_to_dict(conn.execute('SELECT * FROM parc_general WHERE client_id=?', (cid,)).fetchone() or {})
    total = conn.execute('SELECT COUNT(*) FROM identifiants WHERE client_id=?', (cid,)).fetchone()[0]
    conn.close()
    wifi_parc = []
    if parc.get('wifi_ssid'):
        wifi_parc.append({
            'nom': parc['wifi_ssid'] + ' (Parc général)',
            'login': parc.get('wifi_ssid', ''),
            'mot_de_passe': parc.get('wifi_password', ''),
            'description': 'Réseau principal — depuis le Parc général',
        })
    if parc.get('wifi_ssid2'):
        wifi_parc.append({
            'nom': parc['wifi_ssid2'] + ' (Parc général)',
            'login': parc.get('wifi_ssid2', ''),
            'mot_de_passe': parc.get('wifi_password2', ''),
            'description': 'Réseau invités — depuis le Parc général',
        })
    return render_template('mobile/identifiants.html', identifiants=ids_, wifi_parc=wifi_parc, client=client,
                           clients=get_clients(), client_actif_id=cid,
                           categories=get_liste_cached('categories_identifiants'), total=total)


@app.route('/m/diag-reseau')
@login_required
def mobile_diag_reseau():
    """Diagnostic réseau — consultation mobile, lecture seule (évènements actifs)."""
    cid = get_client_id()
    if not cid:
        return redirect(url_for('nouveau_client'))
    conn = get_db()
    client = row_to_dict(conn.execute('SELECT * FROM clients WHERE id=?', (cid,)).fetchone() or {})
    rows = conn.execute(
        "SELECT e.*, a.nom_machine AS appareil_nom FROM diag_reseau_evenements e "
        "LEFT JOIN appareils a ON a.id = e.appareil_id "
        "WHERE e.client_id=? AND e.resolu=0 "
        "ORDER BY CASE e.gravite WHEN 'critique' THEN 0 WHEN 'avertissement' THEN 1 ELSE 2 END, "
        "e.derniere_occurrence DESC LIMIT 100", (cid,)).fetchall()
    conn.close()
    evenements = []
    for r in rows:
        d = row_to_dict(r)
        d['categorie_libelle'] = network_diag.libelle_categorie(d.get('categorie', ''))
        evenements.append(d)
    return render_template('mobile/diag_reseau.html', evenements=evenements,
                           etat_moniteur=network_diag.etat_moniteur(),
                           client=client, clients=get_clients(), client_actif_id=cid)


if __name__ == '__main__':
    init_db()

    # Démarrer le scheduler pour les cron jobs
    # Cron job: régénérer les occurrences maintenances tous les jours à 2h du matin
    scheduler.add_job(_regenerate_all_maintenance_occurrences, 'cron', hour=2, minute=0)
    # Cron job: notifier maintenances à venir tous les jours à 8h du matin
    scheduler.add_job(_notify_upcoming_maintenances, 'cron', hour=8, minute=0)
    # Cron job: rafraîchir la base OUI si elle a plus de 30 jours (no-op sinon) —
    # nécessaire en plus du téléchargement au démarrage pour une instance
    # qui tourne en continu (Docker) sans jamais redémarrer.
    scheduler.add_job(lambda: _oui_telecharger(force=False), 'cron', hour=3, minute=30)
    scheduler.start()
    logger.info("Cron scheduler démarré (régénération à 02:00, notifications à 08:00, base OUI à 03:30)")

    # Charger la base OUI en arrière-plan pour ne pas bloquer le démarrage —
    # et la télécharger automatiquement si elle est absente ou dépassée
    # (30 jours). Le scan reste utilisable entre-temps avec la table
    # embarquée (~930 préfixes) ou l'ancienne base déjà chargée.
    def _oui_demarrage():
        _oui_load_full()
        _oui_telecharger(force=False)
    threading.Thread(target=_oui_demarrage, daemon=True).start()
    # Lancer la synchronisation des uploads (local ↔ Turso)
    start_sync_thread(interval=60)
    # Sauvegarde périodique de la base, avec rotation sur les 3 dernières
    demarrer_sauvegardes()
    print("="*50)
    print("  ParcInfo Multi-Clients")
    print(f"  OS      : {platform.system()}")
    print(f"  DB      : {DB_PATH}")
    print(f"  Uploads : {UPLOAD_FOLDER}")
    print("  URL     : http://localhost:3456")
    print("="*50)

    # Register mDNS service (parcinfo.local)
    _register_mdns()
    print("  [mDNS] : http://parcinfo.local:3456")

    if not os.environ.get('RUNNING_IN_DOCKER'):
        import webbrowser
        def _open_browser():
            import time; time.sleep(1.5)
            webbrowser.open('http://localhost:3456')
        threading.Thread(target=_open_browser, daemon=True).start()

    debug = os.environ.get('FLASK_DEBUG', '0') == '1'

    # Configuration pour Docker (Synology, etc.)
    if os.environ.get('RUNNING_IN_DOCKER'):
        # En Docker, utiliser Werkzeug directement (plus stable que Gunicorn)
        print("🚀 Lancement avec Werkzeug (multi-threaded)")
        print(f"   Host: 0.0.0.0:3456")
        print(f"   Threaded: True")
        print(f"   Debug: {debug}")
        print(f"   Reloader: False")
        try:
            app.run(debug=debug, host='0.0.0.0', port=3456,
                   use_reloader=False, threaded=True)
        except Exception as e:
            print(f"❌ Erreur Flask: {e}")
            import traceback
            traceback.print_exc()
            raise
    else:
        # Mode développement (ouverture automatique du navigateur)
        print("[*] Lancement en mode developpement")
        app.run(debug=debug, host='0.0.0.0', port=3456, threaded=True)
